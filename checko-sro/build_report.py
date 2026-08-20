#!/usr/bin/env python3
"""Собирает финальный отчёт по членам СРО из результата первого прогона.

Исходный файл не изменяется — результат всегда пишется в новый.

    python3 build_report.py --source final_report_20260820.xlsx

Если рядом есть кэш ответов Чеко (cache/checko/*.json), точные поля
берутся оттуда, а эвристики применяются только там, где данных нет.
"""
import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib.address import detect_shared_addresses, pick_address, split_addresses
from lib.checko import extract
from lib.names import director_from_ip_name, shorten_name

COLUMNS = [
    ("Название компании", 46),
    ("ИНН", 15),
    ("Телефон", 22),
    ("Email", 30),
    ("Адрес", 62),
    ("ФИО руководителя", 30),
    ("Дата вступления в СРО", 15),
    ("Статус членства", 16),
]

# Колонки исходного файла первого прогона.
SRC_NAME, SRC_INN, SRC_PHONE, SRC_EMAIL, SRC_ADDR, SRC_HEAD, SRC_JOINED = 0, 1, 2, 3, 4, 5, 6
SRC_STATUS = 8

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
STRIPE_FILL = PatternFill("solid", fgColor="F2F5FA")
EXCLUDED_FILL = PatternFill("solid", fgColor="FCE9E9")
THIN = Side(style="thin", color="D0D7E5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def read_source(path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.worksheets[0]
    return [
        [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, SRC_INN + 1).value
    ]


def load_cache(cache_dir):
    """ИНН -> разобранный ответ Чеко, если кэш существует."""
    cache = {}
    directory = Path(cache_dir)
    if not directory.exists():
        return cache
    for path in directory.glob("*.json"):
        if path.name.startswith("_"):
            continue
        import json

        fields = extract(json.loads(path.read_text(encoding="utf-8")))
        if fields:
            cache[path.stem] = fields
    return cache


def join_values(existing, extra):
    """Склеивает телефоны/почты из файла и из кэша, сохраняя порядок."""
    seen, result = set(), []
    for value in list(existing) + list(extra):
        value = str(value).strip()
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def build_rows(source_rows, cache):
    shared = detect_shared_addresses(source_rows, address_index=SRC_ADDR)
    rows, needs_review = [], []

    for source in source_rows:
        inn = str(source[SRC_INN]).strip()
        name = source[SRC_NAME]
        checko = cache.get(inn, {})

        # Адрес: точный ЮрАдрес из Чеко приоритетнее выбора из слипшейся ячейки.
        if checko.get("address"):
            address, ambiguous = checko["address"], False
        else:
            address, ambiguous = pick_address(source[SRC_ADDR], inn, shared)

        # Руководитель: у ИП это он сам, ФИО уже есть в наименовании.
        director = source[SRC_HEAD] or checko.get("director") or director_from_ip_name(name)

        phones = join_values(
            [p.strip() for p in str(source[SRC_PHONE] or "").split(";") if p.strip()],
            checko.get("phones", []),
        )
        emails = join_values(
            [e.strip() for e in str(source[SRC_EMAIL] or "").split(";") if e.strip()],
            checko.get("emails", []),
        )

        rows.append(
            {
                "name": shorten_name(name),
                "inn": inn,
                "phone": "\n".join(phones),
                "email": "\n".join(emails),
                "address": address or "",
                "director": director or "",
                "joined": source[SRC_JOINED],
                "status": source[SRC_STATUS],
            }
        )

        missing = []
        if not address:
            missing.append("адрес")
        elif ambiguous:
            missing.append("адрес выбран из нескольких — проверить")
        if not director:
            missing.append("руководитель")
        if not phones:
            missing.append("телефон")
        if not emails:
            missing.append("email")
        if missing:
            needs_review.append((inn, shorten_name(name), ", ".join(missing)))

    rows.sort(key=lambda r: r["name"].lower())
    return rows, needs_review


def style_header(sheet, columns):
    for index, (title, width) in enumerate(columns, start=1):
        cell = sheet.cell(1, index, title)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def write_report(rows, needs_review, output_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Члены СРО"
    style_header(sheet, COLUMNS)

    for offset, row in enumerate(rows):
        excel_row = offset + 2
        values = [
            row["name"], row["inn"], row["phone"], row["email"],
            row["address"], row["director"], row["joined"], row["status"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(excel_row, column, value)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=column in (1, 3, 4, 5, 6),
                horizontal="center" if column in (2, 7, 8) else "left",
            )
            # Исключённых подсвечиваем целиком: это первое, что ищут в отчёте.
            if row["status"] == "исключён":
                cell.fill = EXCLUDED_FILL
            elif offset % 2:
                cell.fill = STRIPE_FILL
        sheet.cell(excel_row, 2).number_format = "@"
        sheet.cell(excel_row, 7).number_format = "DD.MM.YYYY"

    if needs_review:
        review = workbook.create_sheet("Требуют уточнения")
        style_header(review, [("ИНН", 15), ("Название компании", 46), ("Чего не хватает", 46)])
        for offset, item in enumerate(needs_review):
            for column, value in enumerate(item, start=1):
                cell = review.cell(offset + 2, column, value)
                cell.font = Font(name=FONT, size=10)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=column != 1)
            review.cell(offset + 2, 1).number_format = "@"

    workbook.save(output_path)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", required=True, help="xlsx первого прогона")
    parser.add_argument("--cache", default="cache/checko", help="кэш ответов Чеко")
    parser.add_argument("--out", default="Отчёт_СРО.xlsx", help="куда записать результат")
    args = parser.parse_args()

    if Path(args.out).resolve() == Path(args.source).resolve():
        parser.error("исходный файл перезаписывать нельзя — укажите другой --out")

    source_rows = read_source(args.source)
    cache = load_cache(args.cache)
    rows, needs_review = build_rows(source_rows, cache)
    write_report(rows, needs_review, args.out)

    filled = sum(1 for r in rows if r["address"])
    print(f"Строк в отчёте: {len(rows)}")
    print(f"Данные из кэша Чеко: {len(cache)} ИНН")
    print(f"С адресом: {filled}   без адреса: {len(rows) - filled}")
    print(f"На лист «Требуют уточнения»: {len(needs_review)}")
    print(f"Записано: {args.out}")


if __name__ == "__main__":
    main()
