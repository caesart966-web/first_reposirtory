import yaml
from openpyxl import load_workbook

import run
from core.export import build_export
from core.models import EXCLUDED_FROM_SRO, JOINED_SRO, SUSPENDED, TENDER_NO_SRO_DESIGN, Org, Signal
from core.scoring import rescore_all

TODAY = "2026-03-01"


def seed(db):
    db.add_signals([
        Signal("0105012345", EXCLUDED_FROM_SRO, "2026-02-27", "nostroy", "http://r/1", {"name": "ООО Ноль", "sro_name": "СРО А"}),
        Signal("1000000002", SUSPENDED, "2026-02-01", "nostroy", "http://r/2", {"name": "ООО Пауза"}),
        Signal("1000000003", TENDER_NO_SRO_DESIGN, "2026-02-10", "tenderguru", None, {"name": "ООО Проект", "sum": 1e6}),
        Signal("1000000004", EXCLUDED_FROM_SRO, "2026-02-01", "nostroy"),   # ликвидирован
        Signal("1000000005", EXCLUDED_FROM_SRO, "2026-02-01", "nostroy"),   # уже обзвонен
        Signal("1000000006", JOINED_SRO, "2026-02-01", "nostroy"),          # не лид
        Signal("1000000007", SUSPENDED, "2026-02-20", "nostroy", None, {"name": "ООО Спор", "event_date": "2026-02-20"}),
        Signal("1000000007", JOINED_SRO, "2026-02-25", "nostroy", None, {"event_date": None}),  # конфликт дат
    ])
    db.upsert_org(Org(inn="1000000004", name="ООО Труп", status="LIQUIDATED"))
    db.upsert_org(Org(inn="0105012345", region="г Санкт-Петербург", phone="+7 (812) 000-00-00",
                      site="https://nol.ru/", site_verified="unverified", phone_unverified="+7 (812) 999-99-99"))
    rescore_all(db, {"scoring": yaml.safe_load(open("config.yaml", encoding="utf-8"))["scoring"]}, TODAY)
    db.set_outreach("1000000005", "called", "не берут трубку")
    db.commit()


def test_export_sheets_and_formats(cfg, db):
    seed(db)
    path = build_export(db, cfg, TODAY)
    assert path.name == "Лиды_2026-03-01.xlsx"
    wb = load_workbook(path)
    assert wb.sheetnames == ["Горячие", "Все лиды", "История сигналов"]
    ws = wb["Все лиды"]
    header = [c.value for c in ws[1]]
    assert header[:3] == ["Приоритет", "Скор", "ИНН"] and header[-2:] == ["Статус обзвона", "Комментарий"]
    inns = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
    assert inns == ["0105012345", "1000000003", "1000000002", "1000000007"]  # по скору вниз (130, 90, 70, 70 — при равенстве по ИНН)
    conflict_col = header.index("Конфликт дат") + 1
    assert ws.cell(row=5, column=conflict_col).value == "да" and ws.cell(row=2, column=conflict_col).value is None
    c = ws.cell(row=2, column=3)
    assert c.number_format == "@" and isinstance(c.value, str)   # ИНН текстом, ведущий ноль на месте
    assert ws.cell(row=2, column=1).fill.fgColor.rgb.endswith("FFC7CE")  # приоритет 1 — красный
    assert ws.cell(row=3, column=1).fill.fgColor.rgb.endswith("FFEB9C")  # приоритет 2 — жёлтый
    assert ws.cell(row=2, column=4).value == "ООО Ноль"            # имя подтянулось из сигнала
    assert ws.cell(row=2, column=7).value == "Исключён из СРО"
    col = {h: i + 1 for i, h in enumerate(header)}
    assert ws.cell(row=2, column=col["Сайт проверен"]).value == "unverified"
    assert ws.cell(row=2, column=col["Телефон"]).value == "+7 (812) 000-00-00"
    assert ws.cell(row=2, column=col["Телефон (не подтверждён)"]).value == "+7 (812) 999-99-99"
    assert ws.cell(row=2, column=col["Телефон (не подтверждён)"]).number_format == "@"
    hot = wb["Горячие"]
    assert [hot.cell(row=r, column=3).value for r in range(2, hot.max_row + 1)] == ["0105012345"]
    hist = wb["История сигналов"]
    assert hist.max_row == 6 and "sro_name=СРО А" in hist.cell(row=2, column=9).value


def test_export_region_filter(cfg, db):
    seed(db)
    cfg["export"]["regions"] = ["Санкт-Петербург"]
    wb = load_workbook(build_export(db, cfg, TODAY))
    ws = wb["Все лиды"]
    assert [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)] == ["0105012345"]


def test_run_main_export_only_and_mark(tmp_path):
    cfg = yaml.safe_load(open("config.yaml", encoding="utf-8"))
    for k in cfg["paths"]:
        cfg["paths"][k] = str(tmp_path / cfg["paths"][k])
    cfg_path = tmp_path / "config.yaml"
    cfg_path.write_text(yaml.safe_dump(cfg, allow_unicode=True), encoding="utf-8")
    assert run.main(["--mark", "1000000001", "called", "перезвонить", "--config", str(cfg_path)]) == 0
    assert run.main(["--export-only", "--config", str(cfg_path)]) == 0
    assert list((tmp_path / "output").glob("Лиды_*.xlsx"))
    assert list((tmp_path / "logs").glob("*.log"))
    # неизвестный коллектор не роняет прогон
    assert run.main(["--only", "nope", "--no-enrich", "--config", str(cfg_path)]) == 0


def test_run_check_api_writes_nothing(tmp_path, monkeypatch):
    import collectors.nostroy_registry as nr
    cfg = yaml.safe_load(open("config.yaml", encoding="utf-8"))
    for k in cfg["paths"]:
        cfg["paths"][k] = str(tmp_path / cfg["paths"][k])
    cfg_path = tmp_path / "config.yaml"
    cfg_path.write_text(yaml.safe_dump(cfg, allow_unicode=True), encoding="utf-8")
    monkeypatch.setattr(nr.NostroyRegistry, "check_api", lambda self, **kw: ("ok", "ОТЧЁТ"))
    assert run.main(["--check-api", "nostroy", "--config", str(cfg_path)]) == 0
    monkeypatch.setattr(nr.NostroyRegistry, "check_api", lambda self, **kw: ("warn", "ВНИМАНИЕ"))
    assert run.main(["--check-api", "nostroy", "--config", str(cfg_path)]) == 0   # предупреждение — не отказ
    monkeypatch.setattr(nr.NostroyRegistry, "check_api", lambda self, **kw: ("error", "ОШИБКА"))
    assert run.main(["--check-api", "nostroy", "--config", str(cfg_path)]) == 1
    assert not (tmp_path / "data" / "sro_leads.db").exists()          # БД не создаётся и не пишется
    assert run.main(["--check-api", "nope", "--config", str(cfg_path)]) == 2


def test_export_filters_max_rows_min_score_age(cfg, db):
    seed(db)
    # 6 горячих: по одному свежему исключению; свежие сигналы -> 130 баллов, приоритет 1
    for i in range(10, 16):
        db.add_signals([Signal(f"20000000{i}", EXCLUDED_FROM_SRO, "2026-02-28", "nostroy")])
    db.add_signals([Signal("3000000001", EXCLUDED_FROM_SRO, "2025-06-01", "nostroy")])   # свежайший сигнал старше 180 дней
    db.commit()
    cfg["export"]["max_rows"] = 3
    wb = load_workbook(build_export(db, cfg, TODAY))
    hot, allv = wb["Горячие"], wb["Все лиды"]
    assert hot.max_row - 1 == 3                                          # потолок только на «Горячие»
    all_inns = [allv.cell(row=r, column=3).value for r in range(2, allv.max_row + 1)]
    assert len(all_inns) == 4 + 6                                        # «Все лиды» без потолка (4 из сида + 6)
    assert "3000000001" not in all_inns                                  # устаревший отсеян
    assert "1000000006" not in all_inns                                  # только joined_sro — не лид
    cfg["export"]["signal_max_age_days"] = None
    cfg["export"]["min_score"] = 100
    wb = load_workbook(build_export(db, cfg, TODAY))
    allv = wb["Все лиды"]
    all_inns = [allv.cell(row=r, column=3).value for r in range(2, allv.max_row + 1)]
    assert "3000000001" not in all_inns and "1000000002" not in all_inns  # 50 и 70 баллов — ниже отсечки
    assert "0105012345" in all_inns and len(all_inns) == 1 + 6
