"""Экспорт в Excel: output/Лиды_YYYY-MM-DD.xlsx (Горячие / Все лиды / История сигналов)."""
from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .db import Database
from .enrich import LIQUIDATED_STATUSES
from .models import SIGNAL_TITLES
from .scoring import score_org
from .utils import resolve_path, today_str

log = logging.getLogger("sro_leads")

PRIORITY_FILL = {
    1: PatternFill("solid", fgColor="FFC7CE"),  # красный
    2: PatternFill("solid", fgColor="FFEB9C"),  # жёлтый
    3: PatternFill("solid", fgColor="E7E6E6"),  # серый
}
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")

LEAD_COLUMNS = [
    "Приоритет", "Скор", "ИНН", "Название", "Регион", "ОКВЭД", "Сигналы", "Дата последнего сигнала",
    "Сайт", "Телефон", "Почта", "Руководитель", "Ссылка на источник", "Конфликт дат", "Статус обзвона", "Комментарий",
]
SIGNAL_COLUMNS = ["ИНН", "Название", "Тип сигнала", "Описание", "Дата", "Источник", "Ссылка", "Детали"]
TEXT_COLUMNS = {"ИНН", "Телефон"}
WIDTHS = {"Приоритет": 10, "Скор": 8, "ИНН": 14, "Название": 40, "Регион": 24, "ОКВЭД": 10, "Сигналы": 34,
          "Дата последнего сигнала": 14, "Сайт": 28, "Телефон": 24, "Почта": 30, "Руководитель": 28,
          "Ссылка на источник": 40, "Конфликт дат": 12, "Статус обзвона": 14, "Комментарий": 30, "Тип сигнала": 26,
          "Описание": 34, "Дата": 12, "Источник": 12, "Ссылка": 40, "Детали": 60}


def _write_sheet(ws: Worksheet, columns: list[str], rows: list[list[Any]], fills: Optional[list[int]] = None) -> None:
    ws.append(columns)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
    text_idx = [i for i, name in enumerate(columns, 1) if name in TEXT_COLUMNS]
    for r_i, row in enumerate(rows, 2):
        ws.append(row)
        for ci in text_idx:  # ИНН строго текстом: иначе Excel съест ведущие нули
            cell = ws.cell(row=r_i, column=ci)
            cell.number_format = "@"
            if cell.value is not None:
                cell.value = str(cell.value)
        if fills:
            fill = PRIORITY_FILL.get(fills[r_i - 2])
            if fill:
                for ci in range(1, len(columns) + 1):
                    ws.cell(row=r_i, column=ci).fill = fill
    for i, name in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(name, 16)
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"


def _short_details(raw_json: Optional[str]) -> str:
    if not raw_json:
        return ""
    try:
        raw = json.loads(raw_json)
    except ValueError:
        return raw_json[:200]
    parts = []
    for k in ("sro_name", "status", "customer", "sum", "okpd", "subject", "number", "file"):
        v = raw.get(k)
        if v not in (None, "", []):
            parts.append(f"{k}={v}")
    return "; ".join(parts)[:500]


def build_export(db: Database, cfg: dict[str, Any], date: Optional[str] = None) -> Path:
    xcfg = cfg.get("export", {})
    scfg = cfg.get("scoring", {})
    date = date or today_str()
    regions = [r.lower() for r in xcfg.get("regions", []) or []]
    allowed_status = set(xcfg.get("outreach_statuses", ["new"]) or [])
    min_score = float(xcfg.get("min_score", 1))
    exclude_liq = bool(xcfg.get("exclude_liquidated", True))

    outreach = db.outreach_map()
    signals_by_inn = db.signals_by_inn()
    leads: list[dict[str, Any]] = []
    skipped = {"score": 0, "liquidated": 0, "outreach": 0, "region": 0}

    for org in db.orgs_rows():
        inn = org["inn"]
        res = score_org(inn, signals_by_inn.get(inn, []), scfg, date)
        if res.score < min_score:
            skipped["score"] += 1
            continue
        if exclude_liq and (org["status"] or "") in LIQUIDATED_STATUSES:
            skipped["liquidated"] += 1
            continue
        o = outreach.get(inn)
        o_status = o["status"] if o else "new"
        if allowed_status and o_status not in allowed_status:
            skipped["outreach"] += 1
            continue
        region = org["region"] or ""
        if regions and not any(r in region.lower() for r in regions):
            skipped["region"] += 1
            continue
        sigs = [s for s in signals_by_inn.get(inn, []) if s["signal_type"] in res.types]
        url = next((s["url"] for s in sigs if s["url"]), None)
        name = org["name"] or next((json.loads(s["raw_json"]).get("name") for s in sigs
                                    if s["raw_json"] and json.loads(s["raw_json"]).get("name")), None)
        leads.append({
            "priority": res.priority,
            "score": res.score,
            "inn": inn,
            "name": name,
            "region": org["region"],
            "okved": org["okved"],
            "signals": ", ".join(SIGNAL_TITLES.get(t, t) for t in res.types),
            "last": res.last_signal_date,
            "site": org["site"],
            "phone": org["phone"],
            "email": org["email"],
            "director": org["director"],
            "url": url,
            "date_conflict": "да" if res.date_conflict else None,
            "status": o_status,
            "note": o["note"] if o else None,
        })

    leads.sort(key=lambda x: (-x["score"], x["inn"]))

    def to_row(l: dict[str, Any]) -> list[Any]:
        return [l["priority"], l["score"], l["inn"], l["name"], l["region"], l["okved"], l["signals"], l["last"],
                l["site"], l["phone"], l["email"], l["director"], l["url"], l["date_conflict"], l["status"], l["note"]]

    wb = Workbook()
    ws_hot = wb.active
    ws_hot.title = "Горячие"
    hot = [l for l in leads if l["priority"] == 1 and l["status"] == "new"]
    _write_sheet(ws_hot, LEAD_COLUMNS, [to_row(l) for l in hot], [l["priority"] for l in hot])

    ws_all = wb.create_sheet("Все лиды")
    _write_sheet(ws_all, LEAD_COLUMNS, [to_row(l) for l in leads], [l["priority"] for l in leads])

    ws_sig = wb.create_sheet("История сигналов")
    exported = {l["inn"]: l["name"] for l in leads}
    sig_rows = []
    for inn, name in exported.items():
        for s in signals_by_inn.get(inn, []):
            sig_rows.append([inn, name, s["signal_type"], SIGNAL_TITLES.get(s["signal_type"], s["signal_type"]),
                             s["signal_date"], s["source"], s["url"], _short_details(s["raw_json"])])
    _write_sheet(ws_sig, SIGNAL_COLUMNS, sig_rows)

    out_dir = resolve_path(cfg, "output_dir", "output")
    path = out_dir / xcfg.get("filename", "Лиды_{date}.xlsx").format(date=date)
    wb.save(path)
    log.info("Экспорт: %s — горячих %d, всего лидов %d, сигналов %d; отсеяно %s",
             path, len(hot), len(leads), len(sig_rows), skipped)
    return path
