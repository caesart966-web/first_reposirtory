# -*- coding: utf-8 -*-
"""
stage2_merge.py — вернуть найденные почты в исходную таблицу.

Берёт исходный xlsx и result/progress.csv (его пишет stage1_sites.py) и
собирает итоговый файл: там, где почта была пустой, а парсер её нашёл —
подставляет. Плюс три служебные колонки: источник, статус, комментарий.

Ничего не выдумывает: подставляется только то, что реально снято с сайта.
Ячейки, для которых почта не нашлась, помечаются текстом, а не заполняются
наугад.

Запуск:
    python stage2_merge.py
    python stage2_merge.py --input "Компании (4).xlsx" --progress result/progress.csv
"""

import argparse
import csv
import os
from collections import Counter

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from stage1_sites import find_input

NOT_FOUND_TEXT = "нет в открытых источниках"
NOT_CHECKED_TEXT = "не проверялось"


def main():
    ap = argparse.ArgumentParser(description="Склейка найденных почт с исходной таблицей")
    here = os.path.dirname(os.path.abspath(__file__))
    ap.add_argument("--input", default="")
    ap.add_argument("--progress", default=os.path.join(here, "result", "progress.csv"))
    ap.add_argument("--output", default="")
    args = ap.parse_args()

    inp = find_input(args.input)
    out = args.output or os.path.join(os.path.dirname(inp),
                                      os.path.splitext(os.path.basename(inp))[0]
                                      + " — с почтами.xlsx")

    found = {}
    if os.path.exists(args.progress):
        with open(args.progress, "r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                inn = (row.get("inn") or "").strip()
                if inn:
                    found[inn] = row
        print(f"Из progress.csv прочитано записей: {len(found)}")
    else:
        print(f"!! {args.progress} не найден — почты подставлять неоткуда.")
        print("   Сначала прогоните stage1_sites.py")

    wb = load_workbook(inp)
    ws = wb[wb.sheetnames[0]]
    header = [str(ws.cell(1, c).value or "").strip().lower()
              for c in range(1, ws.max_column + 1)]

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n) + 1
        raise SystemExit(f"!! в файле нет колонки {names}")

    c_inn, c_mail = col("инн"), col("email", "e-mail", "почта")
    c_src, c_st, c_note = ws.max_column + 1, ws.max_column + 2, ws.max_column + 3
    for c, title in ((c_src, "Источник email"), (c_st, "Статус"), (c_note, "Комментарий")):
        ws.cell(1, c, title).font = Font(bold=True)

    green = PatternFill("solid", fgColor="D9EAD3")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    stats = Counter()

    def blank(v):
        return v is None or str(v).strip() == ""

    for r in range(2, ws.max_row + 1):
        inn = str(ws.cell(r, c_inn).value or "").strip()
        if not blank(ws.cell(r, c_mail).value):
            ws.cell(r, c_st, "была в исходном файле")
            stats["исходные"] += 1
            continue
        rec = found.get(inn)
        if rec and (rec.get("emails") or "").strip():
            ws.cell(r, c_mail, rec["emails"].strip())
            ws.cell(r, c_src, rec.get("source", ""))
            ws.cell(r, c_st, "найдено парсером")
            ws.cell(r, c_note, rec.get("site", ""))
            for c in (c_mail, c_src, c_st):
                ws.cell(r, c).fill = green
            stats["найдено"] += 1
        elif rec:
            ws.cell(r, c_mail, NOT_FOUND_TEXT)
            ws.cell(r, c_st, f"проверено — {rec.get('status', 'не найдено')}")
            ws.cell(r, c_note, rec.get("note", ""))
            ws.cell(r, c_mail).fill = yellow
            stats["проверено-пусто"] += 1
        else:
            ws.cell(r, c_mail, NOT_CHECKED_TEXT)
            ws.cell(r, c_st, "в очереди на проверку")
            stats["не проверено"] += 1

    ws.column_dimensions["K"].width = 34
    wb.save(out)

    print("\n" + "=" * 56)
    print(f"Готово: {out}")
    for k in ("исходные", "найдено", "проверено-пусто", "не проверено"):
        if stats[k]:
            print(f"  {k:18} {stats[k]}")
    print(f"  {'пустых ячеек':18} 0")

    try:
        os.startfile(os.path.dirname(os.path.abspath(out)))
    except AttributeError:
        pass


if __name__ == "__main__":
    main()
