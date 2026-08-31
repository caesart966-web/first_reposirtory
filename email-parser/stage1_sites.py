# -*- coding: utf-8 -*-
"""
stage1_sites.py — сбор email с сайтов компаний.

Берёт из xlsx строки, где email ПУСТОЙ и сайт ЗАПОЛНЕН, обходит главную и
типовые страницы контактов, достаёт почты из mailto: и из текста, включая
обфусцированные написания. Ничего не выдумывает: в результат попадает только
то, что реально снято со страницы.

Запуск:
    python stage1_sites.py                       # весь подходящий вход
    python stage1_sites.py --limit 30            # первые 30 строк (тест)
    python stage1_sites.py --input "path.xlsx" --outdir result

Перезапуск безопасен: обработанные компании берутся из progress.csv и
пропускаются.
"""

import argparse
import csv
import os
import re
import sys
import threading
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from urllib.parse import urlparse, urljoin

import requests
import urllib3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ─────────────────────────── настройки ───────────────────────────

THREADS = 15
TIMEOUT = 10
RETRIES = 2                      # 2 ретрая = до 3 попыток на URL
PATHS = ["", "/contacts", "/contact", "/kontakty", "/about", "/o-kompanii"]

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
HEADERS = {
    "User-Agent": UA,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate",
    "Connection": "close",
}

# мусор, который никогда не является рабочей почтой компании
BAD_SUBSTRINGS = [
    "example.com", "noreply", "no-reply", "sentry", "wixpress", "cloudflare",
    "@2x", "domain.com", "yourdomain", "site.com", "mail.example",
    "your@", "email@email", "test@test", "@sentry.io", "@wix.com",
]
BAD_SUFFIXES = (".png", ".jpg", ".jpeg", ".svg", ".webp", ".gif",
                ".bmp", ".ico", ".css", ".js")
PRIORITY_LOCALS = ["info", "mail", "office"]

# ────────────────────── извлечение адресов ───────────────────────

# обычный email
RE_PLAIN = re.compile(
    r"[A-Za-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[A-Za-z0-9!#$%&'*+/=?^_`{|}~-]+)*"
    r"@[A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?"
    r"(?:\.[A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?)+"
)

# обфускация: info (at) domain.ru / info[собака]domain.ru / info @ domain.ru
_AT = (r"(?:\s*[\(\[\{<]\s*(?:at|собака|собак|dog|ат|а)\s*[\)\]\}>]\s*"
       r"|\s+(?:at|собака|ат)\s+"
       r"|\s*@\s*)")
_DOT = (r"(?:\s*[\(\[\{<]\s*(?:dot|точка|тчк)\s*[\)\]\}>]\s*"
        r"|\s+(?:dot|точка)\s+"
        r"|\s*\.\s*)")
RE_OBF_FULL = re.compile(
    r"[A-Za-z0-9][A-Za-z0-9._%+-]{0,63}"
    + _AT +
    r"(?:[A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?)"
    + r"(?:" + _DOT + r"(?:[A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?))+",
    re.IGNORECASE,
)
RE_MAILTO = re.compile(r"""mailto:\s*([^"'>\s?]+)""", re.IGNORECASE)
RE_TAGS = re.compile(r"<(script|style)[^>]*>.*?</\1>", re.IGNORECASE | re.DOTALL)
RE_ANY_TAG = re.compile(r"<[^>]+>")
RE_META_CHARSET = re.compile(
    rb"""<meta[^>]+charset\s*=\s*["']?\s*([A-Za-z0-9_\-]+)""", re.IGNORECASE)


def normalize_obfuscated(raw: str) -> str:
    """'info (at) domain (dot) ru' -> 'info@domain.ru'"""
    s = raw.strip()
    s = re.sub(_AT, "@", s, count=1, flags=re.IGNORECASE)
    s = re.sub(_DOT, ".", s, flags=re.IGNORECASE)
    return re.sub(r"\s+", "", s)


def is_garbage(email: str) -> bool:
    e = email.lower().strip()
    if not e or e.count("@") != 1:
        return True
    if e.endswith(BAD_SUFFIXES):
        return True
    if any(bad in e for bad in BAD_SUBSTRINGS):
        return True
    local, _, domain = e.partition("@")
    if not local or not domain or "." not in domain:
        return True
    tld = domain.rsplit(".", 1)[-1]
    if len(tld) < 2 or not tld.isalpha():
        return True
    # 'logo@2x.png' и подобные артефакты вёрстки
    if re.fullmatch(r"\d+x", local):
        return True
    return False


def extract_emails(html: str):
    """-> {email: источник}. Источник: 'mailto' | 'текст' | 'текст (обфускация)'"""
    found = {}
    for m in RE_MAILTO.findall(html):
        e = m.split("?")[0].strip().strip(".,;:").lower()
        if e and not is_garbage(e):
            found.setdefault(e, "mailto")

    text = RE_ANY_TAG.sub(" ", RE_TAGS.sub(" ", html))
    text = (text.replace("&#64;", "@").replace("&commat;", "@")
                .replace("&#46;", ".").replace("&nbsp;", " "))

    for m in RE_PLAIN.findall(text):
        e = m.strip().strip(".,;:").lower()
        if not is_garbage(e):
            found.setdefault(e, "текст")

    for raw in RE_OBF_FULL.findall(text):
        e = normalize_obfuscated(raw).lower().strip(".,;:")
        if not is_garbage(e):
            found.setdefault(e, "текст (обфускация)")
    return found


def strip_www(host):
    """Именно префикс. lstrip('www.') съел бы 'w' у домена вроде wood.ru."""
    h = (host or "").lower().strip()
    return h[4:] if h.startswith("www.") else h


def rank_emails(emails, site_host):
    """Домен сайта вперёд, затем info@/mail@/office@, затем остальное."""
    host = strip_www(site_host)

    def key(e):
        local, _, dom = e.partition("@")
        dom = strip_www(dom)
        same = 0 if (host and (dom == host or dom.endswith("." + host)
                               or host.endswith("." + dom))) else 1
        try:
            pri = PRIORITY_LOCALS.index(local.lower())
        except ValueError:
            pri = len(PRIORITY_LOCALS)
        return (same, pri, e)

    return sorted(emails, key=key)

# ─────────────────────── сеть и кодировки ────────────────────────


def decode_response(resp) -> str:
    """Русские сайты часто в windows-1251. На дефолт requests не полагаемся."""
    raw = resp.content
    enc = None
    # 1) charset из самого html — на русских сайтах он вернее заголовка,
    #    сервер нередко отдаёт utf-8 в Content-Type, а страницу в cp1251
    m = RE_META_CHARSET.search(raw[:4096]) or RE_META_CHARSET.search(raw)
    if m:
        try:
            enc = m.group(1).decode("ascii", "ignore")
        except Exception:
            enc = None
    # 2) charset из заголовка
    if not enc:
        ctype = resp.headers.get("Content-Type", "")
        m = re.search(r"charset\s*=\s*([A-Za-z0-9_\-]+)", ctype, re.IGNORECASE)
        if m:
            enc = m.group(1)
    # 3) определение по содержимому
    if not enc:
        enc = resp.apparent_encoding
    if not enc or enc.lower() in ("iso-8859-1", "ascii"):
        enc = resp.apparent_encoding or "utf-8"
    try:
        return raw.decode(enc, errors="replace")
    except (LookupError, TypeError):
        return raw.decode("utf-8", errors="replace")


def fetch(session, url, log):
    """-> (html, None) или (None, причина). RETRIES ретраев, ошибки не роняют прогон."""
    last = "неизвестная ошибка"
    for attempt in range(RETRIES + 1):
        try:
            r = session.get(url, headers=HEADERS, timeout=TIMEOUT,
                            allow_redirects=True, verify=False)
            if r.status_code >= 400:
                last = f"HTTP {r.status_code}"
            else:
                ctype = r.headers.get("Content-Type", "").lower()
                if "html" not in ctype and "text" not in ctype and ctype:
                    last = f"не HTML ({ctype.split(';')[0]})"
                else:
                    return decode_response(r), None
        except requests.exceptions.SSLError as e:
            last = f"SSL: {str(e)[:70]}"
        except requests.exceptions.ConnectTimeout:
            last = "таймаут соединения"
        except requests.exceptions.ReadTimeout:
            last = "таймаут чтения"
        except requests.exceptions.TooManyRedirects:
            last = "цикл редиректов"
        except requests.exceptions.ConnectionError as e:
            last = f"нет соединения: {str(e)[:70]}"
        except Exception as e:
            last = f"{type(e).__name__}: {str(e)[:70]}"
        if attempt < RETRIES:
            time.sleep(0.7 * (attempt + 1))
    log(f"    {url} -> {last}")
    return None, last


def split_sites(cell):
    if not cell:
        return []
    parts = re.split(r"[,;\s]+", str(cell).strip())
    return [p.strip().strip(".,;") for p in parts if p.strip().strip(".,;")]


def normalize_site(raw):
    s = str(raw).strip().strip("/")
    if not s:
        return None, None
    if not s.startswith(("http://", "https://")):
        s = "https://" + s
    host = urlparse(s).netloc
    return s, host

# ───────────────────────── чекпоинты ─────────────────────────────

PROGRESS_FIELDS = ["inn", "name", "site", "emails", "source", "status", "note", "ts"]
_lock = threading.Lock()


def load_progress(path):
    done = {}
    if not os.path.exists(path):
        return done
    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                if row.get("inn"):
                    done[str(row["inn"]).strip()] = row
    except Exception:
        print("!! progress.csv повреждён, начинаю с нуля")
    return done


def append_progress(path, row):
    with _lock:
        new = not os.path.exists(path) or os.path.getsize(path) == 0
        with open(path, "a", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=PROGRESS_FIELDS)
            if new:
                w.writeheader()
            w.writerow(row)

# ──────────────────────── обработка компании ─────────────────────


def process(company, log):
    inn, name, sites_raw = company["inn"], company["name"], company["site"]
    sites = split_sites(sites_raw)
    if not sites:
        return dict(inn=inn, name=name, site="", emails="", source="",
                    status="нет сайта", note="колонка «Сайт» пустая",
                    ts=datetime.now().isoformat(timespec="seconds"))

    all_found, opened_any, errors = {}, False, []
    for site in sites:
        base, host = normalize_site(site)
        if not base:
            continue
        session = requests.Session()
        try:
            for path in PATHS:
                url = urljoin(base + "/", path.lstrip("/")) if path else base
                html, err = fetch(session, url, log)
                if html is None:
                    # https не встал — пробуем http этой же страницы
                    if url.startswith("https://"):
                        html, err2 = fetch(session, "http://" + url[8:], log)
                        if html is None:
                            errors.append(f"{url}: {err}")
                            continue
                    else:
                        errors.append(f"{url}: {err}")
                        continue
                opened_any = True
                page = path or "/"
                for mail, src in extract_emails(html).items():
                    if mail not in all_found:
                        all_found[mail] = f"{src} — {host}{page}"
        finally:
            session.close()

    ts = datetime.now().isoformat(timespec="seconds")
    if all_found:
        _, host0 = normalize_site(sites[0])
        ordered = rank_emails(list(all_found), host0)
        return dict(inn=inn, name=name, site=", ".join(sites),
                    emails=";".join(ordered),
                    source=" | ".join(all_found[e] for e in ordered),
                    status="найдено", note="", ts=ts)
    if opened_any:
        return dict(inn=inn, name=name, site=", ".join(sites), emails="",
                    source="", status="почт нет",
                    note="страницы открылись, адресов на них не найдено", ts=ts)
    return dict(inn=inn, name=name, site=", ".join(sites), emails="", source="",
                status="сайт не открылся", note="; ".join(errors[:4]), ts=ts)

# ────────────────────────── чтение входа ─────────────────────────


def load_extra_sites(path):
    """ИНН -> сайт из stage0_discover.py. Берём только подтверждённые по ИНН."""
    m = {}
    if not path:
        return m
    if not os.path.exists(path):
        print(f"!! --extra-sites: файл не найден: {path}")
        return m
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            inn = (row.get("inn") or "").strip()
            site = (row.get("site") or "").strip()
            status = (row.get("status") or "").strip().lower()
            # берём только подтверждённые по ИНН; «не подтверждён» отбрасываем
            accepted = (not status) or status.startswith("найден и подтвержд")
            if inn and site and accepted:
                m[inn] = site
    return m


def read_input(path, extra=None):
    extra = extra or {}
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip().lower() if h else "" for h in next(rows)]

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        raise SystemExit(f"!! в файле нет колонки {names}; есть: {header}")

    c_name, c_inn, c_mail, c_site = (col("название", "наименование"), col("инн"),
                                     col("email", "e-mail", "почта"),
                                     col("сайт", "site", "вебсайт"))

    def blank(v):
        return v is None or str(v).strip() == ""

    scope, skipped = [], []
    for i, r in enumerate(rows, start=2):
        rec = dict(row=i,
                   name="" if blank(r[c_name]) else str(r[c_name]).strip(),
                   inn="" if blank(r[c_inn]) else str(r[c_inn]).strip(),
                   site="" if blank(r[c_site]) else str(r[c_site]).strip())
        if not blank(r[c_mail]):
            continue                                   # почта уже есть
        if blank(r[c_site]):
            found = extra.get(rec["inn"])
            if found:
                rec["site"] = found
                rec["from_stage0"] = True
                scope.append(rec)
                continue
            rec["reason"] = ("почта пустая и сайта нет — парсить нечего "
                             "(запустите stage0_discover.py, чтобы найти сайт)")
            skipped.append(rec)
            continue
        scope.append(rec)
    wb.close()
    return scope, skipped

# ─────────────────────────── выгрузка ────────────────────────────


def write_xlsx(out, results, skipped):
    wb = Workbook()
    ws = wb.active
    ws.title = "Результат"
    head = ["Название", "ИНН", "Сайт", "Найденные почты", "Источник", "Статус", "Комментарий"]
    ws.append(head)
    fills = {"найдено": PatternFill("solid", fgColor="D9EAD3"),
             "почт нет": PatternFill("solid", fgColor="FFF2CC"),
             "сайт не открылся": PatternFill("solid", fgColor="F4CCCC"),
             "нет сайта": PatternFill("solid", fgColor="EFEFEF")}
    for r in results:
        ws.append([r["name"], r["inn"], r["site"], r["emails"],
                   r["source"], r["status"], r["note"]])
        f = fills.get(r["status"])
        if f:
            ws.cell(ws.max_row, 6).fill = f

    ws2 = wb.create_sheet("Необработанные")
    ws2.append(["Название", "ИНН", "Сайт", "Причина"])
    for r in skipped:
        ws2.append([r["name"], r["inn"], r.get("site", ""), r["reason"]])

    for sheet, widths in ((ws, [40, 14, 30, 44, 52, 20, 46]), (ws2, [40, 14, 24, 56])):
        for c in sheet[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(vertical="center")
        sheet.freeze_panes = "A2"
        for i, w in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = w
    wb.save(out)

# ──────────────────────────── main ───────────────────────────────


def main():
    ap = argparse.ArgumentParser(description="Сбор email с сайтов компаний")
    ap.add_argument("--input", default=os.path.join(os.path.dirname(__file__) or ".",
                                                    "..", "Компании (4).xlsx"))
    ap.add_argument("--outdir", default=os.path.join(os.path.dirname(__file__) or ".", "result"))
    ap.add_argument("--limit", type=int, default=0, help="обработать только первые N (0 = все)")
    ap.add_argument("--threads", type=int, default=THREADS)
    ap.add_argument("--extra-sites", default="",
                    help="CSV из stage0_discover.py: подтверждённые по ИНН сайты")
    args = ap.parse_args()

    inp = os.path.abspath(args.input)
    outdir = os.path.abspath(args.outdir)
    os.makedirs(outdir, exist_ok=True)
    progress_csv = os.path.join(outdir, "progress.csv")
    log_path = os.path.join(outdir, "errors.log")
    out_xlsx = os.path.join(outdir, "stage1_emails.xlsx")

    if not os.path.exists(inp):
        raise SystemExit(f"!! не найден входной файл: {inp}")

    logfile = open(log_path, "a", encoding="utf-8")

    def log(msg):
        with _lock:
            logfile.write(f"{datetime.now():%H:%M:%S} {msg}\n")
            logfile.flush()

    print(f"Вход : {inp}")
    extra = load_extra_sites(args.extra_sites)
    if extra:
        print(f"Из stage0 подхвачено подтверждённых по ИНН сайтов: {len(extra)}")
    scope, skipped = read_input(inp, extra)
    from_s0 = sum(1 for c in scope if c.get("from_stage0"))
    print(f"Подходит под условие «email пустой И сайт заполнен»: {len(scope) - from_s0}"
          + (f" (+{from_s0} из stage0)" if from_s0 else ""))
    print(f"Пропущено (почта пустая, сайта нет): {len(skipped)}")

    if args.limit:
        scope = scope[:args.limit]
        print(f"Ограничение --limit {args.limit}: беру {len(scope)}")

    done = load_progress(progress_csv)
    todo = [c for c in scope if c["inn"] not in done]
    print(f"Уже в progress.csv: {len(done)} | к обработке сейчас: {len(todo)}")
    if done:
        print("  (перезапуск: обработанные пропускаю)")

    results = [done[c["inn"]] for c in scope if c["inn"] in done]
    t0 = time.time()
    if todo:
        with ThreadPoolExecutor(max_workers=args.threads) as pool:
            futures = {pool.submit(process, c, log): c for c in todo}
            for n, fut in enumerate(as_completed(futures), start=1):
                c = futures[fut]
                try:
                    row = fut.result()
                except Exception:
                    log(f"!! ИНН {c['inn']}\n{traceback.format_exc()}")
                    row = dict(inn=c["inn"], name=c["name"], site=c["site"],
                               emails="", source="", status="сайт не открылся",
                               note="исключение при обработке, см. errors.log",
                               ts=datetime.now().isoformat(timespec="seconds"))
                append_progress(progress_csv, row)      # чекпоинт после каждой
                results.append(row)
                mark = "OK " if row["status"] == "найдено" else "   "
                print(f"[{n}/{len(todo)}] {mark}{row['name'][:34]:36} {row['status']}"
                      + (f" -> {row['emails']}" if row["emails"] else ""))

    write_xlsx(out_xlsx, results, skipped)
    logfile.close()

    found = sum(1 for r in results if r["status"] == "найдено")
    print("\n" + "=" * 60)
    print(f"Обработано : {len(results)}   за {time.time()-t0:.1f} c")
    print(f"Найдено    : {found}")
    print(f"Почт нет   : {sum(1 for r in results if r['status'] == 'почт нет')}")
    print(f"Не открылся: {sum(1 for r in results if r['status'] == 'сайт не открылся')}")
    print(f"Результат  : {out_xlsx}")
    print(f"Чекпоинт   : {progress_csv}")
    print(f"Лог ошибок : {log_path}")

    try:
        os.startfile(outdir)                            # Windows
    except AttributeError:
        pass                                            # не Windows — молча пропускаем


if __name__ == "__main__":
    main()
