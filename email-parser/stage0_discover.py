# -*- coding: utf-8 -*-
"""
stage0_discover.py — поиск официального сайта компании по ИНН.

Нужен потому, что у 1 194 строк из 1 199 колонка «Сайт» пустая: stage1 нечего
парсить, пока не появится URL.

Принцип, исключающий «левые» компании: кандидат принимается ТОЛЬКО если на его
страницах найден тот самый ИНН из таблицы. Совпадение названия не засчитывается —
одноимённых юрлиц в РФ десятки. Не подтверждено ИНН — сайт не берём.

Результат discovered_sites.csv скармливается stage1:
    python stage1_sites.py --extra-sites result/discovered_sites.csv

Запуск:
    python stage0_discover.py --limit 30
    python stage0_discover.py                      # все строки без почты и сайта
"""

import argparse
import csv
import os
import re
import sys
import threading
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from urllib.parse import urlparse, quote_plus, unquote, parse_qs, urljoin

import requests
import urllib3
from openpyxl import load_workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

from stage1_sites import (HEADERS, TIMEOUT, RETRIES, decode_response, fetch,
                          normalize_site, strip_www, _lock)

SEARCH_THREADS = 2          # поисковик не любит параллелизм — держим низким
VERIFY_THREADS = 12
SEARCH_PAUSE = 4.0          # сек между запросами; меняется ключом --pause
SEARCH_RETRY_WAIT = 25      # пауза после отказа поисковика (rate limit)
CHECK_PATHS = ["", "/contacts", "/kontakty", "/about", "/o-kompanii",
               "/rekvizity", "/requisites"]

# агрегаторы и справочники: там ИНН найдётся всегда, но это не сайт компании
AGGREGATORS = {
    "rusprofile.ru", "checko.ru", "list-org.com", "zachestnyibiznes.ru",
    "spark-interfax.ru", "sbis.ru", "saby.ru", "focus.kontur.ru", "kontur.ru",
    "audit-it.ru", "companies.rbc.ru", "rbc.ru", "tbank.ru", "casebook.ru",
    "fbc.ru", "1cont.ru", "xfirm.ru", "klerk.ru", "synapsenet.ru", "upfox.ru",
    "injust.pro", "reabiz.ru", "excheck.pro", "companium.ru", "vbankcenter.ru",
    "b2b.house", "star-pro.ru", "moedelo.org", "buhonline.ru", "zakgo.ru",
    "kontragent.vbr.ru", "check.tochka.com", "globas.credinform.ru",
    "nalog.gov.ru", "nalog.ru", "egrul.nalog.ru", "zakupki.gov.ru",
    "2gis.ru", "yandex.ru", "google.com", "wikipedia.org", "youtube.com",
    "vk.com", "facebook.com", "instagram.com", "avito.ru", "hh.ru",
    "reestr-sro.ru", "all-sro.ru", "orgpage.ru", "zoon.ru", "flamp.ru",
    "spravkaru.info", "cataloxy.ru", "tek-all.ru", "bbnt.ru", "catalogfactory.org",
    "duckduckgo.com", "bing.com", "mail.ru", "rambler.ru", "livelib.ru",
}


def host_of(url):
    try:
        return strip_www(urlparse(url).netloc)
    except Exception:
        return ""


def is_aggregator(url):
    h = host_of(url)
    if not h:
        return True
    return any(h == a or h.endswith("." + a) for a in AGGREGATORS)


# ─────────────────────────── поисковик ───────────────────────────

RE_DDG_LINK = re.compile(r'<a[^>]+class="result__a"[^>]+href="([^"]+)"', re.I)
RE_DDG_ANY = re.compile(r'href="(/l/\?uddg=[^"]+)"', re.I)
_search_lock = threading.Lock()
BLOCKED = [0]
_last_search = [0.0]


def _throttle():
    with _search_lock:
        wait = SEARCH_PAUSE - (time.time() - _last_search[0])
        if wait > 0:
            time.sleep(wait)
        _last_search[0] = time.time()


def ddg_search(query, log, max_results=10):
    """DuckDuckGo HTML — без ключа. Возвращает список URL."""
    url = "https://html.duckduckgo.com/html/?q=" + quote_plus(query)
    html = None
    for attempt in range(2):
        _throttle()
        session = requests.Session()
        try:
            html, err = fetch(session, url, log)
        finally:
            session.close()
        if html is not None:
            break
        # почти всегда это rate limit — ждём заметно дольше и пробуем ещё раз
        BLOCKED[0] += 1
        log(f"  поиск отбит [{query}]: {err}; жду {SEARCH_RETRY_WAIT} c")
        time.sleep(SEARCH_RETRY_WAIT)
    if html is None:
        log(f"  поиск не удался [{query}]")
        return []
    if "anomaly" in html[:4000].lower() or "captcha" in html[:4000].lower():
        BLOCKED[0] += 1
        log(f"  капча вместо выдачи [{query}]")
        return []

    urls = []
    for raw in RE_DDG_LINK.findall(html) + RE_DDG_ANY.findall(html):
        u = raw
        if u.startswith("/l/?") or "uddg=" in u:
            q = parse_qs(urlparse(u if u.startswith("http") else "https://x" + u).query)
            if "uddg" in q:
                u = unquote(q["uddg"][0])
        if u.startswith("//"):
            u = "https:" + u
        if u.startswith("http") and u not in urls:
            urls.append(u)
    return urls[:max_results]


def serpapi_search(query, log, max_results=10):
    """Опционально: если задан SERPAPI_KEY — берём выдачу через SerpAPI."""
    key = os.environ.get("SERPAPI_KEY")
    if not key:
        return None
    try:
        r = requests.get("https://serpapi.com/search.json",
                         params={"q": query, "api_key": key, "num": max_results,
                                 "hl": "ru", "gl": "ru"},
                         headers=HEADERS, timeout=TIMEOUT + 10)
        data = r.json()
        return [x["link"] for x in data.get("organic_results", []) if x.get("link")]
    except Exception as e:
        log(f"  serpapi упал: {e}")
        return None


def search(query, log):
    return serpapi_search(query, log) or ddg_search(query, log)


# ──────────────────── проверка кандидата по ИНН ──────────────────

def inn_on_site(base_url, inn, log):
    """Ищем ИНН на страницах кандидата. -> (найдено?, на какой странице)"""
    variants = {inn, " ".join([inn[i:i + 4] for i in range(0, len(inn), 4)])}
    session = requests.Session()
    try:
        for path in CHECK_PATHS:
            url = urljoin(base_url + "/", path.lstrip("/")) if path else base_url
            html, _ = fetch(session, url, log)
            if html is None:
                continue
            flat = re.sub(r"[\s ]+", "", html)
            if any(re.sub(r"\s+", "", v) in flat for v in variants):
                return True, (path or "/")
    finally:
        session.close()
    return False, None


def discover(company, log):
    inn, name = company["inn"], company["name"]
    ts = datetime.now().isoformat(timespec="seconds")
    base = dict(inn=inn, name=name, site="", verified_by="", candidates="",
                status="", note="", ts=ts)

    queries = [f'"{inn}"', f'{name} {inn}', f'{name} Санкт-Петербург официальный сайт']
    seen, candidates = set(), []
    for q in queries:
        for u in search(q, log):
            if is_aggregator(u):
                continue
            h = host_of(u)
            if h and h not in seen:
                seen.add(h)
                candidates.append("https://" + h)
        if len(candidates) >= 6:
            break

    base["candidates"] = ", ".join(host_of(c) for c in candidates[:6])
    if not candidates:
        base.update(status="сайт не найден",
                    note="в выдаче только агрегаторы либо пусто")
        return base

    for cand in candidates[:6]:
        ok, page = inn_on_site(cand, inn, log)
        if ok:
            base.update(site=host_of(cand), verified_by=f"ИНН найден на {page}",
                        status="найден и подтверждён")
            return base

    base.update(status="не подтверждён",
                note="кандидаты есть, но ИНН на них не найден — не берём")
    return base


# ─────────────────────────── ввод/вывод ──────────────────────────

FIELDS = ["inn", "name", "site", "verified_by", "candidates", "status", "note", "ts"]


def find_input(explicit):
    """Ищем xlsx: явный путь -> рядом со скриптом -> на уровень выше -> любой xlsx рядом."""
    if explicit:
        p = os.path.abspath(explicit)
        if os.path.exists(p):
            return p
        raise SystemExit(f"!! не найден файл: {p}")
    here = os.path.dirname(os.path.abspath(__file__))
    names = ["Компании (4).xlsx", "Компании(4).xlsx", "Компании.xlsx"]
    for d in (here, os.path.dirname(here)):
        for n in names:
            p = os.path.join(d, n)
            if os.path.exists(p):
                return p
    for d in (here, os.path.dirname(here)):
        try:
            cands = [f for f in os.listdir(d)
                     if f.lower().endswith(".xlsx") and not f.startswith("~$")
                     and "stage" not in f.lower() and "результат" not in f.lower()]
        except OSError:
            continue
        if len(cands) == 1:
            return os.path.join(d, cands[0])
    raise SystemExit(
        "!! Не нашёл входной xlsx.\n"
        "   Положите файл рядом со скриптом или укажите путь:\n"
        '   python stage1_sites.py --input "C:\\parser\\Компании (4).xlsx"')


def read_input(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip().lower() if h else "" for h in next(rows)]

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        raise SystemExit(f"!! нет колонки {names}")

    ci, cn, cm, cs = (col("инн"), col("название", "наименование"),
                      col("email", "e-mail", "почта"), col("сайт", "site"))

    def blank(v):
        return v is None or str(v).strip() == ""

    out = []
    for r in rows:
        if blank(r[cm]) and blank(r[cs]):
            out.append(dict(inn=str(r[ci]).strip(),
                            name="" if blank(r[cn]) else str(r[cn]).strip()))
    wb.close()
    return out


def load_done(path):
    done = {}
    if os.path.exists(path):
        try:
            with open(path, encoding="utf-8-sig", newline="") as f:
                for row in csv.DictReader(f):
                    if row.get("inn"):
                        done[row["inn"].strip()] = row
        except Exception:
            print("!! discovered_sites.csv повреждён, начинаю заново")
    return done


def append_row(path, row):
    with _lock:
        new = not os.path.exists(path) or os.path.getsize(path) == 0
        with open(path, "a", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=FIELDS)
            if new:
                w.writeheader()
            w.writerow(row)


def main():
    global SEARCH_PAUSE
    ap = argparse.ArgumentParser(description="Поиск сайта компании по ИНН с проверкой по ИНН")
    here = os.path.dirname(os.path.abspath(__file__))
    ap.add_argument("--input", default="", help="путь к xlsx (по умолчанию ищется рядом)")
    ap.add_argument("--outdir", default=os.path.join(here, "result"))
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--threads", type=int, default=SEARCH_THREADS)
    ap.add_argument("--pause", type=float, default=SEARCH_PAUSE,
                    help="секунд между запросами к поисковику (больше = меньше блокировок)")
    args = ap.parse_args()
    SEARCH_PAUSE = args.pause

    inp, outdir = find_input(args.input), os.path.abspath(args.outdir)
    os.makedirs(outdir, exist_ok=True)
    csv_path = os.path.join(outdir, "discovered_sites.csv")
    log_path = os.path.join(outdir, "stage0_errors.log")
    logfile = open(log_path, "a", encoding="utf-8")

    def log(msg):
        with _lock:
            logfile.write(f"{datetime.now():%H:%M:%S} {msg}\n")
            logfile.flush()

    scope = read_input(inp)
    print(f"Вход: {inp}")
    print(f"Строк без почты И без сайта: {len(scope)}")
    if args.limit:
        scope = scope[:args.limit]
        print(f"--limit {args.limit}: беру {len(scope)}")

    done = load_done(csv_path)
    todo = [c for c in scope if c["inn"] not in done]
    print(f"Уже обработано: {len(done)} | сейчас: {len(todo)}")
    if os.environ.get("SERPAPI_KEY"):
        print("Поиск: SerpAPI (ключ найден)")
    else:
        print(f"Поиск: DuckDuckGo HTML (без ключа), пауза {SEARCH_PAUSE} c, "
              f"потоков {args.threads}.")
        print("       Много отбитых запросов в итоге -> увеличьте --pause.")

    ok = 0
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=args.threads) as pool:
        futures = {pool.submit(discover, c, log): c for c in todo}
        for n, fut in enumerate(as_completed(futures), start=1):
            c = futures[fut]
            try:
                row = fut.result()
            except Exception:
                log(f"!! ИНН {c['inn']}\n{traceback.format_exc()}")
                row = dict(inn=c["inn"], name=c["name"], site="", verified_by="",
                           candidates="", status="ошибка",
                           note="исключение, см. stage0_errors.log",
                           ts=datetime.now().isoformat(timespec="seconds"))
            append_row(csv_path, row)
            if row["status"] == "найден и подтверждён":
                ok += 1
            mark = "OK " if row["status"] == "найден и подтверждён" else "   "
            print(f"[{n}/{len(todo)}] {mark}{row['name'][:32]:34} {row['status']}"
                  + (f" -> {row['site']}" if row["site"] else ""))

    logfile.close()
    print("\n" + "=" * 60)
    print(f"Обработано: {len(todo)} за {time.time()-t0:.1f} c")
    print(f"Сайт найден и подтверждён по ИНН: {ok}")
    print(f"Поисковик отбивал запросы: {BLOCKED[0]} раз"
          + ("   <-- много: результат недостоверен, увеличьте --pause" if BLOCKED[0] > 5 else ""))
    print(f"Файл: {csv_path}")
    print("Дальше:  python stage1_sites.py --extra-sites result/discovered_sites.csv")


if __name__ == "__main__":
    main()
