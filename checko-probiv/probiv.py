#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Пробив контактов членов СРО через API Checko.

Программа берёт реестр, отбирает нужные компании, запрашивает по каждой ИНН
контакты в Checko и складывает результат в Excel. Суточный лимит запросов
учитывается между запусками: за один день программа не потратит больше, чем
разрешено, а на следующий день продолжит с того места, где остановилась,
и никого не пробьёт дважды.

Запуск:  python probiv.py            обычный прогон (одна «сотня»)
         python probiv.py --demo     показать формат отчёта, не тратя запросы
         python probiv.py --limit 20 потратить не больше 20 запросов
         python probiv.py --status   что сделано, что осталось, ничего не запрашивая
"""

import sys
sys.dont_write_bytecode = True

import argparse
import configparser
import datetime
import json
import os
import re
import shutil
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print('Не установлена библиотека openpyxl. Выполните в командной строке:')
    print('    pip install openpyxl')
    sys.exit(1)

try:
    from phones import norm as norm_phone
except ImportError:
    norm_phone = lambda s: (s or '', 'raw')

ВЕРСИЯ = '6 от 18.08.2026'
BASE = Path(__file__).resolve().parent
INPUT_DIR = BASE / 'input'
STATE_DIR = BASE / 'состояние'
CONFIG = BASE / 'config.ini'

DEFAULT_CONFIG = """\
[checko]
; Ключ API из личного кабинета: https://checko.ru/user/account/api
key =
; Сколько запросов можно потратить за сутки (тариф Checko).
daily_limit = 100
; Адреса методов API. Меняйте, только если Checko изменит их у себя.
base_url = https://api.checko.ru/v2
company_path = /company
entrepreneur_path = /entrepreneur
; Пауза между запросами в секундах — чтобы не долбить сервер.
pause = 0.7

[выборка]
; Кого пробивать:
;   active         — только действующие члены СРО (дата прекращения пустая)
;   active_no_phone— действующие, у которых в реестре нет своего телефона
;   no_phone       — все, у кого в реестре нет своего телефона
;   all            — все записи реестра
scope = active
; Номер считается «не своим», если он встречается в реестре у стольких компаний и более.
общий_номер_от = 2
"""

# --- утилиты ---------------------------------------------------------------

def log(msg=''):
    print(msg, flush=True)


def today():
    return datetime.date.today().isoformat()


def read_config():
    if not CONFIG.exists():
        CONFIG.write_text(DEFAULT_CONFIG, encoding='utf-8')
        log('Создан файл настроек config.ini.')
    cp = configparser.ConfigParser()
    cp.read(CONFIG, encoding='utf-8')
    key = cp.get('checko', 'key', fallback='').strip()
    if not key:
        log('')
        log('В config.ini не указан ключ API Checko.')
        log('Возьмите его здесь: https://checko.ru/user/account/api')
        try:
            key = input('Вставьте ключ и нажмите Enter: ').strip()
        except EOFError:
            key = ''
        if not key:
            log('Ключ не введён — запускать нечего.')
            sys.exit(1)
        cp.set('checko', 'key', key)
        with CONFIG.open('w', encoding='utf-8') as fh:
            cp.write(fh)
        log('Ключ сохранён в config.ini, больше спрашивать не буду.')
    return cp


# --- чтение реестра --------------------------------------------------------

def _date(v):
    if v in (None, ''):
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v).strip()
    for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _clean(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        v = int(v)                                   # ИНН-число: 7801234567.0 -> 7801234567
    return re.sub(r'\s+', ' ', str(v).replace('\xa0', ' ')).strip()


def read_compiled(path):
    """Сводная таблица, собранная ранее (лист «Реестр»)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if 'Реестр' not in wb.sheetnames:
        wb.close()
        return None
    ws = wb['Реестр']
    if _clean(ws.cell(row=3, column=1).value) != '№':
        wb.close()
        return None
    out = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not isinstance(r[0], (int, float)) or not _clean(r[1]):
            continue
        out.append({'n': int(r[0]), 'full': _clean(r[1]), 'short': _clean(r[2]),
                    'inn': _clean(r[3]), 'phone': _clean(r[4]),
                    'date_in': _date(r[5]), 'date_out': _date(r[6])})
    wb.close()
    return out


def read_export(path):
    """Исходная выгрузка с сайта СРО (register-*.xlsx)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=1, values_only=True)
    header = None
    for i, r in enumerate(rows):
        if i == 1:
            header = r
            break
    if not header or _clean(header[1]) != 'N п/п':
        wb.close()
        return None
    out = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not isinstance(r[1], (int, float)) or not _clean(r[2]):
            continue
        out.append({'n': int(r[1]), 'full': _clean(r[2]), 'short': _clean(r[3]),
                    'inn': _clean(r[10]), 'phone': norm_phone(_clean(r[11]))[0],
                    'date_in': _date(r[5]), 'date_out': _date(r[7])})
    wb.close()
    return out


HEADERS = {
    'inn':      (('инн',), ('инн',)),
    'n':        (('n п/п', '№ п/п', '№'), ('п/п', 'номер')),
    'full':     (('полное наименование',), ('наименование', 'организац', 'компан', 'член')),
    'short':    (('сокращенное наименование', 'сокращённое наименование'), ('сокращ',)),
    'phone':    (('контактные телефоны', 'телефон'), ('телефон', 'контакт')),
    'date_in':  (('дата регистрации в реестре сро', 'дата вступления'),
                 ('дата регистрации', 'дата вступ', 'дата прием', 'дата приём',
                  'дата включ', 'вступ', 'включ', 'приём в член', 'прием в член')),
    'date_out': (('дата прекращения членства', 'дата прекращения'),
                 ('дата прекращ', 'дата исключ', 'прекращ', 'исключ', 'выбыт')),
}


def _match_columns(header):
    """Подбирает номера колонок по названиям. Сначала точное совпадение, потом частичное;
    из нескольких подходящих берём самую левую — «ИНН страховой компании» так не перебьёт «ИНН»."""
    cells = [_clean(c).lower() for c in header]
    found = {}
    for field, (exact, loose) in HEADERS.items():
        for want in exact:
            hit = [i for i, c in enumerate(cells) if c == want]
            if hit:
                found[field] = hit[0]
                break
        if field in found:
            continue
        for want in loose:
            hit = [i for i, c in enumerate(cells) if want in c]
            if hit:
                found[field] = hit[0]
                break
    return found


def _inn_like(v):
    """Похоже ли значение ячейки на ИНН: только цифры, 10 или 12 знаков."""
    t = _clean(v).replace(' ', '')
    if not t or re.search(r'\D', t):
        return False
    return len(t) in (10, 12)


def _guess_by_data(grid, inn_col, start):
    """Когда шапки нет — угадываем колонки по содержимому строк с данными."""
    cols, date_cols = {}, []
    body = [r for r in grid[start:] if len(r) > inn_col and _inn_like(r[inn_col])]
    if not body:
        return cols
    width = max(len(r) for r in body)

    def column(j):
        return [r[j] for r in body if j < len(r)]

    best_len = 0
    for j in range(width):
        if j == inn_col:
            continue
        vals = [_clean(v) for v in column(j)]
        texts = [v for v in vals if re.search(r'[А-Яа-яA-Za-z]', v)]
        phones = [v for v in vals if len(re.sub(r'\D', '', v)) >= 6 and re.search(r'[+()\-]|^8', v)]
        dates = [v for v in column(j) if _date(v)]
        if len(texts) > len(vals) * 0.5:                       # колонка с наименованием — самая «длинная»
            avg = sum(len(t) for t in texts) / len(texts)
            if avg > best_len:
                best_len, cols['full'] = avg, j
        if len(phones) > len(vals) * 0.3 and 'phone' not in cols:
            cols['phone'] = j
        if dates:
            date_cols.append((j, len(dates), len(vals)))

    # дата вступления заполнена почти везде, дата прекращения — редкая, она следующая
    dense = [j for j, d, n in date_cols if d > n * 0.5]
    if dense:
        cols.setdefault('date_in', dense[0])
    rest = [j for j, d, n in date_cols if j != cols.get('date_in')]
    if rest:
        cols.setdefault('date_out', rest[0])
    return cols


def read_any(path):
    """Любая таблица, где есть ИНН: выгрузка НОСТРОЙ, список из Excel и т.п.
    Сначала ищем шапку по названиям колонок, а если её нет — находим ИНН по самим данным."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    best = None
    for ws in wb.worksheets:
        grid = list(ws.iter_rows(min_row=1, max_row=1000, values_only=True))
        if not grid:
            continue

        for i, r in enumerate(grid):                            # 1. шапка с названиями
            cols = _match_columns(r)
            if 'inn' in cols:
                best = (ws, i + 1, cols, 'по названиям колонок')
                break
        if best:
            break

        hits, first_row = {}, {}                                # 2. поиск ИНН по содержимому
        for i, r in enumerate(grid):
            for j, v in enumerate(r):
                if _inn_like(v):
                    hits[j] = hits.get(j, 0) + 1
                    first_row.setdefault(j, i)
        if not hits:
            continue
        inn_col = max(hits, key=lambda k: hits[k])
        start = first_row[inn_col]

        cols, how = {}, 'по данным'
        for i in range(start - 1, max(-1, start - 11), -1):     # шапка чуть выше данных?
            named = _match_columns(grid[i])
            if len([c for c in grid[i] if _clean(c)]) >= 3 and named:
                cols, how = named, 'по данным и шапке выше'
                break
        cols.update({k: v for k, v in _guess_by_data(grid, inn_col, start).items() if k not in cols})
        cols['inn'] = inn_col
        best = (ws, start, cols, how)
        break

    if not best:
        wb.close()
        return None, None

    ws, hdr_row, cols, how = best
    out, seen, num = [], set(), 0
    for r in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if cols['inn'] >= len(r):
            continue
        inn = re.sub(r'\D', '', _clean(r[cols['inn']]))
        if len(inn) not in (10, 12) or inn in seen:
            continue
        seen.add(inn)
        num += 1
        get = lambda f: _clean(r[cols[f]]) if f in cols and cols[f] < len(r) else ''
        n = r[cols['n']] if 'n' in cols and cols['n'] < len(r) else None
        full = get('full') or get('short')
        if not full:
            continue
        out.append({'n': int(n) if isinstance(n, (int, float)) else num,
                    'full': full, 'short': get('short') or full, 'inn': inn,
                    'phone': norm_phone(get('phone'))[0],
                    'date_in': _date(get('date_in')), 'date_out': _date(get('date_out'))})
    wb.close()
    names = {'inn': 'ИНН', 'full': 'Наименование', 'short': 'Сокращённое',
             'phone': 'Телефон', 'date_in': 'Дата вступления', 'date_out': 'Дата прекращения'}
    descr = '%s — %s' % (how, ', '.join('%s=%s' % (names[f], get_column_letter(cols[f] + 1))
                                        for f in ('inn', 'full', 'short', 'phone', 'date_in', 'date_out')
                                        if f in cols))
    return out, descr


def load_register(input_dir):
    """Читает всё, что положили в input: zip-архив, выгрузки, сводную таблицу."""
    input_dir.mkdir(parents=True, exist_ok=True)
    files = [p for p in sorted(input_dir.iterdir()) if p.is_file()]
    if not files:
        log('Папка input пуста. Положите туда реестр: zip-архив с выгрузкой,')
        log('файлы register-*.xlsx или сводную таблицу Реестр_членов_СРО_ПОС.xlsx.')
        sys.exit(1)

    tmp = STATE_DIR / 'распаковка'
    if tmp.exists():
        shutil.rmtree(tmp)
    sheets = []
    for p in files:
        if p.suffix.lower() == '.zip':
            tmp.mkdir(parents=True, exist_ok=True)
            with zipfile.ZipFile(p) as z:
                for name in z.namelist():
                    if name.lower().endswith('.xlsx') and not name.startswith('__MACOSX'):
                        target = tmp / Path(name).name
                        target.write_bytes(z.read(name))
                        sheets.append(target)
        elif p.suffix.lower() == '.xlsx' and not p.name.startswith('~$'):
            sheets.append(p)

    records, seen_source = {}, []
    for p in sorted(sheets):
        got, how = read_compiled(p), 'сводная таблица'
        if not got:
            got, how = read_export(p), 'выгрузка с сайта СРО'
        if not got:
            got, how = read_any(p)           # любая таблица с колонкой ИНН
            how = 'колонки распознаны %s' % how if got else None
        if not got:
            log('  ПРОПУСКАЮ %s — не нашёл в нём колонку с ИНН' % p.name)
            continue
        seen_source.append('%s: %d записей, %s' % (p.name, len(got), how))
        for rec in got:
            key = (rec['inn'], rec['n'])
            records[key] = rec               # одна и та же запись из разных файлов — не дублируем
    if not records:
        log('Ни в одном файле не нашёл реестр. Проверьте, что лежит в input.')
        sys.exit(1)
    for line in seen_source:
        log('  ' + line)
    return sorted(records.values(), key=lambda r: r['n'])


# --- выборка ---------------------------------------------------------------

def phone_counts(records):
    cnt = {}
    for rec in records:
        for num in filter(None, (rec['phone'] or '').split(', ')):
            num = re.sub(r'\s*доб\..*', '', num)
            cnt[num] = cnt.get(num, 0) + 1
    return cnt


def select(records, scope, shared_from):
    cnt = phone_counts(records)

    def own_phone(rec):
        nums = [re.sub(r'\s*доб\..*', '', n) for n in (rec['phone'] or '').split(', ') if n]
        return any(cnt.get(n, 0) < shared_from for n in nums)

    if scope == 'active':
        return [r for r in records if not r['date_out']]
    if scope == 'active_no_phone':
        return [r for r in records if not r['date_out'] and not own_phone(r)]
    if scope == 'no_phone':
        return [r for r in records if not own_phone(r)]
    if scope == 'all':
        return list(records)
    log('Неизвестное значение scope=%s в config.ini' % scope)
    sys.exit(1)


# --- состояние -------------------------------------------------------------

class State:
    def __init__(self, path):
        self.path = path
        self.data = {'requests_by_day': {}, 'processed': {}, 'runs': 0}
        if path.exists():
            try:
                self.data = json.loads(path.read_text(encoding='utf-8'))
            except (ValueError, OSError):
                log('Файл состояния повреждён, начинаю учёт заново.')
        self.data.setdefault('requests_by_day', {})
        self.data.setdefault('processed', {})
        self.data.setdefault('runs', 0)

    def save(self):
        self.path.parent.mkdir(parents=True, exist_ok=True)
        tmp = self.path.with_suffix('.tmp')
        tmp.write_text(json.dumps(self.data, ensure_ascii=False, indent=1), encoding='utf-8')
        tmp.replace(self.path)

    def spent_today(self):
        return self.data['requests_by_day'].get(today(), 0)

    def spend(self, k=1):
        self.data['requests_by_day'][today()] = self.spent_today() + k

    MAX_ATTEMPTS = 3

    def done(self, inn):
        """Обработанной считаем компанию с окончательным ответом. Сбой связи или
        ошибку сервера повторяем в следующие прогоны, но не больше трёх раз."""
        rec = self.data['processed'].get(inn)
        if not rec:
            return False
        if rec.get('статус') == 'ошибка':
            return rec.get('попыток', 1) >= self.MAX_ATTEMPTS
        return True

    def attempts(self, inn):
        return self.data['processed'].get(inn, {}).get('попыток', 0)

    def mark(self, inn, status, run):
        self.data['processed'][inn] = {'статус': status, 'дата': today(), 'прогон': run,
                                       'попыток': self.attempts(inn) + 1}


# --- запрос в Checko -------------------------------------------------------

DEMO_ANSWER = {
    'meta': {'status': 'ok'},
    'data': {
        'ИНН': '7810733295', 'ОГРН': '1187847182137',
        'НаимСокр': 'ООО «Север Град»',
        'Статус': {'Наим': 'Действующая'},
        'ЮрАдрес': {'АдресРФ': '196191, г. Санкт-Петербург, Новоизмайловский пр-кт, д. 20, к. 1'},
        'Контакты': {'Тел': ['+7 (812) 642-03-75', '+79219876543'],
                     'Емэйл': ['info@severgrad.ru'],
                     'ВебСайт': 'severgrad.ru'},
    },
}


class Quota(Exception):
    """Checko сказал, что запросы на сегодня кончились."""


def fetch(cfg, inn, demo=False):
    """Возвращает (данные, потрачен_ли_запрос, текст_ошибки)."""
    if demo:
        answer = json.loads(json.dumps(DEMO_ANSWER))
        answer['data']['ИНН'] = inn
        return answer, False, None

    path = cfg.get('checko', 'entrepreneur_path') if len(inn) == 12 \
        else cfg.get('checko', 'company_path')
    url = cfg.get('checko', 'base_url').rstrip('/') + path + '?' + urllib.parse.urlencode(
        {'key': cfg.get('checko', 'key'), 'inn': inn})
    req = urllib.request.Request(url, headers={'User-Agent': 'sro-probiv/1.0'})

    for attempt in (1, 2):
        try:
            with urllib.request.urlopen(req, timeout=40) as resp:
                body = resp.read().decode('utf-8', 'replace')
            break
        except urllib.error.HTTPError as e:
            body = e.read().decode('utf-8', 'replace')
            if e.code == 429:
                raise Quota('Checko ответил 429 — лимит запросов исчерпан')
            # Ответ от сервера получен, значит запрос, скорее всего, засчитан.
            return None, True, 'HTTP %s: %s' % (e.code, body[:200])
        except (urllib.error.URLError, OSError, TimeoutError) as e:
            if attempt == 2:
                # До сервера не достучались — запрос не засчитан.
                return None, False, 'нет связи: %s' % e
            time.sleep(3)

    try:
        answer = json.loads(body)
    except ValueError:
        return None, True, 'сервер вернул не JSON: %s' % body[:200]

    meta = answer.get('meta') or {}
    message = str(meta.get('message') or '')
    if re.search(r'лимит|limit|quota|исчерпан', message, re.I):
        raise Quota('Checko: %s' % message)
    return answer, True, None


# --- разбор ответа ---------------------------------------------------------

def walk(node, want, out, depth=0):
    """Собирает значения из любых веток JSON, чьи ключи похожи на нужные."""
    if depth > 8:
        return
    if isinstance(node, dict):
        for k, v in node.items():
            kl = str(k).lower()
            if any(w in kl for w in want):
                collect_leaves(v, out)
            walk(v, want, out, depth + 1)
    elif isinstance(node, list):
        for v in node:
            walk(v, want, out, depth + 1)


def collect_leaves(node, out, depth=0):
    if depth > 6:
        return
    if isinstance(node, (str, int, float)):
        s = str(node).strip()
        if s and s.lower() not in ('none', 'null') and s not in out:
            out.append(s)
    elif isinstance(node, dict):
        for v in node.values():
            collect_leaves(v, out, depth + 1)
    elif isinstance(node, list):
        for v in node:
            collect_leaves(v, out, depth + 1)


def dedupe_phones(vals):
    """«+7 (812) 642-03-75» и «88126420375» — один номер; оставляем первое написание."""
    out, seen = [], set()
    for v in vals:
        d = re.sub(r'\D', '', v)
        key = d[1:] if len(d) == 11 and d[0] in '78' else d
        if key in seen:
            continue
        seen.add(key)
        out.append(v)
    return out


def parse(answer):
    data = answer.get('data') if isinstance(answer, dict) else None
    if isinstance(data, list):
        data = data[0] if data else None
    if not isinstance(data, dict) or not data:
        return None

    phones, emails, sites, status, address = [], [], [], [], []
    walk(data, ('тел', 'phone'), phones)
    walk(data, ('емэйл', 'емейл', 'email', 'почт', 'e-mail'), emails)
    walk(data, ('сайт', 'site', 'вебсайт', 'website'), sites)

    st = data.get('Статус')
    if isinstance(st, dict):
        status = [str(st.get('Наим') or st.get('Наименование') or '').strip()]
    elif st:
        status = [str(st).strip()]
    else:
        walk(data, ('статус',), status)

    addr = data.get('ЮрАдрес') or data.get('Адрес')
    if isinstance(addr, dict):
        address = [str(addr.get('АдресРФ') or '').strip()] if addr.get('АдресРФ') else []
        if not address:
            collect_leaves(addr, address)
            address = [', '.join(address)]
    elif addr:
        address = [str(addr).strip()]

    def clean_list(vals, keep):
        out = []
        for v in vals:
            v = v.strip().strip(',;')
            if not v or not keep(v):
                continue
            if v not in out:
                out.append(v)
        return out

    # у телефона считаем цифры во всей строке: «+7 (812) 642-03-75» подряд их не имеет
    phones = clean_list(phones, lambda v: len(re.sub(r'\D', '', v)) >= 6)
    emails = clean_list(emails, lambda v: '@' in v)
    sites = clean_list(sites, lambda v: re.search(r'\.[a-zA-Zа-яА-Я]{2,}', v) is not None)
    phones = dedupe_phones(phones)
    phones = [norm_phone(x)[0] for x in phones]      # +78123202982 -> +7 (812) 320-29-82
    return {
        'phones': ', '.join(phones),
        'emails': ', '.join(emails),
        'site': ', '.join(sites),
        'status': ' '.join(x for x in status if x).strip(),
        'address': ' '.join(x for x in address if x).strip(),
        'наим': str(data.get('НаимСокр') or data.get('НаимПолн') or data.get('ФИО') or '').strip(),
    }


# --- отчёты ----------------------------------------------------------------

FONT = 'Arial'
COLS = [
    ('№', 7, 'center'),
    ('Наименование', 34, 'left'),
    ('ИНН', 14, 'center'),
    ('Телефоны', 30, 'left'),
    ('Почта', 26, 'left'),
    ('Сайт', 22, 'left'),
    ('Статус', 18, 'left'),
    ('Адрес', 44, 'left'),
    ('Дата вступления', 15, 'center'),
    ('Дата прекращения', 16, 'center'),
    ('Дата пробива', 13, 'center'),
]


def write_report(path, title, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Результаты'
    ws['A1'] = title
    ws['A1'].font = Font(name=FONT, size=13, bold=True,
                        color='B00020' if title.startswith('ДЕМО') else '1F3864')
    ws.merge_cells('A1:K1')
    ws.row_dimensions[1].height = 24

    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, (name, width, _) in enumerate(COLS, start=1):
        c = ws.cell(row=2, column=i, value=name)
        c.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F3864')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[2].height = 30

    for r, row in enumerate(rows, start=3):
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT, size=10)
            c.border = border
            c.alignment = Alignment(horizontal=COLS[i - 1][2], vertical='top',
                                    wrap_text=i in (2, 4, 5, 6, 8))
            if r % 2 == 0:
                c.fill = PatternFill('solid', fgColor='F2F5FA')
        ws.cell(row=r, column=3).number_format = '@'
        for col in (9, 10, 11):
            ws.cell(row=r, column=col).number_format = 'DD.MM.YYYY'

    ws.auto_filter.ref = 'A2:K%d' % max(2, len(rows) + 2)
    ws.freeze_panes = 'A3'
    ws.sheet_view.showGridLines = False
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def make_row(rec, res, probe_date):
    phones = (res or {}).get('phones', '')
    if res is None:
        phones = 'не найдено в Checko'
    return [rec['n'], rec['short'] or rec['full'], rec['inn'],
            phones, (res or {}).get('emails', ''), (res or {}).get('site', ''),
            (res or {}).get('status', ''), (res or {}).get('address', ''),
            rec['date_in'], rec['date_out'], probe_date]


def dump_input():
    """Печатает структуру файлов из input — чтобы было видно, где что лежит."""
    files = [p for p in sorted(INPUT_DIR.iterdir()) if p.is_file()] if INPUT_DIR.exists() else []
    if not files:
        log('Папка input пуста.')
        return
    for f in files:
        log('')
        log('=' * 70)
        log('ФАЙЛ: %s   (%d КБ)' % (f.name, f.stat().st_size // 1024))
        if f.suffix.lower() != '.xlsx':
            log('  не .xlsx — пропускаю')
            continue
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception as e:
            log('  не открывается: %s' % e)
            continue
        for ws in wb.worksheets:
            log('')
            log('  ЛИСТ «%s»: строк %s, колонок %s' % (ws.title, ws.max_row, ws.max_column))
            rows_seen = []
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=500, values_only=True), start=1):
                cells = [(get_column_letter(j + 1), _clean(v)) for j, v in enumerate(row) if _clean(v)]
                if cells:
                    rows_seen.append((i, cells))
            # интересны строки таблицы, а не одиночные заголовки-заглушки сверху
            table = [x for x in rows_seen if len(x[1]) >= 2][:12] or rows_seen[:12]
            for i, cells in table:
                log('    строка %d: %s' % (i, ' | '.join('%s=%s' % (c, v[:38]) for c, v in cells[:9])))
            if not rows_seen:
                log('    лист пустой')
        wb.close()
    log('')
    log('=' * 70)
    log('Пришлите этот текст — по нему видно, какая колонка где.')


# --- главный сценарий ------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description='Пробив контактов членов СРО через Checko')
    ap.add_argument('--limit', type=int, help='сколько запросов потратить за этот прогон')
    ap.add_argument('--demo', action='store_true', help='показать формат отчёта без обращения к API')
    ap.add_argument('--status', action='store_true', help='показать, что сделано, и выйти')
    ap.add_argument('--reset-day', action='store_true', help='обнулить счётчик запросов за сегодня')
    ap.add_argument('--dump', action='store_true', help='показать, что лежит в файлах папки input')
    args = ap.parse_args()

    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    log('Версия программы: %s' % ВЕРСИЯ)

    if args.dump:                                   # показать содержимое файлов, ключ не нужен
        dump_input()
        return

    cfg = read_config() if not args.demo else configparser.ConfigParser()
    if args.demo:
        cfg.read_string(DEFAULT_CONFIG)
        cfg.set('checko', 'key', 'demo')

    state_dir = STATE_DIR / ('демо' if args.demo else '')
    state = State(state_dir / 'состояние.json')
    cache_dir = state_dir / 'ответы'

    if args.reset_day:
        state.data['requests_by_day'].pop(today(), None)
        state.save()
        log('Счётчик запросов за %s обнулён.' % today())
        return

    records = load_register(INPUT_DIR)
    scope = cfg.get('выборка', 'scope', fallback='active')
    shared_from = cfg.getint('выборка', 'общий_номер_от', fallback=2)
    targets = select(records, scope, shared_from)

    todo = [r for r in targets if not state.done(r['inn'])]
    daily = cfg.getint('checko', 'daily_limit', fallback=100)
    left_today = max(0, daily - state.spent_today())
    limit = (args.limit or 5) if args.demo else (args.limit or left_today)

    log('')
    log('Всего в реестре: %d   |   в выборке «%s»: %d' % (len(records), scope, len(targets)))
    log('Уже пробито: %d   |   осталось: %d' % (len(targets) - len(todo), len(todo)))
    log('Потрачено сегодня: %d из %d   |   доступно сейчас: %d' % (state.spent_today(), daily, left_today))

    if args.status:
        return
    if not todo:
        log('')
        log('Вся выборка пробита — новых запросов не требуется.')
        return
    if left_today <= 0 and not args.demo:
        log('')
        log('Суточный лимит исчерпан. Запустите программу завтра — продолжит с №%d.' % todo[0]['n'])
        return

    limit = min(limit, len(todo)) if args.demo else min(limit, left_today, len(todo))
    state.data['runs'] += 1
    run = state.data['runs']
    log('')
    log('Прогон №%d: беру %d компаний, начиная с №%d.' % (run, limit, todo[0]['n']))
    log('')

    done_now, found, errors = [], 0, 0
    by_inn = {}
    stopped = None
    for rec in todo[:limit]:
        inn = rec['inn']
        if inn in by_inn:                                  # тот же ИНН в реестре дважды
            res, note = by_inn[inn]
            done_now.append(rec)
            state.mark(inn, 'ok', run)
            continue
        try:
            answer, spent, err = fetch(cfg, inn, demo=args.demo)
        except Quota as e:
            stopped = str(e)
            break
        if spent:
            state.spend()
        if err:
            errors += 1
            res, note = None, 'ошибка: %s' % err
            state.mark(inn, 'ошибка', run)
            if state.attempts(inn) >= State.MAX_ATTEMPTS:
                note += ' (больше не повторяю)'
        else:
            cache_dir.mkdir(parents=True, exist_ok=True)
            (cache_dir / ('%s.json' % inn)).write_text(
                json.dumps(answer, ensure_ascii=False, indent=1), encoding='utf-8')
            res = parse(answer)
            if args.demo:
                note = 'ДЕМО — данные вымышленные'
                found += 1
            elif res and (res['phones'] or res['emails'] or res['site']):
                found += 1
                note = 'найдено'
            elif res:
                note = 'компания найдена, контактов нет'
            else:
                note = 'в Checko не найдено'
            state.mark(inn, 'ok', run)
        by_inn[inn] = (res, note)
        done_now.append(rec)
        state.save()
        log('  №%-5d %-34s %-13s %s' % (rec['n'], (rec['short'] or rec['full'])[:34], inn,
                                        (res or {}).get('phones', '') or note))
        time.sleep(cfg.getfloat('checko', 'pause', fallback=0.7) if not args.demo else 0)

    state.save()

    # Отчёт один на всё: пересобирается из сохранённых ответов после каждого прогона.
    rows_out = []
    for rec in targets:
        if not state.done(rec['inn']):
            continue
        cached = cache_dir / ('%s.json' % rec['inn'])
        if cached.exists():
            try:
                res = parse(json.loads(cached.read_text(encoding='utf-8')))
            except ValueError:
                res = None
        else:
            res = None
        rows_out.append(make_row(rec, res, _date(state.data['processed'][rec['inn']].get('дата'))))

    mark = 'ДЕМО, ДАННЫЕ ВЫМЫШЛЕННЫЕ — образец. ' if args.demo else ''
    report = BASE / ('ДЕМО_Отчет.xlsx' if args.demo else 'Отчет.xlsx')
    title = '%sКонтакты членов СРО из Checko — на %s, пробито %d компаний' % (
        mark, datetime.datetime.now().strftime('%d.%m.%Y'), len(rows_out))
    try:
        write_report(report, title, rows_out)
    except PermissionError:
        report = report.with_name(report.stem + '_новый.xlsx')
        write_report(report, title, rows_out)
        log('')
        log('ВНИМАНИЕ: файл Отчет.xlsx открыт в Excel, поэтому записал рядом %s' % report.name)

    left = len(todo) - len(done_now)
    log('')
    if stopped:
        log('ОСТАНОВЛЕНО: %s' % stopped)
    log('Обработано за прогон: %d   |   с контактами: %d   |   ошибок: %d' % (len(done_now), found, errors))
    log('Потрачено сегодня: %d из %d' % (state.spent_today(), daily))
    log('Осталось пробить: %d %s' % (left, ('(это ещё %d дн.)' % -(-left // max(1, daily))) if left else ''))
    log('')
    log('')
    log('ОТЧЁТ: %s' % report)


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        log('\nПрервано пользователем. Прогресс сохранён, ничего не потеряно.')
