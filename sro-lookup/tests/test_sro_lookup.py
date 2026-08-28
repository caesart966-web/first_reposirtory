"""Тесты поиска СРО. Сеть не нужна."""

from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from sro_lookup.reader import read_companies
from sro_lookup.registry import parse_member_payload
from sro_lookup.report import write_report

PAYLOAD = {
    "data": {"data": [{
        "inn": "7813692784",
        "ogrn": "1237800000000",
        "full_description": "Общество с ограниченной ответственностью «ОЛЛИ ИТ»",
        "member_status": {"id": 1, "title": "Является членом СРО"},
        "registry_registration_date": "2023-04-12T00:00:00+03:00",
        "sro": {
            "id": 410,
            "full_description": "Ассоциация «СРО «Балтийский строительный комплекс»",
            "registration_number": "СРО-С-410-16122014",
            "inn": "7825489730",
        },
    }]}
}


class TestParseMember(unittest.TestCase):
    def test_sro_is_extracted(self) -> None:
        info = parse_member_payload(PAYLOAD, "7813692784")
        self.assertTrue(info.found)
        self.assertIn("Балтийский строительный комплекс", info.sro_name)
        self.assertEqual(info.sro_number, "СРО-С-410-16122014")
        self.assertEqual(info.status_text, "Является членом СРО")
        self.assertEqual(info.date_join, date(2023, 4, 12))

    def test_sro_inn_is_not_company_inn(self) -> None:
        """ИНН СРО и ИНН компании нельзя перепутать — это разные лица."""
        info = parse_member_payload(PAYLOAD, "7813692784")
        self.assertEqual(info.sro_inn, "7825489730")
        self.assertNotEqual(info.sro_inn, info.inn)

    def test_other_company_is_not_matched(self) -> None:
        """Запись чужой компании не должна выдаваться за искомую."""
        info = parse_member_payload(PAYLOAD, "7817158012")
        self.assertFalse(info.found)
        self.assertTrue(info.error)


class TestReader(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = Path(tempfile.mkdtemp(prefix="sro-test-"))

    def test_columns_are_detected_by_content(self) -> None:
        path = self.tmp / "companies.xlsx"
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.append(["Наименование", "ИНН"])            # шапка — пропускается
        sheet.append(['ООО "ОЛЛИ ИТ"', "7813692784"])
        sheet.append(["7817158012", 'ООО "ФАКТСТРОЙ"'])  # колонки наоборот
        sheet.append(['ООО "ДУБЛЬ"', "7813692784"])      # повтор ИНН
        sheet.append(["мусор", "123"])                    # не ИНН
        book.save(path)

        companies = read_companies(path)
        self.assertEqual([inn for _, inn in companies], ["7813692784", "7817158012"])
        self.assertIn("ОЛЛИ", companies[0][0])
        self.assertIn("ФАКТСТРОЙ", companies[1][0])

    def test_invalid_inn_is_rejected(self) -> None:
        """Десять цифр — ещё не ИНН: контрольная сумма должна сойтись."""
        path = self.tmp / "bad.xlsx"
        book = openpyxl.Workbook()
        book.active.append(["ООО «Ромашка»", "1234567890"])
        book.save(path)
        self.assertEqual(read_companies(path), [])


class TestReport(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = Path(tempfile.mkdtemp(prefix="sro-test-"))

    def test_unchecked_is_not_reported_as_absence(self) -> None:
        """Недоступный реестр не должен выглядеть как «не состоит в СРО»."""
        rows = [
            {"name": "ООО «А»", "inn": "7813692784", "sro": "Ассоциация «СРО»",
             "number": "СРО-С-410-16122014", "registry": "НОСТРОЙ",
             "status": "Является членом СРО", "join": date(2023, 4, 12),
             "address": "190000, г. Санкт-Петербург, ул. Примерная, д. 1",
             "note": "", "unchecked": False},
            {"name": "ООО «Б»", "inn": "7817158012", "sro": "", "number": "",
             "registry": "", "status": "", "join": None,
             "address": "",
             "note": "ПРОВЕРКА НЕ ВЫПОЛНЕНА — нет связи с реестром: НОСТРОЙ, НОПРИЗ",
             "unchecked": True},
        ]
        path = self.tmp / "out.xlsx"
        write_report(rows, path, date(2026, 8, 20))

        sheet = openpyxl.load_workbook(path)["Членство в СРО"]
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(sheet.cell(1, 3).value, "СРО")
        self.assertEqual(sheet.cell(2, 3).value, "Ассоциация «СРО»")
        # Колонку ищем по заголовку: их состав меняется, номер — нет опора.
        note = sheet.cell(3, headers.index("Результат проверки") + 1).value
        self.assertIn("НЕ ВЫПОЛНЕНА", note)
        self.assertNotIn("не состоит", note)


class TestTargetSro(unittest.TestCase):
    """Отдельный ответ на вопрос «состоит ли в этой конкретной СРО»."""

    def setUp(self) -> None:
        self.tmp = Path(tempfile.mkdtemp(prefix="sro-test-"))

    def test_abbreviation_does_not_match_inside_words(self) -> None:
        """«ОРС» не должно находиться внутри «ПОМОРСКОГО»."""
        from sro_lookup.target import build_matcher

        matches = build_matcher("ОРС")
        self.assertTrue(matches("Ассоциация «Объединение Ростовских Строителей»"))
        self.assertTrue(matches("СРО Ассоциация ОРС"))
        self.assertFalse(matches("Ассоциация строителей ПОМОРСКОГО края"))
        self.assertFalse(matches("Ассоциация «Балтийский строительный комплекс»"))

    def test_unchecked_never_becomes_no(self) -> None:
        """Реестр не ответил — ответ «не проверено», а не «Нет»."""
        from sro_lookup.target import NO, UNKNOWN, YES, mark_target

        rows = [
            {"inn": "1", "sro": "Ассоциация «Объединение Ростовских Строителей»",
             "number": "", "unchecked": False},
            {"inn": "2", "sro": "Ассоциация «Балтийский комплекс»", "number": "",
             "unchecked": False},
            {"inn": "3", "sro": "", "number": "", "unchecked": True},
        ]
        marked = {row["inn"]: row["target"] for row in mark_target(rows, "ОРС")}
        self.assertEqual(marked["1"], YES)
        self.assertEqual(marked["2"], NO)
        self.assertEqual(marked["3"], UNKNOWN)

    def test_answer_is_same_for_all_rows_of_one_company(self) -> None:
        """Компания в двух СРО занимает две строки — ответ в них одинаковый."""
        from sro_lookup.target import YES, mark_target

        rows = [
            {"inn": "7700000000", "sro": "Ассоциация «Объединение Ростовских Строителей»",
             "number": "", "unchecked": False},
            {"inn": "7700000000", "sro": "Ассоциация проектировщиков «Другая»",
             "number": "", "unchecked": False},
        ]
        self.assertEqual([row["target"] for row in mark_target(rows, "ОРС")], [YES, YES])

    def test_column_is_added_and_filled(self) -> None:
        rows = [{
            "name": "ИП Иванов Иван Иванович", "inn": "616519746524",
            "sro": "Ассоциация «Объединение Ростовских Строителей»",
            "number": "СРО-С-999-01012015", "registry": "НОСТРОЙ",
            "status": "Является членом СРО", "join": date(2022, 3, 1),
            "note": "", "unchecked": False, "target": "Да",
        }]
        path = self.tmp / "target.xlsx"
        write_report(rows, path, date(2026, 8, 21), target_label="ОРС")

        sheet = openpyxl.load_workbook(path)["Членство в СРО"]
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(headers[2], "Состоит в ОРС")
        self.assertEqual(sheet.cell(2, 3).value, "Да")
        # Колонки после вставленной не должны разъехаться.
        self.assertEqual(headers[3], "СРО")
        self.assertEqual(sheet.cell(2, 4).value, "Ассоциация «Объединение Ростовских Строителей»")


class TestAllNegativeWarning(unittest.TestCase):
    """Поголовно отрицательный результат — повод усомниться, а не вывод."""

    def test_warns_when_nothing_found(self) -> None:
        from sro_lookup.report import all_negative_warning

        rows = [{"inn": str(i), "sro": "", "unchecked": False} for i in range(26)]
        self.assertIn("ни одна компания не найдена", all_negative_warning(rows).lower())

    def test_no_warning_when_at_least_one_found(self) -> None:
        from sro_lookup.report import all_negative_warning

        rows = [{"inn": str(i), "sro": "", "unchecked": False} for i in range(25)]
        rows.append({"inn": "99", "sro": "Ассоциация «СРО»", "unchecked": False})
        self.assertEqual(all_negative_warning(rows), "")

    def test_unchecked_rows_do_not_trigger_warning(self) -> None:
        """Если реестр не отвечал, это уже видно по жёлтым строкам."""
        from sro_lookup.report import all_negative_warning

        rows = [{"inn": str(i), "sro": "", "unchecked": True} for i in range(26)]
        self.assertEqual(all_negative_warning(rows), "")


class TestReaderRobustness(unittest.TestCase):
    """Файлы заказчика приходят разной степени аккуратности."""

    def setUp(self) -> None:
        self.tmp = Path(tempfile.mkdtemp(prefix="sro-test-"))

    def test_leading_zero_is_restored(self) -> None:
        """Excel теряет ведущий ноль в ИНН, если ячейка не помечена текстом."""
        from sro_lookup.reader import restore_inn

        self.assertEqual(restore_inn("274147157"), "0274147157")
        self.assertEqual(restore_inn("26910573286"), "026910573286")
        # Вслепую нули не дописываем: иначе любое число стало бы «ИНН».
        self.assertEqual(restore_inn("12310"), "")
        self.assertEqual(restore_inn("1234567890"), "")

    def test_data_sheet_is_chosen_over_instructions(self) -> None:
        """Данные могут лежать не на первом листе, а на втором."""
        import openpyxl
        from sro_lookup.reader import read_companies

        path = self.tmp / "two_sheets.xlsx"
        book = openpyxl.Workbook()
        first = book.active
        first.title = "Как пользоваться"
        first.append(["Инструкция: файл содержит 2 листа"])
        first.append(["Цвета: зелёный — есть СРО"])

        data = book.create_sheet("Лиды")
        data.append(["Приоритет", "Поставщик", "ИНН"])
        data.append(["1. Высокий", 'ООО "МЕГАМЕЙД"', "7806479303"])
        data.append(["2. Средний", 'АО "НЕВСКИЙ"', "7802148357"])
        book.save(path)

        companies = read_companies(path)
        self.assertEqual(len(companies), 2)
        # Наименование берётся из «Поставщик», а не из «Приоритет».
        self.assertIn("МЕГАМЕЙД", companies[0][0])
        self.assertNotIn("Высокий", companies[0][0])

    def test_name_column_comes_from_header(self) -> None:
        """При нескольких текстовых колонках заголовок решает, какая — имя."""
        import openpyxl
        from sro_lookup.reader import read_companies

        path = self.tmp / "many_text.xlsx"
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.append(["Приоритет", "Наименование", "ИНН", "Регион", "Должность"])
        sheet.append(["1. Срочно", "ТИМЕРКАЕВ РУСЛАН РУЗИЛЕВИЧ", "026910573286",
                      "Республика Башкортостан", "ИП"])
        book.save(path)

        companies = read_companies(path)
        self.assertEqual(companies, [("ТИМЕРКАЕВ РУСЛАН РУЗИЛЕВИЧ", "026910573286")])


class TestSimpleTable(unittest.TestCase):
    """Короткая таблица: одна строка на компанию, прямой ответ."""

    def setUp(self) -> None:
        self.tmp = Path(tempfile.mkdtemp(prefix="sro-test-"))

    def test_one_row_per_company_with_all_sro_listed(self) -> None:
        rows = [
            {"name": "ООО «А»", "inn": "7806479303", "sro": "Ассоциация «Проектировщики»",
             "number": "", "registry": "НОПРИЗ", "status": "", "join": None,
             "note": "", "unchecked": False},
            {"name": "ООО «А»", "inn": "7806479303", "sro": "Ассоциация «Строители»",
             "number": "", "registry": "НОСТРОЙ", "status": "", "join": None,
             "note": "", "unchecked": False},
            {"name": "ООО «Б»", "inn": "7802148357", "sro": "", "number": "",
             "registry": "", "status": "", "join": None, "note": "", "unchecked": False},
            {"name": "ООО «В»", "inn": "7839058004", "sro": "", "number": "",
             "registry": "", "status": "", "join": None, "note": "", "unchecked": True},
        ]
        path = self.tmp / "simple.xlsx"
        write_report(rows, path, date(2026, 8, 25), simple=True)

        sheet = openpyxl.load_workbook(path)["Членство в СРО"]
        self.assertEqual([cell.value for cell in sheet[1]][:5],
                         ["Поставщик", "ИНН", "Есть СРО", "В какой СРО", "Реестр"])
        # Компания в двух СРО занимает одну строку, обе СРО перечислены.
        self.assertEqual(sheet.cell(2, 3).value, "Да")
        self.assertIn("Проектировщики", sheet.cell(2, 4).value)
        self.assertIn("Строители", sheet.cell(2, 4).value)
        self.assertEqual(sheet.cell(3, 3).value, "Нет")
        # Недоступный реестр остаётся «не проверено», а не «Нет».
        self.assertEqual(sheet.cell(4, 3).value, "не проверено")


class TestMerge(unittest.TestCase):
    """Перенос колонок из исходного списка в отчёт по СРО."""

    def setUp(self) -> None:
        self.tmp = Path(tempfile.mkdtemp(prefix="sro-test-"))

    def _make(self, name: str, rows: list[list]) -> Path:
        import openpyxl

        path = self.tmp / name
        book = openpyxl.Workbook()
        for row in rows:
            book.active.append(row)
        book.save(path)
        return path

    def test_columns_are_matched_by_inn(self) -> None:
        import subprocess

        sro = self._make("sro.xlsx", [
            ["Название компании", "ИНН", "СРО"],
            ["ООО «А»", "7806479303", "Ассоциация «Первая»"],
            ["ООО «А»", "7806479303", "Ассоциация «Вторая»"],   # две СРО — две строки
            ["ООО «Б»", "0274147157", "Ассоциация «Третья»"],   # ИНН с ведущим нулём
        ])
        # В источнике тот же ИНН записан числом, без ведущего нуля.
        source = self._make("source.xlsx", [
            ["Поставщик", "ИНН", "Телефон", "Email"],
            ["ООО «А»", "7806479303", "+78124387788", "a@example.ru"],
            ["ООО «Б»", 274147157, "+73472000000", "b@example.ru"],
        ])
        out = self.tmp / "merged.xlsx"

        result = subprocess.run(
            [sys.executable, str(Path(__file__).resolve().parent.parent / "merge_sro.py"),
             "--sro", str(sro), "--source", str(source), "--output", str(out)],
            capture_output=True, text=True,
        )
        self.assertEqual(result.returncode, 0, result.stderr)

        sheet = openpyxl.load_workbook(out)["Членство в СРО"]
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(headers[-2:], ["Телефон", "Email"])
        # Обе строки одной компании получают её телефон.
        self.assertEqual(sheet.cell(2, 4).value, "+78124387788")
        self.assertEqual(sheet.cell(3, 4).value, "+78124387788")
        # Ведущий ноль не мешает сопоставлению.
        self.assertEqual(sheet.cell(4, 4).value, "+73472000000")


if __name__ == "__main__":
    unittest.main(verbosity=2)
