#!/usr/bin/env python3
"""
Переносит колонки из исходного файла в отчёт по СРО, сопоставляя по ИНН.

Отчёт по СРО содержит ответ про членство, но не телефоны и почты — они
остались в исходном списке. Скрипт склеивает одно с другим.

    python3 merge_sro.py --sro лиды_СРО.xlsx --source лиды.xlsx

По умолчанию добавляются «Телефон» и «Email», если они есть в источнике.
Другой набор: --columns "Телефон,Email,Руководитель,Приоритет".

Сеть не нужна. Оба исходных файла остаются нетронутыми.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

from sro_lookup.reader import restore_inn

DEFAULT_COLUMNS = ("Телефон", "Email")
INN_HEADERS = ("инн",)


def _sheet_with_inns(book) -> object:
    """Лист с данными — тот, где больше строк с корректным ИНН."""
    def score(sheet) -> int:
        found = 0
        for row in sheet.iter_rows(values_only=True):
            if any(restore_inn(value) for value in row if value not in (None, "")):
                found += 1
        return found

    return max(book.worksheets, key=score)


def read_table(path: Path) -> tuple[list[str], list[list], int]:
    """Возвращает (заголовки, строки, индекс колонки ИНН)."""
    sheet = _sheet_with_inns(openpyxl.load_workbook(path, data_only=True))
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    if not rows:
        raise SystemExit(f"Пустой файл: {path}")

    headers = [str(value or "").strip() for value in rows[0]]
    inn_at = next(
        (i for i, name in enumerate(headers) if name.lower() in INN_HEADERS), None
    )
    if inn_at is None:
        raise SystemExit(f"В файле нет колонки «ИНН»: {path}")
    return headers, rows[1:], inn_at


def main() -> None:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--sro", required=True, help="отчёт по СРО (куда добавляем)")
    parser.add_argument("--source", required=True, help="исходный список (откуда берём)")
    parser.add_argument("--columns", help="какие колонки перенести, через запятую")
    parser.add_argument("--output", help="куда записать результат")
    args = parser.parse_args()

    sro_path, source_path = Path(args.sro), Path(args.source)
    for path in (sro_path, source_path):
        if not path.exists():
            raise SystemExit(f"Файл не найден: {path}")

    target = Path(args.output) if args.output else sro_path.with_name(
        f"{sro_path.stem}_с_контактами.xlsx"
    )

    sro_headers, sro_rows, sro_inn_at = read_table(sro_path)
    src_headers, src_rows, src_inn_at = read_table(source_path)

    wanted = (
        [name.strip() for name in args.columns.split(",") if name.strip()]
        if args.columns
        else [name for name in DEFAULT_COLUMNS if name in src_headers]
    )
    missing = [name for name in wanted if name not in src_headers]
    if missing:
        raise SystemExit(
            f"В исходном файле нет колонок: {', '.join(missing)}\n"
            f"Есть: {', '.join(name for name in src_headers if name)}"
        )
    # Колонку, которая уже есть в отчёте, второй раз не добавляем.
    wanted = [name for name in wanted if name not in sro_headers]
    if not wanted:
        raise SystemExit("Все запрошенные колонки уже есть в отчёте — добавлять нечего.")

    src_at = {name: src_headers.index(name) for name in wanted}
    by_inn: dict[str, list] = {}
    for row in src_rows:
        inn = restore_inn(row[src_inn_at] if src_inn_at < len(row) else "")
        if inn and inn not in by_inn:
            by_inn[inn] = row

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Членство в СРО"

    headers = sro_headers + wanted
    for index, title in enumerate(headers, start=1):
        cell = sheet.cell(1, index, title)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    matched = 0
    for line, row in enumerate(sro_rows, start=2):
        inn = restore_inn(row[sro_inn_at] if sro_inn_at < len(row) else "")
        source = by_inn.get(inn)
        if source:
            matched += 1

        values = list(row) + [
            (source[src_at[name]] if source and src_at[name] < len(source) else "")
            for name in wanted
        ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(line, index, value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=index <= 2)
        sheet.cell(line, sro_inn_at + 1).number_format = "@"

    for index, title in enumerate(headers, start=1):
        width = max(
            (len(str(sheet.cell(line, index).value or "")) for line in range(1, sheet.max_row + 1)),
            default=12,
        )
        sheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 12), 52)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    book.save(target)

    print(f"Строк в отчёте по СРО:   {len(sro_rows)}")
    print(f"Строк в исходнике:       {len(src_rows)}")
    print(f"Перенесены колонки:      {', '.join(wanted)}")
    print(f"Совпало по ИНН:          {matched}")
    if matched < len(sro_rows):
        print(f"Не нашлось в исходнике:  {len(sro_rows) - matched} — эти ячейки пустые")
    print(f"\nГотово: {target}")


if __name__ == "__main__":
    main()
