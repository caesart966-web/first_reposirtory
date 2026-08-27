"""Чтение списка компаний из Excel — без требований к формату файла."""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from .textutils import is_valid_inn


def restore_inn(value: str) -> str:
    """Возвращает ИНН из значения ячейки, восстанавливая ведущие нули.

    Excel хранит ИНН числом, если его не пометили текстом, и тогда ведущий
    ноль пропадает: 0274147157 превращается в 274147157. Такой ИНН не
    проходит проверку и компания молча выпала бы из отчёта. Дополняем нулём
    до 10 или 12 знаков и принимаем только то, что сходится по контрольной
    сумме — вслепую ноль не дописываем.
    """
    digits = re.sub(r"\D", "", str(value or ""))
    if not digits:
        return ""
    candidates = [digits]
    # Дополняем только когда потерян ровно один разряд: 9 -> 10, 11 -> 12.
    # Дописывать нули произвольному короткому числу нельзя — так любое
    # «12310» из текста превратилось бы в «валидный» ИНН.
    if len(digits) == 9:
        candidates.append(digits.zfill(10))
    elif len(digits) == 11:
        candidates.append(digits.zfill(12))
    for candidate in candidates:
        if len(candidate) in (10, 12) and is_valid_inn(candidate):
            return candidate
    return ""


#: Заголовки колонки с наименованием — как их пишут в выгрузках.
_NAME_HEADER_RE = re.compile(
    r"наимен|назван|поставщик|компан|организац|контрагент|фирм", re.IGNORECASE
)
_INN_HEADER_RE = re.compile(r"^\s*инн\s*$|\bинн\b", re.IGNORECASE)


def _find_header(sheet, limit: int = 5) -> tuple[int, int] | None:
    """Ищет строку заголовков и возвращает (колонка названия, колонка ИНН).

    Когда заголовки есть, они надёжнее любой эвристики: в файле бывает
    несколько текстовых колонок (приоритет, регион, должность), и угадывать,
    какая из них наименование, незачем.
    """
    for row in sheet.iter_rows(min_row=1, max_row=limit):
        name_at = inn_at = None
        for cell in row:
            text = str(cell.value or "")
            if inn_at is None and _INN_HEADER_RE.search(text):
                inn_at = cell.column - 1
            elif name_at is None and _NAME_HEADER_RE.search(text):
                name_at = cell.column - 1
        if name_at is not None and inn_at is not None:
            return name_at, inn_at
    return None


def _count_inns(sheet) -> int:
    """Сколько корректных ИНН на листе."""
    found = 0
    for row in sheet.iter_rows(values_only=True):
        for value in row:
            if value not in (None, "") and restore_inn(value):
                found += 1
                break
    return found


def _sheet_with_data(book) -> object:
    """Лист с компаниями — тот, где больше всего ИНН.

    В книге бывают листы с инструкцией, легендой и служебными пометками.
    Брать первый попавшийся нельзя: данные могут лежать на втором.
    """
    return max(book.worksheets, key=_count_inns)


def read_companies(path: Path) -> list[tuple[str, str]]:
    """Пары «название, ИНН» с первого листа книги.

    Формат файла заранее не известен, поэтому колонки определяются по
    содержимому, а не по заголовкам: ИНН — то значение, что проходит
    проверку контрольной суммы. Так работает файл с шапкой и без неё,
    с колонками в любом порядке.

    Повторяющиеся ИНН отбрасываются: запрашивать одну компанию дважды
    незачем.
    """
    book = openpyxl.load_workbook(path, data_only=True)
    sheet = _sheet_with_data(book)
    companies: list[tuple[str, str]] = []
    seen: set[str] = set()

    header = _find_header(sheet)

    for row in sheet.iter_rows(values_only=True):
        if header is not None:
            name_at, inn_at = header
            inn = restore_inn(row[inn_at] if inn_at < len(row) else "")
            name = str(row[name_at] or "").strip() if name_at < len(row) else ""

            if not inn:
                # В колонке ИНН его нет: строка заголовка — или колонки
                # в этой строке переставлены. Терять компанию из-за сбитой
                # разметки нельзя, поэтому просматриваем строку целиком.
                for index, value in enumerate(row):
                    restored = restore_inn(value)
                    if restored:
                        inn = restored
                        if not restore_inn(name) and name:
                            break
                        name = next(
                            (
                                str(other).strip()
                                for position, other in enumerate(row)
                                if position != index
                                and other not in (None, "")
                                and len(str(other).strip()) > 3
                            ),
                            "",
                        )
                        break
            if not inn:
                continue

            if inn not in seen:
                seen.add(inn)
                companies.append((name, inn))
            continue

        values = [str(value).strip() for value in row if value not in (None, "")]
        if not values:
            continue

        inn = ""
        for value in values:
            restored = restore_inn(value)
            if restored:
                inn = restored
                break
        if not inn:
            continue  # строка шапки или посторонние данные

        name = next(
            (value for value in values if restore_inn(value) != inn and len(value) > 3),
            "",
        )
        if inn not in seen:
            seen.add(inn)
            companies.append((name, inn))

    return companies
