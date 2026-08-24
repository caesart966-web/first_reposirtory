"""Запись результата: одна строка на членство в СРО."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_COLUMNS: list[tuple[str, int]] = [
    ("Название компании", 44),
    ("ИНН", 15),
    ("СРО", 52),
    ("Рег. номер СРО", 22),
    ("Реестр", 12),
    ("Статус членства", 24),
    ("Дата вступления", 16),
    ("Результат проверки", 34),
]

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
#: Компания не состоит ни в одной СРО — розовая заливка.
NO_SRO_FILL = PatternFill("solid", fgColor="FCE9E9")
#: Проверка не состоялась — жёлтая: это не ответ, а повод перезапустить.
UNCHECKED_FILL = PatternFill("solid", fgColor="FFF4CE")
THIN = Side(style="thin", color="D0D7E5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
#: Заливки колонки «состоит в искомой СРО»: зелёная — да, серая — нет.
YES_FILL = PatternFill("solid", fgColor="D8EFD8")
NO_FILL = PatternFill("solid", fgColor="EFEFEF")


def all_negative_warning(rows: list[dict]) -> str:
    """Предупреждение, если ни одна компания не нашлась.

    Поголовно отрицательный результат бывает правдой, но бывает и признаком
    того, что поиск сломался: реестр отвечает, а записи не находятся. Молча
    выдавать 26 «не состоит» в такой ситуации нельзя — читатель поверит.
    """
    checked = [row for row in rows if not row.get("unchecked")]
    if len(checked) < 5 or any(row.get("sro") for row in checked):
        return ""
    return (
        "ВНИМАНИЕ: ни одна компания не найдена в реестрах. Это может быть верно "
        "(членство в СРО нужно не всем), но может означать и сбой поиска. "
        "Проверьте контрольным файлом: добавьте пару компаний, заведомо состоящих "
        "в СРО, и запустите снова. Если и они не найдутся — результату верить нельзя."
    )


def write_report(
    rows: list[dict], path: Path, checked_on: date, target_label: str = ""
) -> None:
    """Пишет отчёт. target_label добавляет колонку «Состоит в <СРО>» первой
    после ИНН — это главный ответ, и искать его в конце таблицы неудобно."""
    columns = list(BASE_COLUMNS)
    target_at = 0
    if target_label:
        target_at = 3
        columns.insert(target_at - 1, (f"Состоит в {target_label}", 18))

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Членство в СРО"

    for index, (title, width) in enumerate(columns, start=1):
        cell = sheet.cell(1, index, title)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 28

    for line, row in enumerate(rows, start=2):
        values = [
            row["name"], row["inn"], row["sro"], row["number"],
            row["registry"], row["status"], row["join"], row["note"],
        ]
        if target_label:
            values.insert(target_at - 1, row.get("target", ""))
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(line, index, value)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=index in (1, 3, 8),
                horizontal="center" if index in (2, 4, 5, 7) else "left",
            )
            if row.get("unchecked"):
                cell.fill = UNCHECKED_FILL
            elif not row["sro"]:
                cell.fill = NO_SRO_FILL
        sheet.cell(line, 2).number_format = "@"
        date_at = 8 if target_label else 7
        sheet.cell(line, date_at).number_format = "DD.MM.YYYY"

        if target_label:
            answer = sheet.cell(line, target_at)
            answer.alignment = Alignment(horizontal="center", vertical="top")
            answer.font = Font(name=FONT, size=10, bold=True)
            if row.get("target") == "Да":
                answer.fill = YES_FILL
            elif row.get("target") == "Нет":
                answer.fill = NO_FILL

    warning = all_negative_warning(rows)
    if warning:
        cell = sheet.cell(len(rows) + 3, 1, warning)
        cell.font = Font(name=FONT, size=10, bold=True, color="9C0006")

    note = sheet.cell(
        len(rows) + (5 if warning else 3), 1,
        "Источник: открытые реестры reestr.nostroy.ru (строители) и "
        f"reestr.nopriz.ru (проектировщики, изыскатели). Проверено {checked_on.strftime('%d.%m.%Y')}. "
        "Жёлтые строки — проверка не выполнена, реестр не ответил: их нужно перезапустить.",
    )
    note.font = Font(name=FONT, size=9, italic=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    book.save(path)
