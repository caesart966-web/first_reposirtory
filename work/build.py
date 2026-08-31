import openpyxl, json, collections
from openpyxl.styles import Font, PatternFill

wb = openpyxl.load_workbook("Компании (4).xlsx")
ws = wb["Таблица"]
found = json.load(open("work/found.json"))
done  = json.load(open("work/done.json"))
notes = json.load(open("work/notes.json"))
EMAIL, INN = 11, 2          # 1-based столбцы
def blank(v): return v is None or str(v).strip()==""

# 1) восстановление из самого файла: тот же ИНН с почтой в другой строке
by_inn = collections.defaultdict(list)
for r in range(2, ws.max_row+1):
    by_inn[str(ws.cell(r,INN).value).strip()].append(r)
recovered = 0
for r in range(2, ws.max_row+1):
    if not blank(ws.cell(r,EMAIL).value): continue
    inn = str(ws.cell(r,INN).value).strip()
    for other in by_inn[inn]:
        v = ws.cell(other,EMAIL).value
        if not blank(v):
            found.setdefault(inn, {"email": str(v).strip(),
                                   "source": "дубль того же ИНН в этом же файле (строка %d)" % other})
            recovered += 1
            break

# 2) новые колонки
c_src, c_st, c_note = ws.max_column+1, ws.max_column+2, ws.max_column+3
ws.cell(1,c_src,  "Источник email")
ws.cell(1,c_st,   "Статус поиска email")
ws.cell(1,c_note, "Примечание")
for c in (c_src,c_st,c_note): ws.cell(1,c).font = Font(bold=True)

yellow = PatternFill("solid", fgColor="FFF2CC")
green  = PatternFill("solid", fgColor="D9EAD3")
stats = collections.Counter()
for r in range(2, ws.max_row+1):
    inn = str(ws.cell(r,INN).value).strip()
    if not blank(ws.cell(r,EMAIL).value):
        ws.cell(r,c_st,"был в исходном файле"); stats["исходные"] += 1
    elif inn in found:
        ws.cell(r,EMAIL, found[inn]["email"])
        ws.cell(r,c_src, found[inn]["source"])
        ws.cell(r,c_st,  "найдено и проверено")
        for c in (EMAIL,c_src,c_st): ws.cell(r,c).fill = green
        stats["найдено"] += 1
    elif inn in done:
        ws.cell(r,EMAIL,"нет в открытых источниках")
        ws.cell(r,c_st, "проверено — не найдено")
        ws.cell(r,EMAIL).fill = yellow
        stats["проверено-пусто"] += 1
    else:
        ws.cell(r,EMAIL,"не проверялось")
        ws.cell(r,c_st, "в очереди на проверку")
        stats["не проверено"] += 1
    if inn in notes: ws.cell(r,c_note, notes[inn])

ws.column_dimensions['K'].width = 34
for col,w in (("V",46),("W",24),("X",60)): ws.column_dimensions[col].width = w
out = "Компании (4) — с почтами.xlsx"
wb.save(out)
print("Сохранено:", out)
for k,v in stats.items(): print(f"  {k}: {v}")
print(f"  (восстановлено из дублей внутри файла: {recovered})")
