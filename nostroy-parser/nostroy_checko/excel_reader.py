"""
Чтение табличных файлов: все листы, все строки, ничего не пропускаем.

Поддерживаются:

* ``.xlsx`` / ``.xlsm`` / ``.xltx`` / ``.xltm`` — через ``openpyxl``;
* ``.xls``  — через ``xlrd`` (включая старые файлы 1С);
* ``.xlsb`` — через ``pyxlsb`` (если библиотека установлена);
* ``.csv`` / ``.tsv`` / ``.txt`` — с автоопределением кодировки и разделителя;
* «псевдо-Excel» из 1С и веб-выгрузок: HTML-таблица с расширением ``.xls``
  или ``.xlsx`` (очень частый случай) — разбирается встроенным HTML-парсером;
* файлы с «неправильным» расширением: содержимое определяется по сигнатуре,
  а не по имени.

Если специализированный ридер по какой-то причине не справился, включается
запасной путь через ``pandas.read_excel(..., sheet_name=None, header=None)``,
который читает разом все листы.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterator

from .config import CSV_EXTENSIONS, EXCEL_EXTENSIONS, IGNORED_DIR_NAMES, IGNORED_NAME_PREFIXES
from .logging_setup import get_logger

logger = get_logger("excel")

#: Сигнатуры начала файла, по которым определяем реальный формат.
_SIGNATURE_OLE2 = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"   # старый .xls (OLE2)
_SIGNATURE_ZIP = b"PK\x03\x04"                          # .xlsx/.xlsm (ZIP-контейнер)

#: Кодировки, которые перебираем для текстовых файлов (по популярности в РФ).
_TEXT_ENCODINGS: tuple[str, ...] = ("utf-8-sig", "cp1251", "utf-8", "cp866", "koi8-r", "latin-1")


@dataclass(slots=True)
class SheetData:
    """Содержимое одного листа одной книги."""

    file_path: Path              # абсолютный путь к файлу
    display_path: str            # путь для отчёта (относительно корня разбора)
    sheet_name: str              # имя листа
    rows: list[list[Any]]        # значения ячеек «как есть»
    date_mode: int = 0           # система дат Excel: 0 — 1900, 1 — 1904
    reader: str = ""             # каким ридером прочитан лист (для диагностики)
    warnings: list[str] = field(default_factory=list)

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def column_count(self) -> int:
        return max((len(row) for row in self.rows), default=0)


# --------------------------------------------------------------------------- #
#                            Поиск файлов на диске                             #
# --------------------------------------------------------------------------- #

def _is_ignored_path(path: Path) -> bool:
    """Служебные файлы ОС, временные файлы Excel и системные папки."""
    if any(part in IGNORED_DIR_NAMES for part in path.parts):
        return True
    return any(path.name.startswith(prefix) for prefix in IGNORED_NAME_PREFIXES)


def find_data_files(root: Path, include_csv: bool = True) -> list[Path]:
    """
    Рекурсивно ищет во всех подпапках таблицы, пригодные для разбора.

    :param root:        корневая папка (или одиночный файл);
    :param include_csv: учитывать ли ``.csv``/``.tsv``/``.txt``;
    :return:            отсортированный список путей.
    """
    extensions = set(EXCEL_EXTENSIONS)
    if include_csv:
        extensions |= set(CSV_EXTENSIONS)

    if root.is_file():
        return [root] if root.suffix.lower() in extensions and not _is_ignored_path(root) else []

    found: list[Path] = []
    for path in sorted(root.rglob("*")):
        if not path.is_file() or _is_ignored_path(path):
            continue
        if path.suffix.lower() in extensions:
            found.append(path)
    return found


def sniff_format(path: Path) -> str:
    """
    Определяет реальный формат файла по первым байтам.

    Возвращает одно из: ``xls``, ``xlsx``, ``html``, ``text``, ``unknown``.
    Нужно потому, что выгрузки часто имеют «неправильное» расширение:
    HTML-таблица с именем ``Реестр.xls`` — классика 1С и сайтов СРО.
    """
    try:
        with open(path, "rb") as handle:
            head = handle.read(4096)
    except OSError as exc:
        logger.error("Не удалось прочитать %s: %s", path, exc)
        return "unknown"

    if head.startswith(_SIGNATURE_OLE2):
        return "xls"
    if head.startswith(_SIGNATURE_ZIP):
        return "xlsx"

    lowered = head.lstrip().lower()
    if (
        lowered.startswith(b"<?xml")
        or lowered.startswith(b"<html")
        or lowered.startswith(b"<!doctype html")
        or b"<table" in lowered
        or b"urn:schemas-microsoft-com:office" in lowered
    ):
        return "html"
    if head:
        return "text"
    return "unknown"


# --------------------------------------------------------------------------- #
#                                 .xlsx / .xlsm                                #
# --------------------------------------------------------------------------- #

def _read_xlsx(path: Path, display_path: str) -> list[SheetData]:
    """Читает все листы книги формата OOXML через openpyxl."""
    from openpyxl import load_workbook

    sheets: list[SheetData] = []
    workbook = load_workbook(
        filename=str(path),
        read_only=True,      # потоковое чтение: файлы реестров бывают на сотни МБ
        data_only=True,      # берём значения формул, а не сами формулы
        keep_links=False,
    )
    try:
        # В файле 1904-й системы дат сдвиг на 4 года — учитываем при разборе чисел-дат.
        date_mode = 1 if getattr(workbook, "epoch", None) is not None and \
            str(getattr(workbook, "epoch")).startswith("1904") else 0
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            # У некоторых генераторов книг размеры листа указаны неверно.
            try:
                worksheet.reset_dimensions()
            except AttributeError:
                pass
            rows: list[list[Any]] = []
            for row in worksheet.iter_rows(values_only=True):
                rows.append(list(row))
            rows = _trim_trailing_empty(rows)
            sheets.append(
                SheetData(
                    file_path=path,
                    display_path=display_path,
                    sheet_name=sheet_name,
                    rows=rows,
                    date_mode=date_mode,
                    reader="openpyxl",
                )
            )
    finally:
        workbook.close()
    return sheets


# --------------------------------------------------------------------------- #
#                                     .xls                                     #
# --------------------------------------------------------------------------- #

def _read_xls(path: Path, display_path: str) -> list[SheetData]:
    """Читает все листы старого формата .xls через xlrd."""
    import xlrd

    sheets: list[SheetData] = []
    workbook = xlrd.open_workbook(str(path), on_demand=True, formatting_info=False)
    try:
        date_mode = workbook.datemode
        for sheet_name in workbook.sheet_names():
            worksheet = workbook.sheet_by_name(sheet_name)
            rows: list[list[Any]] = []
            for row_index in range(worksheet.nrows):
                values: list[Any] = []
                for column_index in range(worksheet.ncols):
                    cell = worksheet.cell(row_index, column_index)
                    values.append(_convert_xls_cell(cell, workbook, date_mode))
                rows.append(values)
            rows = _trim_trailing_empty(rows)
            sheets.append(
                SheetData(
                    file_path=path,
                    display_path=display_path,
                    sheet_name=sheet_name,
                    rows=rows,
                    date_mode=date_mode,
                    reader="xlrd",
                )
            )
            workbook.unload_sheet(sheet_name)
    finally:
        workbook.release_resources()
    return sheets


def _convert_xls_cell(cell: Any, workbook: Any, date_mode: int) -> Any:
    """Преобразует ячейку xlrd в питоновское значение (даты — в ``datetime``)."""
    import xlrd

    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate.xldate_as_datetime(cell.value, date_mode)
        except (ValueError, OverflowError, xlrd.XLDateError):
            return cell.value
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return None
    return cell.value


# --------------------------------------------------------------------------- #
#                                    .xlsb                                     #
# --------------------------------------------------------------------------- #

def _read_xlsb(path: Path, display_path: str) -> list[SheetData]:
    """Читает бинарный формат .xlsb (нужна необязательная библиотека pyxlsb)."""
    from pyxlsb import open_workbook as open_xlsb

    sheets: list[SheetData] = []
    with open_xlsb(str(path)) as workbook:
        for sheet_name in workbook.sheets:
            with workbook.get_sheet(sheet_name) as worksheet:
                rows = [[cell.v for cell in row] for row in worksheet.rows()]
            sheets.append(
                SheetData(
                    file_path=path,
                    display_path=display_path,
                    sheet_name=sheet_name,
                    rows=_trim_trailing_empty(rows),
                    reader="pyxlsb",
                )
            )
    return sheets


# --------------------------------------------------------------------------- #
#                                HTML-таблицы                                  #
# --------------------------------------------------------------------------- #

class _TableExtractor(HTMLParser):
    """
    Минимальный парсер HTML-таблиц.

    Используется для выгрузок, которые называются ``.xls``, но внутри являются
    HTML (1С, сайты СРО, экспорт из веб-реестров). Не тянет зависимостей и
    устойчив к «грязной» разметке.
    """

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[list[list[str]]] = []
        self._table_stack: list[list[list[str]]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    # --- обработка тегов ---------------------------------------------- #
    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag == "table":
            self._table_stack.append([])
        elif tag == "tr" and self._table_stack:
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._cell = []
        elif tag == "br" and self._cell is not None:
            self._cell.append(" ")

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in ("td", "th") and self._cell is not None and self._row is not None:
            self._row.append(" ".join("".join(self._cell).split()))
            self._cell = None
        elif tag == "tr" and self._row is not None:
            if self._table_stack:
                self._table_stack[-1].append(self._row)
            self._row = None
        elif tag == "table" and self._table_stack:
            table = self._table_stack.pop()
            if table:
                self.tables.append(table)

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)

    def close(self) -> None:                      # noqa: D102 — докстрока в базовом классе
        super().close()
        # Незакрытые теги <table> тоже сохраняем — разметка бывает битой.
        while self._table_stack:
            table = self._table_stack.pop()
            if table:
                self.tables.append(table)


def _decode_bytes(data: bytes) -> str:
    """Декодирует байты, перебирая распространённые кодировки."""
    lowered = data[:2048].lower()
    for marker, encoding in ((b"windows-1251", "cp1251"), (b"utf-8", "utf-8"), (b"koi8-r", "koi8-r")):
        if marker in lowered:
            try:
                return data.decode(encoding)
            except UnicodeDecodeError:
                break
    for encoding in _TEXT_ENCODINGS:
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _read_html(path: Path, display_path: str) -> list[SheetData]:
    """Читает все таблицы из HTML-файла; каждая таблица становится «листом»."""
    text = _decode_bytes(path.read_bytes())
    parser = _TableExtractor()
    try:
        parser.feed(text)
        parser.close()
    except Exception as exc:  # noqa: BLE001 — HTML бывает совершенно произвольным
        logger.warning("Ошибка разбора HTML %s: %s", display_path, exc)

    sheets: list[SheetData] = []
    for index, table in enumerate(parser.tables, start=1):
        rows = _trim_trailing_empty([list(row) for row in table])
        if not rows:
            continue
        sheets.append(
            SheetData(
                file_path=path,
                display_path=display_path,
                sheet_name=f"HTML-таблица {index}",
                rows=rows,
                reader="html",
            )
        )
    return sheets


# --------------------------------------------------------------------------- #
#                                CSV / TSV / TXT                               #
# --------------------------------------------------------------------------- #

def _read_csv(path: Path, display_path: str) -> list[SheetData]:
    """Читает текстовую таблицу с автоопределением кодировки и разделителя."""
    text = _decode_bytes(path.read_bytes())
    if not text.strip():
        return []

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        # Русский Excel по умолчанию пишет CSV с точкой с запятой.
        counts = {sep: sample.count(sep) for sep in (";", "\t", ",", "|")}
        delimiter = max(counts, key=lambda key: counts[key]) if any(counts.values()) else ";"

    rows = [list(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]
    rows = _trim_trailing_empty(rows)
    if not rows:
        return []
    return [
        SheetData(
            file_path=path,
            display_path=display_path,
            sheet_name=path.stem[:31] or "CSV",
            rows=rows,
            reader=f"csv({delimiter!r})",
        )
    ]


# --------------------------------------------------------------------------- #
#                          Запасной путь через pandas                          #
# --------------------------------------------------------------------------- #

def _read_with_pandas(path: Path, display_path: str) -> list[SheetData]:
    """
    Запасной ридер: ``pandas.read_excel(..., sheet_name=None, header=None)``.

    Читает сразу все листы книги. Используется, если специализированный ридер
    не смог открыть файл (нестандартный генератор, повреждённые метаданные).
    """
    import pandas as pd

    frames = pd.read_excel(str(path), sheet_name=None, header=None, dtype=object)
    sheets: list[SheetData] = []
    for sheet_name, frame in frames.items():
        rows = [
            [None if _is_pandas_na(value) else value for value in row]
            for row in frame.itertuples(index=False, name=None)
        ]
        sheets.append(
            SheetData(
                file_path=path,
                display_path=display_path,
                sheet_name=str(sheet_name),
                rows=_trim_trailing_empty(rows),
                reader="pandas",
            )
        )
    return sheets


def _is_pandas_na(value: Any) -> bool:
    """Безопасная проверка на NaN/NaT/None, не зависящая от версии pandas."""
    try:
        import pandas as pd

        result = pd.isna(value)
        return bool(result) if not hasattr(result, "__len__") else False
    except Exception:  # noqa: BLE001 — pandas может отсутствовать
        return value is None


# --------------------------------------------------------------------------- #
#                                   Общее                                      #
# --------------------------------------------------------------------------- #

def _trim_trailing_empty(rows: list[list[Any]]) -> list[list[Any]]:
    """
    Убирает пустые строки в конце листа.

    Внутренние пустые строки НЕ удаляются: они разделяют логические блоки
    и нужны при определении границ таблиц. Ни одна содержательная строка
    при этом не теряется.
    """
    end = len(rows)
    while end > 0 and all(value is None or str(value).strip() == "" for value in rows[end - 1]):
        end -= 1
    return rows[:end]


def read_tabular_file(path: Path, display_path: str | None = None) -> list[SheetData]:
    """
    Читает ВСЕ листы одного файла.

    Порядок действий: определяем реальный формат по сигнатуре -> вызываем
    профильный ридер -> при ошибке пробуем остальные варианты и, в последнюю
    очередь, pandas. Ошибка на одном файле никогда не останавливает разбор
    остальных: наверх поднимается пустой список и запись в лог.
    """
    display = display_path or path.name
    suffix = path.suffix.lower()
    detected = sniff_format(path)

    # Порядок попыток: сначала самый вероятный ридер для этого файла.
    attempts: list[tuple[str, Any]] = []
    if detected == "xlsx" or suffix in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        attempts.append(("openpyxl", _read_xlsx))
    if detected == "xls" or suffix == ".xls":
        attempts.append(("xlrd", _read_xls))
    if suffix == ".xlsb":
        attempts.append(("pyxlsb", _read_xlsb))
    if detected == "html":
        attempts.append(("html", _read_html))
    if detected == "text" or suffix in CSV_EXTENSIONS:
        attempts.append(("csv", _read_csv))
    # Универсальные запасные варианты (в порядке убывания вероятности успеха).
    for name, reader in (
        ("openpyxl", _read_xlsx),
        ("xlrd", _read_xls),
        ("html", _read_html),
        ("csv", _read_csv),
        ("pandas", _read_with_pandas),
    ):
        if name not in {item[0] for item in attempts}:
            attempts.append((name, reader))

    errors: list[str] = []
    for name, reader in attempts:
        try:
            sheets = reader(path, display)
        except ImportError as exc:
            errors.append(f"{name}: библиотека не установлена ({exc})")
            continue
        except Exception as exc:  # noqa: BLE001 — ридеры кидают разные типы исключений
            errors.append(f"{name}: {type(exc).__name__}: {exc}")
            continue
        if sheets and any(sheet.rows for sheet in sheets):
            if name != attempts[0][0]:
                logger.info("Файл %s прочитан запасным ридером %s", display, name)
            return sheets
        errors.append(f"{name}: данных не найдено")

    logger.error("Не удалось прочитать файл %s. Попытки: %s", display, "; ".join(errors))
    return []


def iter_sheets(paths: list[Path], root: Path | None = None) -> Iterator[SheetData]:
    """
    Проходит по всем файлам и отдаёт их листы по одному.

    :param paths: список файлов (результат :func:`find_data_files`);
    :param root:  корень, относительно которого строится путь для отчёта.
    """
    for path in paths:
        try:
            display = str(path.relative_to(root)) if root else path.name
        except ValueError:
            display = path.name
        logger.info("Читаю файл: %s", display)
        sheets = read_tabular_file(path, display)
        if not sheets:
            continue
        logger.info("  листов: %d, строк всего: %d",
                    len(sheets), sum(sheet.row_count for sheet in sheets))
        for sheet in sheets:
            yield sheet
