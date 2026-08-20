"""
Формирование итогового Excel-отчёта и промежуточных выгрузок.

Итоговый файл ``output/final_report_YYYY-MM-DD.xlsx`` содержит три обязательные
вкладки:

1. ``checko_contacts``  — только контакты, полученные с checko.ru;
2. ``nostroy_contacts`` — только контакты из самого реестра НОСТРОЙ
   (по одной строке на каждую исходную запись — ничего не теряется);
3. ``combined``         — сводная таблица: название, ИНН, все телефоны,
   все email, адреса, руководители, даты вступления и исключения.

По ключу ``--with-diagnostics`` добавляются служебные вкладки
``unrecognized_rows`` и ``run_summary``.
"""

from __future__ import annotations

import csv
import json
from datetime import date
from pathlib import Path
from typing import Any, Sequence

from .logging_setup import get_logger
from .models import CompanyGroup, RegistryRecord, UnrecognizedRow
from .address_pick import detect_shared_addresses, full_address_list, pick_address
from .textutils import (
    director_fio,
    format_date,
    join_unique,
    shorten_company_name,
    truncate,
)

logger = get_logger("report")

#: Максимальная ширина колонки в символах (иначе отчёт нечитаем).
_MAX_COLUMN_WIDTH = 60
_MIN_COLUMN_WIDTH = 10

#: Формат отображения дат в отчёте.
_DATE_FORMAT = "DD.MM.YYYY"

#: Единственный лист отчёта.
SHEET_NAME = "Контакты"


# --------------------------------------------------------------------------- #
#                          Подготовка строк отчёта                             #
# --------------------------------------------------------------------------- #

def _address_candidates(group: CompanyGroup) -> list[str]:
    """Все адреса группы из всех источников — для поиска общих адресов СРО.

    Из карточки checko берётся ИСХОДНЫЙ список, а не уже очищенный: адрес СРО
    опознаётся по частоте, и если смотреть на очищенные карточки, частота
    падает ниже порога и адрес перестаёт распознаваться в запасных источниках.
    """
    candidates: list[str] = []
    if group.checko:
        candidates.extend(full_address_list(group.checko.addresses, group.checko.extra))
    candidates.extend(group.registry_addresses)
    candidates.extend(group.nostroy_addresses)
    return candidates


def _pick_group_address(group: CompanyGroup, shared: set[str]) -> str:
    """Один адрес компании: checko свежее, реестр и НОСТРОЙ — запасные.

    Пустая ячейка честнее чужого адреса: если у компании остался только
    адрес СРО, лучше показать пустоту, чем ввести читателя в заблуждение.
    """
    for values in (
        group.checko.addresses if group.checko else [],
        group.registry_addresses,
        group.nostroy_addresses,
    ):
        remaining = [value for value in values if value not in shared]
        if remaining:
            return pick_address(remaining, group.inn, shared)[0]
    return ""


def build_report_rows(groups: Sequence[CompanyGroup]) -> tuple[list[str], list[list[Any]]]:
    """
    Формирует таблицу отчёта — только по УЖЕ ОБРАБОТАННЫМ компаниям.

    В отчёт попадает компания, по которой запрос к checko.ru доведён до
    окончательного ответа (карточка получена либо сервис ответил «не найдено»).
    Остальные не показываются вовсе: обработали сто — в файле сто строк,
    назавтра обработались следующие сто — стало двести. Так файл накапливается
    день за днём и в нём никогда нет «пустых» компаний, до которых очередь
    ещё не дошла.

    Контакты берутся с checko.ru, а если сервис по компании ничего не дал —
    из самого реестра НОСТРОЙ. Даты вступления/исключения и статус членства —
    из реестра. В колонке «Руководитель» — только ФИО, без должности.

    Строки отсортированы по названию, чтобы файл был стабильным между днями.
    """
    headers = [
        "Название компании",
        "ИНН",
        "Телефон",
        "Email",
        "Адрес",
        "Руководитель",
        "Дата вступления в СРО",
        "Статус членства",
    ]
    processed = [
        group for group in groups if group.checko is not None and group.checko.is_final
    ]
    processed.sort(key=lambda group: group.name)
    # Адрес СРО общий для сотен её членов. Из карточек checko он уже отсеян,
    # но остаётся в запасных источниках — реестре и выгрузке НОСТРОЙ, — откуда
    # подставляется, когда у компании нет своего адреса. Отсеиваем его и там.
    shared_addresses = detect_shared_addresses(
        _address_candidates(group) for group in processed
    )
    rows: list[list[Any]] = []
    for group in processed:
        rows.append(
            [
                shorten_company_name(group.name),
                group.inn,
                join_unique(group.best_phones),
                join_unique(group.best_emails),
                _pick_group_address(group, shared_addresses),
                join_unique(
                    shorten_company_name(director_fio(value))
                    for value in group.best_directors
                ),
                group.date_join,
                group.membership_status,
            ]
        )
    return headers, rows


# --------------------------------------------------------------------------- #
#                              Запись Excel-файла                              #
# --------------------------------------------------------------------------- #

def _write_sheet(
    workbook: Any,
    title: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
) -> None:
    """
    Записывает лист: шапка, данные, автофильтр, закрепление строки, ширины.

    Оформление намеренно скупое: жирная шапка и ничего больше. Заливку и цвета
    не используем — таблицу читают и фильтруют, а не разглядывают.

    ИНН и телефоны пишутся текстом: иначе Excel съедает ведущие нули и
    показывает ``7,70708E+09`` вместо номера. Даты пишутся настоящими датами,
    поэтому по ним работают сортировка и фильтр по периоду.
    """
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    worksheet = workbook.create_sheet(title=title[:31])

    worksheet.append(list(headers))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    text_columns = {
        index
        for index, header in enumerate(headers, start=1)
        if any(marker in header.upper() for marker in ("ИНН", "ОГРН", "ТЕЛЕФОН"))
    }
    date_columns = {
        index for index, header in enumerate(headers, start=1) if "ДАТА" in header.upper()
    }

    widths = [len(str(header)) for header in headers]
    for row in rows:
        prepared: list[Any] = []
        for index, value in enumerate(row, start=1):
            if isinstance(value, date):
                prepared.append(value)
            elif value is None:
                prepared.append("")
            elif isinstance(value, (int, float)) and index not in text_columns:
                prepared.append(value)
            else:
                prepared.append(truncate(str(value)))
        worksheet.append(prepared)
        row_index = worksheet.max_row
        for index, value in enumerate(prepared, start=1):
            cell = worksheet.cell(row=row_index, column=index)
            if index in date_columns and isinstance(value, date):
                cell.number_format = _DATE_FORMAT
            elif index in text_columns:
                cell.number_format = "@"
            length = len(format_date(value) if isinstance(value, date) else str(value))
            if index - 1 < len(widths):
                widths[index - 1] = max(widths[index - 1], min(length, _MAX_COLUMN_WIDTH))

    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = max(
            _MIN_COLUMN_WIDTH, min(width + 2, _MAX_COLUMN_WIDTH)
        )

    worksheet.freeze_panes = "A2"
    if rows:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"


def write_report(path: Path, groups: Sequence[CompanyGroup]) -> Path:
    """
    Формирует итоговый Excel-файл — один лист «Контакты», по строке на
    обработанную компанию (см. :func:`build_report_rows`).

    Всё, что в таблицу не вошло, сохраняется рядом в ``output/parsed/``:
    построчные данные реестра — в ``records.json``, нераспознанные строки —
    в ``unrecognized_rows.csv``.
    """
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)          # убираем пустой лист по умолчанию

    headers, rows = build_report_rows(groups)
    _write_sheet(workbook, SHEET_NAME, headers, rows)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(path))
    workbook.close()
    logger.info(
        "Отчёт сохранён: %s (в файле %d обработанных компаний из %d известных)",
        path, len(rows), len(groups),
    )
    return path


# --------------------------------------------------------------------------- #
#                          Промежуточные выгрузки                              #
# --------------------------------------------------------------------------- #

def write_parsed_artifacts(
    parsed_dir: Path,
    records: Sequence[RegistryRecord],
    groups: Sequence[CompanyGroup],
    unrecognized: Sequence[UnrecognizedRow],
    files_index: Sequence[dict[str, Any]],
    column_report: Sequence[dict[str, Any]] = (),
) -> None:
    """
    Сохраняет промежуточные результаты разбора в ``output/parsed/``.

    Эти файлы нужны и для отладки, и как «страховка»: если формирование
    итогового Excel по какой-то причине не завершится, данные уже на диске.
    """
    parsed_dir.mkdir(parents=True, exist_ok=True)

    _write_json(parsed_dir / "records.json", [record.to_dict() for record in records])
    _write_json(parsed_dir / "companies.json", [group.to_dict() for group in groups])
    _write_csv(
        parsed_dir / "unrecognized_rows.csv",
        ["Файл", "Лист", "Строка", "Причина", "Содержимое строки"],
        [
            [item.file_path, item.sheet_name, item.row_number, item.reason, item.preview]
            for item in unrecognized
        ],
    )
    _write_csv(
        parsed_dir / "files_index.csv",
        ["Файл", "Листов", "Строк", "Записей", "Нераспознано"],
        [
            [
                item.get("file", ""),
                item.get("sheets", 0),
                item.get("rows", 0),
                item.get("records", 0),
                item.get("unrecognized", 0),
            ]
            for item in files_index
        ],
    )
    if column_report:
        # Разбор столбцов: по нему видно, чем распознан каждый столбец файла.
        # Первое, что нужно смотреть, если в отчёте пусто там, где в исходнике
        # данные есть.
        _write_csv(
            parsed_dir / "columns_detected.csv",
            ["Файл", "Лист", "Колонка", "Заголовок", "Распознано как", "Примеры значений"],
            [
                [item.get(key, "") for key in
                 ("Файл", "Лист", "Колонка", "Заголовок", "Распознано как", "Примеры значений")]
                for item in column_report
            ],
        )
    logger.info("Промежуточные выгрузки сохранены в %s", parsed_dir)


def _write_json(path: Path, payload: Any) -> None:
    """Сохраняет JSON в UTF-8 с читаемым форматированием."""
    try:
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    except OSError as exc:
        logger.error("Не удалось записать %s: %s", path, exc)


def _write_csv(path: Path, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> None:
    """Сохраняет CSV с BOM (чтобы Excel сразу открывал его в правильной кодировке)."""
    try:
        with open(path, "w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle, delimiter=";")
            writer.writerow(headers)
            writer.writerows(rows)
    except OSError as exc:
        logger.error("Не удалось записать %s: %s", path, exc)
