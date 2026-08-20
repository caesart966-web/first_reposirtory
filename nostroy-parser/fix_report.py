#!/usr/bin/env python3
"""
Чинит уже готовый отчёт: один адрес, короткие названия, без лишней колонки.

Ничего не скачивает и не запрашивает — работает с файлом на диске.
Исходный файл не изменяется, результат пишется рядом с пометкой «_исправлен».

    python3 fix_report.py output/final_report_2026-08-20.xlsx

Без аргументов берёт самый свежий final_report_*.xlsx из папки output.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

from nostroy_checko.address_pick import detect_shared_addresses, pick_address
from nostroy_checko.textutils import shorten_company_name, tidy_address

#: Колонка, которую заказчик просил убрать.
DROP_COLUMN = "Дата исключения из СРО"
NAME_COLUMN = "Название компании"
INN_COLUMN = "ИНН"
ADDRESS_COLUMN = "Адрес"
HEAD_COLUMN = "Руководитель"


def newest_report(directory: Path) -> Path | None:
    reports = sorted(directory.glob("final_report_*.xlsx"), key=lambda p: p.stat().st_mtime)
    return reports[-1] if reports else None


def column_index(headers: list[str], title: str) -> int | None:
    for index, value in enumerate(headers):
        if str(value or "").strip().lower() == title.lower():
            return index
    return None


def fix(source: Path, target: Path) -> dict:
    sheet = openpyxl.load_workbook(source).worksheets[0]
    headers = [cell.value for cell in sheet[1]]
    rows = [
        [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]
        for row in range(2, sheet.max_row + 1)
    ]

    address_at = column_index(headers, ADDRESS_COLUMN)
    inn_at = column_index(headers, INN_COLUMN)
    name_at = column_index(headers, NAME_COLUMN)
    head_at = column_index(headers, HEAD_COLUMN)
    drop_at = column_index(headers, DROP_COLUMN)

    if address_at is None or inn_at is None:
        raise SystemExit(
            f"В файле нет колонок «{ADDRESS_COLUMN}» и «{INN_COLUMN}» — это не отчёт по СРО."
        )

    # Адрес СРО повторяется у сотен членов — по этому и опознаётся.
    shared = detect_shared_addresses(
        [part.strip() for part in str(row[address_at] or "").split(";") if part.strip()]
        for row in rows
    )

    stats = {"rows": len(rows), "shared": len(shared), "fixed": 0, "empty": 0}
    for row in rows:
        candidates = [
            part.strip() for part in str(row[address_at] or "").split(";") if part.strip()
        ]
        inn = str(row[inn_at] or "")
        chosen, _ = pick_address(candidates, inn, shared)
        if chosen != (row[address_at] or ""):
            stats["fixed"] += 1
        if not chosen:
            stats["empty"] += 1
        row[address_at] = chosen or ""

        if name_at is not None:
            row[name_at] = shorten_company_name(row[name_at])
        if head_at is not None and row[head_at]:
            row[head_at] = shorten_company_name(row[head_at])

    keep = [index for index in range(len(headers)) if index != drop_at]

    book = openpyxl.Workbook()
    out = book.active
    out.title = "Контакты"
    for position, index in enumerate(keep, start=1):
        cell = out.cell(1, position, headers[index])
        cell.font = Font(name="Arial", size=10, bold=True)
    for line, row in enumerate(rows, start=2):
        for position, index in enumerate(keep, start=1):
            cell = out.cell(line, position, row[index])
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=index in (name_at, address_at))
    for position, index in enumerate(keep, start=1):
        width = max((len(str(row[index] or "")) for row in rows), default=10)
        out.column_dimensions[get_column_letter(position)].width = min(max(width + 2, 12), 60)
    out.freeze_panes = "A2"
    out.auto_filter.ref = f"A1:{get_column_letter(len(keep))}1"
    book.save(target)
    return stats


def main() -> None:
    here = Path(__file__).resolve().parent
    if len(sys.argv) > 1:
        source = Path(sys.argv[1])
    else:
        source = newest_report(here / "output") or Path()
        if not source.name:
            raise SystemExit("Не найден ни один final_report_*.xlsx в папке output.")

    if not source.exists():
        raise SystemExit(f"Файл не найден: {source}")

    target = source.with_name(f"{source.stem}_исправлен.xlsx")
    stats = fix(source, target)

    print(f"Исходный файл:  {source}")
    print(f"Строк:          {stats['rows']}")
    print(f"Адресов СРО отброшено: {stats['shared']}")
    print(f"Адрес исправлен в строках: {stats['fixed']}")
    print(f"Осталось без адреса:       {stats['empty']}")
    print(f"\nГотово: {target}")


if __name__ == "__main__":
    main()
