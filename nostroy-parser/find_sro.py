#!/usr/bin/env python3
"""
По списку компаний определяет, в каких СРО они состоят.

Источник — открытые реестры: НОСТРОЙ (строители) и НОПРИЗ (проектировщики
и изыскатели). Оба бесплатны, ключей не требуют. Компания может состоять
сразу в нескольких СРО, поэтому проверяются оба реестра, а не первый попавшийся.

    python3 find_sro.py компании.xlsx

Во входном файле нужны колонки с названием и ИНН — в любом порядке, с шапкой
или без. Результат: <имя файла>_СРО.xlsx рядом с исходным.
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

from nostroy_checko.config import Settings
from nostroy_checko.logging_setup import get_logger
from nostroy_checko.nostroy_api import NOPRIZ_BASE, NOSTROY_BASE, NostroyClient
from nostroy_checko.textutils import is_valid_inn, shorten_company_name

logger = get_logger("find-sro")

COLUMNS = [
    ("Название компании", 44),
    ("ИНН", 15),
    ("СРО", 52),
    ("Рег. номер СРО", 22),
    ("Реестр", 12),
    ("Статус членства", 22),
    ("Дата вступления", 16),
    ("Результат проверки", 30),
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
NOT_FOUND_FILL = PatternFill("solid", fgColor="FCE9E9")
FONT = "Arial"


def read_companies(path: Path) -> list[tuple[str, str]]:
    """Читает пары «название, ИНН» из любого листа, с шапкой или без.

    Колонки определяются по содержимому: ИНН — тот, что проходит проверку
    контрольной суммы. Так файл заказчика не обязан иметь заданный формат.
    """
    sheet = openpyxl.load_workbook(path, data_only=True).worksheets[0]
    companies: list[tuple[str, str]] = []
    seen: set[str] = set()

    for row in sheet.iter_rows(values_only=True):
        values = [str(value).strip() for value in row if value not in (None, "")]
        if not values:
            continue

        inn = ""
        for value in values:
            digits = re.sub(r"\D", "", value)
            if len(digits) in (10, 12) and is_valid_inn(digits):
                inn = digits
                break
        if not inn:
            continue  # шапка или пустая строка

        name = next(
            (value for value in values if re.sub(r"\D", "", value) != inn and len(value) > 3),
            "",
        )
        if inn not in seen:
            seen.add(inn)
            companies.append((name, inn))

    return companies


def lookup_all(companies: list[tuple[str, str]], settings: Settings) -> list[dict]:
    """Опрашивает оба реестра по каждой компании."""
    registries = [
        ("НОСТРОЙ", NOSTROY_BASE, settings.parsed_dir / "nostroy_sample.json"),
        ("НОПРИЗ", NOPRIZ_BASE, settings.parsed_dir / "nopriz_sample.json"),
    ]
    clients = [
        (title, NostroyClient(settings, sample_path=sample, base_url=base, title=title))
        for title, base, sample in registries
    ]

    rows: list[dict] = []
    try:
        for position, (name, inn) in enumerate(companies, start=1):
            memberships = []
            unreachable = []
            for title, client in clients:
                info = client.lookup(inn, name=name)
                if info.found:
                    memberships.append((title, info))
                elif "недоступен" in info.error or "сетевая" in info.error:
                    # Недоступный реестр — это НЕ отрицательный ответ.
                    # Смешать одно с другим значило бы выдать «не состоит в СРО»
                    # там, где проверка просто не состоялась.
                    unreachable.append(title)

            if memberships:
                for title, info in memberships:
                    rows.append(
                        {
                            "name": shorten_company_name(name),
                            "inn": inn,
                            "sro": info.sro_name,
                            "number": info.sro_number,
                            "registry": title,
                            "status": info.status_text,
                            "join": info.date_join,
                            "note": "" if info.sro_name else "СРО в ответе реестра не указана",
                            "failed": False,
                        }
                    )
            else:
                if unreachable:
                    note = "ПРОВЕРКА НЕ ВЫПОЛНЕНА — нет связи с реестром: " + ", ".join(unreachable)
                else:
                    note = "не состоит в СРО (нет записи ни в НОСТРОЙ, ни в НОПРИЗ)"
                rows.append(
                    {
                        "name": shorten_company_name(name),
                        "inn": inn,
                        "sro": "",
                        "number": "",
                        "registry": "",
                        "status": "",
                        "join": None,
                        "note": note,
                        "failed": bool(unreachable),
                    }
                )
            logger.info("%d/%d  %s — записей: %d", position, len(companies), inn, len(memberships))
    finally:
        for _, client in clients:
            close = getattr(client, "close", None)
            if callable(close):
                close()
    return rows


def write(rows: list[dict], path: Path, checked_on: date) -> None:
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Членство в СРО"

    for index, (title, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(1, index, title)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 28

    for line, row in enumerate(rows, start=2):
        values = [
            row["name"], row["inn"], row["sro"], row["number"],
            row["registry"], row["status"], row["join"], row["note"],
        ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(line, index, value)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=index in (1, 3, 8),
                horizontal="center" if index in (2, 4, 5, 7) else "left",
            )
            if not row["sro"]:
                cell.fill = NOT_FOUND_FILL
        sheet.cell(line, 2).number_format = "@"
        sheet.cell(line, 7).number_format = "DD.MM.YYYY"

    note_row = len(rows) + 3
    sheet.cell(
        note_row, 1,
        f"Источник: открытые реестры reestr.nostroy.ru и reestr.nopriz.ru. "
        f"Проверено {checked_on.strftime('%d.%m.%Y')}.",
    ).font = Font(name=FONT, size=9, italic=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    book.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("input", help="xlsx со списком компаний (название и ИНН)")
    parser.add_argument("--output", help="куда записать результат")
    args = parser.parse_args()

    source = Path(args.input)
    if not source.exists():
        raise SystemExit(f"Файл не найден: {source}")

    target = Path(args.output) if args.output else source.with_name(f"{source.stem}_СРО.xlsx")

    settings = Settings(input_path=source, output_dir=target.parent / "output")
    settings.ensure_dirs()

    companies = read_companies(source)
    if not companies:
        raise SystemExit("В файле не нашлось ни одного корректного ИНН.")

    logger.info("Компаний к проверке: %d", len(companies))
    rows = lookup_all(companies, settings)
    write(rows, target, date.today())

    found = sum(1 for row in rows if row["sro"])
    not_member = {row["inn"] for row in rows if not row["sro"] and not row.get("failed")}
    failed = {row["inn"] for row in rows if row.get("failed")}
    print(f"\nКомпаний в файле:      {len(companies)}")
    print(f"Записей о членстве:    {found}")
    print(f"Не состоят в СРО:      {len(not_member)}")
    if failed:
        print(f"ПРОВЕРКА НЕ ВЫПОЛНЕНА: {len(failed)} — реестр не ответил, запустите позже")
    print(f"\nГотово: {target}")


if __name__ == "__main__":
    main()
