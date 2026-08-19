#!/usr/bin/env python3
"""
Автотесты парсера реестра НОСТРОЙ.

Запуск (из корня проекта ``nostroy-parser``)::

    python -m unittest discover -s tests -v
    # или
    python tests/test_suite.py

Сеть не требуется: обращения к checko.ru эмулируются локальным HTTP-сервером,
который отдаёт заготовленные страницы и коды ответов (200/403/404/429).
"""

from __future__ import annotations

import json
import os
import shutil
import sys
import tempfile
import threading
import unittest
import zipfile
from datetime import date, datetime, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

# Позволяем запускать файл напрямую: python tests/test_suite.py
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from nostroy_checko import archives, checko_client, column_mapper, textutils  # noqa: E402
from nostroy_checko.config import Settings  # noqa: E402
from nostroy_checko.excel_reader import find_data_files, read_tabular_file  # noqa: E402
from nostroy_checko.logging_setup import setup_logging  # noqa: E402
from nostroy_checko.models import (  # noqa: E402
    STATUS_FORBIDDEN,
    STATUS_NOT_FOUND,
    STATUS_OK,
    STATUS_RATE_LIMITED,
    CheckoContacts,
)
from nostroy_checko.quota import DailyQuota  # noqa: E402
from nostroy_checko.registry_parser import group_records, parse_sheet  # noqa: E402
from nostroy_checko.state import StateStore  # noqa: E402

from make_fixtures import COMPANIES, build_fixtures  # noqa: E402


# --------------------------------------------------------------------------- #
#                         Эмулятор сайта checko.ru                             #
# --------------------------------------------------------------------------- #

COMPANY_PAGE = """<!DOCTYPE html>
<html lang="ru"><head><title>ООО "Строительная компания Ромашка" — Проверка контрагента</title>
<script type="application/ld+json">
{"@context":"https://schema.org","@type":"Organization",
 "name":"ООО \\"Строительная компания Ромашка\\"","taxID":"7707083893",
 "telephone":"+7 (495) 123-45-67","email":"info@romashka-stroy.ru",
 "address":{"@type":"PostalAddress","postalCode":"123456","addressLocality":"г. Москва",
 "streetAddress":"ул. Ленина, д. 1, оф. 5"}}
</script></head>
<body>
<h1>ООО "Строительная компания Ромашка"</h1>
<dl>
  <dt>ИНН</dt><dd>7707083893</dd>
  <dt>ОГРН</dt><dd>1027700132195</dd>
  <dt>Руководитель</dt><dd>Иванов Иван Иванович</dd>
  <dt>Адрес</dt><dd>123456, г. Москва, ул. Ленина, д. 1, оф. 5</dd>
  <dt>Телефон</dt><dd><a href="tel:+74951234567">+7 (495) 123-45-67</a></dd>
  <dt>Электронная почта</dt><dd><a href="mailto:info@romashka-stroy.ru">info@romashka-stroy.ru</a></dd>
  <dt>Сайт</dt><dd><a href="https://romashka-stroy.ru">romashka-stroy.ru</a></dd>
</dl>
<footer>Служба поддержки: <a href="mailto:info@checko.ru">info@checko.ru</a>, 8 800 222-56-14</footer>
</body></html>"""

SEARCH_PAGE = """<!DOCTYPE html><html><body><h2>Результаты поиска</h2>
<div class="results">
  <a href="/company/romashka-stroy-1027700132195">ООО "Строительная компания Ромашка"</a>
</div></body></html>"""

EMPTY_SEARCH_PAGE = """<!DOCTYPE html><html><body>
<p>По вашему запросу ничего не найдено</p></body></html>"""

LIMIT_PAGE = """<!DOCTYPE html><html><body>
<h1>Превышен лимит запросов</h1><p>Исчерпан лимит бесплатных запросов на сегодня.</p>
</body></html>"""


class _CheckoHandler(BaseHTTPRequestHandler):
    """Обработчик, воспроизводящий поведение checko.ru, включая ошибки."""

    #: Режим работы сервера, задаётся тестом: ok / not_found / limit / forbidden / flaky
    mode = "ok"
    #: Счётчик запросов — тесты проверяют, что лимит не превышен.
    hits: list[str] = []
    lock = threading.Lock()

    def log_message(self, *args):  # noqa: D102 — глушим служебный вывод сервера
        return

    def _respond(self, code: int, body: str, content_type: str = "text/html; charset=utf-8") -> None:
        payload = body.encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def do_GET(self) -> None:  # noqa: N802 — имя задано базовым классом
        with _CheckoHandler.lock:
            _CheckoHandler.hits.append(self.path)
            count = len(_CheckoHandler.hits)

        mode = _CheckoHandler.mode
        if mode == "forbidden":
            self._respond(403, "<html><body>Access denied</body></html>")
            return
        if mode == "limit":
            self._respond(429, "<html><body>Too Many Requests</body></html>")
            return
        if mode == "limit_page":
            self._respond(200, LIMIT_PAGE)
            return
        if mode == "flaky" and count % 2 == 1:
            # Каждый второй запрос — ошибка сервера, клиент обязан повторить.
            self._respond(503, "<html><body>Service Unavailable</body></html>")
            return

        if self.path.startswith("/company/"):
            self._respond(200, COMPANY_PAGE)
            return
        if self.path.startswith("/search"):
            if mode == "not_found":
                self._respond(200, EMPTY_SEARCH_PAGE)
            else:
                self._respond(200, SEARCH_PAGE)
            return
        self._respond(404, "<html><body>Not Found</body></html>")


#: Ответ официального API checko.ru на карточку организации.
API_COMPANY_PAYLOAD = {
    "meta": {"status": "ok", "message": ""},
    "data": {
        "ИНН": "7707083893", "КПП": "770701001", "ОГРН": "1027700132195",
        "НаимСокрЮЛ": 'ООО "РОМАШКА"',
        "НаимПолнЮЛ": 'ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "РОМАШКА"',
        "ЮрАдрес": {
            "АдресРФ": "123456, г. Москва, ул. Ленина, д. 1, оф. 5",
            "Индекс": "123456",
            "Регион": {"Наим": "Москва", "Код": "77"},
        },
        "Контакты": {
            "Тел": ["+7 (495) 123-45-67", "8 916 000-00-00"],
            "Емэйл": ["info@romashka-stroy.ru"],
            "ВебСайт": "romashka-stroy.ru",
        },
        "Руковод": [{"ФИО": "Иванов Иван Иванович", "НаимДолжн": "Генеральный директор"}],
        "Статус": {"Наим": "Действующее"},
    },
}


class _CheckoApiHandler(BaseHTTPRequestHandler):
    """Эмулятор api.checko.ru: успешный ответ, лимит, неверный ключ."""

    mode = "ok"
    hits: list[str] = []
    methods: list[str] = []
    lock = threading.Lock()

    def log_message(self, *args):  # noqa: D102
        return

    def _json(self, code: int, payload: dict) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_POST(self) -> None:  # noqa: N802
        """Некоторые тарифы API принимают только POST — проверяем и такой режим."""
        length = int(self.headers.get("Content-Length", 0) or 0)
        self.rfile.read(length)
        with _CheckoApiHandler.lock:
            _CheckoApiHandler.methods.append("POST")
        self._dispatch(force_ok=True)

    def do_GET(self) -> None:  # noqa: N802
        with _CheckoApiHandler.lock:
            _CheckoApiHandler.hits.append(self.path)
            _CheckoApiHandler.methods.append("GET")
        if _CheckoApiHandler.mode == "post_only":
            self._json(405, {"meta": {"status": "error", "message": "Method Not Allowed"}})
            return
        self._dispatch()

    def _dispatch(self, force_ok: bool = False) -> None:
        mode = "ok" if force_ok else _CheckoApiHandler.mode
        if mode == "bad_key":
            self._json(401, {"meta": {"status": "error", "message": "Неверный ключ доступа"}})
        elif mode == "limit_http":
            self._json(429, {"meta": {"status": "error", "message": "Too Many Requests"}})
        elif mode == "limit_message":
            self._json(200, {"meta": {"status": "error",
                                      "message": "Превышен лимит запросов на сегодня"}})
        elif mode == "not_found":
            self._json(200, {"meta": {"status": "ok", "message": "Компания не найдена"},
                             "data": {}})
        else:
            self._json(200, API_COMPANY_PAYLOAD)


class _FakeCheckoApiServer:
    """Контекстный менеджер с локальным сервером-заглушкой для API."""

    def __init__(self) -> None:
        self.server = ThreadingHTTPServer(("127.0.0.1", 0), _CheckoApiHandler)
        self.thread = threading.Thread(target=self.server.serve_forever, daemon=True)

    def __enter__(self) -> "_FakeCheckoApiServer":
        self.thread.start()
        return self

    def __exit__(self, *exc_info) -> None:
        self.server.shutdown()
        self.server.server_close()
        self.thread.join(timeout=5)

    @property
    def base_url(self) -> str:
        host, port = self.server.server_address[:2]
        return f"http://{host}:{port}/v2"


#: Ответ внутреннего API реестра НОСТРОЙ (структура как у reestr.nostroy.ru).
NOSTROY_MEMBER_PAYLOAD = {
    "data": {
        "data": [
            {
                "id": 123456,
                "full_description": 'ООО "СибирьСтрой"',
                "inn": "5407123456",
                "ogrn": "1065407123456",
                "member_status": {"id": 1, "title": "Является членом"},
                "registry_registration_date": "2016-04-14T00:00:00+03:00",
                "accession_date": "2016-04-14T00:00:00+03:00",
                "exclusion_date": None,
                "birth_date": "1980-01-01T00:00:00+03:00",
                "created_at": "2021-05-01T10:00:00+03:00",
            }
        ],
        "totalCount": 1,
    }
}

NOSTROY_EXCLUDED_PAYLOAD = {
    "data": {
        "data": [
            {
                "full_description": 'ЗАО "ЮгСтройИнвест"',
                "inn": "2310031475",
                "member_status": {"id": 2, "title": "Исключен"},
                "registry_registration_date": "2011-07-08T00:00:00+03:00",
                "exclusion_date": "2021-06-30T00:00:00+03:00",
            }
        ]
    }
}


class _NostroyHandler(BaseHTTPRequestHandler):
    """Эмулятор внутреннего API реестра НОСТРОЙ."""

    hits: list[str] = []
    lock = threading.Lock()

    def log_message(self, *args):  # noqa: D102
        return

    def do_POST(self) -> None:  # noqa: N802
        length = int(self.headers.get("Content-Length", 0) or 0)
        body = self.rfile.read(length).decode("utf-8", errors="replace")
        with _NostroyHandler.lock:
            _NostroyHandler.hits.append(self.path)
        if "2310031475" in body:
            payload = NOSTROY_EXCLUDED_PAYLOAD
        elif "5407123456" in body:
            payload = NOSTROY_MEMBER_PAYLOAD
        else:
            payload = {"data": {"data": [], "totalCount": 0}}
        raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)


class _FakeNostroyServer:
    def __init__(self) -> None:
        self.server = ThreadingHTTPServer(("127.0.0.1", 0), _NostroyHandler)
        self.thread = threading.Thread(target=self.server.serve_forever, daemon=True)

    def __enter__(self) -> "_FakeNostroyServer":
        self.thread.start()
        return self

    def __exit__(self, *exc_info) -> None:
        self.server.shutdown()
        self.server.server_close()
        self.thread.join(timeout=5)

    @property
    def base_url(self) -> str:
        host, port = self.server.server_address[:2]
        return f"http://{host}:{port}"


class _FakeCheckoServer:
    """Контекстный менеджер, поднимающий локальный сервер-заглушку."""

    def __init__(self) -> None:
        self.server = ThreadingHTTPServer(("127.0.0.1", 0), _CheckoHandler)
        self.thread = threading.Thread(target=self.server.serve_forever, daemon=True)

    def __enter__(self) -> "_FakeCheckoServer":
        self.thread.start()
        return self

    def __exit__(self, *exc_info) -> None:
        self.server.shutdown()
        self.server.server_close()
        self.thread.join(timeout=5)

    @property
    def base_url(self) -> str:
        host, port = self.server.server_address[:2]
        return f"http://{host}:{port}"


# --------------------------------------------------------------------------- #
#                                Базовый класс                                 #
# --------------------------------------------------------------------------- #

class BaseTestCase(unittest.TestCase):
    """Общая подготовка: временная папка и отключение системного прокси."""

    @classmethod
    def setUpClass(cls) -> None:
        cls.tmp_dir = Path(tempfile.mkdtemp(prefix="nostroy-test-"))
        setup_logging(cls.tmp_dir / "logs", "ERROR")
        # Локальный сервер-заглушка не должен ходить через корпоративный прокси.
        cls._saved_env = {
            key: os.environ.pop(key)
            for key in ("HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy")
            if key in os.environ
        }

    @classmethod
    def tearDownClass(cls) -> None:
        os.environ.update(cls._saved_env)
        shutil.rmtree(cls.tmp_dir, ignore_errors=True)


# --------------------------------------------------------------------------- #
#                        1. Утилиты нормализации                               #
# --------------------------------------------------------------------------- #

class TestTextUtils(BaseTestCase):
    """Проверка распознавания ИНН, дат, телефонов и email."""

    def test_inn_checksum(self) -> None:
        self.assertTrue(textutils.is_valid_inn("7707083893"))       # ЮЛ
        self.assertTrue(textutils.is_valid_inn("500100732259"))     # ИП
        self.assertFalse(textutils.is_valid_inn("7707083894"))      # неверная сумма
        self.assertFalse(textutils.is_valid_inn("12345"))           # неверная длина

    def test_inn_extraction(self) -> None:
        self.assertEqual(textutils.normalize_inn("ИНН 7707083893"), "7707083893")
        self.assertEqual(textutils.normalize_inn(7707083893.0), "7707083893")
        self.assertEqual(textutils.normalize_inn("7707083893 / 770701001"), "7707083893")
        self.assertEqual(textutils.normalize_inn("77 07 08 38 93"), "7707083893")
        self.assertEqual(textutils.normalize_inn("нет"), "")

    def test_dates(self) -> None:
        expected = date(2010, 5, 12)
        for value in ("12.05.2010", "2010-05-12", "12/05/2010", "12 мая 2010 г."):
            self.assertEqual(textutils.parse_date(value), expected, msg=value)
        self.assertEqual(textutils.parse_date(40123), date(2009, 11, 6))   # серийное число
        self.assertEqual(textutils.parse_date(2010), date(2010, 1, 1))     # «голый» год
        self.assertIsNone(textutils.parse_date("7707083893"))              # ИНН — не дата
        self.assertIsNone(textutils.parse_date(""))

    def test_phones(self) -> None:
        self.assertEqual(
            textutils.extract_phones("8(495)123-45-67, +7 916 000-00-00"),
            ["+74951234567", "+79160000000"],
        )
        # ИНН и ОГРН не должны опознаваться как телефоны.
        self.assertEqual(textutils.extract_phones("7707083893"), [])
        self.assertEqual(textutils.extract_phones("ОГРН 1027700132195"), [])

    def test_emails(self) -> None:
        self.assertEqual(
            textutils.extract_emails("Ivanov@Mail.RU; test@sub.example.co.uk"),
            ["ivanov@mail.ru", "test@sub.example.co.uk"],
        )
        self.assertEqual(textutils.extract_emails("ivanov @ mail.ru"), ["ivanov@mail.ru"])

    def test_director_fio_without_position(self) -> None:
        """В отчёте нужно ФИО, а не «Генеральный директор: ФИО»."""
        cases = {
            "Генеральный директор: Кузьмина Арина Сергеевна": "Кузьмина Арина Сергеевна",
            "Директор Хватынец Леонид Витальевич": "Хватынец Леонид Витальевич",
            "Врио генерального директора Сидоров С.С.": "Сидоров С.С.",
            "Председатель правления Смирнова Анна Владимировна": "Смирнова Анна Владимировна",
            "Единоличный исполнительный орган: Ковалёв Дмитрий Олегович": "Ковалёв Дмитрий Олегович",
            "Петров Пётр Петрович": "Петров Пётр Петрович",
            "Иванов И.И.": "Иванов И.И.",
            # Фамилию «Главин» нельзя принять за должность «глава».
            "Главин Иван Иванович": "Главин Иван Иванович",
            "": "",
        }
        for source, expected in cases.items():
            self.assertEqual(textutils.director_fio(source), expected, msg=source)

    def test_header_normalization(self) -> None:
        self.assertEqual(
            textutils.normalize_header("  Полное  наименование\nчлена СРО (ЮЛ/ИП) "),
            "полное наименование члена сро юл ип",
        )


# --------------------------------------------------------------------------- #
#                        2. Определение колонок                                #
# --------------------------------------------------------------------------- #

class TestColumnMapper(BaseTestCase):
    """Шапка ищется в любом месте листа, роли — по заголовку и по содержимому."""

    def test_header_below_title_rows(self) -> None:
        rows = [
            ["Реестр членов СРО-С-410", None, None],
            [None, None, None],
            ["Наименование", "ИНН", "Дата вступления в СРО"],
            ['ООО "Ромашка"', "7707083893", "12.05.2010"],
        ]
        mapping = column_mapper.build_column_map(rows)
        self.assertEqual(mapping.header_row, 2)
        self.assertEqual(mapping.data_start_row, 3)
        self.assertEqual(mapping.first(column_mapper.ROLE_NAME), 0)
        self.assertEqual(mapping.first(column_mapper.ROLE_INN), 1)
        self.assertEqual(mapping.first(column_mapper.ROLE_DATE_JOIN), 2)

    def test_two_row_header(self) -> None:
        """Шапка в два этажа: нижняя строка сама по себе неинформативна."""
        rows = [
            ["Наименование", "ИНН", "Дата", None],
            [None, None, "вступления", "исключения"],
            ['АО "Тест"', "6658001234", "03.09.2012", "21.11.2019"],
        ]
        mapping = column_mapper.build_column_map(rows)
        self.assertEqual(mapping.header_row, 0)
        self.assertEqual(mapping.header_span, 2)
        self.assertEqual(mapping.data_start_row, 2)
        self.assertEqual(mapping.first(column_mapper.ROLE_DATE_JOIN), 2)
        self.assertEqual(mapping.first(column_mapper.ROLE_DATE_EXIT), 3)

    def test_no_header_detected_by_content(self) -> None:
        """Без шапки роли определяются по самим значениям."""
        rows = [
            ['ООО "Альфа"', "7707083893", "12.05.2010", "123456, г. Москва, ул. Ленина, д. 1"],
            ['ЗАО "Бета"', "2310031475", "08.07.2011", "350000, г. Краснодар, ул. Красная, д. 100"],
            ["ИП Сидоров Сергей Сергеевич", "500100732259", "17.02.2015",
             "141700, Московская обл., г. Долгопрудный, ул. Первомайская, д. 12"],
        ]
        mapping = column_mapper.build_column_map(rows)
        self.assertEqual(mapping.first(column_mapper.ROLE_INN), 1)
        self.assertEqual(mapping.first(column_mapper.ROLE_NAME), 0)
        self.assertEqual(mapping.first(column_mapper.ROLE_DATE_JOIN), 2)
        self.assertEqual(mapping.first(column_mapper.ROLE_ADDRESS), 3)

    def test_birth_date_is_not_taken_as_join_date(self) -> None:
        """
        «Дата рождения» заполнена только у ИП и не должна стать датой вступления.

        Реальный случай: в выгрузке колонка с датой вступления называется
        нестандартно, а рядом есть «Дата рождения» — заполненная лишь у
        предпринимателей. Без проверки плотности заполнения парсер принимал
        её за дату вступления, и у пары ИП в отчёте появлялись 1970-е годы.
        """
        rows = [["Организация", "Номер", "Дата рождения"]]
        for index in range(12):
            # Дата рождения есть только у двух записей из двенадцати.
            birth = "03.05.1970" if index == 3 else ("10.08.1990" if index == 7 else "")
            rows.append([f'ООО "Компания {index}"', f"78{index:08d}", birth])
        mapping = column_mapper.build_column_map(rows)
        self.assertIsNone(mapping.first(column_mapper.ROLE_DATE_JOIN))

    def test_dense_date_column_is_taken_as_join_date(self) -> None:
        """А заполненная у всех колонка с датами — это дата вступления."""
        rows = [["Организация", "Номер", "Когда"]]
        for index in range(12):
            rows.append([f'ООО "Компания {index}"', f"78{index:08d}", f"{index + 1:02d}.03.2020"])
        mapping = column_mapper.build_column_map(rows)
        self.assertEqual(mapping.first(column_mapper.ROLE_DATE_JOIN), 2)

    def test_unknown_column_names(self) -> None:
        """Совершенно нестандартные заголовки — роли берутся из содержимого."""
        rows = [
            ["Кто", "Номер", "Когда", "Как связаться"],
            ['ООО "Гамма"', "7707083893", "12.05.2010", "8 (495) 123-45-67"],
            ['ООО "Дельта"', "2310031475", "08.07.2011", "+7 861 200-30-40"],
        ]
        mapping = column_mapper.build_column_map(rows)
        self.assertEqual(mapping.first(column_mapper.ROLE_INN), 1)
        self.assertEqual(mapping.first(column_mapper.ROLE_PHONE), 3)


# --------------------------------------------------------------------------- #
#                        3. Архивы и чтение файлов                             #
# --------------------------------------------------------------------------- #

class TestArchives(BaseTestCase):
    """Распаковка ZIP: русские имена, вложенные архивы, защита от zip slip."""

    def test_extract_with_cyrillic_names_and_nesting(self) -> None:
        work = self.tmp_dir / "arch"
        work.mkdir(parents=True, exist_ok=True)

        inner = work / "внутренний.zip"
        with zipfile.ZipFile(inner, "w") as archive:
            archive.writestr("Папка/Файл внутри.txt", "данные")

        outer = work / "внешний.zip"
        with zipfile.ZipFile(outer, "w") as archive:
            archive.writestr("Реестр/Отчёт.txt", "содержимое")
            archive.write(inner, arcname="Архивы/внутренний.zip")

        extracted = archives.extract_archive(outer, work / "out")
        names = sorted(path.name for path in extracted)
        self.assertIn("Отчёт.txt", names)
        self.assertIn("Файл внутри.txt", names)   # вложенный архив тоже распакован

    def test_zip_slip_is_blocked(self) -> None:
        work = self.tmp_dir / "slip"
        work.mkdir(parents=True, exist_ok=True)
        malicious = work / "bad.zip"
        with zipfile.ZipFile(malicious, "w") as archive:
            archive.writestr("../../escaped.txt", "плохо")
        extracted = archives.extract_archive(malicious, work / "out")
        for path in extracted:
            self.assertTrue(str(path).startswith(str((work / "out").resolve())))


class TestExcelReader(BaseTestCase):
    """Чтение всех поддерживаемых форматов, включая «псевдо-Excel»."""

    @classmethod
    def setUpClass(cls) -> None:
        super().setUpClass()
        cls.archive = build_fixtures(cls.tmp_dir / "fixtures")
        cls.source_dir = cls.tmp_dir / "fixtures" / "_source"

    def test_all_formats_are_read(self) -> None:
        files = find_data_files(self.source_dir)
        self.assertGreaterEqual(len(files), 4)
        readers = set()
        total_rows = 0
        for path in files:
            sheets = read_tabular_file(path, path.name)
            self.assertTrue(sheets, msg=f"не прочитан файл {path}")
            for sheet in sheets:
                readers.add(sheet.reader)
                total_rows += sheet.row_count
        self.assertTrue(any(reader.startswith("csv") for reader in readers))
        self.assertIn("html", readers)          # HTML-таблица с расширением .xls
        self.assertIn("openpyxl", readers)
        self.assertGreater(total_rows, 10)

    def test_all_sheets_are_read(self) -> None:
        """У книги с тремя листами должны прочитаться все три."""
        path = self.source_dir / "Реестр членов СРО.xlsx"
        sheets = read_tabular_file(path, path.name)
        self.assertEqual(len(sheets), 3)
        self.assertEqual(
            [sheet.sheet_name for sheet in sheets],
            ["Реестр членов", "Доп. сведения", "Выгрузка"],
        )


# --------------------------------------------------------------------------- #
#                        4. Разбор записей реестра                             #
# --------------------------------------------------------------------------- #

class TestRegistryParser(BaseTestCase):
    """Проверка извлечения обязательных полей и группировки."""

    @classmethod
    def setUpClass(cls) -> None:
        super().setUpClass()
        build_fixtures(cls.tmp_dir / "fixtures2")
        cls.source_dir = cls.tmp_dir / "fixtures2" / "_source"

    def _all_records(self):
        records = []
        for path in find_data_files(self.source_dir):
            for sheet in read_tabular_file(path, path.name):
                parsed, _, _ = parse_sheet(sheet)
                records.extend(parsed)
        return records

    def test_required_fields(self) -> None:
        records = self._all_records()
        self.assertTrue(records)
        # Проверяем объединённые данные компании: одна и та же организация
        # встречается в нескольких файлах, и контакты в них разные.
        groups = {group.inn: group for group in group_records(records, "inn")}
        for company in COMPANIES:
            self.assertIn(company["inn"], groups, msg=f"потеряна компания {company['name']}")

        romashka = groups["7707083893"]
        self.assertIn("Ромашка", romashka.name)
        self.assertEqual(romashka.date_join, date(2010, 5, 12))
        self.assertIsNone(romashka.date_exit)
        self.assertIn("+74951234567", romashka.registry_phones)
        self.assertIn("info@romashka-stroy.ru", romashka.registry_emails)
        self.assertTrue(any("Москва" in address for address in romashka.registry_addresses))
        self.assertIn("Иванов Иван Иванович", romashka.registry_directors)

    def test_exit_date_is_parsed(self) -> None:
        records = self._all_records()
        ural = next(record for record in records if record.inn == "6658001234")
        self.assertEqual(ural.date_exit, date(2019, 11, 21))

    def test_nothing_is_skipped(self) -> None:
        """Число записей должно совпадать с числом строк данных во всех файлах."""
        records = self._all_records()
        # 3 + 2 + 1 (xlsx) + 2 (xls) + 2 (csv) + 2 (html) = 12
        self.assertEqual(len(records), 12)

    def test_membership_status(self) -> None:
        """Короткий статус членства: «исключён», если есть дата исключения."""
        from nostroy_checko.models import CompanyGroup, RegistryRecord

        def group_of(**kwargs) -> CompanyGroup:
            return CompanyGroup(key="k", records=[RegistryRecord(name="Тест", **kwargs)])

        self.assertEqual(group_of().membership_status, "действует")
        self.assertEqual(
            group_of(date_exit=date(2019, 11, 21)).membership_status, "исключён"
        )
        self.assertEqual(
            group_of(status="действующий").membership_status, "действует"
        )
        # Дата исключения отсутствует, но текст статуса говорит об исключении.
        self.assertEqual(
            group_of(status="Прекращено членство").membership_status, "исключён"
        )
        self.assertEqual(
            group_of(status="Исключен из членов СРО").membership_status, "исключён"
        )

    def test_grouping_by_inn(self) -> None:
        records = self._all_records()
        groups = group_records(records, "inn")
        self.assertEqual(len(groups), len(COMPANIES))
        # Дедупликация «название + ИНН» даёт не меньше групп, чем по одному ИНН.
        groups_pairs = group_records(records, "name_inn")
        self.assertGreaterEqual(len(groups_pairs), len(groups))


class TestParserRobustness(BaseTestCase):
    """Мусорные строки не ломают разбор и не исчезают бесследно."""

    def _sheet(self, rows):
        from nostroy_checko.excel_reader import SheetData

        return SheetData(
            file_path=Path("dummy.xlsx"),
            display_path="dummy.xlsx",
            sheet_name="Лист1",
            rows=rows,
        )

    def test_junk_rows_are_reported_not_dropped(self) -> None:
        rows = [
            ["Наименование", "ИНН", "Дата вступления"],
            ['ООО "Альфа"', "7707083893", "12.05.2010"],
            [None, None, None],                       # пустая строка — просто разделитель
            ["Наименование", "ИНН", "Дата вступления"],  # повтор шапки на «странице»
            [12345, None, None],                      # мусор: ни ИНН, ни названия
            ['ЗАО "Бета"', "2310031475", "08.07.2011"],
        ]
        records, unrecognized, _ = parse_sheet(self._sheet(rows))
        self.assertEqual(len(records), 2)
        self.assertEqual(len(unrecognized), 1)
        self.assertEqual(unrecognized[0].row_number, 5)
        self.assertIn("12345", unrecognized[0].preview)

    def test_inn_found_without_inn_column(self) -> None:
        """Если колонка ИНН не опознана, он всё равно находится в строке."""
        rows = [
            ["Организация", "Реквизиты", "Дата вступления"],
            ['ООО "Гамма"', "ИНН 7707083893, КПП 770701001", "12.05.2010"],
        ]
        records, _, _ = parse_sheet(self._sheet(rows))
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].inn, "7707083893")

    def test_contacts_found_in_arbitrary_column(self) -> None:
        """Телефон и почта извлекаются даже из колонки «Примечание»."""
        rows = [
            ["Наименование", "ИНН", "Примечание"],
            ['ООО "Дельта"', "7707083893", "связь: 8 (495) 123-45-67, mail: d@delta.ru"],
        ]
        records, _, _ = parse_sheet(self._sheet(rows))
        self.assertIn("+74951234567", records[0].phones)
        self.assertIn("d@delta.ru", records[0].emails)

    def test_empty_sheet_does_not_crash(self) -> None:
        records, unrecognized, _ = parse_sheet(self._sheet([]))
        self.assertEqual(records, [])
        self.assertEqual(unrecognized, [])


# --------------------------------------------------------------------------- #
#                        5. Квота и файл состояния                             #
# --------------------------------------------------------------------------- #

class TestQuota(BaseTestCase):
    """Суточный лимит: бронирование, возврат, сброс в новые сутки."""

    def test_reserve_respects_limit(self) -> None:
        quota = DailyQuota(limit=3)
        self.assertTrue(all(quota.reserve() for _ in range(3)))
        self.assertFalse(quota.reserve())
        self.assertEqual(quota.remaining, 0)
        self.assertTrue(quota.exhausted)

    def test_release_returns_unit(self) -> None:
        quota = DailyQuota(limit=1)
        self.assertTrue(quota.reserve())
        quota.release()
        self.assertTrue(quota.reserve())

    def test_rollover_to_new_day(self) -> None:
        """Вчерашний счётчик обнуляется — это и есть работа «по дням»."""
        yesterday = date.today() - timedelta(days=1)
        quota = DailyQuota(limit=100, used=100, window_date=yesterday, exhausted=True)
        self.assertEqual(quota.remaining, 100)
        self.assertFalse(quota.exhausted)
        self.assertTrue(quota.reserve())

    def test_forbidden_stop_does_not_burn_the_day(self) -> None:
        """
        Отказ сервиса (403, неверный ключ) не должен «съедать» сутки.

        Реальный случай: пользователь запустил сбор с неработающим доступом,
        получил forbidden — и при перезапуске с исправленным ключом скрипт
        отвечал «квота исчерпана», хотя не потратил ни одного запроса.
        """
        quota = DailyQuota(limit=100)
        quota.reserve(); quota.reserve(); quota.reserve()      # потрачено 3
        quota.mark_exhausted("HTTP 403", spent=False)
        self.assertTrue(quota.exhausted)                       # текущий запуск стоит

        # Следующий запуск в тот же день: остановка не переносится.
        restored = DailyQuota.from_dict(quota.to_dict(), limit=100)
        self.assertFalse(restored.exhausted)
        self.assertEqual(restored.used, 3)                     # но счётчик честный
        self.assertTrue(restored.reserve())

        # А настоящий исчерпанный лимит (429) переносится до полуночи.
        spent = DailyQuota(limit=100)
        spent.mark_exhausted("HTTP 429", spent=True)
        restored_spent = DailyQuota.from_dict(spent.to_dict(), limit=100)
        self.assertTrue(restored_spent.exhausted)
        self.assertFalse(restored_spent.reserve())

    def test_parallel_reservations_do_not_exceed_limit(self) -> None:
        """При работе в потоках лимит не должен «протекать»."""
        quota = DailyQuota(limit=50)
        granted: list[bool] = []
        lock = threading.Lock()

        def worker() -> None:
            allowed = quota.reserve()
            with lock:
                granted.append(allowed)

        threads = [threading.Thread(target=worker) for _ in range(200)]
        for thread in threads:
            thread.start()
        for thread in threads:
            thread.join()
        self.assertEqual(sum(granted), 50)


class TestStateStore(BaseTestCase):
    """Файл состояния переживает перезапуск и повреждение."""

    def test_save_and_load(self) -> None:
        path = self.tmp_dir / "state1" / "state.json"
        store = StateStore(path, daily_limit=10)
        store.load()
        store.quota.reserve()
        store.set_result("inn:7707083893", CheckoContacts(query="7707083893", status=STATUS_OK,
                                                          phones=["+74951234567"]))
        store.save()

        restored = StateStore(path, daily_limit=10)
        restored.load()
        self.assertEqual(restored.quota.used, 1)
        result = restored.get_result("inn:7707083893")
        self.assertIsNotNone(result)
        self.assertEqual(result.phones, ["+74951234567"])
        self.assertTrue(restored.has_final_result("inn:7707083893"))

    def test_corrupted_file_does_not_crash(self) -> None:
        path = self.tmp_dir / "state2" / "state.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("{ это не JSON", encoding="utf-8")
        store = StateStore(path, daily_limit=10)
        store.load()                                   # не должно упасть
        self.assertEqual(store.summary()["total"], 0)

    def test_success_is_not_overwritten_by_error(self) -> None:
        path = self.tmp_dir / "state3" / "state.json"
        store = StateStore(path, daily_limit=10)
        store.load()
        store.set_result("inn:1", CheckoContacts(status=STATUS_OK, phones=["+70000000000"]))
        store.set_result("inn:1", CheckoContacts(status="error", error="сеть"))
        self.assertEqual(store.get_result("inn:1").status, STATUS_OK)


# --------------------------------------------------------------------------- #
#                        6. Разбор страницы checko.ru                          #
# --------------------------------------------------------------------------- #

class TestCheckoParsing(BaseTestCase):
    """Разбор HTML карточки компании без обращения к сети."""

    def test_company_page(self) -> None:
        result = checko_client.parse_company_page(
            COMPANY_PAGE, "https://checko.ru/company/x", expected_inn="7707083893"
        )
        self.assertEqual(result.status, STATUS_OK)
        self.assertEqual(result.inn, "7707083893")
        self.assertEqual(result.ogrn, "1027700132195")
        self.assertIn("+74951234567", result.phones)
        self.assertIn("info@romashka-stroy.ru", result.emails)
        self.assertTrue(any("Москва" in address for address in result.addresses))
        self.assertTrue(any("Иванов" in director for director in result.directors))
        self.assertTrue(result.inn_matched)

    def test_own_contacts_are_filtered(self) -> None:
        """Телефон и почта самого сервиса не должны попасть в контакты компании."""
        result = checko_client.parse_company_page(COMPANY_PAGE, "https://checko.ru/company/x")
        self.assertNotIn("info@checko.ru", result.emails)
        self.assertNotIn("+78002225614", result.phones)

    def test_limit_page_is_detected(self) -> None:
        result = checko_client.parse_company_page(LIMIT_PAGE, "https://checko.ru/search")
        self.assertEqual(result.status, STATUS_RATE_LIMITED)

    def test_search_link_extraction(self) -> None:
        url = checko_client.find_card_link(SEARCH_PAGE, "https://checko.ru/search?query=1")
        self.assertEqual(url, "https://checko.ru/company/romashka-stroy-1027700132195")


# --------------------------------------------------------------------------- #
#                        7. HTTP-клиент: коды ответов                          #
# --------------------------------------------------------------------------- #

class TestCheckoClientHTTP(BaseTestCase):
    """Поведение клиента при 200 / 404 / 403 / 429 и при сбоях сервера."""

    def setUp(self) -> None:
        _CheckoHandler.hits = []
        _CheckoHandler.mode = "ok"

    def _client(self, server: _FakeCheckoServer, quota: DailyQuota, **overrides):
        settings = Settings(
            input_path=Path("."), output_dir=self.tmp_dir / "http",
            rps=0.0, timeout=5.0, max_retries=overrides.pop("max_retries", 1),
            **overrides,
        )
        checko_client.CHECKO_BASE_URL = server.base_url
        return checko_client.CheckoClient(settings, quota)

    def test_successful_lookup(self) -> None:
        with _FakeCheckoServer() as server:
            client = self._client(server, DailyQuota(limit=10))
            result = client.lookup("7707083893", expected_inn="7707083893")
        self.assertEqual(result.status, STATUS_OK)
        self.assertIn("+74951234567", result.phones)
        self.assertIn("info@romashka-stroy.ru", result.emails)
        self.assertEqual(len(_CheckoHandler.hits), 2)     # поиск + карточка

    def test_not_found(self) -> None:
        _CheckoHandler.mode = "not_found"
        with _FakeCheckoServer() as server:
            client = self._client(server, DailyQuota(limit=10))
            result = client.lookup("0000000000")
        self.assertEqual(result.status, STATUS_NOT_FOUND)

    def test_forbidden_marks_quota_after_streak(self) -> None:
        _CheckoHandler.mode = "forbidden"
        quota = DailyQuota(limit=10)
        with _FakeCheckoServer() as server:
            client = self._client(server, quota)
            for _ in range(checko_client.CheckoClient.FORBIDDEN_TOLERANCE):
                result = client.lookup("7707083893")
                self.assertEqual(result.status, STATUS_FORBIDDEN)
        self.assertTrue(quota.exhausted)

    def test_429_exhausts_quota(self) -> None:
        _CheckoHandler.mode = "limit"
        quota = DailyQuota(limit=10)
        with _FakeCheckoServer() as server:
            client = self._client(server, quota)
            result = client.lookup("7707083893")
        self.assertEqual(result.status, STATUS_RATE_LIMITED)
        self.assertTrue(quota.exhausted)

    def test_5xx_is_retried(self) -> None:
        """Ошибка 503 должна привести к повтору, а не к отказу."""
        _CheckoHandler.mode = "flaky"
        with _FakeCheckoServer() as server:
            client = self._client(server, DailyQuota(limit=10), max_retries=3)
            result = client.lookup("7707083893")
        self.assertEqual(result.status, STATUS_OK)
        self.assertGreater(len(_CheckoHandler.hits), 2)


class TestSroContactsNotMixedIn(BaseTestCase):
    """
    Реквизиты самой СРО не должны попадать в контакты её членов.

    В выгрузках НОСТРОЙ рядом с данными члена идут телефон и адрес самой
    саморегулируемой организации. Без защиты у всех компаний в отчёте
    оказался бы один и тот же телефон.
    """

    SRO_ADDRESS = "192007, Санкт-Петербург, наб. Обводного канала, д. 60, лит.А"
    SRO_PHONE = "+7 (812) 251-31-01"

    def _sheet(self):
        from nostroy_checko.excel_reader import SheetData

        headers = ["Полное наименование", "ИНН", "Дата вступления",
                   "Адрес места нахождения", "Руководитель",
                   "Наименование СРО", "Адрес СРО", "Телефон СРО", "Телефон"]
        members = [
            ('ООО "СПС"', "5050150032", "12.01.2026", "193230, СПб, пр. Солидарности, д. 1"),
            ('ООО "Статус"', "7814809202", "21.01.2026", "197341, СПб, ул. Афонская, д. 2"),
            ('ООО "Монтаж"', "7814852857", "09.02.2026", "197348, СПб, Богатырский пр., д. 3"),
            ('ООО "ГеоПром"', "7840441705", "16.02.2026", "193167, СПб, ул. Моисеенко, д. 4"),
            ('ООО "Сервис"', "7810445794", "18.02.2026", "198152, СПб, ул. Автовская, д. 5"),
            ('ООО "Дивис"', "7802146141", "26.02.2026", "194362, СПб, Парголово, д. 6"),
            ('ООО "ТоргДом"', "4705075034", "11.02.2026", "190005, СПб, Измайловский б-р, д. 7"),
            ('ООО "Нептун"', "7802765459", "26.02.2026", "192174, СПб, ул. Кибальчича, д. 8"),
            ('ООО "СКР"', "7840486265", "27.03.2026", "192239, СПб, ул. Бухарестская, д. 9"),
            ('ООО "СтройИ"', "7806313724", "17.04.2026", "197022, СПб, наб. Карповки, д. 10"),
        ]
        rows = [headers] + [
            [name, inn, joined, address, f"Директор {index}",
             "Ассоциация «Строительный комплекс»", self.SRO_ADDRESS,
             self.SRO_PHONE, "8 800 100-21-00"]
            for index, (name, inn, joined, address) in enumerate(members, start=1)
        ]
        return SheetData(Path("t.xlsx"), "t.xlsx", "Лист1", rows)

    def test_sro_phone_and_address_are_excluded(self) -> None:
        records, _, _ = parse_sheet(self._sheet())
        self.assertEqual(len(records), 10)
        for record in records:
            self.assertEqual(record.phones, [], msg=f"{record.name}: телефон СРО попал в контакты")
            self.assertFalse(
                any("Обводного" in address for address in record.addresses),
                msg=f"{record.name}: адрес СРО попал в контакты",
            )
        # Собственные адреса членов, наоборот, должны сохраниться — все разные.
        addresses = {address for record in records for address in record.addresses}
        self.assertEqual(len(addresses), 10)
        # И руководители тоже на месте.
        self.assertTrue(all(record.directors for record in records))

    def test_varying_column_is_kept(self) -> None:
        """Колонка с РАЗНЫМИ телефонами остаётся контактом компании."""
        from nostroy_checko.excel_reader import SheetData

        sheet = self._sheet()
        rows = [list(row) for row in sheet.rows]
        for index, row in enumerate(rows[1:], start=1):
            row[8] = f"8 (812) 100-00-{index:02d}"      # телефоны теперь разные
        records, _, _ = parse_sheet(SheetData(Path("t.xlsx"), "t.xlsx", "Лист1", rows))
        self.assertTrue(all(record.phones for record in records))
        self.assertEqual(len({record.phones[0] for record in records}), 10)


# --------------------------------------------------------------------------- #
#                        7b. Официальный API checko.ru                         #
# --------------------------------------------------------------------------- #

class TestCheckoApi(BaseTestCase):
    """Разбор ответа API и обработка его ошибок."""

    def setUp(self) -> None:
        _CheckoApiHandler.hits = []
        _CheckoApiHandler.methods = []
        _CheckoApiHandler.mode = "ok"

    def _client(self, server, quota, sample_path=None, **overrides):
        from nostroy_checko import checko_api

        settings = Settings(
            input_path=Path("."), output_dir=self.tmp_dir / "api",
            checko_api_key="test-key", rps=0.0, timeout=5.0, max_retries=1,
            **overrides,
        )
        checko_api.CHECKO_API_BASE = server.base_url
        return checko_api.CheckoApiClient(settings, quota, sample_path=sample_path)

    def test_falls_back_to_post_when_get_is_rejected(self) -> None:
        """Если сервер не принимает GET, клиент сам повторяет запрос POST'ом."""
        _CheckoApiHandler.mode = "post_only"
        with _FakeCheckoApiServer() as server:
            client = self._client(server, DailyQuota(limit=10))
            first = client.lookup("7707083893", expected_inn="7707083893")
            second = client.lookup("6658001234", expected_inn="6658001234")
        self.assertEqual(first.status, STATUS_OK)
        self.assertIn("+74951234567", first.phones)
        # Первая компания: неудачный GET, затем POST. Метод запомнен —
        # вторая компания идёт сразу POST'ом, лишних обращений нет.
        self.assertEqual(_CheckoApiHandler.methods, ["GET", "POST", "POST"])
        self.assertEqual(second.status, STATUS_OK)

    def test_raw_response_sample_is_saved_without_the_key(self) -> None:
        """Первый ответ API сохраняется целиком — по нему сверяются имена полей."""
        sample = self.tmp_dir / "sample" / "checko_api_sample.json"
        with _FakeCheckoApiServer() as server:
            client = self._client(server, DailyQuota(limit=10), sample_path=sample)
            client.lookup("7707083893", expected_inn="7707083893")
            client.lookup("6658001234", expected_inn="6658001234")
        self.assertTrue(sample.exists())
        saved = json.loads(sample.read_text(encoding="utf-8"))
        self.assertEqual(saved["ответ"]["http_код"], 200)
        self.assertIn("НаимСокрЮЛ", saved["ответ"]["тело"])
        # Ключ не должен попасть в файл, которым будут делиться для диагностики.
        self.assertNotIn("test-key", json.dumps(saved, ensure_ascii=False))

    def test_payload_parsing_company(self) -> None:
        from nostroy_checko.checko_api import parse_api_payload

        result = parse_api_payload(API_COMPANY_PAYLOAD["data"], expected_inn="7707083893")
        self.assertEqual(result.status, STATUS_OK)
        self.assertIn("РОМАШКА", result.name)
        self.assertEqual(result.inn, "7707083893")
        self.assertEqual(result.ogrn, "1027700132195")
        self.assertIn("+74951234567", result.phones)
        self.assertIn("info@romashka-stroy.ru", result.emails)
        self.assertTrue(any("Ленина" in address for address in result.addresses))
        self.assertTrue(any("Иванов" in director for director in result.directors))
        self.assertTrue(result.inn_matched)
        # Вложенное поле «Регион.Наим» не должно подменять название компании.
        self.assertNotEqual(result.name, "Москва")

    def test_director_is_stored_as_fio_only(self) -> None:
        """В отчёт должно попасть ФИО, а не «Генеральный директор: ФИО»."""
        from nostroy_checko.checko_api import parse_api_payload
        from nostroy_checko.report import build_report_rows
        from nostroy_checko.models import CompanyGroup, RegistryRecord

        parsed = parse_api_payload(API_COMPANY_PAYLOAD["data"])
        self.assertEqual(parsed.directors, ["Иванов Иван Иванович"])

        # И на случай данных из другого источника — чистка при выводе в отчёт.
        group = CompanyGroup(
            key="inn:7707083893",
            records=[RegistryRecord(name="Тест", inn="7707083893")],
        )
        group.checko = parsed
        group.checko.directors = ["Генеральный директор: Петров Пётр Петрович"]
        headers, rows = build_report_rows([group])
        self.assertEqual(rows[0][headers.index("Руководитель")], "Петров Пётр Петрович")

    def test_payload_parsing_entrepreneur(self) -> None:
        from nostroy_checko.checko_api import parse_api_payload

        payload = {
            "ИНН": "500100732259", "ОГРНИП": "304500116000157",
            "ФИО": "Сидоров Сергей Сергеевич",
            "Адрес": {"АдресРФ": "141700, Московская обл., г. Долгопрудный, ул. Первомайская, д. 12"},
            "Контакты": {"Тел": "8-916-000-00-00", "Емэйл": "sidorov@mail.ru"},
        }
        result = parse_api_payload(payload, expected_inn="500100732259")
        self.assertEqual(result.name, "Сидоров Сергей Сергеевич")
        self.assertIn("+79160000000", result.phones)
        self.assertIn("sidorov@mail.ru", result.emails)

    def test_successful_request(self) -> None:
        with _FakeCheckoApiServer() as server:
            client = self._client(server, DailyQuota(limit=10))
            result = client.lookup("7707083893", expected_inn="7707083893")
        self.assertEqual(result.status, STATUS_OK)
        self.assertIn("+74951234567", result.phones)
        self.assertEqual(len(_CheckoApiHandler.hits), 1)   # одна компания = один запрос
        self.assertIn("/v2/company", _CheckoApiHandler.hits[0])

    def test_entrepreneur_endpoint_for_12_digit_inn(self) -> None:
        with _FakeCheckoApiServer() as server:
            client = self._client(server, DailyQuota(limit=10))
            client.lookup("500100732259", expected_inn="500100732259")
        self.assertIn("/v2/entrepreneur", _CheckoApiHandler.hits[0])

    def test_bad_key_stops_work(self) -> None:
        _CheckoApiHandler.mode = "bad_key"
        quota = DailyQuota(limit=10)
        with _FakeCheckoApiServer() as server:
            client = self._client(server, quota)
            result = client.lookup("7707083893", expected_inn="7707083893")
        self.assertEqual(result.status, STATUS_FORBIDDEN)
        self.assertTrue(quota.exhausted)     # смысла продолжать нет

    def test_limit_by_http_and_by_message(self) -> None:
        for mode in ("limit_http", "limit_message"):
            _CheckoApiHandler.mode = mode
            quota = DailyQuota(limit=10)
            with _FakeCheckoApiServer() as server:
                client = self._client(server, quota)
                result = client.lookup("7707083893", expected_inn="7707083893")
            self.assertEqual(result.status, STATUS_RATE_LIMITED, msg=mode)
            self.assertTrue(quota.exhausted, msg=mode)

    def test_not_found(self) -> None:
        _CheckoApiHandler.mode = "not_found"
        with _FakeCheckoApiServer() as server:
            client = self._client(server, DailyQuota(limit=10))
            result = client.lookup("7707083893", expected_inn="7707083893")
        self.assertEqual(result.status, STATUS_NOT_FOUND)

    def test_api_key_is_read_from_file(self) -> None:
        """Ключ можно просто вписать в checko_api_key.txt рядом со скриптом."""
        from nostroy_checko.config import read_api_key_file

        directory = self.tmp_dir / "keyfile"
        directory.mkdir(parents=True, exist_ok=True)
        key_file = directory / "checko_api_key.txt"

        # Файла нет — ключа нет.
        self.assertIsNone(read_api_key_file(directory))

        # Файл с одной лишь заглушкой ключом не считается.
        key_file.write_text("# подсказка\n\nВСТАВЬТЕ_КЛЮЧ_СЮДА\n", encoding="utf-8")
        self.assertIsNone(read_api_key_file(directory))

        # Вписанный ключ читается, комментарии игнорируются.
        key_file.write_text("# подсказка\n\n  abc123SECRET  \n", encoding="utf-8")
        self.assertEqual(read_api_key_file(directory), "abc123SECRET")

    def test_record_without_inn_costs_no_request(self) -> None:
        """Записи без ИНН пропускаются: API ищет только по реквизитам."""
        with _FakeCheckoApiServer() as server:
            client = self._client(server, DailyQuota(limit=10))
            result = client.lookup('ООО "Без ИНН"')
        self.assertEqual(result.status, STATUS_NOT_FOUND)
        self.assertEqual(len(_CheckoApiHandler.hits), 0)


class TestNostroyRegistry(BaseTestCase):
    """Даты вступления/исключения дотягиваются из открытого реестра НОСТРОЙ."""

    def test_payload_parsing(self) -> None:
        from nostroy_checko.nostroy_api import parse_member_payload

        info = parse_member_payload(NOSTROY_MEMBER_PAYLOAD, "5407123456")
        self.assertTrue(info.found)
        self.assertEqual(info.date_join, date(2016, 4, 14))
        self.assertIsNone(info.date_exit)          # действующий член
        # Дата рождения и дата создания записи не приняты за даты членства.

        excluded = parse_member_payload(NOSTROY_EXCLUDED_PAYLOAD, "2310031475")
        self.assertEqual(excluded.date_join, date(2011, 7, 8))
        self.assertEqual(excluded.date_exit, date(2021, 6, 30))
        self.assertIn("Исключен", excluded.status_text)

    def test_lookup_and_cache(self) -> None:
        from nostroy_checko import nostroy_api
        from nostroy_checko.models import CompanyGroup, RegistryRecord
        from nostroy_checko.pipeline import enrich_dates_from_nostroy

        _NostroyHandler.hits = []
        settings = Settings(input_path=Path("."), output_dir=self.tmp_dir / "nostroy",
                            rps=0.0, nostroy_rps=1000.0, timeout=5.0)
        store = StateStore(self.tmp_dir / "nostroy" / "state.json", daily_limit=100)
        store.load()

        groups = [
            CompanyGroup(key="inn:5407123456",
                         records=[RegistryRecord(name='ООО "СибирьСтрой"', inn="5407123456")]),
            CompanyGroup(key="inn:2310031475",
                         records=[RegistryRecord(name='ЗАО "ЮгСтройИнвест"', inn="2310031475")]),
        ]
        with _FakeNostroyServer() as server:
            nostroy_api.NOSTROY_BASE = server.base_url
            enrich_dates_from_nostroy(groups, settings, store)
            hits_first = len(_NostroyHandler.hits)
            self.assertGreater(hits_first, 0)

            # Даты подставились и попали в итоговые свойства группы.
            self.assertEqual(groups[0].date_join, date(2016, 4, 14))
            self.assertEqual(groups[0].membership_status, "действует")
            self.assertEqual(groups[1].date_exit, date(2021, 6, 30))
            self.assertEqual(groups[1].membership_status, "исключён")

            # Повторный запуск берёт всё из кэша — новых запросов нет.
            fresh = [
                CompanyGroup(key="inn:5407123456",
                             records=[RegistryRecord(name="X", inn="5407123456")]),
            ]
            enrich_dates_from_nostroy(fresh, settings, store)
            self.assertEqual(len(_NostroyHandler.hits), hits_first)
            self.assertEqual(fresh[0].date_join, date(2016, 4, 14))

    def test_file_dates_take_precedence(self) -> None:
        """Дата из файла выгрузки главнее даты из реестра НОСТРОЙ."""
        from nostroy_checko.models import CompanyGroup, RegistryRecord

        group = CompanyGroup(
            key="inn:5407123456",
            records=[RegistryRecord(name="X", inn="5407123456", date_join=date(2015, 1, 1))],
        )
        group.nostroy_date_join = date(2016, 4, 14)
        self.assertEqual(group.date_join, date(2015, 1, 1))


class TestGenericDatabases(BaseTestCase):
    """Скрипт принимает любые базы данных, не только выгрузки НОСТРОЙ."""

    def test_arbitrary_client_database(self) -> None:
        """Файл с произвольными колонками: клиенты CRM, а не реестр СРО."""
        from nostroy_checko.excel_reader import SheetData

        rows = [
            ["Клиент", "Реквизиты (ИНН)", "Комментарий менеджера"],
            ['ООО "Альфа Трейд"', "7707083893", "перезвонить в четверг"],
            ['ЗАО "Бета Логистик"', "2310031475", "ждёт КП"],
            ["ИП Сидоров Сергей Сергеевич", "500100732259", "8 916 000-00-00 — личный"],
        ]
        sheet = SheetData(Path("clients.xlsx"), "clients.xlsx", "CRM", rows)
        records, unrecognized, _ = parse_sheet(sheet)
        self.assertEqual(len(records), 3)
        self.assertEqual(unrecognized, [])
        by_inn = {record.inn: record for record in records}
        self.assertIn("7707083893", by_inn)
        # Телефон из «Комментария» тоже подобран — контакты ищутся по всей строке.
        self.assertIn("+79160000000", by_inn["500100732259"].phones)
        # Группировка и подготовка к checko работают без дат — это не ошибка.
        groups = group_records(records, "inn")
        self.assertEqual(len(groups), 3)
        self.assertTrue(all(group.query_for_checko() for group in groups))


# --------------------------------------------------------------------------- #
#              8. Полный цикл: лимит, остановка и продолжение назавтра         #
# --------------------------------------------------------------------------- #

class TestEndToEndAcrossDays(BaseTestCase):
    """
    Главный сценарий из ТЗ.

    День 1: лимит 3 запроса -> обработаны 3 компании, остальные ждут.
    День 2: счётчик обнуляется -> обрабатываются оставшиеся, ранее собранные
    данные не запрашиваются повторно.
    """

    def _settings(self, output: Path, base_url: str, daily_limit: int) -> Settings:
        checko_client.CHECKO_BASE_URL = base_url
        return Settings(
            input_path=self.archive,
            output_dir=output,
            daily_limit=daily_limit,
            workers=3,
            rps=0.0,
            timeout=5.0,
            max_retries=1,
            no_progress_bar=True,
        )

    @classmethod
    def setUpClass(cls) -> None:
        super().setUpClass()
        cls.archive = build_fixtures(cls.tmp_dir / "fixtures3")

    def test_two_day_run(self) -> None:
        from nostroy_checko.pipeline import run

        _CheckoHandler.hits = []
        _CheckoHandler.mode = "ok"
        output = self.tmp_dir / "e2e"

        with _FakeCheckoServer() as server:
            # ------------------------------ день 1 -------------------------- #
            settings = self._settings(output, server.base_url, daily_limit=3)
            day_one = run(settings)

            self.assertEqual(len(day_one.groups), len(COMPANIES))
            self.assertEqual(day_one.checko_done + day_one.checko_failed, 3)
            self.assertEqual(day_one.checko_pending, len(COMPANIES) - 3)
            self.assertTrue(day_one.report_path.exists())

            state_file = settings.state_file
            self.assertTrue(state_file.exists())
            saved = json.loads(state_file.read_text(encoding="utf-8"))
            self.assertEqual(saved["quota"]["used"], 3)
            self.assertEqual(len(saved["results"]), 3)

            # --- имитируем наступление следующих суток ----------------------- #
            saved["quota"]["window_date"] = (date.today() - timedelta(days=1)).isoformat()
            state_file.write_text(json.dumps(saved, ensure_ascii=False), encoding="utf-8")
            hits_after_day_one = len(_CheckoHandler.hits)

            # ------------------------------ день 2 -------------------------- #
            settings = self._settings(output, server.base_url, daily_limit=100)
            day_two = run(settings)

        # Все компании обработаны, повторных запросов по готовым — нет.
        self.assertEqual(day_two.checko_pending, 0)
        self.assertEqual(day_two.checko_done + day_two.checko_failed, len(COMPANIES) - 3)
        self.assertLessEqual(
            len(_CheckoHandler.hits) - hits_after_day_one,
            (len(COMPANIES) - 3) * 2,
        )

        # Отчёт — один лист с нужными колонками.
        from openpyxl import load_workbook

        workbook = load_workbook(day_two.report_path)
        self.assertEqual(workbook.sheetnames, ["Контакты"])
        sheet = workbook["Контакты"]
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(
            headers,
            [
                "Название компании",
                "ИНН",
                "Телефон",
                "Email",
                "Адрес",
                "Руководитель",
                "Дата вступления в СРО",
                "Дата исключения из СРО",
                "Статус членства",
            ],
        )

        # Контакты получены по всем компаниям.
        phone_column = headers.index("Телефон")
        filled = [
            row[phone_column]
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if row[phone_column]
        ]
        self.assertEqual(len(filled), len(COMPANIES))

        # Даты вступления — настоящие даты, взятые из реестра НОСТРОЙ.
        join_column = headers.index("Дата вступления в СРО")
        joins = [row[join_column] for row in sheet.iter_rows(min_row=2, values_only=True)]
        self.assertTrue(all(isinstance(value, datetime) for value in joins))

        # Шапка выделена только жирностью — заливки и цветного текста нет.
        header_cell = sheet["A1"]
        self.assertTrue(header_cell.font.bold)
        self.assertIn(header_cell.fill.fill_type, (None, "none"))

    def test_retry_policy(self) -> None:
        """Ошибки повторяются ограниченное число раз, «не успели» — всегда."""
        from nostroy_checko.pipeline import MAX_AUTO_RETRY_ATTEMPTS, _needs_lookup
        from nostroy_checko.models import CompanyGroup, RegistryRecord

        settings = Settings(input_path=Path("."), output_dir=self.tmp_dir / "retry")
        store = StateStore(self.tmp_dir / "retry" / "state.json", daily_limit=100)
        store.load()
        group = CompanyGroup(key="inn:7707083893",
                             records=[RegistryRecord(name="Тест", inn="7707083893")])

        # Ничего не известно — запрашиваем.
        self.assertTrue(_needs_lookup(group, store, settings))

        # Окончательные ответы повторять нельзя — это трата квоты.
        store.set_result(group.key, CheckoContacts(status=STATUS_OK))
        self.assertFalse(_needs_lookup(group, store, settings))
        store._results[group.key] = CheckoContacts(status=STATUS_NOT_FOUND)
        self.assertFalse(_needs_lookup(group, store, settings))

        # «Не успели из-за лимита» — повторяем всегда.
        store._results[group.key] = CheckoContacts(status=STATUS_RATE_LIMITED)
        self.assertTrue(_needs_lookup(group, store, settings))

        # Ошибки — не более MAX_AUTO_RETRY_ATTEMPTS раз подряд.
        store._attempts.pop(group.key, None)
        for _ in range(MAX_AUTO_RETRY_ATTEMPTS):
            store.set_result(group.key, CheckoContacts(status="error", error="сеть"))
        self.assertFalse(_needs_lookup(group, store, settings))
        settings.retry_failed = True
        self.assertTrue(_needs_lookup(group, store, settings))

    def test_switching_to_api_retries_html_failures(self) -> None:
        """
        Появился ключ API — прошлые неудачи парсера HTML не блокируют повтор.

        Иначе компании, которые не смог найти разбор страниц сайта, навсегда
        остались бы со статусом not_found, и API их бы уже не проверил.
        """
        from nostroy_checko.models import CompanyGroup, RegistryRecord
        from nostroy_checko.pipeline import _needs_lookup

        store = StateStore(self.tmp_dir / "switch" / "state.json", daily_limit=100)
        store.load()
        group = CompanyGroup(key="inn:7707083893",
                             records=[RegistryRecord(name="Тест", inn="7707083893")])

        # Вчера парсер сайта ничего не нашёл — результат «окончательный».
        store.set_result(group.key, CheckoContacts(status=STATUS_NOT_FOUND, source="html"))

        without_key = Settings(input_path=Path("."), output_dir=self.tmp_dir / "switch")
        self.assertFalse(_needs_lookup(group, store, without_key))

        with_key = Settings(input_path=Path("."), output_dir=self.tmp_dir / "switch",
                            checko_api_key="new-key")
        self.assertTrue(_needs_lookup(group, store, with_key))

        # А вот удачно собранные данные перезапрашивать незачем — квота дорога.
        store._results[group.key] = CheckoContacts(status=STATUS_OK, source="html",
                                                   phones=["+74951234567"])
        self.assertFalse(_needs_lookup(group, store, with_key))

    def test_forbidden_stops_run_but_report_is_written(self) -> None:
        """Блокировка останавливает запросы, но отчёт всё равно формируется."""
        from nostroy_checko.pipeline import run

        _CheckoHandler.hits = []
        _CheckoHandler.mode = "forbidden"
        output = self.tmp_dir / "e2e-forbidden"
        with _FakeCheckoServer() as server:
            settings = self._settings(output, server.base_url, daily_limit=100)
            result = run(settings)
        _CheckoHandler.mode = "ok"

        self.assertTrue(result.interrupted)
        self.assertEqual(result.checko_done, 0)
        self.assertTrue(result.report_path.exists())
        # Данные реестра собраны полностью, несмотря на недоступность checko.ru.
        self.assertEqual(len(result.groups), len(COMPANIES))
        self.assertEqual(len(result.records), 12)

    def test_report_accumulates_processed_companies_only(self) -> None:
        """
        В отчёте только обработанные компании — файл копится день за днём.

        Обработали две — в файле две строки, с контактами, датами и статусом
        членства. Необработанных компаний в таблице нет вовсе: «пустых» строк,
        до которых не дошла квота, пользователь видеть не должен.
        """
        from nostroy_checko.pipeline import run
        from openpyxl import load_workbook

        _CheckoApiHandler.hits = []
        _CheckoApiHandler.mode = "ok"
        output = self.tmp_dir / "e2e-accumulate"
        with _FakeCheckoApiServer() as server:
            from nostroy_checko import checko_api

            checko_api.CHECKO_API_BASE = server.base_url
            settings = Settings(
                input_path=self.archive, output_dir=output, checko_api_key="k",
                daily_limit=2, workers=2, rps=0.0, timeout=5.0, max_retries=1,
                no_progress_bar=True,
            )
            result = run(settings)

        sheet = load_workbook(result.report_path)["Контакты"]
        headers = [cell.value for cell in sheet[1]]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))

        # Квоты хватило на две компании — в отчёте ровно две строки.
        self.assertEqual(len(rows), 2)
        self.assertEqual(result.in_report, 2)
        self.assertNotIn("Статус запроса", headers)

        # У вошедших строк заполнено всё: контакты, даты, статус членства.
        phone_column = headers.index("Телефон")
        join_column = headers.index("Дата вступления в СРО")
        status_column = headers.index("Статус членства")
        for row in rows:
            self.assertTrue(row[phone_column])
            self.assertTrue(row[join_column])
            self.assertIn(row[status_column], ("действует", "исключён"))

    def test_quota_is_never_exceeded(self) -> None:
        """При лимите N выполняется ровно N запросов к сайту, не больше."""
        from nostroy_checko.pipeline import run

        _CheckoHandler.hits = []
        _CheckoHandler.mode = "ok"
        output = self.tmp_dir / "e2e-limit"
        with _FakeCheckoServer() as server:
            settings = self._settings(output, server.base_url, daily_limit=2)
            result = run(settings)
        # 2 компании * (поиск + карточка) = максимум 4 обращения.
        self.assertLessEqual(len(_CheckoHandler.hits), 4)
        self.assertEqual(result.checko_done, 2)
        self.assertEqual(result.quota_used, 2)


if __name__ == "__main__":
    unittest.main(verbosity=2)
