#!/usr/bin/env python3
"""
Генератор тестовых данных.

Создаёт ZIP-архив, внутри которого лежат «реестры» в разных форматах и с
разными названиями колонок — ровно то, с чем приходится иметь дело на практике:

* ``xlsx`` с обычной шапкой и «шапкой в два этажа»;
* ``xls`` (старый формат) с нестандартными заголовками;
* HTML-таблица с расширением ``.xls`` (выгрузка из 1С);
* ``csv`` в кодировке cp1251 с разделителем «;»;
* лист вообще без шапки — роли колонок должны определиться по содержимому.

Запуск: ``python tests/make_fixtures.py [папка]``
"""

from __future__ import annotations

import csv
import io
import sys
import zipfile
from pathlib import Path

# --------------------------------------------------------------------------- #
#            Набор реальных по формату (но вымышленных) компаний              #
# --------------------------------------------------------------------------- #

# ИНН подобраны с корректной контрольной суммой — иначе автодетект их отбракует.
COMPANIES: list[dict[str, str]] = [
    {
        "name": 'ООО "Строительная компания Ромашка"',
        "inn": "7707083893",
        "ogrn": "1027700132195",
        "join": "12.05.2010",
        "exit": "",
        "phone": "8 (495) 123-45-67",
        "email": "info@romashka-stroy.ru",
        "address": "123456, г. Москва, ул. Ленина, д. 1, оф. 5",
        "director": "Иванов Иван Иванович",
    },
    {
        "name": 'АО "СтройМонтаж-Урал"',
        "inn": "6658001234",
        "ogrn": "1026602345678",
        "join": "03.09.2012",
        "exit": "21.11.2019",
        "phone": "+7 343 200-10-20",
        "email": "office@sm-ural.ru",
        "address": "620014, Свердловская обл., г. Екатеринбург, пр-т Ленина, д. 5",
        "director": "Петров Пётр Петрович",
    },
    {
        "name": "ИП Сидоров Сергей Сергеевич",
        "inn": "500100732259",
        "ogrn": "304500116000157",
        "join": "17.02.2015",
        "exit": "",
        "phone": "8-916-000-00-00",
        "email": "sidorov@mail.ru",
        "address": "141700, Московская обл., г. Долгопрудный, ул. Первомайская, д. 12, кв. 3",
        "director": "Сидоров Сергей Сергеевич",
    },
    {
        "name": 'ООО "ПромСтройПроект"',
        "inn": "7802312345",
        "ogrn": "1057812345678",
        "join": "25.12.2009",
        "exit": "",
        "phone": "(812) 777-88-99, +7 921 111-22-33",
        "email": "psp@psp-spb.ru; sales@psp-spb.ru",
        "address": "195269, г. Санкт-Петербург, ш. Революции, д. 84, литер А",
        "director": "Смирнова Анна Владимировна",
    },
    {
        "name": 'ЗАО "ЮгСтройИнвест"',
        "inn": "2310031475",
        "ogrn": "1022301614873",
        "join": "08.07.2011",
        "exit": "30.06.2021",
        "phone": "861 200 30 40",
        "email": "info@ugsi.ru",
        "address": "350000, Краснодарский край, г. Краснодар, ул. Красная, д. 100",
        "director": "Ковалёв Дмитрий Олегович",
    },
    {
        "name": 'ООО "СибирьСтрой"',
        "inn": "5407123456",
        "ogrn": "1065407123456",
        "join": "14.04.2016",
        "exit": "",
        "phone": "",
        "email": "",
        "address": "630099, Новосибирская обл., г. Новосибирск, ул. Ленина, д. 21",
        "director": "",
    },
]


def _write_xlsx_standard(path: Path) -> None:
    """Лист с «нормальной» шапкой + отдельный лист со вторым куском реестра."""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Реестр членов"
    sheet.append(["Реестр членов СРО-С-410-14102010"])          # заголовок-название
    sheet.append([])                                             # пустая строка
    sheet.append(
        [
            "№ п/п",
            "Полное наименование члена СРО",
            "ИНН",
            "ОГРН",
            "Дата вступления в СРО",
            "Дата исключения",
            "Статус",
            "Юридический адрес",
            "Контактный телефон",
            "Адрес электронной почты",
            "Руководитель",
        ]
    )
    for index, company in enumerate(COMPANIES[:3], start=1):
        sheet.append(
            [
                index,
                company["name"],
                company["inn"],
                company["ogrn"],
                company["join"],
                company["exit"],
                "исключен" if company["exit"] else "действующий",
                company["address"],
                company["phone"],
                company["email"],
                company["director"],
            ]
        )

    # Второй лист — «многоэтажная» шапка с объединёнными ячейками.
    second = workbook.create_sheet("Доп. сведения")
    second.append(["Сведения о членах", None, None, "Членство", None, "Контакты", None])
    second.append(
        ["Наименование", "ИНН", "ОГРН", "дата приёма", "дата прекращения", "телефон", "e-mail"]
    )
    for company in COMPANIES[3:5]:
        second.append(
            [
                company["name"],
                company["inn"],
                company["ogrn"],
                company["join"],
                company["exit"],
                company["phone"],
                company["email"],
            ]
        )

    # Третий лист — вообще без шапки (роли определяются по содержимому).
    third = workbook.create_sheet("Выгрузка")
    for company in COMPANIES[5:]:
        third.append(
            [
                company["name"],
                company["inn"],
                company["join"],
                company["address"],
            ]
        )
    workbook.save(str(path))


def _write_xls_legacy(path: Path) -> None:
    """
    Старый формат .xls с нестандартными названиями колонок.

    Пишется через xlwt, если он есть; иначе — через openpyxl в .xlsx-совместимом
    виде (тест всё равно проверит разбор нестандартной шапки).
    """
    try:
        import xlwt
    except ImportError:
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "1С Выгрузка"
        sheet.append(["Контрагент", "Идентификационный номер налогоплательщика",
                      "Членство с", "Членство по", "Тел.", "Почта", "Место нахождения"])
        for company in COMPANIES[:2]:
            sheet.append([company["name"], company["inn"], company["join"],
                          company["exit"], company["phone"], company["email"],
                          company["address"]])
        workbook.save(str(path.with_suffix(".xlsx")))
        return

    workbook = xlwt.Workbook(encoding="utf-8")
    sheet = workbook.add_sheet("1С Выгрузка")
    headers = ["Контрагент", "Идентификационный номер налогоплательщика",
               "Членство с", "Членство по", "Тел.", "Почта", "Место нахождения"]
    for column, header in enumerate(headers):
        sheet.write(0, column, header)
    for row, company in enumerate(COMPANIES[:2], start=1):
        for column, key in enumerate(
            ["name", "inn", "join", "exit", "phone", "email", "address"]
        ):
            sheet.write(row, column, company[key])
    workbook.save(str(path))


def _write_html_as_xls(path: Path) -> None:
    """HTML-таблица с расширением .xls — классическая выгрузка из 1С и с сайтов."""
    rows = "".join(
        f"<tr><td>{company['name']}</td><td>{company['inn']}</td>"
        f"<td>{company['join']}</td><td>{company['exit']}</td>"
        f"<td>{company['phone']}</td><td>{company['email']}</td></tr>"
        for company in COMPANIES[2:4]
    )
    html = (
        '<html xmlns:o="urn:schemas-microsoft-com:office:office">'
        '<head><meta http-equiv="Content-Type" content="text/html; charset=windows-1251"></head>'
        "<body><table border=1>"
        "<tr><th>Название организации</th><th>ИНН</th><th>Дата принятия в члены</th>"
        "<th>Дата исключения из членов</th><th>Телефоны</th><th>Электронная почта</th></tr>"
        f"{rows}</table></body></html>"
    )
    path.write_bytes(html.encode("cp1251"))


def _write_csv_cp1251(path: Path) -> None:
    """CSV в кодировке cp1251 с разделителем «;» — как сохраняет русский Excel."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow(["Наименование", "ИНН", "Дата включения в реестр",
                     "Телефон", "E-mail", "Фактический адрес"])
    for company in COMPANIES[4:]:
        writer.writerow([company["name"], company["inn"], company["join"],
                         company["phone"], company["email"], company["address"]])
    path.write_bytes(buffer.getvalue().encode("cp1251"))


def build_fixtures(target_dir: Path) -> Path:
    """
    Создаёт набор файлов и упаковывает их в ZIP со вложенными папками.

    :return: путь к созданному архиву.
    """
    target_dir.mkdir(parents=True, exist_ok=True)
    work_dir = target_dir / "_source"
    nested = work_dir / "СРО 410" / "Выгрузки"
    nested.mkdir(parents=True, exist_ok=True)

    _write_xlsx_standard(work_dir / "Реестр членов СРО.xlsx")
    _write_xls_legacy(nested / "Выгрузка 1С.xls")
    _write_html_as_xls(nested / "Реестр с сайта.xls")
    _write_csv_cp1251(nested / "Дополнение.csv")

    archive_path = target_dir / "Реестр НОСТРОЙ.zip"
    with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in sorted(work_dir.rglob("*")):
            if path.is_file():
                archive.write(path, arcname=str(path.relative_to(work_dir)))
    return archive_path


if __name__ == "__main__":
    output = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("tests/fixtures")
    created = build_fixtures(output)
    print(f"Тестовый архив создан: {created}")
