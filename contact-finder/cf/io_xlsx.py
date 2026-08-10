"""Чтение входного и запись выходного xlsx.

Чтение работает даже без openpyxl: .xlsx — это zip с XML, и минимальный
парсер вытаскивает первые два столбца (название, ИНН). Запись требует
openpyxl (красивое форматирование); если его нет — пишем CSV.
"""
from __future__ import annotations

import csv
import os
import re
import xml.etree.ElementTree as ET
import zipfile

from .models import Company

_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


# --------------------------------------------------------------------- чтение

def _col_to_idx(ref: str) -> int:
    """A1 -> 0, B2 -> 1. Индекс столбца из адреса ячейки."""
    letters = re.match(r"[A-Z]+", ref or "")
    if not letters:
        return 0
    idx = 0
    for ch in letters.group(0):
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def _read_xlsx_raw(path: str) -> list[list[str]]:
    """Минимальный парсер xlsx без внешних зависимостей."""
    with zipfile.ZipFile(path) as zf:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall("m:si", _NS):
                shared.append("".join(t.text or "" for t in si.iter("{%s}t" % _NS["m"])))

        # Берём первый лист по порядку.
        sheet_name = "xl/worksheets/sheet1.xml"
        names = [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet")]
        if sheet_name not in zf.namelist() and names:
            sheet_name = sorted(names)[0]

        root = ET.fromstring(zf.read(sheet_name))
        rows: list[list[str]] = []
        for row in root.iter("{%s}row" % _NS["m"]):
            cells: dict[int, str] = {}
            for c in row.findall("m:c", _NS):
                idx = _col_to_idx(c.get("r", ""))
                v = c.find("m:v", _NS)
                if v is None:
                    is_node = c.find("m:is", _NS)
                    text = "".join(t.text or "" for t in is_node.iter("{%s}t" % _NS["m"])) if is_node is not None else ""
                elif c.get("t") == "s":
                    text = shared[int(v.text)] if v.text and v.text.isdigit() else ""
                else:
                    text = v.text or ""
                cells[idx] = text
            width = max(cells) + 1 if cells else 0
            rows.append([cells.get(i, "") for i in range(width)])
        return rows


def read_companies(path: str) -> list[Company]:
    """Читает файл: 1-й столбец — название, 2-й — ИНН. Первая строка — шапка."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".tsv"):
        delim = "\t" if ext == ".tsv" else ","
        with open(path, encoding="utf-8-sig") as fh:
            rows = [list(r) for r in csv.reader(fh, delimiter=delim)]
    else:
        try:
            import openpyxl  # type: ignore
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            rows = [[("" if c is None else str(c)) for c in r]
                    for r in wb.active.iter_rows(values_only=True)]
        except Exception:
            rows = _read_xlsx_raw(path)

    companies: list[Company] = []
    for i, row in enumerate(rows, 1):
        if i == 1:  # шапка
            continue
        name = (row[0] if len(row) > 0 else "").strip()
        inn = (row[1] if len(row) > 1 else "").strip()
        inn = re.sub(r"\.0$", "", inn)  # 7802708778.0 -> 7802708778
        if name and inn:
            companies.append(Company(row=i, name=name, inn=inn))
    return companies


# ---------------------------------------------------------------------- запись

_HEADERS = [
    "№", "Организация", "ИНН", "Телефон(ы)", "E-mail", "Сайт",
    "Адрес", "Директор", "Источник(и)", "Достоверность", "Отброшено (проверить)",
]


def _row_for(idx: int, c: Company) -> list[str]:
    return [
        idx, c.name, c.inn, c.phones_display, "; ".join(c.emails), c.site,
        c.address, c.director, c.sources, c.confidence, c.rejected_note,
    ]


def write_results(path: str, companies: list[Company]) -> str:
    """Пишет xlsx (openpyxl) либо CSV, если openpyxl недоступен."""
    try:
        return _write_xlsx(path, companies)
    except Exception:
        csv_path = os.path.splitext(path)[0] + ".csv"
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(_HEADERS)
            for i, c in enumerate(companies, 1):
                w.writerow(_row_for(i, c))
        return csv_path


def _write_xlsx(path: str, companies: list[Company]) -> str:
    import openpyxl  # type: ignore
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Контакты"
    ws.append(_HEADERS)

    head_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ok = PatternFill("solid", fgColor="E2EFDA")
    mid = PatternFill("solid", fgColor="FFF2CC")
    no = PatternFill("solid", fgColor="FCE4E4")

    for i, c in enumerate(companies, 1):
        ws.append(_row_for(i, c))
        r = ws.max_row
        has_phone = bool(c.phones)
        fill = ok if has_phone else (mid if c.rejected_note else no)
        for col in range(1, len(_HEADERS) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=col in (2, 5, 6, 7, 8, 9, 11))
            cell.fill = fill
            if col == 4:
                cell.font = Font(bold=has_phone)
        ws.cell(row=r, column=3).number_format = "@"

    for col, width in zip("ABCDEFGHIJK", [5, 34, 14, 30, 26, 24, 40, 26, 40, 30, 34]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{ws.max_row}"

    found = sum(1 for c in companies if c.phones)
    s = wb.create_sheet("Сводка")
    for row in [
        ["Показатель", "Значение"],
        ["Всего организаций", len(companies)],
        ["Телефон найден", found],
        ["Без телефона", len(companies) - found],
        ["", ""],
        ["Как работает", "Многоисточниковый пробив: агрегаторы по ИНН → сайт компании → гео-справочники"],
        ["Проверка ИНН", "Телефон принимается, только если на странице найден искомый ИНН"],
        ["Столбец «Отброшено»", "Контакты, где ИНН не подтверждён — кандидаты на ручную сверку"],
    ]:
        s.append(row)
    for cell in s[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
    s.column_dimensions["A"].width = 26
    s.column_dimensions["B"].width = 96
    for r in s.iter_rows(min_row=2):
        r[1].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)
    return path
