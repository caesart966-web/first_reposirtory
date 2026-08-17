#!/usr/bin/env python3
"""Сбор реестра членов СРО «ПОС» с sropos.ru.

Запускать НА СВОЁЙ МАШИНЕ: в среде, где собирался этот репозиторий,
внешний трафик закрыт прокси, поэтому сайт оттуда недоступен.

    pip install requests beautifulsoup4 openpyxl
    python3 sropos_scrape.py

Что делает:
  1. находит адреса карточек членов (страница реестра, sitemap, перебор ID);
  2. качает каждую карточку, вытаскивает ИНН, ОГРН, дату регистрации,
     рег.№ и ЛЮБЫЕ телефоны, какие на ней есть;
  3. кладёт результат в sropos_raw.json и sropos.xlsx;
  4. сохраняет 3 карточки целиком в sropos_sample/ — если разбор что-то
     упустит, по ним видно, что именно поправить.

Скрипт намеренно не завязан на конкретную вёрстку: структура страницы
заранее неизвестна, поэтому поля ищутся регулярками по всему тексту, а
подписи «поле — значение» дополнительно снимаются с таблиц и списков.
"""
import json
import os
import re
import sys
import time
from collections import OrderedDict

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    sys.exit('Сначала: pip install requests beautifulsoup4 openpyxl')

BASE = 'https://sropos.ru'
OUT_JSON = 'sropos_raw.json'
OUT_XLSX = 'sropos.xlsx'
SAMPLE_DIR = 'sropos_sample'
DELAY = 0.7                      # пауза между запросами, чтобы не долбить сайт
TIMEOUT = 20
RETRIES = 3

S = requests.Session()
S.headers['User-Agent'] = ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                           'AppleWebKit/537.36 (KHTML, like Gecko) '
                           'Chrome/124.0 Safari/537.36')

# ---------------------------------------------------------------- регулярки
# Телефон: +7/8, код в скобках или без, разделители любые.
RE_PHONE = re.compile(r'(?:\+7|\b8)[\s\-—(]*\d{3}[\s\-—)]*\d{3}[\s\-—]*\d{2}[\s\-—]*\d{2}\b')
RE_INN = re.compile(r'\b(\d{12}|\d{10})\b')
RE_OGRN = re.compile(r'\b(\d{15}|\d{13})\b')
RE_DATE = re.compile(r'\b(\d{2}\.\d{2}\.\d{4})\b')
RE_EMAIL = re.compile(r'\b[\w.+-]+@[\w-]+\.[\w.]+\b')
RE_CARD = re.compile(r'/(?:reestr|memberinfo)/(\d+)/?')


def get(url):
    for i in range(RETRIES):
        try:
            r = S.get(url, timeout=TIMEOUT)
            if r.status_code == 404:
                return None
            r.raise_for_status()
            r.encoding = r.apparent_encoding or 'utf-8'
            return r.text
        except Exception as e:
            if i == RETRIES - 1:
                print(f'  !! {url}: {e}')
                return None
            time.sleep(2 ** i)


def discover():
    """Адреса карточек: со страницы реестра, из sitemap, из самих карточек."""
    urls = OrderedDict()

    def harvest(html, where):
        if not html:
            return 0
        n = 0
        for m in RE_CARD.finditer(html):
            u = f'{BASE}{m.group(0)}'.rstrip('/') + '/'
            if u not in urls:
                urls[u] = True
                n += 1
        if n:
            print(f'  {where}: +{n}')
        return n

    print('Ищу карточки...')
    for path in ('/register/', '/reestr/', '/sitemap.xml', '/sitemap_index.xml'):
        harvest(get(BASE + path), path)

    # Реестр часто разбит на страницы — пробуем обычные варианты пагинации.
    for tpl in ('/register/?PAGEN_1={}', '/register/page/{}/', '/register/?page={}'):
        empty = 0
        for p in range(2, 200):
            added = harvest(get(BASE + tpl.format(p)), tpl.format(p))
            empty = empty + 1 if not added else 0
            if empty >= 3:
                break
            time.sleep(DELAY)
        if len(urls) > 50:
            break
    print(f'Найдено адресов карточек: {len(urls)}')
    return list(urls)


def probe_ids(known):
    """Если ссылок мало — перебрать ID вокруг уже известных.

    В выдаче встречались /reestr/17676/ … /reestr/18307/, то есть ID идут
    сплошным диапазоном. Перебор закрывает карточки, на которые нет ссылок
    со страницы реестра.
    """
    ids = sorted(int(RE_CARD.search(u).group(1)) for u in known)
    if not ids:
        lo, hi = 17600, 18400
    else:
        lo, hi = max(1, ids[0] - 200), ids[-1] + 200
    print(f'Перебираю ID {lo}..{hi} (это долго, ~{(hi - lo) * DELAY / 60:.0f} мин)')
    return [f'{BASE}/reestr/{i}/' for i in range(lo, hi + 1)]


def labelled_pairs(soup):
    """Пары «подпись: значение» из таблиц и списков определений."""
    out = {}
    for tr in soup.find_all('tr'):
        tds = [td.get_text(' ', strip=True) for td in tr.find_all(['td', 'th'])]
        if len(tds) == 2 and tds[0] and tds[1]:
            out.setdefault(tds[0][:80], tds[1][:300])
    for dl in soup.find_all('dl'):
        dts, dds = dl.find_all('dt'), dl.find_all('dd')
        for dt, dd in zip(dts, dds):
            out.setdefault(dt.get_text(' ', strip=True)[:80],
                           dd.get_text(' ', strip=True)[:300])
    return out


def parse(url, html):
    soup = BeautifulSoup(html, 'html.parser')
    for tag in soup(['script', 'style']):
        tag.decompose()
    text = re.sub(r'\s+', ' ', soup.get_text(' ', strip=True))

    # Телефон и адрес самой СРО стоят в шапке/подвале на каждой странице —
    # их надо выкинуть, иначе один и тот же номер приедет всем 972 членам.
    body = soup.find(['main', 'article']) or soup.find(
        'div', class_=re.compile(r'content|detail|reestr|member', re.I)) or soup
    btext = re.sub(r'\s+', ' ', body.get_text(' ', strip=True))

    inns = RE_INN.findall(btext)
    ogrns = RE_OGRN.findall(btext)
    # ИНН не должен быть куском ОГРН
    inns = [i for i in inns if not any(i in o for o in ogrns)]

    title = soup.find(['h1', 'h2'])
    return {
        'url': url,
        'name': title.get_text(' ', strip=True) if title else '',
        'inn': inns[0] if inns else '',
        'inn_all': sorted(set(inns)),
        'ogrn': ogrns[0] if ogrns else '',
        'dates': sorted(set(RE_DATE.findall(btext))),
        'phones': sorted(set(RE_PHONE.findall(btext))),
        'phones_page': sorted(set(RE_PHONE.findall(text))),   # включая шапку
        'emails': sorted(set(RE_EMAIL.findall(btext))),
        'fields': labelled_pairs(soup),
        'text': btext[:4000],
    }


def main():
    urls = discover()
    if len(urls) < 100:
        print('Ссылок мало — добираю перебором ID.')
        urls = list(OrderedDict.fromkeys(urls + probe_ids(urls)))

    os.makedirs(SAMPLE_DIR, exist_ok=True)
    rows, saved = [], 0
    for i, u in enumerate(urls, 1):
        html = get(u)
        time.sleep(DELAY)
        if not html:
            continue
        rec = parse(u, html)
        if not rec['inn'] and not rec['name']:
            continue
        rows.append(rec)
        if saved < 3:                       # образцы для уточнения разбора
            open(f'{SAMPLE_DIR}/card_{saved + 1}.html', 'w',
                 encoding='utf-8').write(html)
            saved += 1
        if i % 25 == 0 or i == len(urls):
            ph = sum(1 for r in rows if r['phones'])
            print(f'[{i}/{len(urls)}] карточек: {len(rows)}, с телефоном: {ph}')
            json.dump(rows, open(OUT_JSON, 'w'), ensure_ascii=False, indent=1)

    json.dump(rows, open(OUT_JSON, 'w'), ensure_ascii=False, indent=1)
    print(f'\nГотово. Карточек: {len(rows)}, из них с телефоном: '
          f'{sum(1 for r in rows if r["phones"])}')

    # Номер, повторяющийся почти везде, — это телефон самой СРО из шапки.
    from collections import Counter
    cnt = Counter(p for r in rows for p in set(r['phones']))
    junk = {p for p, c in cnt.items() if c > max(5, len(rows) * 0.2)}
    if junk:
        print(f'Сквозные номера (скорее всего, самой СРО), исключены: {sorted(junk)}')

    try:
        from openpyxl import Workbook
    except ImportError:
        print('Для xlsx: pip install openpyxl'); return
    wb = Workbook(); ws = wb.active; ws.title = 'СРО ПОС'
    ws.append(['№', 'Организация', 'ИНН', 'ОГРН', 'Даты на карточке',
               'Телефон(ы)', 'E-mail', 'Ссылка'])
    for i, r in enumerate(rows, 1):
        ws.append([i, r['name'], r['inn'], r['ogrn'], '; '.join(r['dates']),
                   '; '.join(p for p in r['phones'] if p not in junk),
                   '; '.join(r['emails']), r['url']])
        ws.cell(row=ws.max_row, column=3).number_format = '@'
    for col, w in zip('ABCDEFGH', [6, 52, 15, 18, 34, 46, 30, 44]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'
    wb.save(OUT_XLSX)
    print(f'Файлы: {OUT_XLSX}, {OUT_JSON}, образцы в {SAMPLE_DIR}/')
    print('\nПришлите мне sropos_raw.json и карточку из sropos_sample/ — '
          'подгоню разбор под реальную вёрстку и сведу с текущей таблицей.')


if __name__ == '__main__':
    main()
