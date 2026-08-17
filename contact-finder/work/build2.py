import json, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import datetime
recs = json.load(open('nostroy.json'))
res  = json.load(open('nostroy_results.json'))
# Только действующие записи: у 5 компаний осталась ещё и старая строка
# «Исключен», и она затирала действующую — рег.№ и статус вставали чужие.
active_rows = [r for r in recs if r['status'] == 'Является членом']
by_inn = {r['inn']: r for r in active_rows}
assert len(by_inn) == 972, len(by_inn)

joined = {k: datetime.date.fromisoformat(v[:10])
          for k, v in json.load(open('join_dates.json')).items()}
TODAY = datetime.date(2026, 8, 17)
CUT6 = TODAY - datetime.timedelta(days=183)

wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Телефоны СРО-410'
ws.append(['№','Организация (реестр НОСТРОЙ)','ИНН','Телефон(ы)','Доп. контакты / реквизиты',
           'Источник','Достоверность','Рег.№ в СРО','Статус в СРО','КФ ВВ, ₽',
           'Дата вступления в СРО','Новичок (≤6 мес)',
           'Дозвонились?','Дата звонка','Кто звонил','Комментарий'])
hf = PatternFill('solid', fgColor='1F4E79'); thin = Side(style='thin', color='BFBFBF')
bd = Border(left=thin,right=thin,top=thin,bottom=thin)
for c in ws[1]:
    c.font=Font(bold=True,color='FFFFFF',size=11); c.fill=hf
    c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=bd

ok=PatternFill('solid',fgColor='E2EFDA'); mid=PatternFill('solid',fgColor='FFF2CC'); no=PatternFill('solid',fgColor='FCE4E4')

# сначала найденные, потом остальные
items = sorted(res.items(), key=lambda kv: (not kv[1]['phone'], kv[1]['name']))
n=0
for inn, r in items:
    n+=1; src = by_inn.get(inn, {})
    kf = src.get('kf_vv') or ''
    jd = joined.get(inn)
    ws.append([n, r['name'], inn, r['phone'], r['extra'], r['src'], r['conf'],
               src.get('reg',''), src.get('status',''), kf,
               jd, 'да' if jd and jd >= CUT6 else '', '', '', '', ''])
    row=ws.max_row
    fill = ok if r['phone'] and r['conf'].startswith('высокая') else (mid if r['phone'] else no)
    for col in range(1,17):
        cell=ws.cell(row=row,column=col); cell.border=bd
        cell.alignment=Alignment(vertical='top',wrap_text=(col in (2,4,5,6,7)))
        cell.fill=fill
        if col==4: cell.font=Font(bold=bool(r['phone']))
        if col==11 and jd: cell.number_format='DD.MM.YYYY'
        # Новичка видно с первого взгляда — по нему обзвон обычно приоритетный.
        if col==12 and jd and jd >= CUT6:
            cell.font=Font(bold=True, color='C00000'); cell.alignment=Alignment(horizontal='center')
    ws.cell(row=row,column=3).number_format='@'
for col,w in zip('ABCDEFGHIJKLMNOP',[5,44,14,34,52,42,26,10,16,14,17,15,14,13,16,40]):
    ws.column_dimensions[col].width=w
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type='list', formula1='"да, дозвонились,не отвечает,номер не тот,отказ,перезвонить"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f'M2:M{ws.max_row}')
ws.freeze_panes='A2'; ws.auto_filter.ref=f"A1:P{ws.max_row}"


# ---------------------------------------------------------------- лист «Новички»
nw = wb.create_sheet('Новички за 6 мес')
nw.append(['Дата вступления','Организация','ИНН','Телефон(ы)','Доп. контакты / реквизиты',
           'Достоверность','Тип','Дозвонились?','Дата звонка','Кто звонил','Комментарий'])
for c in nw[1]:
    c.font=Font(bold=True,color='FFFFFF',size=11); c.fill=hf
    c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=bd

newbies = sorted((r for r in recs
                  if r['status']=='Является членом'
                  and joined.get(r['inn']) and joined[r['inn']] >= CUT6),
                 key=lambda r: joined[r['inn']], reverse=True)
for r in newbies:
    v = res.get(r['inn'], {})
    nw.append([joined[r['inn']], v.get('name') or r.get('short') or r.get('full'), r['inn'],
               v.get('phone',''), v.get('extra',''), v.get('conf',''), r['typ'], '', '', '', ''])
    row = nw.max_row
    fill = ok if v.get('phone') else no
    for col in range(1, 12):
        cell = nw.cell(row=row, column=col); cell.border = bd; cell.fill = fill
        cell.alignment = Alignment(vertical='top', wrap_text=(col in (2,4,5,6,11)))
    nw.cell(row=row, column=1).number_format = 'DD.MM.YYYY'
    nw.cell(row=row, column=3).number_format = '@'
    nw.cell(row=row, column=4).font = Font(bold=bool(v.get('phone')))
for col,w in zip('ABCDEFGHIJK',[16,46,14,34,52,26,7,14,13,16,40]):
    nw.column_dimensions[col].width = w
nw.freeze_panes='A2'; nw.auto_filter.ref=f"A1:K{nw.max_row}"
dv2 = DataValidation(type='list', formula1='"да, дозвонились,не отвечает,номер не тот,отказ,перезвонить"', allow_blank=True)
nw.add_data_validation(dv2); dv2.add(f'H2:H{nw.max_row}')

found=sum(1 for v in res.values() if v['phone'])
act=sum(1 for r in recs if r['status']=='Является членом')
s=wb.create_sheet('Сводка')
for row in [
 ['Показатель','Значение'],
 ['Источник','Реестр НОСТРОЙ, СРО № 410 — Ассоциация «ПОС», Санкт-Петербург'],
 ['Всего записей в реестре', len(recs)],
 ['Из них действующих членов', act],
 ['Исключённых', len(recs)-act],
 ['','' ],
 ['Обработано в этой сессии', len(res)],
 ['Вступили в СРО за последние 6 месяцев', sum(1 for r in recs if r['status']=='Является членом' and joined.get(r['inn']) and joined[r['inn']] >= CUT6)],
 ['   из них с найденным телефоном', sum(1 for r in recs if r['status']=='Является членом' and joined.get(r['inn']) and joined[r['inn']] >= CUT6 and res.get(r['inn'],{}).get('phone'))],
 ['Откуда дата вступления','Столбец registry_registration_date первой выгрузки НОСТРОЙ — во второй он был удалён, подшит обратно по ИНН'],
 ['Телефон найден', found],
 ['Телефон не найден', len(res)-found],
 ['Как отбирали','Проверены ВСЕ действующие члены СРО, без исключений'],
 ['','' ],
 ['ВАЖНО: телефонов в самом реестре НЕТ','В выгрузке 37 столбцов: ИНН, ОГРН, наименование, статус, уровни ответственности, взносы. Полей телефона и e-mail нет вообще'],
 ['Как искали','Поисковая выдача — единственный доступный канал сети в этой среде'],
 ['Проверка','Телефон записан только там, где в источнике совпал ИНН'],
 ['ИП в реестре (29 шт.)','Проверены; телефонов ИП в открытых бизнес-реестрах нет ни у одного'],
 ['Если нужно добрать ещё','Прогнать программой contact-finder с ключом Checko API — она берёт контакты из платного источника'],
 ['Самый быстрый путь ко всем 972','СРО «ПОС» (Ассоциация «Петровское объединение строителей»), +7 812 335-36-86, office@sropos.ru — у СРО контакты всех своих членов'],
]:
    s.append(row)
for c in s[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=hf
s.column_dimensions['A'].width=42; s.column_dimensions['B'].width=100
for r in s.iter_rows(min_row=2): r[1].alignment=Alignment(wrap_text=True,vertical='top')

wb.save('/home/user/first_reposirtory/Телефоны_СРО410_НОСТРОЙ.xlsx')
print('found',found,'of',len(res))
