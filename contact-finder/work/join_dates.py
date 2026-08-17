"""Даты вступления в СРО из выгрузки реестра НОСТРОЙ.

Ключевой момент: компания может стоять в реестре дважды — исключена и
вступила заново (таких 5). Дата вступления — у ДЕЙСТВУЮЩЕЙ записи,
поэтому исключённые строки отбрасываются ДО построения словаря, а не
после. Иначе более старая дата исключённой записи затирает верную.

Источник — первая выгрузка реестра (ff3ce902): во второй колонок с
датами уже нет.
"""
import json

from openpyxl import load_workbook

SRC = ('/root/.claude/uploads/8b3c1f71-8d10-5171-ae76-8d4f758fc4f3/'
       'ff3ce902-reestr_nostroy_sro_410_20260811.xlsx')
INN, DATE_ISO, STATUS = 1, 4, 15

ws = load_workbook(SRC, read_only=True).active
act = [r for r in ws.iter_rows(min_row=2, values_only=True)
       if r[STATUS] == 'Является членом']

dates = {}
for r in act:
    inn = str(r[INN])
    assert inn not in dates, f'две действующие записи у ИНН {inn}'
    dates[inn] = r[DATE_ISO][:10]

assert len(dates) == 972, len(dates)
json.dump(dates, open('join_dates.json', 'w'), ensure_ascii=False, indent=1)
print(f'дат вступления: {len(dates)}  (действующих записей в реестре: {len(act)})')
print(f'диапазон: {min(dates.values())} .. {max(dates.values())}')
