"""Итоговая таблица: всё, где телефон найден хоть одним источником.

Объединяются два независимых набора:
  Checko       — платный агрегатор, прогнаны первые 486 компаний из 972;
  ручной пробив — веб-поиск по всем 972, телефон засчитан только если
                  на странице-источнике стоял тот же ИНН, что в реестре.
Совпадение номера в обоих наборах — самый сильный признак, что номер
живой, поэтому такие компании помечены отдельно.

Даты вступления — из join_dates.py (только действующие записи реестра).
"""
import datetime
import json
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from flags import flag
from parse_checko import parse, phones

TODAY = datetime.date(2026, 8, 17)
CUT6 = TODAY - datetime.timedelta(days=183)


def key(p):
    """Ключ для сравнения номеров: 8(812)… и +7 812 … — один номер."""
    d = re.sub(r'\D', '', p)
    return '7' + d[-10:] if len(d) == 11 and d[0] in '78' else d


rows, bad = parse('checko_raw.txt')
assert not bad
CK_DONE = {r['inn'] for r in rows}
ck = {r['inn']: phones(r['tail']) for r in rows if phones(r['tail'])}

manual = json.load(open('nostroy_results.json'))
mine = {k: [p.strip() for p in re.split(r'[;,]', v['phone']) if p.strip()]
        for k, v in manual.items() if v.get('phone')}

recs = json.load(open('nostroy.json'))
# Только действующие записи: у 5 компаний в реестре есть ещё и старая
# строка «Исключен». Если строить словарь по всем строкам, она затирает
# действующую, а при обходе строк компания попадает в таблицу дважды.
active_rows = [r for r in recs if r['status'] == 'Является членом']
by = {r['inn']: r for r in active_rows}
act = set(by)
assert len(by) == len(active_rows) == 972, (len(by), len(active_rows))
joined = {k: datetime.date.fromisoformat(v)
          for k, v in json.load(open('join_dates.json')).items()}

data = []
for inn in set(ck) | set(mine):
    assert inn in act, f'{inn} не действующий член'
    a, b = ck.get(inn, []), mine.get(inn, [])
    ka, kb = {key(p) for p in a}, {key(p) for p in b}
    merged = a + [p for p in b if key(p) not in ka]      # Checko вперёд, ручное дополняет
    if a and b:
        src = 'оба источника, номера совпали' if ka & kb else 'оба источника, номера разные'
    elif a:
        src = 'Checko'
    else:
        src = ('ручной пробив' if inn in CK_DONE
               else 'ручной пробив (Checko не прогонял)')
    r = by[inn]
    data.append({
        'inn': inn, 'name': r.get('short') or r.get('full') or '',
        'typ': r['typ'], 'jd': joined[inn], 'ph': merged, 'src': src,
        'notes': '; '.join(sorted({f'{p} — {flag(p)}' for p in merged if flag(p)})),
        'conf': manual.get(inn, {}).get('conf', ''),
    })
data.sort(key=lambda d: d['jd'], reverse=True)

HEAD = ['№ п/п', 'Организация (реестр НОСТРОЙ)', 'ИНН', 'Тип',
        'Дата вступления в СРО', 'Новичок (≤6 мес)', 'Телефон 1', 'Телефон 2',
        'Телефон 3', 'Остальные номера', 'Всего номеров', 'Источник данных',
        'Замечания по номерам', 'Дозвонились?', 'Дата звонка', 'Кто звонил',
        'Комментарий']
WID = [7, 52, 14, 7, 20, 17, 20, 20, 20, 38, 13, 30, 44, 20, 14, 16, 38]
WRAP = {2, 10, 12, 13, 17}

HFILL = PatternFill('solid', fgColor='1F4E78')
HFONT = Font(bold=True, color='FFFFFF', size=11)
BOTH = PatternFill('solid', fgColor='E2EFDA')      # подтверждено двумя источниками
WARN = PatternFill('solid', fgColor='FCE4D6')

wb = Workbook()
ws = wb.active
ws.title = 'Телефоны — все источники'


def head(s, hd, wd):
    s.append(hd)
    for i, w in enumerate(wd, 1):
        s.column_dimensions[get_column_letter(i)].width = w
    for c in s[1]:
        c.fill, c.font = HFILL, HFONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    s.row_dimensions[1].height = 34


head(ws, HEAD, WID)
for i, d in enumerate(data, 1):
    p = d['ph']
    ws.append([i, d['name'], d['inn'], d['typ'], d['jd'],
               'ДА' if d['jd'] >= CUT6 else '',
               p[0] if len(p) > 0 else '', p[1] if len(p) > 1 else '',
               p[2] if len(p) > 2 else '', '; '.join(p[3:]), len(p),
               d['src'], d['notes'], '', '', '', ''])
    row = ws[ws.max_row]
    for c in row:
        c.alignment = Alignment(vertical='top', wrap_text=c.column in WRAP)
    row[2].number_format = '@'
    row[2].alignment = Alignment(horizontal='left', vertical='top')
    row[4].number_format = 'DD.MM.YYYY'
    for i2 in (3, 4, 10):
        row[i2].alignment = Alignment(horizontal='center', vertical='top')
    if d['jd'] >= CUT6:
        row[5].font = Font(bold=True, color='C00000')
        row[5].alignment = Alignment(horizontal='center', vertical='top')
    if d['src'] == 'оба источника, номера совпали':
        row[11].fill = BOTH
    if d['notes']:
        row[12].fill = WARN

dv = DataValidation(type='list', allow_blank=True,
                    formula1='"да, дозвонились,не отвечает,номер не тот,отказ,перезвонить"')
ws.add_data_validation(dv)
dv.add(f'N2:N{ws.max_row}')
ws.freeze_panes = 'C2'
ws.auto_filter.ref = f'A1:Q{ws.max_row}'

# --- Новички: у них решение о подрядчиках чаще ещё не принято ---
nb = [d for d in data if d['jd'] >= CUT6]
s2 = wb.create_sheet('Новички за 6 мес')
head(s2, ['№', 'Организация', 'ИНН', 'Тип', 'Дата вступления', 'Телефон(ы)',
          'Всего', 'Источник данных', 'Дозвонились?', 'Дата звонка',
          'Кто звонил', 'Комментарий'],
     [6, 50, 14, 7, 18, 62, 8, 30, 20, 14, 16, 38])
for i, d in enumerate(nb, 1):
    s2.append([i, d['name'], d['inn'], d['typ'], d['jd'], '; '.join(d['ph']),
               len(d['ph']), d['src'], '', '', '', ''])
    row = s2[s2.max_row]
    for c in row:
        c.alignment = Alignment(vertical='top', wrap_text=c.column in (2, 6, 8, 12))
    row[2].number_format = '@'
    row[2].alignment = Alignment(horizontal='left', vertical='top')
    row[4].number_format = 'DD.MM.YYYY'
    for i2 in (3, 4, 6):
        row[i2].alignment = Alignment(horizontal='center', vertical='top')
dv2 = DataValidation(type='list', allow_blank=True,
                     formula1='"да, дозвонились,не отвечает,номер не тот,отказ,перезвонить"')
s2.add_data_validation(dv2)
dv2.add(f'I2:I{s2.max_row}')
s2.freeze_panes = 'C2'
s2.auto_filter.ref = f'A1:L{s2.max_row}'

# --- Без телефона: чтобы было видно, что осталось закрыть ---
nof = sorted((r for r in active_rows
              if r['inn'] not in ck and r['inn'] not in mine),
             key=lambda r: joined[r['inn']], reverse=True)
assert len(nof) + len(data) == 972, (len(nof), len(data))
s3 = wb.create_sheet('Телефон не найден')
head(s3, ['№', 'Организация', 'ИНН', 'Тип', 'Дата вступления',
          'Checko прогонял?', 'Что известно по ручному пробиву'],
     [6, 52, 14, 7, 18, 20, 56])
for i, r in enumerate(nof, 1):
    s3.append([i, r.get('short') or r.get('full') or '', r['inn'], r['typ'],
               joined[r['inn']],
               'да, телефона нет' if r['inn'] in CK_DONE else 'ещё не прогонял',
               manual.get(r['inn'], {}).get('conf', '')])
    row = s3[s3.max_row]
    for c in row:
        c.alignment = Alignment(vertical='top', wrap_text=c.column in (2, 7))
    row[2].number_format = '@'
    row[2].alignment = Alignment(horizontal='left', vertical='top')
    row[4].number_format = 'DD.MM.YYYY'
    for i2 in (3, 4, 5):
        row[i2].alignment = Alignment(horizontal='center', vertical='top')
s3.freeze_panes = 'C2'
s3.auto_filter.ref = f'A1:G{s3.max_row}'

# --- Сводка ---
tot = sum(len(d['ph']) for d in data)
uniq = len({key(p) for d in data for p in d['ph']})
flg = sum(1 for d in data for p in d['ph'] if flag(p))
cnt = lambda s: sum(1 for d in data if d['src'] == s)
sm = wb.create_sheet('Сводка')
sm.column_dimensions['A'].width = 60
sm.column_dimensions['B'].width = 76
BLOCK = [
    ('ОХВАТ', ''),
    ('Действующих членов СРО № 410', '972'),
    ('Из них с найденным телефоном', f'{len(data)}  ({len(data) * 100 // 972}%)'),
    ('Телефон не найден', f'{972 - len(data)}  (отдельный лист)'),
    ('Всего номеров / уникальных', f'{tot} / {uniq}'),
    ('', ''),
    ('ОТКУДА НОМЕРА', ''),
    ('Оба источника, номера совпали',
     f'{cnt("оба источника, номера совпали")} — самые надёжные, залиты зелёным'),
    ('Оба источника, номера разные',
     f'{cnt("оба источника, номера разные")} — оба могут быть верны (офис и мобильный)'),
    ('Только Checko', str(cnt('Checko'))),
    ('Только ручной пробив', str(cnt('ручной пробив'))),
    ('Только ручной (Checko эти не прогонял)',
     str(cnt('ручной пробив (Checko не прогонял)'))),
    ('', ''),
    ('ЧТО ЕЩЁ НЕ СДЕЛАНО', ''),
    ('Checko прогнан частично',
     f'{len(CK_DONE)} компаний из 972 — вторую половину надо догнать, '
     f'охват вырастет'),
    ('', ''),
    ('НОВИЧКИ', ''),
    ('Отсечка', f'вступил {CUT6.strftime("%d.%m.%Y")} или позже'),
    ('Новичков с телефоном', f'{len(nb)} (отдельный лист)'),
    ('', ''),
    ('КАЧЕСТВО', ''),
    ('Помечено браком', f'{flg} номеров — несуществующие коды и заглушки'),
    ('Привязка к компании',
     'номер засчитан только там, где источник указывает тот же ИНН, что в реестре'),
    ('Номера НЕ прозвонены',
     'проверена принадлежность, но не то, что номер сегодня отвечает'),
    ('', ''),
    ('ИСТОЧНИК ДАТ', ''),
    ('Даты вступления',
     'выгрузка реестра НОСТРОЙ СРО № 410 от 11.08.2026, только действующие записи'),
    ('Повторное вступление',
     '5 компаний вступали дважды — взята дата действующей записи, не исключённой'),
]
for k, v in BLOCK:
    sm.append([k, v])
    c = sm[sm.max_row][0]
    c.font = (Font(bold=True, color='1F4E78', size=12) if v == '' and k
              else Font(bold=True))
    sm[sm.max_row][1].alignment = Alignment(wrap_text=True, vertical='top')

OUT = '/home/user/first_reposirtory/Телефоны_СРО410_ИТОГ.xlsx'
wb.save(OUT)
print(f'сохранено: {OUT}')
for s in wb.worksheets:
    print(f'  «{s.title}» — {s.max_row - 1} строк')
print(f'номеров {tot}, уникальных {uniq}, помечено браком {flg}')
