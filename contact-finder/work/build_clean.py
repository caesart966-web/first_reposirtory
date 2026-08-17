"""Чистая таблица по выгрузке Checko: название, ИНН, телефоны, дата вступления.

Там, где Checko телефона не дал, но он нашёлся ручным пробивом, номер
подставляется — иначе компания попала бы в «без телефона», хотя он есть.
Названия берутся из реестра НОСТРОЙ по ИНН: в выгрузке Checko они
обрезаны по ширине колонки.
"""
import datetime
import json
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from flags import flag
from parse_checko import parse, phones

rows, bad = parse('checko_raw.txt')
assert not bad and len(rows) == 486, (len(rows), bad)

recs = json.load(open('nostroy.json'))
by = {r['inn']: r for r in recs if r['status'] == 'Является членом'}
assert len(by) == 972
joined = {k: datetime.date.fromisoformat(v)
          for k, v in json.load(open('join_dates.json')).items()}
manual = json.load(open('nostroy_results.json'))


def key(p):
    d = re.sub(r'\D', '', p)
    return '7' + d[-10:] if len(d) == 11 and d[0] in '78' else d


data = []
for r in rows:
    ck = phones(r['tail'])
    mn = [p.strip() for p in re.split(r'[;,]', manual.get(r['inn'], {}).get('phone') or '')
          if p.strip()]
    merged = ck + [p for p in mn if key(p) not in {key(q) for q in ck}]
    e = by[r['inn']]
    data.append({
        'name': e.get('short') or e.get('full') or '',
        'inn': r['inn'],
        'ph': merged,
        'jd': joined[r['inn']],
        'src': 'Checko' if ck else ('ручной пробив' if mn else ''),
        'note': '; '.join(sorted({f'{p} — {flag(p)}' for p in merged if flag(p)})),
    })
data.sort(key=lambda d: d['jd'], reverse=True)
have = [d for d in data if d['ph']]
none = [d for d in data if not d['ph']]

# ---------------------------------------------------------------- оформление
NAVY = '1F3864'
HEAD_F = PatternFill('solid', fgColor=NAVY)
HEAD_T = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
ZEBRA = PatternFill('solid', fgColor='F4F6FA')
NEWF = Font(bold=True, color='C00000', size=11)
GREY = Font(color='909090', italic=True)
hair = Side(style='thin', color='D6DCE4')
BD = Border(left=hair, right=hair, top=hair, bottom=hair)
TODAY = datetime.date(2026, 8, 17)
CUT6 = TODAY - datetime.timedelta(days=183)

wb = Workbook()


def build(ws, items, cols, widths, title):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(bold=True, size=14, color=NAVY)
    t.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 30
    ws.append(cols)
    for c in ws[2]:
        c.fill, c.font, c.border = HEAD_F, HEAD_T, BD
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 32
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


# ---------------------------------------------------------------- лист 1
ws = wb.active
ws.title = 'Телефоны'
build(ws, have,
      ['№', 'Название компании', 'ИНН', 'Номера телефонов', 'Дата вступления в СРО'],
      [6, 56, 16, 26, 21],
      f'Члены СРО № 410 «ПОС» с найденными телефонами — {len(have)} компаний '
      f'из {len(data)} обработанных')
for i, d in enumerate(have, 1):
    # каждый номер на своей строке — так его удобно набирать и копировать
    ws.append([i, d['name'], d['inn'], '\n'.join(d['ph']), d['jd']])
    row = ws[ws.max_row]
    ws.row_dimensions[ws.max_row].height = max(16, 15 * len(d['ph']))
    for c in row:
        c.border = BD
        if i % 2 == 0:
            c.fill = ZEBRA
    row[0].alignment = Alignment(horizontal='center', vertical='center')
    row[1].alignment = Alignment(vertical='center', wrap_text=True)
    row[2].alignment = Alignment(horizontal='center', vertical='center')
    row[2].number_format = '@'
    row[3].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row[3].font = Font(name='Consolas', size=11)
    row[4].alignment = Alignment(horizontal='center', vertical='center')
    row[4].number_format = 'DD.MM.YYYY'
    if d['jd'] >= CUT6:
        row[4].font = NEWF
ws.freeze_panes = 'A3'
ws.auto_filter.ref = f'A2:E{ws.max_row}'

# ---------------------------------------------------------------- лист 2
w2 = wb.create_sheet('Без телефона')
build(w2, none, ['№', 'Название компании', 'ИНН', 'Дата вступления в СРО'],
      [6, 56, 16, 21],
      f'Телефон не найден ни Checko, ни ручным пробивом — {len(none)} компаний')
for i, d in enumerate(none, 1):
    w2.append([i, d['name'], d['inn'], d['jd']])
    row = w2[w2.max_row]
    for c in row:
        c.border = BD
        c.font = GREY
        if i % 2 == 0:
            c.fill = ZEBRA
    row[0].alignment = Alignment(horizontal='center', vertical='center')
    row[1].alignment = Alignment(vertical='center', wrap_text=True)
    row[2].alignment = Alignment(horizontal='center', vertical='center')
    row[2].number_format = '@'
    row[3].alignment = Alignment(horizontal='center', vertical='center')
    row[3].number_format = 'DD.MM.YYYY'
w2.freeze_panes = 'A3'
w2.auto_filter.ref = f'A2:D{w2.max_row}'

# ---------------------------------------------------------------- лист 3
w3 = wb.create_sheet('Проверить номер')
flagged = [(d, p) for d in data for p in d['ph'] if flag(p)]
build(w3, flagged, ['№', 'Название компании', 'ИНН', 'Номер', 'Что не так'],
      [6, 50, 16, 22, 30],
      f'Номера, которые почти наверняка не сработают — {len(flagged)} шт.')
for i, (d, p) in enumerate(flagged, 1):
    w3.append([i, d['name'], d['inn'], p, flag(p)])
    row = w3[w3.max_row]
    for c in row:
        c.border = BD
        c.alignment = Alignment(vertical='center', wrap_text=c.column == 2)
    row[2].number_format = '@'
    row[2].alignment = Alignment(horizontal='center', vertical='center')
w3.freeze_panes = 'A3'

OUT = '/home/user/first_reposirtory/Телефоны_486_компаний.xlsx'
wb.save(OUT)
tot = sum(len(d['ph']) for d in have)
print(f'{OUT}')
print(f'  с телефоном : {len(have)}  (номеров {tot}, уникальных '
      f'{len({key(p) for d in have for p in d["ph"]})})')
print(f'  без телефона: {len(none)}')
print(f'  подставлено из ручного пробива: {sum(1 for d in have if d["src"] == "ручной пробив")}')
print(f'  помечено браком: {len(flagged)}')
