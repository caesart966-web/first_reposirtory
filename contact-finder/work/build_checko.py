"""Сборка таблицы по выгрузке Checko.

Названия берутся не из выгрузки (там они обрезаны по ширине колонки),
а из реестра НОСТРОЙ по ИНН — иначе в таблице окажутся огрызки вроде
«АО «Балтийская инжиниринговая компания» без закрывающей кавычки.
Даты вступления — из первой версии реестра: во второй выгрузке
колонки с датами уже не было.
"""
import datetime
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from flags import flag
from parse_checko import parse, phones

TODAY = datetime.date(2026, 8, 17)
CUT6 = TODAY - datetime.timedelta(days=183)

rows, bad = parse('checko_raw.txt')
assert not bad, bad
recs = json.load(open('nostroy.json'))
# Только действующие записи — исключённая строка затирала действующую.
by = {r['inn']: r for r in recs if r['status'] == 'Является членом'}
assert len(by) == 972, len(by)
mine = json.load(open('nostroy_results.json'))
joined = {k: datetime.date.fromisoformat(v[:10])
          for k, v in json.load(open('join_dates.json')).items()}

TOTAL = rows[0]['total']


def nm(inn):
    r = by[inn]
    return r.get('short') or r.get('full') or ''


data = []
for r in rows:
    ph = phones(r['tail'])
    notes = sorted({f'{p} — {flag(p)}' for p in ph if flag(p)})
    data.append({
        'n': r['n'], 'inn': r['inn'], 'name': nm(r['inn']),
        'ph': ph, 'jd': joined.get(r['inn']),
        'notes': '; '.join(notes),
        'mine': (mine.get(r['inn']) or {}).get('phone') or '',
    })
# сортировка по дате вступления, свежие сверху — их обзванивают в первую очередь
data.sort(key=lambda d: (d['jd'] or datetime.date(1900, 1, 1)), reverse=True)

HEAD = ['№ п/п', 'Организация (реестр НОСТРОЙ)', 'ИНН', 'Дата вступления в СРО',
        'Новичок (≤6 мес)', 'Телефон 1', 'Телефон 2', 'Телефон 3',
        'Остальные номера', 'Всего номеров', 'Замечания по номерам',
        '№ в выгрузке Checko', 'Дозвонились?', 'Дата звонка', 'Кто звонил',
        'Комментарий']
WID = [7, 52, 14, 20, 17, 20, 20, 20, 40, 14, 46, 20, 20, 14, 16, 40]

HFILL = PatternFill('solid', fgColor='1F4E78')
HFONT = Font(bold=True, color='FFFFFF', size=11)
NOPH = PatternFill('solid', fgColor='F2F2F2')
WARN = PatternFill('solid', fgColor='FCE4D6')

wb = Workbook()
ws = wb.active
ws.title = f'Checko 1-{len(rows)}'
ws.append(HEAD)
for i, w in enumerate(WID, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for c in ws[1]:
    c.fill, c.font = HFILL, HFONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[1].height = 34

for i, d in enumerate(data, 1):
    p = d['ph']
    ws.append([i, d['name'], d['inn'], d['jd'],
               'ДА' if d['jd'] and d['jd'] >= CUT6 else '',
               p[0] if len(p) > 0 else '', p[1] if len(p) > 1 else '',
               p[2] if len(p) > 2 else '', '; '.join(p[3:]), len(p),
               d['notes'], d['n'], '', '', '', ''])
    row = ws[ws.max_row]
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=cell.column in (2, 9, 11, 16))
    row[2].alignment = Alignment(horizontal='left', vertical='top')
    row[2].number_format = '@'                       # ИНН — текст, ведущий ноль не съесть
    if d['jd']:
        row[3].number_format = 'DD.MM.YYYY'
        row[3].alignment = Alignment(horizontal='center', vertical='top')
    if d['jd'] and d['jd'] >= CUT6:
        row[4].font = Font(bold=True, color='C00000')
        row[4].alignment = Alignment(horizontal='center', vertical='top')
    row[9].alignment = Alignment(horizontal='center', vertical='top')
    row[11].alignment = Alignment(horizontal='center', vertical='top')
    if not p:
        for cell in row:
            cell.fill = NOPH
    if d['notes']:
        row[10].fill = WARN

dv = DataValidation(type='list', allow_blank=True,
                    formula1='"да, дозвонились,не отвечает,номер не тот,отказ,перезвонить"')
ws.add_data_validation(dv)
dv.add(f'M2:M{ws.max_row}')
ws.freeze_panes = 'C2'
ws.auto_filter.ref = f'A1:P{ws.max_row}'

def sheet(title, head, wid, body):
    s = wb.create_sheet(title)
    s.append(head)
    for i, w in enumerate(wid, 1):
        s.column_dimensions[get_column_letter(i)].width = w
    for c in s[1]:
        c.fill, c.font = HFILL, HFONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    s.row_dimensions[1].height = 34
    for r in body:
        s.append(r)
        row = s[s.max_row]
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=cell.column == 2)
        row[2].number_format = '@'
        row[2].alignment = Alignment(horizontal='left', vertical='top')
        if isinstance(row[3].value, datetime.date):
            row[3].number_format = 'DD.MM.YYYY'
            row[3].alignment = Alignment(horizontal='center', vertical='top')
    s.freeze_panes = 'C2'
    s.auto_filter.ref = f'A1:{get_column_letter(len(head))}{s.max_row}'
    return s


CALL = ['Дозвонились?', 'Дата звонка', 'Кто звонил', 'Комментарий']
CW = [20, 14, 16, 40]

newb = [d for d in data if d['jd'] and d['jd'] >= CUT6]
s2 = sheet('Новички за 6 мес',
           ['№', 'Организация', 'ИНН', 'Дата вступления', 'Телефон(ы)',
            'Всего номеров', 'Замечания по номерам'] + CALL,
           [6, 52, 14, 18, 60, 14, 40] + CW,
           [[i, d['name'], d['inn'], d['jd'], '; '.join(d['ph']) or '—',
             len(d['ph']), d['notes'], '', '', '', '']
            for i, d in enumerate(newb, 1)])
dv2 = DataValidation(type='list', allow_blank=True,
                     formula1='"да, дозвонились,не отвечает,номер не тот,отказ,перезвонить"')
s2.add_data_validation(dv2)
dv2.add(f'H2:H{s2.max_row}')

# Датасеты дополняют друг друга: тут телефон нашёлся вручную там,
# где у Checko пусто. Терять эти компании при переходе на Checko нельзя.
only = [d for d in data if not d['ph'] and d['mine']]
sheet('Нет у Checko — есть у меня',
      ['№', 'Организация', 'ИНН', 'Дата вступления', 'Телефон(ы) (ручной пробив)',
       'Достоверность'] + CALL,
      [6, 52, 14, 18, 60, 34] + CW,
      [[i, d['name'], d['inn'], d['jd'], d['mine'],
        mine[d['inn']].get('conf', ''), '', '', '', '']
       for i, d in enumerate(only, 1)])

nph = sum(1 for d in data if d['ph'])
tot = sum(len(d['ph']) for d in data)
uniq = len({''.join(ch for ch in p if ch.isdigit()) for d in data for p in d['ph']})
flagged = sum(1 for d in data for p in d['ph'] if flag(p))

sm = wb.create_sheet('Сводка')
sm.column_dimensions['A'].width = 62
sm.column_dimensions['B'].width = 74
for k, v in [
    ('ЧТО ЭТО', ''),
    ('Источник', 'выгрузка Checko по реестру СРО № 410 (Ассоциация «ПОС», СПб)'),
    ('Обработано в этой выгрузке', f'{len(data)} компаний из {TOTAL} действующих членов'),
    ('ВАЖНО: выгрузка неполная',
     f'остальные {TOTAL - len(data)} компаний ещё не прогнаны — таблицу нужно будет дополнить'),
    ('Дата сборки', TODAY.strftime('%d.%m.%Y')),
    ('', ''),
    ('НОМЕРА', ''),
    ('Компаний с телефоном', f'{nph} из {len(data)} ({nph * 100 // len(data)}%)'),
    ('Компаний без телефона', f'{len(data) - nph} (строки залиты серым)'),
    ('Всего номеров', str(tot)),
    ('Уникальных номеров', str(uniq)),
    ('Помечено как брак', f'{flagged} (колонка «Замечания по номерам», залита оранжевым)'),
    ('', ''),
    ('НОВИЧКИ', ''),
    ('Отсечка «≤6 мес»', f'вступил {CUT6.strftime("%d.%m.%Y")} или позже'),
    ('Новичков в этой выгрузке', f'{len(newb)} (отдельный лист)'),
    ('', ''),
    ('СВЕРКА С РУЧНЫМ ПРОБИВОМ', ''),
    ('Нет у Checko, но нашлось вручную', f'{len(only)} компаний (отдельный лист)'),
    ('', ''),
    ('ЧЕГО ЭТА ТАБЛИЦА НЕ ГАРАНТИРУЕТ', ''),
    ('Номера не прозвонены',
     'проверена принадлежность компании, но не то, что номер сейчас отвечает'),
    ('Привязка к ИНН', 'названия подставлены из реестра НОСТРОЙ по ИНН, не из выгрузки'),
]:
    sm.append([k, v])
    c = sm[sm.max_row][0]
    if v == '' and k:
        c.font = Font(bold=True, color='1F4E78', size=12)
    else:
        c.font = Font(bold=True)
    sm[sm.max_row][1].alignment = Alignment(wrap_text=True, vertical='top')
sm['B4'].font = Font(bold=True, color='C00000')

OUT = '/home/user/first_reposirtory/Телефоны_СРО410_Checko.xlsx'
wb.save(OUT)
print(f'сохранено: {OUT}')
print(f'  лист 1 «{ws.title}»          : {ws.max_row - 1} строк')
print(f'  лист 2 «Новички за 6 мес»    : {s2.max_row - 1} строк')
print(f'  компаний с телефоном         : {nph} | номеров {tot} | уникальных {uniq}')
print(f'  помечено браком              : {flagged}')
print(f'  есть у меня, нет у Checko    : {len(only)}')
