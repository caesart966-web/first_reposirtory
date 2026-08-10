"""Экспорт результатов в XLSX.

Листы:
- «Результат» — все компании с полным набором колонок по ТЗ;
- «Найдено полностью» / «Частично» / «Не найдено» — те же колонки, фильтр по полноте;
- «Источники» — каждый контакт с URL источника (для проверки, ТЗ п. 8);
- «Недоступные источники» — что было закрыто/отключено в прогоне.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.models import Company, ContactType, MergedContact
from storage.db import Database
from validators.phone import format_phone_display, phone_type

HEADERS = [
    "ИНН", "ОГРН", "Полное наименование", "Краткое наименование",
    "Статус", "Дата регистрации", "ОКВЭД основной", "Регион", "Юр. адрес",
    "Сайт", "E-mail основной", "E-mail все (через ;)",
    "Телефон основной", "Телефоны все", "Руководитель ФИО", "Должность",
    "Соцсети/мессенджеры", "Источник каждого контакта",
    "Оценка достоверности (0-100)", "Дата сбора",
]

_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_COL_WIDTHS = [14, 15, 45, 30, 14, 14, 12, 20, 45, 28, 28, 40, 18, 34, 28, 24, 40, 55, 12, 18]


def completeness_of(merged: list[MergedContact]) -> str:
    """full — есть e-mail и телефон; partial — что-то есть; none — ничего."""
    has_email = any(m.type == ContactType.EMAIL for m in merged)
    has_phone = any(m.type == ContactType.PHONE for m in merged)
    if has_email and has_phone:
        return "full"
    if merged:
        return "partial"
    return "none"


def _primary(merged: list[MergedContact], ctype: ContactType) -> Optional[MergedContact]:
    typed = [m for m in merged if m.type == ctype]
    if not typed:
        return None
    primaries = [m for m in typed if m.is_primary]
    return primaries[0] if primaries else max(typed, key=lambda m: m.confidence)


def _company_row(company: Company, merged: list[MergedContact],
                 collected_at: Optional[str]) -> list[Any]:
    emails = [m for m in merged if m.type == ContactType.EMAIL]
    phones = [m for m in merged if m.type == ContactType.PHONE]
    socials = [m for m in merged if m.type in (ContactType.SOCIAL, ContactType.MESSENGER)]

    primary_email = _primary(merged, ContactType.EMAIL)
    primary_phone = _primary(merged, ContactType.PHONE)

    phones_all = "; ".join(
        f"{format_phone_display(m.value)} ({phone_type(m.value)})" for m in phones
    )
    source_notes: list[str] = []
    for m in merged:
        label = m.value if m.type != ContactType.PHONE else format_phone_display(m.value)
        source_notes.append(f"{label} — {', '.join(m.sources)}")

    site = company.website or (
        f"{company.website_candidate} (не подтверждён)" if company.website_candidate else ""
    )
    confidences = [m.confidence for m in (primary_email, primary_phone) if m]
    if not confidences:
        confidences = [m.confidence for m in merged]
    total_confidence = max(confidences) if confidences else 0

    return [
        company.inn or "", company.ogrn or "",
        company.name_full or "", company.name_short or "",
        company.entity_status or "неизвестно",
        company.reg_date or "", company.okved or "",
        company.region or "", company.address or "",
        site,
        primary_email.value if primary_email else "",
        "; ".join(m.value for m in emails),
        format_phone_display(primary_phone.value) if primary_phone else "",
        phones_all,
        company.director_name or "", company.director_position or "",
        "; ".join(m.value for m in socials),
        "; ".join(source_notes),
        total_confidence,
        collected_at or "",
    ]


def _style_sheet(ws: Worksheet, n_cols: int) -> None:
    for idx, cell in enumerate(ws[1][:n_cols]):
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for i, width in enumerate(_COL_WIDTHS[:n_cols], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def export_results(db: Database, output_path: str | Path) -> dict[str, int]:
    """Собрать итоговый XLSX из базы. Возвращает счётчики по листам."""
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "Результат"
    ws_full = wb.create_sheet("Найдено полностью")
    ws_partial = wb.create_sheet("Частично")
    ws_none = wb.create_sheet("Не найдено")
    ws_sources = wb.create_sheet("Источники")

    for ws in (ws_all, ws_full, ws_partial, ws_none):
        ws.append(HEADERS)

    ws_sources.append([
        "ИНН", "Компания", "Тип контакта", "Значение", "Источник",
        "URL источника", "Уровень", "Достоверность", "Дата сбора",
    ])

    counts = {"full": 0, "partial": 0, "none": 0, "total": 0}

    for company in db.fetch_all_companies():
        merged = db.merged_contacts(company.key)
        row_meta = db.company_row(company.key)
        collected_at = row_meta["processed_at"] if row_meta else None
        row = _company_row(company, merged, collected_at)
        ws_all.append(row)
        bucket = completeness_of(merged)
        counts[bucket] += 1
        counts["total"] += 1
        {"full": ws_full, "partial": ws_partial, "none": ws_none}[bucket].append(row)

        for m in merged:
            for origin in m.origins:
                ws_sources.append([
                    company.inn or "", company.display_name,
                    m.type.value, m.value, origin.source,
                    origin.source_url or "", origin.level.value,
                    m.confidence, origin.collected_at,
                ])

    for ws in (ws_all, ws_full, ws_partial, ws_none):
        _style_sheet(ws, len(HEADERS))
    _style_sheet(ws_sources, 9)
    for i, width in enumerate([14, 40, 14, 34, 16, 70, 10, 14, 18], start=1):
        ws_sources.column_dimensions[get_column_letter(i)].width = width

    unavailable = db.unavailable_sources()
    if unavailable:
        ws_un = wb.create_sheet("Недоступные источники")
        ws_un.append(["Источник", "Примечание"])
        for state in unavailable:
            ws_un.append([state["source"], state.get("note") or ""])
        _style_sheet(ws_un, 2)
        ws_un.column_dimensions["A"].width = 24
        ws_un.column_dimensions["B"].width = 70

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return counts
