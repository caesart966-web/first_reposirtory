"""Чтение входных файлов (XLSX/CSV/TXT), автоопределение колонок,
валидация ИНН/ОГРН, отбраковка мусора и дедупликация по ИНН.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from core.models import Company, company_key_for
from validators.inn_ogrn import (
    clean_digits,
    looks_like_inn,
    looks_like_ogrn,
    validate_inn,
    validate_ogrn,
)

# Известные названия колонок (нормализуются к lower без пробелов/подчёркиваний)
_HEADER_ALIASES: dict[str, set[str]] = {
    "inn": {"инн", "inn", "tin", "иннкомпании", "иннорганизации"},
    "ogrn": {"огрн", "ogrn", "огрнип"},
    "name": {"наименование", "название", "компания", "организация", "name",
             "company", "фирма", "юрлицо", "полноенаименование",
             "краткоенаименование", "контрагент"},
    "website": {"сайт", "site", "website", "домен", "domain", "url", "вебсайт",
                "webсайт", "адрессайта"},
    "region": {"регион", "область", "город", "субъект", "region", "city",
               "субъектрф"},
}

_URL_RE = re.compile(r"^(https?://)?[\wа-яё\-]+(\.[\wа-яё\-]+)+(/.*)?$", re.IGNORECASE)


@dataclass
class InputError:
    """Отбракованная строка входного файла."""

    row_number: int
    value: str
    reason: str


@dataclass
class InputResult:
    companies: list[Company] = field(default_factory=list)
    errors: list[InputError] = field(default_factory=list)
    duplicates: int = 0


def _normalize_header(cell: object) -> str:
    return re.sub(r"[\s_\-.:]+", "", str(cell or "").strip().lower())


def _read_rows(path: Path) -> list[list[str]]:
    """Прочитать файл в таблицу строк независимо от формата."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = [
            ["" if c is None else str(c).strip() for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
        wb.close()
        return rows
    if suffix == ".csv":
        raw = path.read_bytes()
        text = None
        for enc in ("utf-8-sig", "utf-8", "cp1251"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            text = raw.decode("utf-8", errors="replace")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.reader(io.StringIO(text), dialect)
        return [[c.strip() for c in row] for row in reader]
    # TXT: список значений построчно (обычно ИНН)
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    return [[line.strip()] for line in lines if line.strip()]


def _detect_columns(rows: list[list[str]]) -> tuple[dict[str, int], bool]:
    """Определить, какая колонка что содержит.

    Возвращает (маппинг поле->индекс, есть_ли_строка_заголовка).
    """
    if not rows:
        return {}, False

    mapping: dict[str, int] = {}
    header = rows[0]
    for idx, cell in enumerate(header):
        norm = _normalize_header(cell)
        for fieldname, aliases in _HEADER_ALIASES.items():
            if norm in aliases and fieldname not in mapping:
                mapping[fieldname] = idx
    if mapping:
        return mapping, True

    # Заголовка нет — определяем по содержимому на выборке до 200 строк
    sample = rows[:200]
    n_cols = max(len(r) for r in sample)
    scores: dict[int, dict[str, int]] = {i: {"inn": 0, "ogrn": 0, "url": 0, "text": 0, "total": 0}
                                         for i in range(n_cols)}
    for row in sample:
        for i in range(n_cols):
            value = row[i].strip() if i < len(row) else ""
            if not value:
                continue
            s = scores[i]
            s["total"] += 1
            if looks_like_inn(value):
                s["inn"] += 1
            elif looks_like_ogrn(value):
                s["ogrn"] += 1
            elif _URL_RE.match(value) and "." in value and "@" not in value:
                s["url"] += 1
            elif re.search(r"[а-яёa-z]{3,}", value, re.IGNORECASE):
                s["text"] += 1

    def best_col(kind: str, threshold: float = 0.5) -> Optional[int]:
        candidates = [
            (s[kind] / s["total"], i) for i, s in scores.items()
            if s["total"] > 0 and i not in mapping.values()
        ]
        candidates = [(ratio, i) for ratio, i in candidates if ratio >= threshold]
        return max(candidates)[1] if candidates else None

    for kind, fieldname in (("inn", "inn"), ("ogrn", "ogrn"), ("url", "website")):
        col = best_col(kind)
        if col is not None:
            mapping[fieldname] = col
    text_col = best_col("text", threshold=0.4)
    if text_col is not None:
        mapping["name"] = text_col
    return mapping, False


def _clean_website(value: str) -> Optional[str]:
    value = (value or "").strip().rstrip("/")
    if not value or "@" in value:
        return None
    if not _URL_RE.match(value):
        return None
    if not value.lower().startswith(("http://", "https://")):
        value = "https://" + value
    return value


def read_companies(path: str | Path) -> InputResult:
    """Прочитать входной файл, отбраковать мусор, дедуплицировать по ИНН."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Входной файл не найден: {path}")

    rows = _read_rows(path)
    mapping, has_header = _detect_columns(rows)
    data_rows = rows[1:] if has_header else rows

    result = InputResult()
    seen_keys: set[str] = set()

    for offset, row in enumerate(data_rows):
        row_number = offset + (2 if has_header else 1)
        raw_joined = " | ".join(c for c in row if c)
        if not raw_joined:
            continue

        def cell(fieldname: str) -> str:
            idx = mapping.get(fieldname)
            if idx is None or idx >= len(row):
                return ""
            return row[idx].strip()

        inn_raw = cell("inn")
        ogrn_raw = cell("ogrn")
        name = cell("name")
        website_raw = cell("website")
        region = cell("region")

        # Если колонки не размечены (одноколоночный TXT) — пробуем понять значение
        if not mapping:
            value = row[0].strip()
            if looks_like_inn(value):
                inn_raw = value
            elif looks_like_ogrn(value):
                ogrn_raw = value
            elif _URL_RE.match(value) and "." in value:
                website_raw = value
            else:
                name = value

        inn: Optional[str] = None
        if inn_raw:
            ok, detail = validate_inn(inn_raw)
            if not ok:
                result.errors.append(InputError(row_number, raw_joined, detail))
                continue
            inn = clean_digits(inn_raw)

        ogrn: Optional[str] = None
        if ogrn_raw:
            ok, detail = validate_ogrn(ogrn_raw)
            if not ok:
                # ОГРН с ошибкой не бракует строку целиком, если есть другой идентификатор
                if not (inn or name or website_raw):
                    result.errors.append(InputError(row_number, raw_joined, detail))
                    continue
            else:
                ogrn = clean_digits(ogrn_raw)

        website = _clean_website(website_raw)

        if not any((inn, ogrn, name, website)):
            result.errors.append(InputError(
                row_number, raw_joined,
                "не найдено ни одного идентификатора (ИНН/ОГРН/название/сайт)",
            ))
            continue

        key = company_key_for(inn, name, region, website)
        if key in seen_keys:
            result.duplicates += 1
            continue
        seen_keys.add(key)

        result.companies.append(Company(
            key=key, inn=inn, ogrn=ogrn,
            name_full=name or None, region=region or None, website=website,
            raw={"row": row_number, "source_file": path.name, "name": name or None},
        ))

    return result


def write_errors_xlsx(errors: list[InputError], path: str | Path) -> None:
    """Сохранить отбракованные строки в errors.xlsx с причиной."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Ошибки входа"
    ws.append(["Строка входного файла", "Значение", "Причина отбраковки"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for err in errors:
        ws.append([err.row_number, err.value, err.reason])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 50
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
