#!/usr/bin/env python3
"""
Проверка контрагентов на членство в СРО по выгрузкам контрактов tenderguru.ru.

Пайплайн:
  extract  — разобрать выгрузки, свести поставщиков в реестр (дедуп по ИНН)
  report   — собрать XLSX-отчёт
  inns     — выгрузить список ИНН для пакетной проверки в реестре
  merge    — влить обратно результаты проверки из реестра (CSV/XLSX)

Правовая база (по состоянию на 2026 год):
  * Строительство, реконструкция, капремонт — ч. 2 ст. 52 ГрК РФ; членство
    в СРО (НОСТРОЙ) не требуется, если размер обязательств по КАЖДОМУ договору
    не превышает 10 млн руб. (ч. 2.1 ст. 52 ГрК РФ, порог с 01.05.2022).
  * Проектирование — ч. 4 ст. 48 ГрК РФ; изыскания — ч. 2 ст. 47 ГрК РФ.
    Стоимостного порога НЕТ: членство в СРО (НОПРИЗ) обязательно при любой
    сумме договора. С 01.03.2026 (309-ФЗ от 31.07.2025) требование
    распространено и на разработку рабочей документации.
  * Освобождены (ч. 2.2 ст. 52, ч. 4.1 ст. 48 ГрК РФ) ГУП/МУП, госучреждения
    и ряд юрлиц с госучастием — при работе по договору с учредившим их органом.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

POROG_STROY = 10_000_000  # ч. 2.1 ст. 52 ГрК РФ

NOT_CHECKED = "НЕ ПРОВЕРЕНО"

# --------------------------------------------------------------------------
# ИНН
# --------------------------------------------------------------------------

_W10 = (2, 4, 10, 3, 5, 9, 4, 6, 8)
_W11 = (7, 2, 4, 10, 3, 5, 9, 4, 6, 8)
_W12 = (3, 7, 2, 4, 10, 3, 5, 9, 4, 6, 8)


def _csum(digits, weights):
    return sum(d * w for d, w in zip(digits, weights)) % 11 % 10


def normalize_inn(raw: str) -> tuple[str, bool, str]:
    """Вернуть (нормализованный ИНН, валиден, комментарий).

    Excel срезает ведущие нули, поэтому 11-значный ИНН — почти всегда
    12-значный ИНН физлица с потерянным нулём. Восстанавливаем и проверяем
    контрольную сумму, а не угадываем.
    """
    s = re.sub(r"\D", "", str(raw or ""))
    note = ""
    if len(s) == 11:
        cand = "0" + s
        if validate_inn(cand):
            note = "восстановлен ведущий ноль (Excel срезал)"
            s = cand
    elif len(s) == 9:
        cand = "0" + s
        if validate_inn(cand):
            note = "восстановлен ведущий ноль (Excel срезал)"
            s = cand
    return s, validate_inn(s), note


def validate_inn(s: str) -> bool:
    if not s.isdigit():
        return False
    d = [int(c) for c in s]
    if len(d) == 10:
        return _csum(d[:9], _W10) == d[9]
    if len(d) == 12:
        return _csum(d[:10], _W11) == d[10] and _csum(d[:11], _W12) == d[11]
    return False


def inn_kind(s: str) -> str:
    if len(s) == 10:
        return "Юрлицо"
    if len(s) == 12:
        return "ИП / физлицо"
    return "не определено"


# --------------------------------------------------------------------------
# ОКПД2 -> вид СРО
# --------------------------------------------------------------------------

STROY = "СРО строителей (НОСТРОЙ)"
PROEKT = "СРО проектировщиков (НОПРИЗ)"
IZYSK = "СРО изыскателей (НОПРИЗ)"
UNKNOWN = "не определено по ОКПД2"

# Ключ — префикс ОКПД2 (группа). Значение — (вид СРО, расшифровка, примечание).
OKPD_MAP: dict[str, tuple[str, str, str]] = {
    "41.10": (PROEKT, "Документация проектная для строительства", ""),
    "41.20": (STROY, "Строительство жилых и нежилых зданий", ""),
    "42.11": (STROY, "Строительство автодорог и автомагистралей", ""),
    "42.12": (STROY, "Строительство железных дорог и метро", ""),
    "42.13": (STROY, "Строительство мостов и тоннелей", ""),
    "42.21": (STROY, "Строительство инженерных коммуникаций", ""),
    "42.22": (STROY, "Строительство коммунальных объектов электроснабжения и связи", ""),
    "42.91": (STROY, "Строительство водных сооружений", ""),
    "42.99": (STROY, "Строительство прочих инженерных сооружений", ""),
    "43.11": (STROY, "Разборка и снос зданий", ""),
    "43.12": (STROY, "Подготовка строительной площадки", ""),
    "43.21": (STROY, "Электромонтажные работы", ""),
    "43.22": (STROY, "Санитарно-технические работы, монтаж отопления и вентиляции", ""),
    "43.29": (STROY, "Прочие строительно-монтажные работы", ""),
    "43.31": (STROY, "Штукатурные работы", ""),
    "43.32": (STROY, "Столярные и плотничные работы", ""),
    "43.33": (STROY, "Работы по устройству покрытий полов и облицовке стен", ""),
    "43.34": (STROY, "Малярные и стекольные работы", ""),
    "43.39": (STROY, "Прочие отделочные и завершающие работы", ""),
    "43.91": (STROY, "Производство кровельных работ", ""),
    "43.99": (STROY, "Прочие специализированные строительные работы", ""),
    "71.11": (PROEKT, "Услуги архитектурные", ""),
    "71.12.11": (PROEKT, "Проектирование зданий", ""),
    "71.12.12": (PROEKT, "Проектирование промышленных объектов и сооружений", ""),
    "71.12.13": (PROEKT, "Проектирование инженерных сетей и систем", ""),
    "71.12.14": (PROEKT, "Проектирование транспортных объектов", ""),
    "71.12.15": (PROEKT, "Проектирование объектов водоснабжения и водоотведения", ""),
    "71.12.16": (PROEKT, "Проектирование в области благоустройства", ""),
    "71.12.17": (PROEKT, "Услуги по разработке проектов промышленных процессов", ""),
    "71.12.18": (PROEKT, "Проектирование прочих объектов", ""),
    "71.12.19": (PROEKT, "Прочие услуги инженерно-технического проектирования",
                 "Уточнить предмет: экспертиза и обследование СРО не требуют, "
                 "нужна аккредитация Минстроя"),
    "71.12.3": (IZYSK, "Инженерные изыскания", ""),
    "71.20": (UNKNOWN, "Технические испытания и сертификация", "СРО, как правило, не требуется"),
}

GOS_FORMS = (
    "УНИТАРНОЕ ПРЕДПРИЯТИЕ", "ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ", "МУНИЦИПАЛЬНОЕ БЮДЖЕТНОЕ",
    "КАЗЕННОЕ", "ГОСУДАРСТВЕННОЕ АВТОНОМНОЕ", "МУНИЦИПАЛЬНОЕ АВТОНОМНОЕ",
    "ГБУ", "МБУ", "ГУП", "МУП", "ФГУП", "ФГБУ",
)


# Предмет договора может опровергнуть код ОКПД2: под 71.12 заказчики проводят
# и экспертизу, и обследование, и паспортизацию — а членства в СРО они не требуют.
SUBJECT_FLAGS: tuple[tuple[str, str], ...] = (
    ("экспертиз", "В предмете — экспертиза проектной документации: нужна аккредитация "
                  "Минстроя, членство в СРО не требуется"),
    ("обследовани", "В предмете — обследование конструкций: уточнить, выполняется ли оно "
                    "в составе проектных работ (тогда СРО нужно) или отдельно"),
    ("паспортизац", "В предмете — паспортизация: членство в СРО, как правило, не требуется"),
    ("изыскани", "В предмете — инженерные изыскания: нужно СРО изыскателей (НОПРИЗ)"),
    ("сметн", "В предмете — сметная документация: если разрабатывается отдельно от "
              "проекта, членство в СРО может не требоваться"),
    ("субподряд", "В предмете упомянут субподряд: у субподрядчика членство не требуется"),
)


def flag_subjects(subjects: list[str]) -> str:
    blob = " ".join(subjects).lower()
    hits = [note for kw, note in SUBJECT_FLAGS if kw in blob]
    return "; ".join(hits)


def classify_okpd(code: str) -> tuple[str, str, str]:
    c = re.sub(r"[^\d.]", "", str(code or ""))
    for length in (8, 7, 5):  # 71.12.19 -> 71.12 -> 41.10
        key = c[:length].rstrip(".")
        if key in OKPD_MAP:
            return OKPD_MAP[key]
    return (UNKNOWN, "", "ОКПД2 не распознан — определить вид работ по предмету договора")


def membership_required(sro_type: str, max_sum: int, name: str) -> tuple[str, str]:
    """Вернуть (обязательно ли членство, правовое основание)."""
    if any(f in (name or "").upper() for f in GOS_FORMS):
        return ("УТОЧНИТЬ", "ч. 2.2 ст. 52 / ч. 4.1 ст. 48 ГрК РФ: ГУП/МУП и госучреждения "
                            "освобождены при договоре с учредившим их органом")
    if sro_type == STROY:
        if max_sum > POROG_STROY:
            return ("ДА", f"ч. 2 ст. 52 ГрК РФ: договор {max_sum:,} руб. > порога 10 млн руб."
                    .replace(",", " "))
        return ("НЕТ", f"ч. 2.1 ст. 52 ГрК РФ: договор {max_sum:,} руб. в пределах порога 10 млн руб."
                .replace(",", " "))
    if sro_type == PROEKT:
        return ("ДА", "ч. 4 ст. 48 ГрК РФ: для проектирования стоимостного порога нет")
    if sro_type == IZYSK:
        return ("ДА", "ч. 2 ст. 47 ГрК РФ: для изысканий стоимостного порога нет")
    return ("УТОЧНИТЬ", "вид работ не определён по ОКПД2")


# --------------------------------------------------------------------------
# Разбор выгрузок tenderguru
# --------------------------------------------------------------------------

def _clean(v) -> str:
    s = "" if v is None else str(v)
    return re.sub(r"\s+", " ", s).strip()


def _strip_fio_suffix(v: str) -> str:
    """'mail@x.ru(ИВАНОВ ИВАН)' -> 'mail@x.ru'"""
    return re.sub(r"\s*\([^)]*\)\s*$", "", v).strip()


def _to_int(v) -> int:
    s = re.sub(r"[^\d-]", "", str(v or ""))
    try:
        return int(s)
    except ValueError:
        return 0


def parse_exports(paths: list[Path]) -> list[dict]:
    contracts = []
    for p in paths:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        if "Контракты" not in wb.sheetnames:
            wb.close()
            continue
        ws = wb["Контракты"]
        rit = ws.iter_rows(values_only=True)
        hdr = [_clean(c) for c in next(rit)]
        for row in rit:
            d = {hdr[i]: _clean(row[i]) for i in range(min(len(hdr), len(row))) if hdr[i]}
            if not d.get("ИНН поставщика"):
                continue
            d["_file"] = p.name
            contracts.append(d)
        wb.close()
    return contracts


def build_registry(contracts: list[dict]) -> list[dict]:
    by_inn: dict[str, list[dict]] = defaultdict(list)
    for c in contracts:
        inn, _, _ = normalize_inn(c["ИНН поставщика"])
        by_inn[inn].append(c)

    out = []
    for inn, cs in by_inn.items():
        _, valid, note = normalize_inn(cs[0]["ИНН поставщика"])
        sums = [_to_int(c.get("Сумма контракта")) for c in cs]
        max_sum, total = max(sums), sum(sums)
        # Самый дорогой контракт — по нему определяем профиль и порог.
        lead = cs[sums.index(max_sum)]
        name = lead.get("Поставщик", "")
        sro_type, okpd_name, okpd_note = classify_okpd(lead.get("Коды", ""))
        required, basis = membership_required(sro_type, max_sum, name)

        def first(field, transform=lambda x: x):
            for c in cs:
                v = transform(c.get(field, ""))
                if v:
                    return v
            return ""

        subjects = [c.get("Закупаемая продукция", "") for c in cs]
        notes = [n for n in (okpd_note, flag_subjects(subjects)) if n]

        out.append({
            "Поставщик": name,
            "ИНН": inn,
            "ИНН валиден": "да" if valid else "НЕТ — проверить",
            "Примечание к ИНН": note,
            "Форма": inn_kind(inn),
            "Руководитель": first("Руководитель поставшика"),
            "Должность": first("Должность"),
            "E-mail": first("E-mail", _strip_fio_suffix),
            "Телефон": first("Телефон", _strip_fio_suffix),
            "Адрес": first("Адрес поставщика"),
            "Дата регистрации": first("Дата регистрации"),
            "Регион": first("Регион"),
            "Контрактов в выгрузке": len(cs),
            "Макс. сумма контракта, руб.": max_sum,
            "Сумма контрактов, руб.": total,
            "ОКПД2": lead.get("Коды", ""),
            "Вид работ по ОКПД2": okpd_name,
            "Требуемое СРО": sro_type,
            "Членство обязательно": required,
            "Правовое основание": basis,
            "Предмет крупнейшего договора": lead.get("Закупаемая продукция", ""),
            "Примечание": "; ".join(notes),
            "Статус в реестре СРО": NOT_CHECKED,
            "Наименование СРО": "",
            "Рег. номер члена": "",
            "Дата вступления": "",
            "Уровень КФ ВВ": "",
            "Уровень КФ ОДО": "",
            "Источник проверки": "",
            "Дата проверки": "",
            "Достоверность": "источник: выгрузка ЕИС/tenderguru",
        })

    order = {STROY: 0, PROEKT: 1, IZYSK: 2, UNKNOWN: 3}
    out.sort(key=lambda r: (order.get(r["Требуемое СРО"], 9),
                            -r["Макс. сумма контракта, руб."]))
    return out


# --------------------------------------------------------------------------
# Слияние результатов проверки из реестра
# --------------------------------------------------------------------------

MERGE_FIELDS = ("Статус в реестре СРО", "Наименование СРО", "Рег. номер члена",
                "Дата вступления", "Уровень КФ ВВ", "Уровень КФ ОДО", "Источник проверки")


def load_lookup(path: Path) -> dict[str, dict]:
    """Прочитать выгрузку из реестра СРО (CSV или XLSX).

    Нужны колонки: ИНН + любые из MERGE_FIELDS. Регистр и порядок не важны.
    """
    rows: list[dict] = []
    if path.suffix.lower() in (".csv", ".txt"):
        with path.open(encoding="utf-8-sig", newline="") as fh:
            sample = fh.read(4096)
            fh.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
            rows = list(csv.DictReader(fh, dialect=dialect))
    else:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rit = ws.iter_rows(values_only=True)
        hdr = [_clean(c) for c in next(rit)]
        for r in rit:
            rows.append({hdr[i]: _clean(r[i]) for i in range(min(len(hdr), len(r))) if hdr[i]})
        wb.close()

    lookup: dict[str, dict] = {}
    for r in rows:
        key = next((k for k in r if "ИНН" in str(k).upper() or str(k).upper() == "INN"), None)
        if not key:
            continue
        inn, _, _ = normalize_inn(r[key])
        if inn:
            lookup[inn] = r
    return lookup


def merge_status(registry: list[dict], lookup: dict[str, dict], today: str) -> tuple[int, int]:
    hit = miss = 0
    for row in registry:
        found = lookup.get(row["ИНН"])
        if not found:
            miss += 1
            continue
        hit += 1
        for f in MERGE_FIELDS:
            for k, v in found.items():
                if _clean(k).lower() == f.lower() and _clean(v):
                    row[f] = _clean(v)
        if row["Статус в реестре СРО"] == NOT_CHECKED:
            row["Статус в реестре СРО"] = "не найден в реестре"
        row["Дата проверки"] = today
        row["Достоверность"] = "High — данные официального реестра"
    return hit, miss


# --------------------------------------------------------------------------
# Отчёт
# --------------------------------------------------------------------------

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FILL = PatternFill("solid", fgColor="F8CBAD")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WIDTHS = {
    "Поставщик": 46, "ИНН": 14, "ИНН валиден": 13, "Примечание к ИНН": 26, "Форма": 13,
    "Руководитель": 26, "Должность": 20, "E-mail": 28, "Телефон": 16, "Адрес": 40,
    "Дата регистрации": 15, "Регион": 22, "Контрактов в выгрузке": 11,
    "Макс. сумма контракта, руб.": 18, "Сумма контрактов, руб.": 18, "ОКПД2": 14,
    "Вид работ по ОКПД2": 40, "Требуемое СРО": 24, "Членство обязательно": 12,
    "Правовое основание": 52, "Предмет крупнейшего договора": 60, "Примечание": 56, "Статус в реестре СРО": 20,
    "Наименование СРО": 30, "Рег. номер члена": 16, "Дата вступления": 14,
    "Уровень КФ ВВ": 12, "Уровень КФ ОДО": 12, "Источник проверки": 30,
    "Дата проверки": 13, "Достоверность": 30,
}


def _write_sheet(ws, rows: list[dict], cols: list[str]) -> None:
    ws.freeze_panes = "C2"
    for j, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = HDR_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(j)].width = WIDTHS.get(col, 18)
    ws.row_dimensions[1].height = 42

    for i, row in enumerate(rows, 2):
        for j, col in enumerate(cols, 1):
            c = ws.cell(row=i, column=j, value=row.get(col, ""))
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=col in
                                    ("Поставщик", "Адрес", "Правовое основание",
                                     "Вид работ по ОКПД2", "Примечание",
                                     "Предмет крупнейшего договора"))
            if "руб." in col:
                c.number_format = "# ##0"
        if row.get("Статус в реестре СРО") == NOT_CHECKED:
            ws.cell(row=i, column=cols.index("Статус в реестре СРО") + 1).fill = WARN_FILL
        if row.get("ИНН валиден", "").startswith("НЕТ"):
            ws.cell(row=i, column=cols.index("ИНН валиден") + 1).fill = BAD_FILL
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"


def build_report(registry: list[dict], contracts: list[dict], out: Path, today: str) -> None:
    cols = list(WIDTHS.keys())
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Все поставщики"
    _write_sheet(ws, registry, cols)

    need = [r for r in registry if r["Членство обязательно"] in ("ДА", "УТОЧНИТЬ")]
    _write_sheet(wb.create_sheet("Членство обязательно"), need, cols)

    exempt = [r for r in registry if r["Членство обязательно"] == "НЕТ"]
    _write_sheet(wb.create_sheet("Членство не требуется"), exempt, cols)

    # Первичные контракты — для прослеживаемости каждой строки реестра.
    cws = wb.create_sheet("Контракты (исходные)")
    ccols = ["Дата", "Поставщик", "ИНН поставщика", "Заказчик", "Закупаемая продукция",
             "Закон/Номер контракта", "Сумма контракта", "Коды", "Регион",
             "Дата исполнения", "Дата окончания исполнения", "_file"]
    cws.freeze_panes = "A2"
    for j, col in enumerate(ccols, 1):
        c = cws.cell(row=1, column=j, value="Файл-источник" if col == "_file" else col)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = HDR_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        cws.column_dimensions[get_column_letter(j)].width = 46 if col in (
            "Поставщик", "Заказчик", "Закупаемая продукция") else 18
    for i, row in enumerate(contracts, 2):
        for j, col in enumerate(ccols, 1):
            v = _to_int(row.get(col)) if col == "Сумма контракта" else row.get(col, "")
            c = cws.cell(row=i, column=j, value=v)
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="top")
            if col == "Сумма контракта":
                c.number_format = "# ##0"
    cws.auto_filter.ref = f"A1:{get_column_letter(len(ccols))}{len(contracts) + 1}"

    _write_methodology(wb.create_sheet("Методика"), len(registry), len(contracts), today)
    wb.save(out)


def _write_methodology(ws, n_sup: int, n_con: int, today: str) -> None:
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 96
    lines = [
        ("ОТЧЁТ О ПРОВЕРКЕ КОНТРАГЕНТОВ НА ЧЛЕНСТВО В СРО", ""),
        ("Дата формирования", today),
        ("Поставщиков (уникальных ИНН)", None),
        ("Строк-контрактов в исходных выгрузках", None),
        ("Источник исходных данных", "Выгрузки контрактов 44-ФЗ, система tenderguru.ru"),
        ("", ""),
        ("СТАТУС ПРОВЕРКИ ПО РЕЕСТРУ", ""),
        ("Статус в реестре СРО", "НЕ ПРОВЕРЕНО. Реестры НОСТРОЙ и НОПРИЗ недоступны "
                                 "из окружения (блокировка сетевой политикой). Колонка "
                                 "заполняется командой merge после выгрузки из реестра."),
        ("Что уже проверено на 100%", "Состав поставщиков, дедупликация по ИНН, контрольная "
                                      "сумма ИНН, форма (юрлицо / ИП), суммы контрактов, "
                                      "контактные данные — всё из первичной выгрузки."),
        ("Что рассчитано", "Вид требуемого СРО (по ОКПД2) и обязательность членства "
                           "(по сумме договора и виду работ). Это расчёт по закону, "
                           "а НЕ факт наличия членства."),
        ("", ""),
        ("ПРАВОВАЯ БАЗА", ""),
        ("Строительство, реконструкция, капремонт", "ч. 2 ст. 52 ГрК РФ — членство в СРО "
                                                    "обязательно"),
        ("Порог для строительства", "ч. 2.1 ст. 52 ГрК РФ — членство НЕ требуется, если размер "
                                    "обязательств по каждому договору не превышает 10 млн руб. "
                                    "(порог действует с 01.05.2022)"),
        ("Проектирование", "ч. 4 ст. 48 ГрК РФ — членство обязательно, стоимостного порога НЕТ"),
        ("Инженерные изыскания", "ч. 2 ст. 47 ГрК РФ — членство обязательно, порога НЕТ"),
        ("Рабочая документация", "С 01.03.2026 (309-ФЗ от 31.07.2025) требование о членстве "
                                 "распространено на разработку рабочей документации"),
        ("Освобождения", "ч. 2.2 ст. 52 и ч. 4.1 ст. 48 ГрК РФ — ГУП/МУП, госучреждения и ряд "
                         "юрлиц с госучастием при договоре с учредившим их органом"),
        ("Субподряд", "Членство требуется только у лица, заключившего договор с застройщиком, "
                      "техзаказчиком или региональным оператором. У субподрядчика — не требуется"),
        ("", ""),
        ("ВАЖНО ПРИ ТРАКТОВКЕ", ""),
        ("«Допуск СРО» не существует с 01.07.2017", "Свидетельства о допуске отменены 372-ФЗ. "
                                                    "Проверяется факт членства и уровень "
                                                    "ответственности по компфондам (КФ ВВ, КФ ОДО)"),
        ("Статус «приостановлено»", "Компания числится членом СРО, но права выполнять работы "
                                    "не имеет. Бинарная колонка «да/нет» это теряет — "
                                    "смотреть точный статус из реестра"),
        ("Отсутствие членства ≠ нарушение", "По договорам в пределах порога, на субподряде и "
                                            "на некапитальных объектах членство не требуется"),
        ("ОКПД2 — косвенный признак", "Вид работ определён по коду закупки. Окончательно вид "
                                      "работ определяется предметом договора"),
    ]
    r = 1
    for k, v in lines:
        a = ws.cell(row=r, column=1, value=k)
        a.font = Font(name=FONT, bold=True, size=11 if r == 1 else 10)
        a.alignment = Alignment(vertical="top", wrap_text=True)
        if v is None:
            # Литералы, а не формулы: книга — снимок выгрузки, а не модель.
            # Формула без кэшированного значения читалась бы как пустая ячейка
            # в pandas и превьюшках, а пересчитать её здесь нечем.
            b = ws.cell(row=r, column=2, value=n_sup if "Поставщиков" in k else n_con)
        else:
            b = ws.cell(row=r, column=2, value=v)
        b.font = Font(name=FONT, size=10)
        b.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1
    ws.cell(row=1, column=1).fill = HDR_FILL
    ws.cell(row=1, column=1).font = Font(name=FONT, bold=True, size=11, color="FFFFFF")


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------

def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("inputs", nargs="+", type=Path, help="выгрузки tenderguru (.xlsx)")
    ap.add_argument("-o", "--out", type=Path, default=Path("sro-report.xlsx"))
    ap.add_argument("--merge", type=Path, help="выгрузка из реестра СРО (CSV/XLSX) с колонкой ИНН")
    ap.add_argument("--inns", type=Path, help="выгрузить список ИНН для пакетной проверки")
    ap.add_argument("--json", type=Path, help="выгрузить реестр поставщиков в JSON")
    args = ap.parse_args(argv)

    today = date.today().strftime("%d.%m.%Y")
    contracts = parse_exports(args.inputs)
    if not contracts:
        print("Не найдено ни одной строки контракта.", file=sys.stderr)
        return 1
    registry = build_registry(contracts)
    print(f"Контрактов: {len(contracts)}   Уникальных поставщиков: {len(registry)}")

    if args.merge:
        hit, miss = merge_status(registry, load_lookup(args.merge), today)
        print(f"Слияние с реестром: сопоставлено {hit}, без совпадения {miss}")

    if args.inns:
        args.inns.write_text("\n".join(r["ИНН"] for r in registry), encoding="utf-8")
        print(f"Список ИНН: {args.inns}")

    if args.json:
        args.json.write_text(json.dumps(registry, ensure_ascii=False, indent=1), encoding="utf-8")
        print(f"JSON: {args.json}")

    build_report(registry, contracts, args.out, today)
    print(f"Отчёт: {args.out}")

    bad = [r for r in registry if r["ИНН валиден"].startswith("НЕТ")]
    if bad:
        print(f"ВНИМАНИЕ: ИНН не прошли контрольную сумму: {len(bad)}")
        for r in bad:
            print(f"  {r['ИНН']}  {r['Поставщик'][:60]}")
    unchecked = sum(1 for r in registry if r["Статус в реестре СРО"] == NOT_CHECKED)
    if unchecked:
        print(f"НЕ ПРОВЕРЕНО по реестру СРО: {unchecked} из {len(registry)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
