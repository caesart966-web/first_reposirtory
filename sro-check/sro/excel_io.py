"""Чтение исходного Excel и запись результатов с сохранением оформления.

Исходный файл НИКОГДА не перезаписывается — результат всегда пишется в новый.
Оформление (шрифты, заливки, ширины колонок, объединённые ячейки) сохраняется
за счёт того, что мы открываем исходный файл через openpyxl и до сохранения
меняем только новые колонки.
"""

from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass, field

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import SOURCE_NOPRIZ, SOURCE_NOSTROY, CheckResult, Verdict

# ---------------------------------------------------------------------------
# Названия колонок
# ---------------------------------------------------------------------------

COL_VERDICT = "Есть СРО"
COL_MEMBER_NAME = "Компания в реестре"
COL_SRO_NAME = "Название СРО"
COL_REG_NUMBER = "Реестровый номер"
COL_STATUS = "Статус членства"
COL_DATE = "Дата проверки"
COL_SOURCE = "Источник"
COL_NOTE = "Примечание"

OUTPUT_COLUMNS = (
    COL_VERDICT,
    COL_MEMBER_NAME,
    COL_SRO_NAME,
    COL_REG_NUMBER,
    COL_STATUS,
    COL_DATE,
    COL_SOURCE,
    COL_NOTE,
)

COLUMN_WIDTHS = {
    COL_VERDICT: 20,
    COL_MEMBER_NAME: 46,
    COL_SRO_NAME: 52,
    COL_REG_NUMBER: 18,
    COL_STATUS: 18,
    COL_DATE: 18,
    COL_SOURCE: 14,
    COL_NOTE: 46,
}

SUMMARY_SHEET = "Итоги проверки"

# Как узнать нужные колонки в исходном файле: подстроки в названии заголовка.
INN_ALIASES = ("инн",)
NAME_ALIASES = ("поставщик", "наименование", "название компании", "название", "компания", "контрагент")
PRIORITY_ALIASES = ("приоритет", "тип", "категория")
VERDICT_ALIASES = ("есть сро", "наличие сро", "сро?")

VERDICT_FILLS = {
    Verdict.YES: PatternFill("solid", fgColor="D6EFD6"),
    Verdict.NO: PatternFill("solid", fgColor="F8D7D7"),
    Verdict.UNKNOWN: PatternFill("solid", fgColor="FCEFC7"),
}


def normalize_header(value) -> str:
    """Привести заголовок к виду, по которому удобно сравнивать."""
    if value is None:
        return ""
    text = str(value).replace("ё", "е").replace("Ё", "Е")
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


@dataclass
class SheetLayout:
    """Где что лежит на листе."""

    header_row: int
    first_data_row: int
    last_row: int
    columns: dict[str, int] = field(default_factory=dict)  # логическое имя -> номер колонки

    def cell_value(self, worksheet: Worksheet, row: int, logical: str):
        column = self.columns.get(logical)
        if not column:
            return None
        return worksheet.cell(row=row, column=column).value


def open_sheet(path: str, sheet_name: str) -> tuple:
    """Открыть книгу и нужный лист. Возвращает (workbook, worksheet, предупреждения)."""
    warnings: list[str] = []
    workbook = load_workbook(path)  # значения формул openpyxl не считает, но текст читает

    worksheet = None
    if sheet_name and sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    elif sheet_name:
        target = normalize_header(sheet_name)
        for name in workbook.sheetnames:
            if normalize_header(name) == target or target in normalize_header(name):
                worksheet = workbook[name]
                warnings.append(f"лист «{sheet_name}» не найден, взят похожий: «{name}»")
                break

    if worksheet is None:
        worksheet = workbook[workbook.sheetnames[0]]
        if sheet_name:
            warnings.append(
                f"лист «{sheet_name}» не найден. Доступные листы: "
                + ", ".join(workbook.sheetnames)
                + f". Взят первый: «{worksheet.title}»"
            )

    if getattr(worksheet, "_charts", None) or getattr(worksheet, "_images", None):
        warnings.append(
            "на листе есть диаграммы или картинки — при пересохранении они могут пропасть "
            "(исходный файл при этом не меняется)"
        )

    return workbook, worksheet, warnings


def find_layout(worksheet: Worksheet, scan_rows: int = 20) -> SheetLayout:
    """Найти строку заголовков и распознать нужные колонки."""
    header_row = 0
    headers: dict[int, str] = {}

    for row in range(1, min(scan_rows, worksheet.max_row or 1) + 1):
        row_headers = {}
        for column in range(1, (worksheet.max_column or 1) + 1):
            text = normalize_header(worksheet.cell(row=row, column=column).value)
            if text:
                row_headers[column] = text
        if any(any(alias in text for alias in INN_ALIASES) for text in row_headers.values()):
            header_row = row
            headers = row_headers
            break

    if not header_row:
        raise ValueError(
            "не нашёл строку заголовков: ни в одной из первых "
            f"{scan_rows} строк нет колонки со словом «ИНН»"
        )

    def match(aliases: tuple[str, ...]) -> int:
        # Сначала точное совпадение, потом вхождение подстроки.
        for alias in aliases:
            for column, text in headers.items():
                if text == alias:
                    return column
        for alias in aliases:
            for column, text in headers.items():
                if alias in text:
                    return column
        return 0

    columns: dict[str, int] = {}
    columns["inn"] = match(INN_ALIASES)
    name_column = match(NAME_ALIASES)
    if name_column:
        columns["name"] = name_column
    priority_column = match(PRIORITY_ALIASES)
    if priority_column:
        columns["priority"] = priority_column

    # Существующая колонка «Есть СРО? (проверить в НОСТРОЙ)» — её заполняем,
    # а не дублируем. Убеждаемся, что не приняли за неё «Название СРО».
    verdict_column = match(VERDICT_ALIASES)
    if verdict_column and verdict_column != columns.get("inn"):
        header_text = headers.get(verdict_column, "")
        if not header_text.startswith(("название", "статус", "реестров")):
            columns[COL_VERDICT] = verdict_column

    last_row = worksheet.max_row or header_row
    return SheetLayout(
        header_row=header_row,
        first_data_row=header_row + 1,
        last_row=last_row,
        columns=columns,
    )


def ensure_output_columns(worksheet: Worksheet, layout: SheetLayout) -> list[str]:
    """Добавить недостающие колонки результата. Возвращает список добавленных."""
    headers = {
        normalize_header(worksheet.cell(row=layout.header_row, column=column).value): column
        for column in range(1, (worksheet.max_column or 1) + 1)
    }

    template_column = layout.columns.get("inn") or 1
    template = worksheet.cell(row=layout.header_row, column=template_column)

    added: list[str] = []
    next_column = (worksheet.max_column or 1) + 1

    for name in OUTPUT_COLUMNS:
        if name in layout.columns:
            continue  # уже нашли (например, исходная колонка «Есть СРО? ...»)
        existing = headers.get(normalize_header(name))
        if existing:
            layout.columns[name] = existing
            continue

        cell = worksheet.cell(row=layout.header_row, column=next_column, value=name)
        # Копируем оформление заголовка соседней колонки, чтобы шапка выглядела единообразно.
        if template.has_style:
            cell._style = copy(template._style)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if not cell.font or not cell.font.bold:
            cell.font = Font(bold=True, size=(template.font.size if template.font else 11))
        worksheet.column_dimensions[get_column_letter(next_column)].width = COLUMN_WIDTHS.get(name, 20)
        layout.columns[name] = next_column
        added.append(name)
        next_column += 1

    return added


def iter_companies(worksheet: Worksheet, layout: SheetLayout):
    """Пройти по строкам с данными. Отдаёт (номер строки, ИНН, название, приоритет)."""
    for row in range(layout.first_data_row, layout.last_row + 1):
        raw_inn = layout.cell_value(worksheet, row, "inn")
        name = layout.cell_value(worksheet, row, "name")
        priority = layout.cell_value(worksheet, row, "priority")

        # Полностью пустые строки пропускаем.
        if raw_inn is None and (name is None or str(name).strip() == ""):
            continue

        yield row, raw_inn, ("" if name is None else str(name).strip()), (
            "" if priority is None else str(priority).strip()
        )


def read_existing_verdict(worksheet: Worksheet, layout: SheetLayout, row: int) -> str:
    """Прочитать уже проставленный вердикт (для режима перепроверки)."""
    value = layout.cell_value(worksheet, row, COL_VERDICT)
    return normalize_header(value)


def write_result(
    worksheet: Worksheet,
    layout: SheetLayout,
    row: int,
    result: CheckResult,
    colorize: bool = True,
    extra_note: str = "",
) -> None:
    """Записать результат проверки в строку.

    `extra_note` дописывается к примечанию: туда попадает то, что зависит от
    самой строки файла, а не от ответа реестра (например, расхождение названий).
    """
    note = "; ".join(part for part in (result.diagnostics, extra_note) if part)
    values = {
        COL_VERDICT: result.verdict.value,
        COL_MEMBER_NAME: result.member_names,
        COL_SRO_NAME: result.sro_names,
        COL_REG_NUMBER: result.registry_numbers,
        COL_STATUS: result.statuses,
        COL_DATE: result.checked_at,
        COL_SOURCE: result.sources,
        COL_NOTE: note,
    }

    wrapped = (COL_MEMBER_NAME, COL_SRO_NAME, COL_NOTE)
    for name, value in values.items():
        column = layout.columns.get(name)
        if not column:
            continue
        cell = worksheet.cell(row=row, column=column, value=value or "")
        cell.alignment = Alignment(vertical="top", wrap_text=name in wrapped)
        if colorize and name == COL_VERDICT:
            cell.fill = VERDICT_FILLS[result.verdict]


def finish_sheet(worksheet: Worksheet, layout: SheetLayout) -> None:
    """Довести лист до удобного вида: закрепить шапку и включить фильтр.

    Делается один раз перед сохранением: 393 строки без фильтра и закреплённой
    шапки читать невозможно.
    """
    try:
        worksheet.freeze_panes = worksheet.cell(row=layout.header_row + 1, column=1)
        last_column = get_column_letter(max(worksheet.max_column, 1))
        last_row = max(worksheet.max_row, layout.header_row)
        worksheet.auto_filter.ref = f"A{layout.header_row}:{last_column}{last_row}"
    except Exception:
        # Оформление — не то, ради чего стоит терять результат проверки.
        pass


def write_summary(workbook, rows: list[tuple[str, object]]) -> None:
    """Отдельный лист с итогами прогона: сколько чего и с какими настройками."""
    if SUMMARY_SHEET in workbook.sheetnames:
        del workbook[SUMMARY_SHEET]
    sheet = workbook.create_sheet(SUMMARY_SHEET)
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 60

    title = sheet.cell(row=1, column=1, value="Итоги проверки СРО")
    title.font = Font(bold=True, size=14)

    for index, (label, value) in enumerate(rows, start=3):
        name_cell = sheet.cell(row=index, column=1, value=label)
        name_cell.font = Font(bold=True)
        name_cell.alignment = Alignment(vertical="top")
        value_cell = sheet.cell(row=index, column=2, value=value)
        value_cell.alignment = Alignment(vertical="top", wrap_text=True)


def registry_order(priority: str) -> list[str]:
    """В каком порядке опрашивать реестры для этой компании.

    Строители — сначала НОСТРОЙ, проектировщики и изыскатели — сначала НОПРИЗ.
    Второй реестр всё равно опрашивается, если в первом ничего не нашлось:
    разметка «Приоритет» в файле может быть неточной.
    """
    text = normalize_header(priority)
    if re.search(r"проект|изыск|нопр|пир\b", text):
        return [SOURCE_NOPRIZ, SOURCE_NOSTROY]
    return [SOURCE_NOSTROY, SOURCE_NOPRIZ]


def save_workbook(workbook, path: str) -> None:
    workbook.save(path)
