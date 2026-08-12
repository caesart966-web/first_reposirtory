"""Тесты. Запуск:  python -m unittest discover -s tests -v

Самая важная группа тестов — TestNeverGuess: она проверяет главное обещание
программы, что при любом сбое в файл попадает «не удалось проверить»,
а не «нет».
"""

from __future__ import annotations

import contextlib
import io
import os
import sys
import tempfile
import threading
import unittest
from unittest.mock import patch

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

import check_sro
import ui_palette
from sro import discover as discover_mod
from sro import excel_io, names
from sro.inn import is_valid_inn, normalize_inn
from sro.models import (
    CheckResult,
    RegistryAnswer,
    SOURCE_NOPRIZ,
    SOURCE_NOSTROY,
    STATUS_ACTIVE,
    STATUS_EXCLUDED,
    STATUS_SUSPENDED,
    Outcome,
    Verdict,
    normalize_status,
)
from sro import registries as registries_mod
from sro.registries import (
    Endpoint,
    RegistryProvider,
    build_providers,
    interpret_payload,
    origin_of,
)
from tests import mock_registry as mock
from tests.mock_registry import MockRegistry


# ---------------------------------------------------------------------------
# ИНН
# ---------------------------------------------------------------------------


class TestInn(unittest.TestCase):
    def test_plain_string(self):
        self.assertEqual(normalize_inn("7702521529"), "7702521529")

    def test_strips_spaces_and_junk(self):
        self.assertEqual(normalize_inn(" 7702 521 529 "), "7702521529")
        self.assertEqual(normalize_inn("ИНН: 7702521529"), "7702521529")

    def test_excel_number(self):
        """Excel часто отдаёт ИНН числом — в том числе с плавающей точкой."""
        self.assertEqual(normalize_inn(7702521529), "7702521529")
        self.assertEqual(normalize_inn(7702521529.0), "7702521529")

    def test_restores_leading_zero(self):
        """0274062111 в Excel превращается в 274062111 — возвращаем ноль."""
        self.assertEqual(normalize_inn(274062111), "0274062111")
        self.assertEqual(len(normalize_inn(27406211101)), 12)

    def test_empty(self):
        self.assertEqual(normalize_inn(None), "")
        self.assertEqual(normalize_inn(""), "")
        self.assertEqual(normalize_inn("нет данных"), "")

    def test_checksum(self):
        self.assertTrue(is_valid_inn("7702521529"))  # реальный контрольный разряд
        self.assertTrue(is_valid_inn("7707083893"))
        self.assertFalse(is_valid_inn("7702521520"))  # испорчен последний разряд
        self.assertFalse(is_valid_inn("123"))
        self.assertFalse(is_valid_inn(""))


class TestStatus(unittest.TestCase):
    def test_active(self):
        for text in ("Является членом СРО", "Действующий", "действует", "Активен"):
            self.assertEqual(normalize_status(text), STATUS_ACTIVE, text)

    def test_suspended(self):
        for text in ("Приостановлено", "Действие членства приостановлено"):
            self.assertEqual(normalize_status(text), STATUS_SUSPENDED, text)

    def test_excluded(self):
        for text in ("Исключен из реестра", "Исключён", "Членство прекращено"):
            self.assertEqual(normalize_status(text), STATUS_EXCLUDED, text)

    def test_unknown_text_is_not_invented(self):
        """Незнакомую формулировку не выдумываем — отдаём пустую строку."""
        self.assertEqual(normalize_status("нечто невнятное"), "")
        self.assertEqual(normalize_status(None), "")


# ---------------------------------------------------------------------------
# Разбор ответов реестра
# ---------------------------------------------------------------------------


class TestInterpret(unittest.TestCase):
    def test_found(self):
        answer = interpret_payload(mock.SCENARIOS[mock.MEMBER_INN], mock.MEMBER_INN, SOURCE_NOSTROY)
        self.assertIs(answer.outcome, Outcome.FOUND)
        self.assertEqual(len(answer.memberships), 1)
        membership = answer.memberships[0]
        self.assertIn("Строители России", membership.sro_name)
        # В колонку идёт номер СРО — тот же, что сайт показывает в «Рег. номер СРО».
        # Внутренний номер записи о членстве человеку ни о чём не говорит.
        self.assertEqual(membership.registry_number, "СРО-С-123-45678910")
        self.assertEqual(membership.status, STATUS_ACTIVE)

    def test_found_does_not_take_company_name_as_sro_name(self):
        """У записи есть и своё название, и название СРО — берём именно СРО."""
        answer = interpret_payload(mock.SCENARIOS[mock.MEMBER_INN], mock.MEMBER_INN, SOURCE_NOSTROY)
        self.assertNotIn("Мостострой", answer.memberships[0].sro_name)

    def test_empty_is_a_real_no(self):
        answer = interpret_payload(
            mock.SCENARIOS[mock.NON_MEMBER_INN], mock.NON_MEMBER_INN, SOURCE_NOSTROY
        )
        self.assertIs(answer.outcome, Outcome.EMPTY)

    def test_suspended_and_excluded_statuses(self):
        suspended = interpret_payload(
            mock.SCENARIOS[mock.SUSPENDED_INN], mock.SUSPENDED_INN, SOURCE_NOSTROY
        )
        self.assertEqual(suspended.memberships[0].status, STATUS_SUSPENDED)
        excluded = interpret_payload(
            mock.SCENARIOS[mock.EXCLUDED_INN], mock.EXCLUDED_INN, SOURCE_NOSTROY
        )
        self.assertEqual(excluded.memberships[0].status, STATUS_EXCLUDED)

    # --- ниже: всё, что должно давать UNKNOWN, а не EMPTY ---

    def test_unknown_json_shape(self):
        answer = interpret_payload(
            mock.SCENARIOS[mock.WEIRD_JSON_INN], mock.WEIRD_JSON_INN, SOURCE_NOSTROY
        )
        self.assertIs(answer.outcome, Outcome.UNKNOWN)

    def test_records_without_inn_field(self):
        answer = interpret_payload(
            mock.SCENARIOS[mock.NO_INN_FIELD_INN], mock.NO_INN_FIELD_INN, SOURCE_NOSTROY
        )
        self.assertIs(answer.outcome, Outcome.UNKNOWN)

    def test_search_filter_ignored(self):
        """Реестр отдал чужие записи — это не доказательство отсутствия."""
        answer = interpret_payload(
            mock.SCENARIOS[mock.UNFILTERED_INN], mock.UNFILTERED_INN, SOURCE_NOSTROY
        )
        self.assertIs(answer.outcome, Outcome.UNKNOWN)

    def test_partial_results(self):
        """Показана часть результатов — компания может быть на другой странице."""
        answer = interpret_payload(
            mock.SCENARIOS[mock.PAGINATED_INN], mock.PAGINATED_INN, SOURCE_NOSTROY
        )
        self.assertIs(answer.outcome, Outcome.UNKNOWN)

    def test_inn_only_in_sro_fields_is_ambiguous(self):
        payload = {
            "success": True,
            "data": {
                "data": [
                    {
                        "id": 1,
                        "full_description": "ООО «Кто-то»",
                        "sro": {"inn": "7702521529", "full_description": "СРО"},
                    }
                ],
                "count": 1,
            },
        }
        answer = interpret_payload(payload, "7702521529", SOURCE_NOSTROY)
        self.assertIs(answer.outcome, Outcome.UNKNOWN)

    def test_garbage_inputs(self):
        for payload in (None, "просто текст", 42, {"success": False, "message": "доступ закрыт"}):
            answer = interpret_payload(payload, "7702521529", SOURCE_NOSTROY)
            self.assertIs(answer.outcome, Outcome.UNKNOWN, payload)


# ---------------------------------------------------------------------------
# Работа по сети (через мок реестра)
# ---------------------------------------------------------------------------


class TestLogColouring(unittest.TestCase):
    """Раскраска журнала: цвет строки должен совпадать с её смыслом."""

    def check(self, line: str, expected: str) -> None:
        self.assertEqual(ui_palette.classify(line), expected, line)

    def test_verdicts(self):
        self.check("Проверено 1 из 393 | ООО «А» | ИНН 7702521529 → да | СРО «Строители»", "ok")
        self.check("Проверено 2 из 393 | АО «Б» | ИНН 7707083893 → нет", "bad")

    def test_unknown_is_amber_not_red(self):
        """Главная ловушка: «не удалось проверить» нельзя красить как «нет»."""
        self.check(
            "Проверено 3 из 393 | ООО «В» | ИНН 7806479303 → не удалось проверить | HTTP 404",
            "warn",
        )

    def test_alerts_and_headings(self):
        self.check("ОШИБКА: файл не найден", "alert")
        self.check("СБОЙ: ConnectionError", "alert")
        self.check("ВНИМАНИЕ: первые 25 компаний — все «нет»", "warn")
        self.check("  ! ВНИМАНИЕ: в реестре другая компания", "warn")
        self.check("=== НОСТРОЙ ===", "head")
        self.check("ИТОГ: рабочий способ проверки есть", "head")
        self.check("=" * 60, "head")

    def test_service_lines_are_dimmed(self):
        self.check("", "muted")
        self.check("[НОСТРОЙ] ищу действующий адрес в скриптах сайта…", "muted")
        self.check("  … промежуточный результат сохранён", "muted")

    def test_company_named_like_a_keyword_keeps_its_verdict(self):
        self.check("Проверено 7 из 9 | ООО «Внимание» | ИНН 7702521529 → да", "ok")

    def test_plain_text_stays_plain(self):
        self.check("Файл: C:/Users/Оператор/Компании.xlsx", "body")

    def test_every_tag_has_a_colour(self):
        for line in ("=== x ===", "→ да", "→ нет", "не удалось проверить", "ОШИБКА", "", "текст"):
            self.assertIn(ui_palette.classify(line), ui_palette.LOG_TAGS)


class TestCompanyNames(unittest.TestCase):
    """Сверка названий ловит опечатку в ИНН: запись найдена, но не про ту компанию."""

    def test_legal_form_does_not_matter(self):
        self.assertTrue(
            names.looks_like_same_company(
                'ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "МОСТОСТРОЙ"', "ООО «Мостострой»"
            )
        )

    def test_shortened_name_still_matches(self):
        self.assertTrue(names.looks_like_same_company('ООО "МЕГ"', "ООО «Мегаполис»"))

    def test_different_companies_are_caught(self):
        self.assertFalse(names.looks_like_same_company('ООО "Ромашка"', "ООО «Лютик»"))

    def test_silence_when_there_is_nothing_to_compare(self):
        """Ложная тревога на каждой строке обесценила бы предупреждение."""
        self.assertTrue(names.looks_like_same_company("", "ООО «Лютик»"))
        self.assertTrue(names.looks_like_same_company('ООО "Ромашка"', None))
        self.assertTrue(names.looks_like_same_company("ООО", "АО"))

    def test_note_quotes_the_registry_name(self):
        note = names.mismatch_note('ООО "Ромашка"', "ООО «Лютик»")
        self.assertIn("Лютик", note)
        self.assertEqual(names.mismatch_note('ООО "Ромашка"', "ООО «Ромашка»"), "")


class TestOrigin(unittest.TestCase):
    """Заголовок Origin: домен реестров сам содержит «reestr», и наивное
    отрезание по подстроке превращало заголовок в обрубок «https:/»."""

    def test_domain_containing_reestr(self):
        self.assertEqual(
            origin_of("https://reestr.nostroy.ru/reestr/chleny-sro"),
            "https://reestr.nostroy.ru",
        )

    def test_plain_domain(self):
        self.assertEqual(origin_of("https://nopriz.ru/nreesters/"), "https://nopriz.ru")

    def test_not_a_url(self):
        self.assertEqual(origin_of("просто текст"), "")


class FakeResponse:
    def __init__(self, text: str, status: int = 200) -> None:
        self.text = text
        self.status_code = status


class FakeSession:
    """Подставной сайт: отдаёт заранее заданные страницы и скрипты."""

    def __init__(self, pages: dict[str, str]) -> None:
        self.pages = pages
        self.requested: list[str] = []

    def get(self, url: str, timeout: float | None = None, headers: dict | None = None) -> FakeResponse:
        self.requested.append(url)
        self.headers_seen = headers or {}
        if url in self.pages:
            return FakeResponse(self.pages[url])
        return FakeResponse("", 404)


class TestDiscovery(unittest.TestCase):
    """Поиск адреса API в скриптах сайта — то же, что человек делает в F12."""

    def test_finds_relative_and_absolute(self):
        text = """
            fetch("/api/sro/all/member/search?inn=" + inn)
            const other = 'https://reestr.nostroy.ru/api/member/list'
        """
        self.assertEqual(
            discover_mod.api_paths(text),
            {"/api/sro/all/member/search", "https://reestr.nostroy.ru/api/member/list"},
        )

    def test_skips_templates(self):
        """Адрес с подстановкой использовать нельзя — подставлять нечего."""
        self.assertEqual(discover_mod.api_paths('url("/api/member/${id}/card")'), set())

    def test_scoring_prefers_member_search(self):
        self.assertGreater(
            discover_mod.score_path("/api/sro/all/member/search"),
            discover_mod.score_path("/api/news/list"),
        )

    def test_scoring_rejects_unrelated(self):
        for path in ("/api/auth/login", "/api/file/upload", "/api/user/password"):
            self.assertLess(discover_mod.score_path(path), 0, path)

    def test_script_urls_are_absolute(self):
        html = '<script src="/static/app.js"></script><script src="https://cdn.example/x.js">'
        self.assertEqual(
            discover_mod.script_urls(html, "https://reestr.nostroy.ru/reestr/chleny-sro"),
            ["https://reestr.nostroy.ru/static/app.js", "https://cdn.example/x.js"],
        )

    def test_discover_end_to_end(self):
        page = discover_mod.PAGES[SOURCE_NOSTROY][0]
        session = FakeSession(
            {
                page: '<html><script src="/static/app.js"></script></html>',
                "https://reestr.nostroy.ru/static/app.js": (
                    'a("/api/auth/login");b("/api/sro/all/member/search");c("/api/news/list")'
                ),
            }
        )
        found = discover_mod.discover(SOURCE_NOSTROY, session, logger=lambda _m: None)
        self.assertEqual(found[0], "https://reestr.nostroy.ru/api/sro/all/member/search")
        self.assertNotIn("https://reestr.nostroy.ru/api/auth/login", found)

    def test_fragments_are_collected_separately(self):
        """Скрипт хранит основу и хвост порознь: BASE + '/member/list'."""
        text = 'const BASE="/api/sro/all"; post(BASE+"/member/list"); img("/static/logo.png")'
        self.assertEqual(discover_mod.path_fragments(text), {"/member/list"})
        self.assertIn("/api/sro/all", discover_mod.api_paths(text))

    def test_combine_glues_base_and_fragment(self):
        self.assertIn(
            "/api/sro/all/member/list",
            discover_mod.combine(["/api/sro/all"], ["/member/list"]),
        )

    def test_discover_finds_address_assembled_in_code(self):
        """Главный случай: целиком адрес в скрипте не написан ни разу."""
        page = discover_mod.PAGES[SOURCE_NOSTROY][0]
        session = FakeSession(
            {
                page: '<html><script src="/js/api.js"></script></html>',
                "https://reestr.nostroy.ru/js/api.js": (
                    'const BASE="/api/sro/all";'
                    'export const members=(q)=>post(BASE+"/member/list",q);'
                    'export const login=()=>post("/api/auth/login");'
                ),
            }
        )
        found = discover_mod.discover(SOURCE_NOSTROY, session, logger=lambda _m: None)
        self.assertIn("https://reestr.nostroy.ru/api/sro/all/member/list", found)

    def test_html_is_requested_as_html(self):
        """Сессия провайдера просит JSON — за страницей нужно идти с другим Accept."""
        page = discover_mod.PAGES[SOURCE_NOSTROY][0]
        session = FakeSession({page: "<html></html>"})
        discover_mod.discover(SOURCE_NOSTROY, session, logger=lambda _m: None)
        self.assertIn("text/html", session.headers_seen.get("Accept", ""))

    def test_foreign_scripts_are_not_downloaded(self):
        """Аналитику и рекламные сети не трогаем — там нечего искать."""
        page = discover_mod.PAGES[SOURCE_NOSTROY][0]
        session = FakeSession({page: '<script src="https://mc.yandex.ru/metrika/tag.js"></script>'})
        discover_mod.discover(SOURCE_NOSTROY, session, logger=lambda _m: None)
        self.assertNotIn("https://mc.yandex.ru/metrika/tag.js", session.requested)


class TestRealRecordShape(unittest.TestCase):
    """Разбор записи в том виде, в каком её отдаёт настоящий реестр.

    Справочные значения приходят объектами: {"id":2,"code":"2","title":"Исключен"}.
    Если взять из такого объекта код вместо «title», статус не распознается,
    и успешно найденная компания получит «не удалось проверить».
    """

    RECORD = {
        "id": 90722,
        "member_type": {"id": 1, "code": "1", "title": "ЮЛ"},
        "member_status": {"id": 2, "code": "2", "title": "Исключен"},
        "inventory_number": "П-192-000278908099-1752",
        "full_description": 'ООО "СтройМонтаж-59"',
        "inn": "5907056036",
        "ogrn": "1135907001801",
        "sro": {
            "id": 906,
            "number": 906,
            "registration_number": "СРО-С-171-13012010",
            "full_description": "Саморегулируемая организация Союз «Строители Урала»",
        },
    }

    def membership(self):
        payload = {"data": {"data": [self.RECORD], "count": 1}}
        answer = interpret_payload(payload, "5907056036", SOURCE_NOSTROY)
        self.assertIs(answer.outcome, Outcome.FOUND)
        return answer.memberships[0]

    def test_status_is_the_word_not_the_code(self):
        self.assertEqual(self.membership().status, STATUS_EXCLUDED)

    def test_registry_number_is_the_one_shown_on_the_site(self):
        self.assertEqual(self.membership().registry_number, "СРО-С-171-13012010")

    def test_sro_and_company_names_are_not_mixed_up(self):
        membership = self.membership()
        self.assertIn("Строители Урала", membership.sro_name)
        self.assertIn("СтройМонтаж-59", membership.member_name)

    def test_verdict_is_no_for_excluded_member(self):
        """Раньше нераспознанный статус давал «не удалось» по найденной компании."""
        result = CheckResult(
            inn="5907056036",
            answers=[
                RegistryAnswer(SOURCE_NOSTROY, Outcome.FOUND, memberships=[self.membership()])
            ],
            basis="active",
        )
        self.assertIs(result.verdict, Verdict.NO)


class TestRealRequestShape(unittest.TestCase):
    """Запрос снят с самого сайта реестра и не должен «улучшаться» по памяти.

    Проверено в панели разработчика на reestr.nostroy.ru:
        POST https://reestr.nostroy.ru/api/sro/all/member/list
        {"filters":{},"page":1,"pageCount":"20","sortBy":{},"searchString":"..."}
    """

    def test_body_is_exactly_what_the_site_sends(self):
        self.assertEqual(
            registries_mod._list_body("5907056036"),
            {
                "filters": {},
                "page": 1,
                "pageCount": "20",  # именно строка: сайт шлёт так
                "sortBy": {},
                "searchString": "5907056036",  # именно в корне, не внутри filters
            },
        )

    def test_first_endpoint_is_the_observed_one(self):
        first = registries_mod.NOSTROY_ENDPOINTS[0]
        self.assertEqual(first.method, "POST")
        self.assertEqual(first.url, "https://reestr.nostroy.ru/api/sro/all/member/list")
        self.assertEqual(first.build("5907056036")["json"]["searchString"], "5907056036")

    def test_origin_matches_what_the_registry_allows(self):
        """Ответ реестра: Access-Control-Allow-Origin: https://reestr.nostroy.ru"""
        self.assertEqual(
            origin_of(registries_mod.NOSTROY_ENDPOINTS[0].referer), "https://reestr.nostroy.ru"
        )


class TestErrorHints(unittest.TestCase):
    """Что показывать человеку, когда реестр ответил ошибкой."""

    class Response:
        def __init__(self, text: str) -> None:
            self.text = text

    def hint(self, text: str) -> str:
        return registries_mod._body_hint(self.Response(text))

    def test_html_page_is_summarised_not_dumped(self):
        """«<!DOCTYPE html> <html> <head>» занимает всю строку и не значит ничего."""
        html = (
            '<!DOCTYPE html><html><head><meta charset="utf-8">'
            "<title>500 Internal Server Error</title></head><body>…</body></html>"
        )
        hint = self.hint(html)
        self.assertIn("HTML-страницу", hint)
        self.assertIn("500 Internal Server Error", hint)
        self.assertNotIn("DOCTYPE", hint)

    def test_html_without_title(self):
        hint = self.hint("<html><body>что-то пошло не так</body></html>")
        self.assertIn("HTML-страницу", hint)
        self.assertNotIn("<html", hint)

    def test_json_error_is_quoted_as_is(self):
        self.assertIn("internal", self.hint('{"error":"internal"}'))

    def test_empty_body_adds_nothing(self):
        self.assertEqual(self.hint(""), "")

    def test_long_reason_does_not_crowd_out_the_second(self):
        """Одна длинная причина не должна вытеснять пояснение второго варианта."""
        endpoints = [
            Endpoint("первый", "GET", "http://a/api", params=lambda inn: {}),
            Endpoint("второй", "GET", "http://b/api", params=lambda inn: {}),
        ]
        provider = RegistryProvider(
            SOURCE_NOSTROY,
            endpoints,
            timeout=1,
            attempts=1,
            logger=lambda _m: None,
            allow_discovery=False,
        )
        answers = iter([(None, "ПЕРВЫЙ " + "ы" * 300), (None, "ВТОРОЙ " + "я" * 300)])
        with patch.object(provider, "_request", side_effect=lambda *_: next(answers)):
            answer = provider.lookup("7702521529")

        self.assertIn("ПЕРВЫЙ", answer.note)
        self.assertIn("ВТОРОЙ", answer.note, "второе пояснение вытеснено первым")


class TestSelfHealing(unittest.TestCase):
    """Прошитый адрес умер — программа должна найти рабочий и продолжить."""

    @classmethod
    def setUpClass(cls):
        cls.mock = MockRegistry()
        cls.mock.__enter__()

    @classmethod
    def tearDownClass(cls):
        cls.mock.__exit__(None, None, None)

    def dead_provider(self, **kwargs) -> RegistryProvider:
        dead = [
            Endpoint(
                "прошитый",
                "GET",
                "http://127.0.0.1:9/api/dead",
                params=lambda inn: {"inn": inn},
            )
        ]
        return RegistryProvider(
            SOURCE_NOSTROY, dead, timeout=2, attempts=1, logger=lambda _m: None, **kwargs
        )

    def test_lookup_uses_discovered_address(self):
        provider = self.dead_provider()
        with patch.object(registries_mod, "discover", return_value=[self.mock.url]):
            answer = provider.lookup(mock.MEMBER_INN)
        self.assertIs(answer.outcome, Outcome.FOUND)
        self.assertIsNotNone(provider.working, "найденный адрес должен запомниться как рабочий")

    def test_discovery_happens_once_not_per_company(self):
        provider = self.dead_provider()
        with patch.object(registries_mod, "discover", return_value=[]) as spy:
            for _ in range(3):
                provider.lookup(mock.MEMBER_INN)
        self.assertEqual(spy.call_count, 1, "393 компании не должны вызвать 393 поиска адреса")

    def test_manual_address_disables_search(self):
        """Адрес задан руками — значит, искать нечего, идём строго по нему."""
        provider = self.dead_provider(allow_discovery=False)
        with patch.object(registries_mod, "discover", return_value=[self.mock.url]) as spy:
            answer = provider.lookup(mock.MEMBER_INN)
        self.assertIs(answer.outcome, Outcome.UNKNOWN)
        spy.assert_not_called()

    def test_adopted_address_is_tried_first(self):
        """Адрес, подсмотренный браузером, должен сразу пойти в дело."""
        provider = self.dead_provider()
        provider.adopt_url(self.mock.url + "?inn=123")
        self.assertTrue(provider.endpoints[0].url.startswith(self.mock.url))
        self.assertNotIn("?", provider.endpoints[0].url, "параметры запроса адресом не являются")

        answer = provider.lookup(mock.MEMBER_INN)
        self.assertIs(answer.outcome, Outcome.FOUND)

    def test_adopting_the_same_address_twice_changes_nothing(self):
        provider = self.dead_provider()
        before = len(provider.endpoints)
        provider.adopt_url(self.mock.url)
        after_first = len(provider.endpoints)
        provider.adopt_url(self.mock.url)
        self.assertGreater(after_first, before)
        self.assertEqual(len(provider.endpoints), after_first)

    def test_session_is_warmed_up_once_not_per_company(self):
        """Реестр ставит PHPSESSID — заходим за ним один раз, а не 393."""
        provider = RegistryProvider(
            SOURCE_NOSTROY,
            list(registries_mod.NOSTROY_ENDPOINTS)[:1],
            timeout=1,
            attempts=1,
            logger=lambda _m: None,
        )
        with patch.object(provider.session, "get") as visit, patch.object(
            provider, "_try_endpoints", return_value=None
        ), patch.object(provider, "find_new_endpoints", return_value=[]):
            provider.lookup("7702521529")
            provider.lookup("7707083893")
        self.assertEqual(visit.call_count, 1)
        self.assertTrue(visit.call_args[0][0].startswith("https://reestr.nostroy.ru"))

    def test_manual_address_skips_the_warm_up(self):
        """Задан свой адрес — ходить на чужой сайт за cookie незачем."""
        provider = self.dead_provider(allow_discovery=False)
        with patch.object(provider.session, "get") as visit:
            provider.warm_up()
        visit.assert_not_called()

    def test_failed_search_still_reports_honestly(self):
        provider = self.dead_provider()
        with patch.object(registries_mod, "discover", side_effect=RuntimeError("сеть легла")):
            answer = provider.lookup(mock.MEMBER_INN)
        self.assertIs(answer.outcome, Outcome.UNKNOWN, "сбой поиска — это «не проверено», а не «нет»")


class TestProviderOverHttp(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.mock = MockRegistry()
        cls.mock.__enter__()

    @classmethod
    def tearDownClass(cls):
        cls.mock.__exit__(None, None, None)

    def provider(self) -> RegistryProvider:
        providers = build_providers(timeout=5, attempts=1, nostroy_url=self.mock.url)
        return providers[SOURCE_NOSTROY]

    def test_found_over_http(self):
        answer = self.provider().lookup(mock.MEMBER_INN)
        self.assertIs(answer.outcome, Outcome.FOUND)

    def test_empty_over_http(self):
        answer = self.provider().lookup(mock.NON_MEMBER_INN)
        self.assertIs(answer.outcome, Outcome.EMPTY)

    def test_server_error(self):
        answer = self.provider().lookup(mock.SERVER_ERROR_INN)
        self.assertIs(answer.outcome, Outcome.UNKNOWN)

    def test_server_error_note_quotes_the_server(self):
        """Без текста ответа «ошибка сервера 500» не даёт ничего для починки."""
        answer = self.provider().lookup(mock.SERVER_ERROR_INN)
        self.assertIn("500", answer.note)
        self.assertIn("internal", answer.note, "в примечание должно попасть тело ответа")

    def test_html_instead_of_json(self):
        """Классическая ловушка: заглушка сайта отдаёт HTML со статусом 200."""
        answer = self.provider().lookup(mock.HTML_INN)
        self.assertIs(answer.outcome, Outcome.UNKNOWN)

    def test_connection_dropped(self):
        answer = self.provider().lookup(mock.TIMEOUT_INN)
        self.assertIs(answer.outcome, Outcome.UNKNOWN)

    def test_unreachable_host(self):
        providers = build_providers(
            timeout=2, attempts=1, nostroy_url="http://127.0.0.1:9/api/search"
        )
        answer = providers[SOURCE_NOSTROY].lookup(mock.MEMBER_INN)
        self.assertIs(answer.outcome, Outcome.UNKNOWN)


# ---------------------------------------------------------------------------
# Сведение результата по двум реестрам
# ---------------------------------------------------------------------------


class TestVerdict(unittest.TestCase):
    def build(self, *outcomes):
        from sro.models import CheckResult, RegistryAnswer

        result = CheckResult(inn="7702521529")
        for source, outcome in outcomes:
            result.answers.append(RegistryAnswer(source, outcome))
        return result

    def test_found_anywhere_is_yes(self):
        result = self.build((SOURCE_NOSTROY, Outcome.EMPTY), (SOURCE_NOPRIZ, Outcome.FOUND))
        self.assertIs(result.verdict, Verdict.YES)

    def test_both_empty_is_no(self):
        result = self.build((SOURCE_NOSTROY, Outcome.EMPTY), (SOURCE_NOPRIZ, Outcome.EMPTY))
        self.assertIs(result.verdict, Verdict.NO)

    def test_one_unknown_spoils_the_no(self):
        """Ключевое правило: не проверили один реестр — значит «не знаю»."""
        result = self.build((SOURCE_NOSTROY, Outcome.EMPTY), (SOURCE_NOPRIZ, Outcome.UNKNOWN))
        self.assertIs(result.verdict, Verdict.UNKNOWN)

    def test_no_answers_is_unknown(self):
        self.assertIs(self.build().verdict, Verdict.UNKNOWN)

    def test_unknown_plus_found_is_yes(self):
        result = self.build((SOURCE_NOSTROY, Outcome.UNKNOWN), (SOURCE_NOPRIZ, Outcome.FOUND))
        self.assertIs(result.verdict, Verdict.YES)

    def with_membership(self, status_raw: str, basis: str):
        from sro.models import CheckResult, Membership, RegistryAnswer

        membership = Membership(source=SOURCE_NOSTROY, sro_name="СРО", status_raw=status_raw)
        answer = RegistryAnswer(SOURCE_NOSTROY, Outcome.FOUND, memberships=[membership])
        return CheckResult(inn="7702521529", answers=[answer], basis=basis)

    def test_basis_found_counts_any_record(self):
        for status in ("Является членом СРО", "Приостановлено", "Исключен"):
            self.assertIs(self.with_membership(status, "found").verdict, Verdict.YES, status)

    def test_basis_active_requires_live_membership(self):
        self.assertIs(self.with_membership("Является членом СРО", "active").verdict, Verdict.YES)
        self.assertIs(self.with_membership("Приостановлено", "active").verdict, Verdict.NO)
        self.assertIs(self.with_membership("Исключен", "active").verdict, Verdict.NO)

    def test_basis_active_does_not_guess_unknown_status(self):
        """Статус не распознан — «нет» ставить нельзя."""
        self.assertIs(self.with_membership("нечто невнятное", "active").verdict, Verdict.UNKNOWN)


class TestRegistryOrder(unittest.TestCase):
    def test_builders_go_to_nostroy_first(self):
        self.assertEqual(excel_io.registry_order("Строители")[0], SOURCE_NOSTROY)

    def test_designers_go_to_nopriz_first(self):
        for priority in ("Проектировщики", "проектировщики/изыскатели", "Изыскатели"):
            self.assertEqual(excel_io.registry_order(priority)[0], SOURCE_NOPRIZ, priority)

    def test_unknown_priority_defaults_to_nostroy(self):
        self.assertEqual(excel_io.registry_order("")[0], SOURCE_NOSTROY)


# ---------------------------------------------------------------------------
# Полный проход по файлу
# ---------------------------------------------------------------------------

COMPANIES = [
    ("ООО «Мостострой»", mock.MEMBER_INN, "Строители"),
    ("АО «Проектстрой»", mock.SUSPENDED_INN, "Проектировщики"),
    ("ООО «Бывший член»", mock.EXCLUDED_INN, "Строители"),
    ("ПАО «Не в СРО»", mock.NON_MEMBER_INN, "Строители"),
    ("ООО «Сервер лёг»", mock.SERVER_ERROR_INN, "Строители"),
    ("ООО «HTML вместо JSON»", mock.HTML_INN, "Строители"),
    ("ООО «Странный JSON»", mock.WEIRD_JSON_INN, "Проектировщики"),
    ("ООО «Фильтр не сработал»", mock.UNFILTERED_INN, "Строители"),
    ("ООО «Битый ИНН»", "12345", "Строители"),
    ("ООО «Без ИНН»", "", "Строители"),
]


def make_source_file(path: str) -> None:
    """Собрать файл, похожий на настоящий: заголовок, шапка, цвета."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Лиды по приоритету"

    worksheet["A1"] = "Список компаний для обзвона"  # строка-заголовок над шапкой
    worksheet["A1"].font = Font(bold=True, size=14)

    headers = ["Поставщик", "ИНН", "Приоритет", "Есть СРО? (проверить в НОСТРОЙ)"]
    for index, title in enumerate(headers, start=1):
        cell = worksheet.cell(row=2, column=index, value=title)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")

    for offset, (name, inn, priority) in enumerate(COMPANIES):
        row = 3 + offset
        worksheet.cell(row=row, column=1, value=name)
        # ИНН специально кладём числом, как это делает Excel
        worksheet.cell(row=row, column=2, value=int(inn) if inn.isdigit() else inn)
        worksheet.cell(row=row, column=3, value=priority)
        if offset == 0:
            worksheet.cell(row=row, column=1).fill = PatternFill("solid", fgColor="FFF2CC")

    workbook.save(path)


class TestEndToEnd(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.mock = MockRegistry()
        cls.mock.__enter__()

    @classmethod
    def tearDownClass(cls):
        cls.mock.__exit__(None, None, None)

    def setUp(self):
        self.folder = tempfile.mkdtemp(prefix="sro-test-")
        self.source = os.path.join(self.folder, "Компании_Список.xlsx")
        self.output = os.path.join(self.folder, "Результат.xlsx")
        make_source_file(self.source)
        self.mock.requests_log.clear()

    def run_tool(self, *extra: str) -> int:
        return self._silently(
            [
                "--input", self.source,
                "--sheet", "Лиды по приоритету",
                "--output", self.output,
                "--cache", os.path.join(self.folder, "кэш.json"),
                "--log", os.path.join(self.folder, "лог.txt"),
                "--nostroy-url", self.mock.url,
                "--nopriz-url", self.mock.url,
                "--browser", "never",
                "--delay-min", "0",
                "--delay-max", "0",
                "--attempts", "1",
                "--timeout", "5",
                *extra,
            ]
        )

    @staticmethod
    def _silently(argv: list[str]) -> int:
        """Запустить программу, не засоряя вывод тестов её прогрессом."""
        buffer = io.StringIO()
        with contextlib.redirect_stdout(buffer):
            return check_sro.main(argv)

    def read_output(self) -> dict[str, dict]:
        workbook = load_workbook(self.output)
        worksheet = workbook["Лиды по приоритету"]
        headers = {
            excel_io.normalize_header(worksheet.cell(row=2, column=column).value): column
            for column in range(1, worksheet.max_column + 1)
        }
        rows: dict[str, dict] = {}
        for row in range(3, worksheet.max_row + 1):
            name = worksheet.cell(row=row, column=1).value
            if not name:
                continue
            rows[name] = {
                title: worksheet.cell(row=row, column=column).value
                for title, column in headers.items()
            }
        return rows

    # -- сами проверки --------------------------------------------------

    def test_full_run(self):
        self.assertEqual(self.run_tool("--verdict-basis", "found"), 0)
        self.assertTrue(os.path.exists(self.output))
        rows = self.read_output()

        # Колонку вердикта ищем так же, как это делает программа
        verdict_key = next(key for key in next(iter(rows.values())) if key.startswith("есть сро"))

        def check(company: str, expected: str) -> dict:
            self.assertIn(company, rows)
            self.assertEqual(rows[company][verdict_key], expected, company)
            return rows[company]

        found = check("ООО «Мостострой»", "да")
        self.assertIn("Строители России", found["название сро"])
        self.assertEqual(found["реестровый номер"], "СРО-С-123-45678910")
        self.assertEqual(found["статус членства"], "действует")
        self.assertEqual(found["источник"], "НОСТРОЙ")
        self.assertTrue(found["дата проверки"])

        self.assertEqual(check("АО «Проектстрой»", "да")["статус членства"], "приостановлено")
        self.assertEqual(check("ООО «Бывший член»", "да")["статус членства"], "исключён")

        # Достоверное «нет»: оба реестра ответили и ничего не нашли
        check("ПАО «Не в СРО»", "нет")

        # Всё, что сломалось, обязано стать «не удалось проверить»
        for company in (
            "ООО «Сервер лёг»",
            "ООО «HTML вместо JSON»",
            "ООО «Странный JSON»",
            "ООО «Фильтр не сработал»",
            "ООО «Битый ИНН»",
            "ООО «Без ИНН»",
        ):
            check(company, "не удалось проверить")

    def test_verdict_basis_active(self):
        """При правиле «active» членство должно быть именно действующим."""
        self.assertEqual(self.run_tool("--verdict-basis", "active"), 0)
        rows = self.read_output()
        verdict_key = next(key for key in next(iter(rows.values())) if key.startswith("есть сро"))

        self.assertEqual(rows["ООО «Мостострой»"][verdict_key], "да")

        # Приостановленные и исключённые — «нет», но сведения о СРО сохраняются,
        # чтобы было видно, на каком основании принято решение.
        for company, status in (
            ("АО «Проектстрой»", "приостановлено"),
            ("ООО «Бывший член»", "исключён"),
        ):
            self.assertEqual(rows[company][verdict_key], "нет", company)
            self.assertEqual(rows[company]["статус членства"], status)
            self.assertIn("Строители России", rows[company]["название сро"])

    def test_broken_inn_never_hits_the_network(self):
        self.run_tool()
        self.assertNotIn("12345", self.mock.requests_log)
        self.assertIn(mock.MEMBER_INN, self.mock.requests_log)

    def test_original_file_untouched(self):
        with open(self.source, "rb") as handle:
            before = handle.read()
        self.run_tool()
        with open(self.source, "rb") as handle:
            self.assertEqual(handle.read(), before)

    def test_formatting_preserved(self):
        self.run_tool()
        workbook = load_workbook(self.output)
        worksheet = workbook["Лиды по приоритету"]
        self.assertEqual(worksheet["A1"].value, "Список компаний для обзвона")
        self.assertTrue(worksheet["A1"].font.bold)
        self.assertTrue(worksheet["A2"].font.bold)
        self.assertEqual(worksheet["A2"].fill.fgColor.rgb[-6:], "DDEBF7")
        self.assertEqual(worksheet["A3"].fill.fgColor.rgb[-6:], "FFF2CC")

    def test_cache_prevents_repeat_requests(self):
        self.run_tool()
        first_pass = len(self.mock.requests_log)
        self.assertGreater(first_pass, 0)

        self.mock.requests_log.clear()
        self.run_tool()
        # Повторный запуск: «да»/«нет» берутся из кэша, заново спрашиваем
        # только те строки, что в прошлый раз не удалось проверить.
        repeated = set(self.mock.requests_log)
        self.assertNotIn(mock.MEMBER_INN, repeated)
        self.assertNotIn(mock.NON_MEMBER_INN, repeated)
        self.assertIn(mock.SERVER_ERROR_INN, repeated)

    def test_retry_failed_only_touches_unresolved_rows(self):
        self.run_tool()
        self.mock.requests_log.clear()

        # Скармливаем результат обратно и просим перепроверить только сбои
        self.source = self.output
        self.output = os.path.join(self.folder, "Результат2.xlsx")
        self.assertEqual(self.run_tool("--retry-failed"), 0)

        touched = set(self.mock.requests_log)
        self.assertNotIn(mock.MEMBER_INN, touched)
        self.assertNotIn(mock.NON_MEMBER_INN, touched)
        self.assertIn(mock.HTML_INN, touched)

        # Ранее полученные ответы в файле сохранились
        rows = self.read_output()
        verdict_key = next(key for key in next(iter(rows.values())) if key.startswith("есть сро"))
        self.assertEqual(rows["ООО «Мостострой»"][verdict_key], "да")
        self.assertEqual(rows["ПАО «Не в СРО»"][verdict_key], "нет")

    def test_checkpoint_file_appears_during_run(self):
        self.run_tool("--checkpoint-every", "2")
        self.assertTrue(os.path.exists(self.output))
        cache = os.path.join(self.folder, "кэш.json")
        self.assertTrue(os.path.exists(cache))

    def test_registry_down_never_produces_a_no(self):
        """Реестр недоступен целиком — ни одной строки «нет» появиться не должно."""
        self.output = os.path.join(self.folder, "Результат_офлайн.xlsx")
        code = self._silently(
            [
                "--input", self.source,
                "--sheet", "Лиды по приоритету",
                "--output", self.output,
                "--cache", os.path.join(self.folder, "кэш_офлайн.json"),
                "--log", os.path.join(self.folder, "лог_офлайн.txt"),
                "--nostroy-url", "http://127.0.0.1:9/api/search",
                "--nopriz-url", "http://127.0.0.1:9/api/search",
                "--browser", "never",
                "--delay-min", "0", "--delay-max", "0",
                "--attempts", "1", "--timeout", "2",
            ]
        )
        self.assertEqual(code, 0)
        rows = self.read_output()
        verdict_key = next(key for key in next(iter(rows.values())) if key.startswith("есть сро"))
        verdicts = {row[verdict_key] for row in rows.values()}
        self.assertEqual(verdicts, {"не удалось проверить"})


class TestGuiPlumbing(unittest.TestCase):
    """Механика, на которой держится оконная версия (gui.py).

    Окно не может пользоваться ни консолью, ни Ctrl+C, поэтому у ядра есть
    приёмник строк (`sink`), событие отмены и обратный вызов прогресса.
    """

    @classmethod
    def setUpClass(cls):
        cls.mock = MockRegistry()
        cls.mock.__enter__()

    @classmethod
    def tearDownClass(cls):
        cls.mock.__exit__(None, None, None)

    def setUp(self):
        self.folder = tempfile.mkdtemp(prefix="sro-gui-")
        self.source = os.path.join(self.folder, "Компании_Список.xlsx")
        self.output = os.path.join(self.folder, "Результат.xlsx")
        make_source_file(self.source)

    def make_args(self, *extra: str):
        return check_sro.parse_args(
            [
                "--input", self.source,
                "--sheet", "Лиды по приоритету",
                "--output", self.output,
                "--cache", os.path.join(self.folder, "кэш.json"),
                "--log", os.path.join(self.folder, "лог.txt"),
                "--nostroy-url", self.mock.url,
                "--nopriz-url", self.mock.url,
                "--browser", "never",
                "--delay-min", "0",
                "--delay-max", "0",
                "--attempts", "1",
                "--timeout", "5",
                *extra,
            ]
        )

    @staticmethod
    def quiet_log() -> check_sro.Log:
        return check_sro.Log(None, sink=lambda _message: None)

    def verdict_column_values(self) -> list:
        workbook = load_workbook(self.output)
        worksheet = workbook["Лиды по приоритету"]
        headers = {
            excel_io.normalize_header(worksheet.cell(row=2, column=column).value): column
            for column in range(1, worksheet.max_column + 1)
        }
        column = next(index for key, index in headers.items() if key.startswith("есть сро"))
        return [
            worksheet.cell(row=row, column=column).value
            for row in range(3, 3 + len(COMPANIES))
        ]

    # -- сами проверки --------------------------------------------------

    def test_sink_receives_lines_instead_of_console(self):
        collected: list[str] = []
        log = check_sro.Log(None, sink=collected.append)
        log("первая строка")
        log()
        self.assertEqual(collected, ["первая строка", ""])

    def test_progress_reports_every_company(self):
        seen: list[tuple] = []
        code = check_sro.run(
            self.make_args(),
            self.quiet_log(),
            progress=lambda processed, total, counters: seen.append(
                (processed, total, dict(counters))
            ),
        )
        self.assertEqual(code, 0)
        self.assertEqual([processed for processed, _, _ in seen], list(range(1, len(COMPANIES) + 1)))

        processed, total, counters = seen[-1]
        self.assertEqual(total, len(COMPANIES))
        # Счётчики в окне должны сходиться с числом обработанных строк
        self.assertEqual(counters["yes"] + counters["no"] + counters["unknown"], processed)

    def test_cancel_stops_early_and_keeps_checked_rows(self):
        cancel = threading.Event()

        def progress(processed: int, _total: int, _counters: dict) -> None:
            if processed == 3:
                cancel.set()  # то же самое, что нажать «Остановить»

        code = check_sro.run(self.make_args(), self.quiet_log(), cancel=cancel, progress=progress)

        self.assertEqual(code, 130)  # тот же код возврата, что и у Ctrl+C
        self.assertTrue(os.path.exists(self.output), "результат должен сохраниться и после отмены")
        filled = [value for value in self.verdict_column_values() if value]
        self.assertEqual(len(filled), 3, "успели проверить ровно три строки — они и должны остаться")

    def test_cancel_before_start_checks_nobody(self):
        cancel = threading.Event()
        cancel.set()
        code = check_sro.run(self.make_args(), self.quiet_log(), cancel=cancel)
        self.assertEqual(code, 130)
        self.assertEqual([value for value in self.verdict_column_values() if value], [])


class TestResultFileIsReadable(unittest.TestCase):
    """Оформление результата: 393 строки без шапки и фильтра читать невозможно."""

    @classmethod
    def setUpClass(cls):
        cls.mock = MockRegistry()
        cls.mock.__enter__()

    @classmethod
    def tearDownClass(cls):
        cls.mock.__exit__(None, None, None)

    def setUp(self):
        self.folder = tempfile.mkdtemp(prefix="sro-view-")
        self.source = os.path.join(self.folder, "Компании_Список.xlsx")
        self.output = os.path.join(self.folder, "Результат.xlsx")
        make_source_file(self.source)
        self.run_tool()

    def run_tool(self) -> None:
        buffer = io.StringIO()
        with contextlib.redirect_stdout(buffer):
            check_sro.main(
                [
                    "--input", self.source,
                    "--sheet", "Лиды по приоритету",
                    "--output", self.output,
                    "--cache", os.path.join(self.folder, "кэш.json"),
                    "--log", os.path.join(self.folder, "лог.txt"),
                    "--nostroy-url", self.mock.url,
                    "--nopriz-url", self.mock.url,
                    "--browser", "never",
                    "--delay-min", "0", "--delay-max", "0",
                    "--attempts", "1", "--timeout", "5",
                ]
            )

    def test_header_is_frozen_and_filter_is_on(self):
        sheet = load_workbook(self.output)["Лиды по приоритету"]
        # Шапка в образце — во второй строке, значит закрепляем по третью
        self.assertEqual(sheet.freeze_panes, "A3")
        self.assertTrue(sheet.auto_filter.ref, "должен быть включён автофильтр")

    def test_summary_sheet_reports_the_run(self):
        workbook = load_workbook(self.output)
        self.assertIn(excel_io.SUMMARY_SHEET, workbook.sheetnames)
        sheet = workbook[excel_io.SUMMARY_SHEET]
        labels = {sheet.cell(row=row, column=1).value for row in range(1, 20)}
        for expected in ("Проверено компаний", "Не удалось проверить", "Есть СРО"):
            self.assertIn(expected, labels)

    def test_summary_survives_second_run(self):
        """Повторный запуск не должен плодить листы «Итоги проверки (2)»."""
        self.run_tool()
        names_after = load_workbook(self.output).sheetnames
        self.assertEqual(names_after.count(excel_io.SUMMARY_SHEET), 1)


class TestWrongInnIsCaught(unittest.TestCase):
    """Опечатка в ИНН: реестр отвечает «да», но про совершенно другую компанию."""

    @classmethod
    def setUpClass(cls):
        cls.mock = MockRegistry()
        cls.mock.__enter__()

    @classmethod
    def tearDownClass(cls):
        cls.mock.__exit__(None, None, None)

    def setUp(self):
        self.folder = tempfile.mkdtemp(prefix="sro-mismatch-")
        self.source = os.path.join(self.folder, "Один.xlsx")
        self.output = os.path.join(self.folder, "Один_результат.xlsx")

    def build(self, company: str) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Лиды по приоритету"
        for index, title in enumerate(["Поставщик", "ИНН", "Приоритет"], start=1):
            worksheet.cell(row=1, column=index, value=title)
        worksheet.cell(row=2, column=1, value=company)
        worksheet.cell(row=2, column=2, value=int(mock.MEMBER_INN))
        worksheet.cell(row=2, column=3, value="Строители")
        workbook.save(self.source)

    def note_for(self, company: str) -> str:
        self.build(company)
        buffer = io.StringIO()
        with contextlib.redirect_stdout(buffer):
            check_sro.main(
                [
                    "--input", self.source,
                    "--sheet", "Лиды по приоритету",
                    "--output", self.output,
                    "--cache", os.path.join(self.folder, "кэш.json"),
                    "--log", os.path.join(self.folder, "лог.txt"),
                    "--nostroy-url", self.mock.url,
                    "--nopriz-url", self.mock.url,
                    "--browser", "never",
                    "--delay-min", "0", "--delay-max", "0",
                    "--attempts", "1", "--timeout", "5",
                ]
            )
        sheet = load_workbook(self.output)["Лиды по приоритету"]
        headers = {
            excel_io.normalize_header(sheet.cell(row=1, column=column).value): column
            for column in range(1, sheet.max_column + 1)
        }
        return str(sheet.cell(row=2, column=headers["примечание"]).value or "")

    def test_wrong_company_is_flagged(self):
        note = self.note_for("ООО «Совершенно другая контора»")
        self.assertIn("ВНИМАНИЕ", note)
        self.assertIn("Мостострой", note, "в примечании должно быть название из реестра")

    def test_matching_company_is_not_flagged(self):
        note = self.note_for('ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "МОСТОСТРОЙ"')
        self.assertNotIn("ВНИМАНИЕ", note)


if __name__ == "__main__":
    unittest.main(verbosity=2)
