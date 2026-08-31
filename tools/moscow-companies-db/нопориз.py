"""Проверка компаний по реестру НОПРИЗ: проектирование и изыскания.

НОПРИЗ — национальное объединение изыскателей и проектировщиков. Его СРО
делятся на два вида деятельности: архитектурно-строительное проектирование
и инженерные изыскания. Компания может состоять в СРО одного вида, другого
или обоих сразу — поэтому в отчёте две отдельные колонки с датами
вступления.

ПОЧЕМУ СКАЧИВАЕМ ВЕСЬ РЕЕСТР, А НЕ ИЩЕМ ПО ОДНОМУ ИНН

Реестры этой платформы молча игнорируют незнакомый формат фильтра и
возвращают общий список вместо отфильтрованного — на НОСТРОЕ это уже
приводило к тихим «0 совпадений». Постраничная выгрузка целиком таких
сюрпризов не даёт: скачиваем один раз, дальше сверяем локально по ИНН,
сколько угодно раз и мгновенно.

ПОРЯДОК РАБОТЫ

1) Скачать реестр (один раз; он большой — 212 тысяч членств):

       .venv\\Scripts\\activate
       python нопориз.py выгрузка --out нопориз_реестр.xlsx

   Прерывать можно: прогресс сохраняется, повтор с --resume продолжит
   с места обрыва.

3) Проверить свои ИНН точечно, без общей выгрузки — надёжнее и быстрее:

       python нопориз.py проверить --file мои_компании.xlsx ^
              --out проверка.xlsx

2) Сверить свой файл (колонка ИНН обязательна):

       python нопориз.py сверить --file реестр_сро.xlsx ^
              --nopriz нопориз_реестр.xlsx --out результат.xlsx

   Допишет колонки «Проектирование: дата вступления» и «Изыскания: дата
   вступления», плюс названия СРО и статус членства.

ЕСЛИ ЧТО-ТО НЕ СХОДИТСЯ

       python нопориз.py образец

   покажет сырую запись реестра и что скрипт из неё извлёк — по этому
   выводу видно, туда ли он смотрит.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from datetime import date
from pathlib import Path
from typing import Any, Iterator

import requests

BASE = "https://reestr.nopriz.ru"
MEMBER_LIST_URL = f"{BASE}/api/sro/all/member/list"


# Дата вступления. Первым — точное имя поля, каким его отдаёт платформа
# (проверено на НОСТРОЕ: registry_registration_date вида
# "2018-02-12T00:00:00+03:00"), дальше запасные варианты.
_START_KEY_HINTS = (
    "registry_registration_date",
    "member_right_start",
    "right_start",
    "вступ",
    "начал",
    "start_date",
)
# Похожи на дату вступления, но ей не являются: госрегистрация самой компании
_STATE_REG_MARKERS = ("state_registration", "ogrn", "егрюл", "егрип")
_STOP_KEY_HINTS = ("member_right_stop", "right_stop", "stop_date", "прекращ", "исключ")
_FORMER_MARKERS = ("исключ", "прекращ")
_ACTIVE_MARKERS = ("является членом", "действ", "член ")

# Ветки записи, где лежат данные самой СРО, а не членство компании
_SKIP_BRANCHES = ("sro", "sro_info", "organization")

DESIGN = "проектирование"
SURVEY = "изыскания"

# Вид деятельности СРО: сначала по регистрационному номеру (он прямо кодирует
# вид буквой после «СРО-»), запасной путь — по словам в наименовании СРО
_SRO_NUMBER = re.compile(r"СРО\s*-\s*([ПИ])\s*-", re.I)
_SRO_NUMBER_KIND = {"П": DESIGN, "И": SURVEY}
# «проект» и «архитект» намеренно короткие: СРО зовутся «Межрегионпроект»,
# «СТРОЙПРОЕКТГАРАНТ», «Гильдия архитекторов» — по «проектир»/«проектн» такие
# не ловились, а это 6 тысяч записей
_DESIGN_WORDS = ("проект", "архитект")
_SURVEY_WORDS = ("изыскан", "изыскател")


# -- разбор ответов ----------------------------------------------------


def _strings(obj: Any) -> Iterator[str]:
    if isinstance(obj, str):
        yield obj
    elif isinstance(obj, dict):
        for value in obj.values():
            yield from _strings(value)
    elif isinstance(obj, list):
        for item in obj:
            yield from _strings(item)


def norm_inn(value: Any) -> str:
    """ИНН к сравнимому виду: только цифры и восстановленный ведущий ноль.

    Excel хранит ИНН числом, поэтому у компаний из регионов с кодом 01–09
    ведущий ноль теряется: 0816034124 приезжает как 816034124. Без
    восстановления такие компании молча не находятся в реестре.
    """
    digits = re.sub(r"\D", "", "" if value is None else str(value))
    if len(digits) == 9:
        return digits.zfill(10)
    if len(digits) == 11:
        return digits.zfill(12)
    return digits


def _parse_date(text: str) -> date | None:
    text = (text or "").strip()
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", text)
    if m:
        day, month, year = map(int, m.groups())
    else:
        m = re.match(r"(\d{4})-(\d{2})-(\d{2})", text)
        if not m:
            return None
        year, month, day = map(int, m.groups())
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _dates_by_hints(record: dict, hints: tuple[str, ...]) -> list[tuple[date, str]]:
    for hint in hints:
        found: list[tuple[date, str]] = []

        def rec(obj: Any, path: str = "") -> None:
            if isinstance(obj, dict):
                for key, value in obj.items():
                    if key.lower() in _SKIP_BRANCHES:
                        continue
                    full = f"{path}.{key}" if path else key
                    low = key.lower()
                    if (isinstance(value, str) and hint in low
                            and not any(m in low for m in _STATE_REG_MARKERS)):
                        parsed = _parse_date(value)
                        if parsed:
                            found.append((parsed, full))
                    elif isinstance(value, (dict, list)):
                        rec(value, full)
            elif isinstance(obj, list):
                for item in obj:
                    rec(item, path)

        rec(record)
        if found:
            return found
    return []


def start_date(record: dict) -> date | None:
    """Дата вступления: самая ранняя из найденных (позже идут продления)."""
    found = _dates_by_hints(record, _START_KEY_HINTS)
    return min(d for d, _k in found) if found else None


def _first_str(record: dict, *keys: str) -> str:
    for key in keys:
        value = record.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def sro_name(record: dict) -> str:
    sro = record.get("sro")
    if isinstance(sro, dict):
        return _first_str(sro, "short_description", "title", "full_description", "name")
    return sro.strip() if isinstance(sro, str) else ""


def sro_id(record: dict) -> str:
    sro = record.get("sro")
    if isinstance(sro, dict):
        for key in ("id", "registry_number", "number"):
            value = sro.get(key)
            if isinstance(value, (str, int)) and str(value).strip():
                return str(value).strip()
    return ""


def member_status(record: dict) -> str:
    status = record.get("member_status")
    if isinstance(status, dict):
        return _first_str(status, "title", "name")
    return status.strip() if isinstance(status, str) else ""


def is_former(record: dict) -> bool:
    """Решает текст статуса: у действующих членов дата окончания права
    бывает плановой (в будущем), и считать её признаком выхода нельзя."""
    status = member_status(record).lower()
    if any(m in status for m in _FORMER_MARKERS):
        return True
    if any(m in status for m in _ACTIVE_MARKERS):
        return False
    stopped = _dates_by_hints(record, _STOP_KEY_HINTS)
    if not stopped:
        return False
    return min(d for d, _k in stopped) <= date.today()


def sro_number(record: dict) -> str:
    """Регистрационный номер СРО, например «СРО-П-147-09032010»."""
    sro = record.get("sro")
    if isinstance(sro, dict):
        return _first_str(sro, "registration_number", "registry_number", "number")
    return ""


def activity_kind(record: dict) -> str:
    """«проектирование», «изыскания», оба через запятую или '' если неясно.

    Главный признак — регистрационный номер СРО: у НОПРИЗ он прямо кодирует
    вид деятельности (СРО-П-… проектные, СРО-И-… изыскательские). Это
    надёжнее любых названий и не зависит от того, как СРО себя назвала.

    Запасной путь — слова в наименовании самой СРО. Раньше смотрели на все
    строки записи подряд, и это было ошибкой: ООО «Проектстрой», состоящее
    в изыскательской СРО, получало вид «проектирование» из собственного
    названия.
    """
    match = _SRO_NUMBER.search(sro_number(record))
    if match:
        return _SRO_NUMBER_KIND[match.group(1).upper()]
    return kind_by_sro_name(sro_name(record))


def kind_by_sro_name(name: str) -> str:
    """Вид деятельности по одному лишь наименованию СРО."""
    text = (name or "").lower()
    kinds = []
    if any(w in text for w in _DESIGN_WORDS):
        kinds.append(DESIGN)
    if any(w in text for w in _SURVEY_WORDS):
        kinds.append(SURVEY)
    return ", ".join(kinds)


def kind_of_row(row: dict, guide: dict[str, str] | None = None) -> str:
    """Вид деятельности для строки готовой выгрузки.

    Порядок доверия: регистрационный номер СРО (если он в выгрузке есть),
    затем ручной справочник по СРО, затем слова в наименовании СРО.
    """
    match = _SRO_NUMBER.search(row.get("Рег. номер СРО") or "")
    if match:
        return _SRO_NUMBER_KIND[match.group(1).upper()]
    if guide:
        for key in (str(row.get("ID СРО") or "").strip(),
                    (row.get("СРО") or "").strip().lower()):
            if key and guide.get(key):
                return guide[key]
    return kind_by_sro_name(row.get("СРО") or "")


def records_of(payload: Any) -> list[dict]:
    if not isinstance(payload, dict):
        return []
    data = payload.get("data")
    if isinstance(data, dict):
        data = data.get("data")
    if isinstance(data, list):
        return [r for r in data if isinstance(r, dict)]
    if isinstance(payload.get("data"), list):
        return [r for r in payload["data"] if isinstance(r, dict)]
    return []


def total_of(payload: Any) -> int | None:
    if not isinstance(payload, dict):
        return None
    data = payload.get("data")
    containers = [data, payload] if isinstance(data, dict) else [payload]
    for container in containers:
        for key in ("count", "total", "totalCount"):
            value = container.get(key)
            if isinstance(value, int):
                return value
    return None


# -- сеть ---------------------------------------------------------------


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "User-Agent": "mosstroybase/0.1 (+open-data harvester)",
        "Content-Type": "application/json",
        "Accept": "application/json",
    })
    return session


def _тело(page: int, page_size: int, search: str = "") -> dict:
    """Тело запроса ровно в том виде, в каком его шлёт сам сайт реестра.

    Подсмотрено в браузере: размер страницы называется pageCount и передаётся
    строкой, а поиск — searchString. Мы полгода слали pageSize, реестр молча
    его игнорировал и отдавал свои 20 записей на страницу.
    """
    return {"filters": {}, "page": page, "pageCount": str(page_size),
            "searchString": search, "sortBy": {}}


def fetch_page(session: requests.Session, page: int, page_size: int,
               timeout: int = 60, attempts: int = 6) -> dict | None:
    """Страница реестра. Реестр периодически отвечает 500 — повторяем."""
    body = _тело(page, page_size)
    for attempt in range(1, attempts + 1):
        try:
            resp = session.post(MEMBER_LIST_URL, json=body, timeout=timeout)
            if resp.status_code == 200:
                return resp.json()
            print(f"[нопориз] страница {page}: HTTP {resp.status_code} "
                  f"(попытка {attempt}/{attempts})", flush=True)
        except (requests.RequestException, ValueError) as exc:
            print(f"[нопориз] страница {page}: {type(exc).__name__} "
                  f"(попытка {attempt}/{attempts})", flush=True)
        time.sleep(min(2 ** attempt, 30))
    return None


def fetch_by_inn(session: requests.Session, inn: str, page_size: int = 50,
                 timeout: int = 60, attempts: int = 4) -> dict | None:
    """Все членства одной компании по ИНН.

    Это тот же поиск, что на сайте реестра, и он даёт то, чего не дала
    постраничная выгрузка: она оборвалась на 147 тысячах записей из
    заявленных 212 тысяч, и часть компаний в неё просто не попала.
    """
    body = _тело(1, page_size, search=inn)
    for attempt in range(1, attempts + 1):
        try:
            resp = session.post(MEMBER_LIST_URL, json=body, timeout=timeout)
            if resp.status_code == 200:
                return resp.json()
            print(f"[нопориз] ИНН {inn}: HTTP {resp.status_code} "
                  f"(попытка {attempt}/{attempts})", flush=True)
        except (requests.RequestException, ValueError) as exc:
            print(f"[нопориз] ИНН {inn}: {type(exc).__name__} "
                  f"(попытка {attempt}/{attempts})", flush=True)
        time.sleep(min(2 ** attempt, 20))
    return None


def членства(payload: Any, inn: str) -> tuple:
    """Из ответа поиска — самое раннее действующее членство каждого вида.

    Поиск идёт по строке, поэтому в ответ может попасть и чужая компания
    (например, если ИНН встретился в другом поле). Сверяем ИНН каждой записи.
    """
    проект = изыск = None
    for record in records_of(payload):
        if norm_inn(record.get("inn")) != inn:
            continue
        if is_former(record):
            continue
        started = start_date(record)
        if started is None:
            continue
        вид = activity_kind(record)
        имя = sro_name(record)
        if DESIGN in вид and (проект is None or started < проект[0]):
            проект = (started, имя)
        if SURVEY in вид and (изыск is None or started < изыск[0]):
            изыск = (started, имя)
    return проект, изыск


# -- файлы --------------------------------------------------------------


_PROGRESS_STEP = 25_000


def _tick(action: str, path: Path, count: int) -> None:
    """Признак жизни на больших файлах.

    В выгрузке НОПРИЗ полтораста тысяч строк: чтение занимает полминуты,
    запись — ещё столько же. Без единой строчки вывода это неотличимо от
    зависшей программы, и её начинают убивать по Ctrl+C.
    """
    if count and count % _PROGRESS_STEP == 0:
        print(f"[файл] {action} {path.name}: {count} строк", flush=True)


def read_rows(path: Path) -> tuple[list[str], list[dict]]:
    if path.suffix.lower() == ".csv":
        import csv
        with open(path, encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh, delimiter=";")
            return list(reader.fieldnames or []), list(reader)
    from openpyxl import load_workbook
    print(f"[файл] открываю {path.name} …", flush=True)
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    rows = []
    for seen, values in enumerate(rows_iter, 1):
        row = {header[i]: ("" if v is None else str(v))
               for i, v in enumerate(values) if i < len(header)}
        if any(row.values()):
            rows.append(row)
        _tick("читаю", path, seen)
    wb.close()
    print(f"[файл] прочитано {path.name}: {len(rows)} строк", flush=True)
    return header, rows


def write_rows(path: Path, columns: list[str], rows: list[dict], sheet: str) -> None:
    if path.suffix.lower() == ".csv":
        import csv
        with open(path, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=columns, delimiter=";",
                                    extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        return
    from openpyxl import Workbook
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(sheet)
    ws.append(columns)
    for written, row in enumerate(rows, 1):
        ws.append([row.get(c, "") for c in columns])
        _tick("пишу", path, written)
    print(f"[файл] сохраняю {path.name} …", flush=True)
    wb.save(path)


# «Рег. номер СРО» держим в выгрузке не для красоты: из него выводится вид
# деятельности, и если классификация вдруг окажется неверной, её можно
# пересчитать по готовому файлу, не выкачивая реестр заново.
def write_sheets(path: Path, sheets: list[tuple[str, list[str], list[dict]]]) -> None:
    """Несколько листов в одном файле.

    У CSV листов нет, поэтому там второй и последующие уезжают в отдельные
    файлы рядом — молча терять их нельзя.
    """
    if path.suffix.lower() == ".csv":
        for i, (name, columns, rows) in enumerate(sheets):
            target = path if i == 0 else path.with_name(f"{path.stem}_{name}.csv")
            write_rows(target, columns, rows, name)
        return
    from openpyxl import Workbook
    wb = Workbook(write_only=True)
    for name, columns, rows in sheets:
        ws = wb.create_sheet(name[:31])
        ws.append(columns)
        for written, row in enumerate(rows, 1):
            ws.append([row.get(c, "") for c in columns])
            _tick("пишу", path, written)
    print(f"[файл] сохраняю {path.name} …", flush=True)
    wb.save(path)


DUMP_COLUMNS = ["ИНН", "Наименование", "ОГРН", "СРО", "ID СРО", "Рег. номер СРО",
                "Вид деятельности", "Дата вступления", "Статус", "Бывший член"]


# -- команды ------------------------------------------------------------


def cmd_dump(args) -> int:
    """Скачивает реестр НОПРИЗ целиком в файл."""
    session = make_session()
    out_path = Path(args.out)
    progress_path = out_path.with_suffix(out_path.suffix + ".progress.json")

    rows: list[dict] = []
    start_page = 1
    if args.resume and progress_path.exists():
        saved = json.loads(progress_path.read_text(encoding="utf-8"))
        rows = saved.get("rows", [])
        start_page = saved.get("next_page", 1)
        print(f"[нопориз] продолжаю со страницы {start_page}, уже собрано {len(rows)}")

    page = start_page
    total = total_pages = None
    actual_size = None
    failed_pages: list[int] = []
    completed = False

    def save_progress(next_page: int) -> None:
        progress_path.write_text(
            json.dumps({"next_page": next_page, "rows": rows}, ensure_ascii=False),
            encoding="utf-8")

    try:
        while True:
            payload = fetch_page(session, page, args.page_size)
            if payload is None:
                failed_pages.append(page)
                print(f"[нопориз] страница {page} не отдалась — пропускаю, "
                      "повторю в конце", file=sys.stderr)
                page += 1
                if total_pages and page > total_pages:
                    break
                if not total_pages and len(failed_pages) >= 3:
                    break
                continue

            records = records_of(payload)
            if not records:
                completed = True
                print(f"[нопориз] страница {page} пустая — конец реестра")
                break

            if actual_size is None:
                # Реестр может отдавать своё число записей, игнорируя pageSize —
                # считать страницы по запрошенному нельзя
                actual_size = len(records)
                total = total_of(payload)
                if actual_size != args.page_size:
                    print(f"[нопориз] реестр отдаёт по {actual_size} записей "
                          f"(запрашивали {args.page_size})")
                if total:
                    total_pages = (total + actual_size - 1) // actual_size
                    print(f"[нопориз] записей в реестре: {total} "
                          f"(~{total_pages} страниц)")

            for record in records:
                started = start_date(record)
                rows.append({
                    "ИНН": norm_inn(record.get("inn")),
                    "Наименование": _first_str(record, "short_description",
                                               "full_description", "name", "title"),
                    "ОГРН": str(record.get("ogrn") or record.get("ogrnip") or "").strip(),
                    "СРО": sro_name(record),
                    "ID СРО": sro_id(record),
                    "Рег. номер СРО": sro_number(record),
                    "Вид деятельности": activity_kind(record),
                    "Дата вступления": started.strftime("%d.%m.%Y") if started else "",
                    "Статус": member_status(record),
                    "Бывший член": "да" if is_former(record) else "",
                })

            if page % 10 == 0 or page == start_page:
                hint = f"/{total_pages}" if total_pages else ""
                print(f"[нопориз] страница {page}{hint}: собрано {len(rows)}", flush=True)
                save_progress(page + 1)

            page += 1
            if total_pages and page > total_pages + 5:
                completed = True
                break
            time.sleep(args.delay)
    except KeyboardInterrupt:
        print("\n[нопориз] прервано — сохраняю прогресс (повтор с --resume)")

    if failed_pages:
        print(f"[нопориз] повторяю {len(failed_pages)} сбойных страниц")
        still: list[int] = []
        for p in failed_pages:
            payload = fetch_page(session, p, args.page_size)
            if payload is None:
                still.append(p)
                continue
            for record in records_of(payload):
                started = start_date(record)
                rows.append({
                    "ИНН": norm_inn(record.get("inn")),
                    "Наименование": _first_str(record, "short_description",
                                               "full_description", "name", "title"),
                    "ОГРН": str(record.get("ogrn") or record.get("ogrnip") or "").strip(),
                    "СРО": sro_name(record), "ID СРО": sro_id(record),
                    "Рег. номер СРО": sro_number(record),
                    "Вид деятельности": activity_kind(record),
                    "Дата вступления": started.strftime("%d.%m.%Y") if started else "",
                    "Статус": member_status(record),
                    "Бывший член": "да" if is_former(record) else "",
                })
            time.sleep(args.delay)
        if still:
            print(f"[нопориз] ВНИМАНИЕ: {len(still)} страниц реестр так и не отдал — "
                  f"выгрузка НЕПОЛНАЯ (страницы {still[:20]})", file=sys.stderr)

    if not rows:
        print("[нопориз] ничего не собрано", file=sys.stderr)
        return 1

    if total and len(rows) < total:
        print(f"[нопориз] ВНИМАНИЕ: реестр заявил {total} записей, собрано "
              f"{len(rows)} — часть данных не попала в выгрузку", file=sys.stderr)

    write_rows(out_path, DUMP_COLUMNS, rows, "НОПРИЗ")
    print(f"[нопориз] готово: {out_path} ({len(rows)} записей)")
    no_kind = sum(1 for r in rows if not r["Вид деятельности"])
    if no_kind:
        print(f"[нопориз] у {no_kind} записей вид деятельности не определился — "
              "если их много, пришлите вывод «образец», поправлю разбор")
    if completed and progress_path.exists():
        progress_path.unlink()
    return 0


AFTER_COLUMNS = ["Проектирование: дней после стройки",
                 "Изыскания: дней после стройки"]

# Пары «колонка с датой вступления в НОПРИЗ -> колонка с разницей в днях»
_AFTER_PAIRS = (("Проектирование: дата вступления", AFTER_COLUMNS[0]),
                ("Изыскания: дата вступления", AFTER_COLUMNS[1]))


def rows_joined_after(rows: list[dict], sro_date_column: str,
                      header: list[str], min_days: int = 1) -> list[dict]:
    """Компании, вступившие в проектную или изыскательскую СРО позже строительной.

    Смысл выборки: сначала компания получила допуск на строительство, а
    проектирование или изыскания добрала потом — то есть расширяла профиль
    уже работающей фирмой, а не заходила в НОПРИЗ сразу.

    Достаточно одного вида: если проектирование добрали позже, а изыскания
    были раньше, компания всё равно попадает в выборку — а разница в днях
    показана по каждому виду отдельно, чтобы это было видно.

    min_days отсекает тех, кто оформил всё одним заходом: на живых данных
    пятая часть выборки вступала в НОПРИЗ через 2–20 дней после строительной
    СРО — это не «расширили профиль потом», а один и тот же поход.
    """
    if sro_date_column not in header:
        print(f"[сверка] колонки «{sro_date_column}» в файле нет — лист "
              f"«вступили после стройки» не строю. Есть: {', '.join(header)}",
              file=sys.stderr)
        return []

    found: list[dict] = []
    for row in rows:
        base = _parse_date(row.get(sro_date_column) or "")
        if base is None:
            continue
        gaps = {}
        for date_column, gap_column in _AFTER_PAIRS:
            joined = _parse_date(row.get(date_column) or "")
            gaps[gap_column] = (joined - base).days if joined else None
        if not any(g is not None and g >= min_days for g in gaps.values()):
            continue
        extra = dict(row)
        for gap_column, days in gaps.items():
            extra[gap_column] = "" if days is None else str(days)
        found.append(extra)
    return found


def cmd_match(args) -> int:
    """Сверяет файл пользователя с выгрузкой НОПРИЗ по ИНН."""
    file_path, nopriz_path = Path(args.file), Path(args.nopriz)
    for p in (file_path, nopriz_path):
        if not p.exists():
            print(f"Файл {p} не найден.", file=sys.stderr)
            return 1

    _h, nopriz_rows = read_rows(nopriz_path)
    # По каждому ИНН может быть несколько членств (проектное и изыскательское,
    # текущее и прошлое) — берём самое раннее действующее по каждому виду
    design: dict[str, tuple[date, str]] = {}
    survey: dict[str, tuple[date, str]] = {}
    for r in nopriz_rows:
        inn = norm_inn(r.get("ИНН"))
        if not inn or (r.get("Бывший член") or "").strip() == "да":
            continue
        started = _parse_date(r.get("Дата вступления") or "")
        if started is None:
            continue
        kind = (r.get("Вид деятельности") or "")
        name = r.get("СРО") or ""
        if DESIGN in kind:
            if inn not in design or started < design[inn][0]:
                design[inn] = (started, name)
        if SURVEY in kind:
            if inn not in survey or started < survey[inn][0]:
                survey[inn] = (started, name)

    header, rows = read_rows(file_path)
    inn_col = args.inn_column
    if inn_col not in header:
        print(f"В файле нет колонки «{inn_col}». Есть: {', '.join(header)}",
              file=sys.stderr)
        return 1

    new_cols = ["Проектирование: дата вступления", "Проектирование: СРО",
                "Изыскания: дата вступления", "Изыскания: СРО", "В НОПРИЗ"]
    columns = header + [c for c in new_cols if c not in header]

    both = only_design = only_survey = neither = no_inn = 0
    for row in rows:
        inn = norm_inn(row.get(inn_col))
        d, s = design.get(inn), survey.get(inn)
        row["Проектирование: дата вступления"] = d[0].strftime("%d.%m.%Y") if d else ""
        row["Проектирование: СРО"] = d[1] if d else ""
        row["Изыскания: дата вступления"] = s[0].strftime("%d.%m.%Y") if s else ""
        row["Изыскания: СРО"] = s[1] if s else ""
        if d and s:
            row["В НОПРИЗ"] = "проектирование и изыскания"; both += 1
        elif d:
            row["В НОПРИЗ"] = "проектирование"; only_design += 1
        elif s:
            row["В НОПРИЗ"] = "изыскания"; only_survey += 1
        elif not inn:
            # Строку не выбрасываем, но и «нет» писать нечестно — не проверяли
            row["В НОПРИЗ"] = "нет ИНН"; no_inn += 1
        else:
            row["В НОПРИЗ"] = "нет"; neither += 1

    after_columns = columns + AFTER_COLUMNS
    after_rows = rows_joined_after(rows, args.дата_стройки, header, args.мин_дней)

    # Лист под обзвон: кому проектное или изыскательское СРО ещё не продано.
    # Строки без ИНН сюда не идут — их не проверяли, и выдавать их за
    # «точно не состоит» нельзя.
    lead_rows = [r for r in rows if r["В НОПРИЗ"] == "нет"]

    out_path = Path(args.out)
    sheets = [("Сверка с НОПРИЗ", columns, rows)]
    if lead_rows:
        sheets.append(("Нет в НОПРИЗ", columns, lead_rows))
    if after_rows:
        sheets.append(("Вступили после стройки", after_columns, after_rows))
    write_sheets(out_path, sheets)
    print(f"[сверка] компаний в файле: {len(rows)}")
    print(f"[сверка]   и проектирование, и изыскания: {both}")
    print(f"[сверка]   только проектирование:          {only_design}")
    print(f"[сверка]   только изыскания:               {only_survey}")
    print(f"[сверка]   не найдены в НОПРИЗ:            {neither}")
    if no_inn:
        print(f"[сверка]   строк без ИНН (не проверялись): {no_inn}")
    if lead_rows:
        print(f"[сверка] НЕ состоят в НОПРИЗ: {len(lead_rows)} — отдельный лист "
              f"«Нет в НОПРИЗ», это и есть база под обзвон")
    if after_rows:
        print(f"[сверка] вступили в НОПРИЗ позже строительной СРО: {len(after_rows)} "
              f"— отдельный лист «Вступили после стройки»")
    print(f"[сверка] готово: {out_path}")
    return 0


def cmd_recheck(args) -> int:
    """Точечная проверка своих ИНН прямо в реестре, без общей выгрузки."""
    file_path = Path(args.file)
    if not file_path.exists():
        print(f"Файл {file_path} не найден.", file=sys.stderr)
        return 1
    header, rows = read_rows(file_path)
    inn_col = args.inn_column
    if inn_col not in header:
        print(f"В файле нет колонки «{inn_col}». Есть: {', '.join(header)}",
              file=sys.stderr)
        return 1

    out_path = Path(args.out)
    progress_path = out_path.with_suffix(out_path.suffix + ".progress.json")
    готово: dict[str, dict] = {}
    if args.resume and progress_path.exists():
        готово = json.loads(progress_path.read_text(encoding="utf-8"))
        print(f"[проверка] продолжаю, уже проверено {len(готово)}")

    session = make_session()
    инны = [norm_inn(r.get(inn_col)) for r in rows]
    к_проверке = [i for i in dict.fromkeys(инны) if i and i not in готово]
    print(f"[проверка] ИНН к проверке: {len(к_проверке)}")

    осечки = 0
    try:
        for номер, инн in enumerate(к_проверке, 1):
            payload = fetch_by_inn(session, инн)
            if payload is None:
                осечки += 1
                print(f"[проверка] ИНН {инн} не проверился — пропускаю",
                      file=sys.stderr)
                continue
            п, и = членства(payload, инн)
            готово[инн] = {
                "Проектирование: дата вступления": п[0].strftime("%d.%m.%Y") if п else "",
                "Проектирование: СРО": п[1] if п else "",
                "Изыскания: дата вступления": и[0].strftime("%d.%m.%Y") if и else "",
                "Изыскания: СРО": и[1] if и else "",
            }
            if номер % 25 == 0:
                print(f"[проверка] {номер}/{len(к_проверке)}", flush=True)
                progress_path.write_text(json.dumps(готово, ensure_ascii=False),
                                         encoding="utf-8")
            time.sleep(args.delay)
    except KeyboardInterrupt:
        print("\n[проверка] прервано — прогресс сохранён "
              "(продолжить: та же команда с --resume)")
        progress_path.write_text(json.dumps(готово, ensure_ascii=False),
                                 encoding="utf-8")
        return 1

    new_cols = ["Проектирование: дата вступления", "Проектирование: СРО",
                "Изыскания: дата вступления", "Изыскания: СРО", "В НОПРИЗ"]
    columns = header + [c for c in new_cols if c not in header]
    both = only_design = only_survey = neither = no_inn = 0
    for row, инн in zip(rows, инны):
        найдено = готово.get(инн)
        if not инн:
            row["В НОПРИЗ"] = "нет ИНН"; no_inn += 1
            continue
        if найдено is None:
            row["В НОПРИЗ"] = "не проверено"
            continue
        row.update(найдено)
        д = bool(найдено["Проектирование: дата вступления"])
        и_ = bool(найдено["Изыскания: дата вступления"])
        if д and и_: row["В НОПРИЗ"] = "проектирование и изыскания"; both += 1
        elif д: row["В НОПРИЗ"] = "проектирование"; only_design += 1
        elif и_: row["В НОПРИЗ"] = "изыскания"; only_survey += 1
        else: row["В НОПРИЗ"] = "нет"; neither += 1

    write_rows(out_path, columns, rows, "Проверка по НОПРИЗ")
    if осечки == 0 and progress_path.exists():
        progress_path.unlink()
    print(f"[проверка] компаний в файле: {len(rows)}")
    print(f"[проверка]   и проектирование, и изыскания: {both}")
    print(f"[проверка]   только проектирование:          {only_design}")
    print(f"[проверка]   только изыскания:               {only_survey}")
    print(f"[проверка]   не найдены в НОПРИЗ:            {neither}")
    if no_inn:
        print(f"[проверка]   без ИНН:                        {no_inn}")
    if осечки:
        print(f"[проверка] ВНИМАНИЕ: {осечки} ИНН реестр не отдал — повторите "
              "с --resume", file=sys.stderr)
    print(f"[проверка] готово: {out_path}")
    return 0


def cmd_sample(args) -> int:
    """Сырая запись реестра и что из неё извлекает скрипт."""
    session = make_session()
    payload = fetch_page(session, 1, 20)
    if payload is None:
        print("Реестр недоступен — проверьте интернет и повторите.", file=sys.stderr)
        return 1
    records = records_of(payload)
    if not records:
        print("Реестр ответил, но записей нет. Сырой ответ:")
        print(json.dumps(payload, ensure_ascii=False)[:4000])
        return 1

    print(f"=== Разбор первых записей (всего в реестре: {total_of(payload)}) ===\n")
    for record in records[:args.count]:
        starts = _dates_by_hints(record, _START_KEY_HINTS)
        print(f"ИНН {record.get('inn')} — {_first_str(record, 'short_description', 'full_description')[:55]}")
        print(f"  СРО:               {sro_name(record)}")
        print(f"  вид деятельности:  {activity_kind(record) or '— НЕ ОПРЕДЕЛЁН —'}")
        print(f"  даты вступления:   {[(f'{d:%d.%m.%Y}', k) for d, k in starts] or '— НЕ НАЙДЕНЫ —'}")
        print(f"  статус:            «{member_status(record)}» (бывший: {is_former(record)})")
        print()

    print("=== Первая запись целиком ===")
    print(json.dumps(records[0], ensure_ascii=False, indent=2)[:6000])
    return 0


GUIDE_COLUMNS = ["ID СРО", "СРО", "Записей", "Вид деятельности"]


def load_guide(path: Path) -> dict[str, str]:
    """Справочник «СРО -> вид деятельности», заполненный руками.

    Ключом годится и ID СРО, и наименование в нижнем регистре — чтобы
    справочник можно было править как удобно, не сверяясь с идентификаторами.
    """
    _h, rows = read_rows(path)
    guide: dict[str, str] = {}
    for row in rows:
        kind = (row.get("Вид деятельности") or "").strip().lower()
        if kind not in (DESIGN, SURVEY, f"{DESIGN}, {SURVEY}"):
            continue
        for key in (str(row.get("ID СРО") or "").strip(),
                    (row.get("СРО") or "").strip().lower()):
            if key:
                guide[key] = kind
    return guide


def cmd_kinds(args) -> int:
    """Список СРО из выгрузки: сколько записей и какой вид определился.

    Нужен, когда вид деятельности определился не у всех. Разных СРО в
    НОПРИЗ пара сотен, поэтому глазами проверяется весь список целиком —
    в отличие от полутора сотен тысяч записей.
    """
    path = Path(args.nopriz)
    if not path.exists():
        print(f"Файл {path} не найден.", file=sys.stderr)
        return 1
    _h, rows = read_rows(path)

    seen: dict[str, dict] = {}
    for row in rows:
        name = (row.get("СРО") or "").strip()
        key = str(row.get("ID СРО") or "").strip() or name.lower()
        item = seen.setdefault(key, {"ID СРО": str(row.get("ID СРО") or "").strip(),
                                     "СРО": name, "Записей": 0,
                                     "Вид деятельности": kind_of_row(row)})
        item["Записей"] += 1

    items = sorted(seen.values(), key=lambda i: -i["Записей"])
    unknown = [i for i in items if not i["Вид деятельности"]]
    for item in items:
        item["Записей"] = str(item["Записей"])

    out_path = Path(args.out)
    write_rows(out_path, GUIDE_COLUMNS, items, "СРО")
    print(f"[виды] разных СРО: {len(items)}, из них без вида: {len(unknown)}")
    for item in unknown[:30]:
        print(f"[виды]   {item['Записей']:>7} записей — {item['СРО'][:60]}")
        if item["ID СРО"]:
            print(f"[виды]           карточка: {BASE}/sro/{item['ID СРО']}")
    if len(unknown) > 30:
        print(f"[виды]   и ещё {len(unknown) - 30}")
    print(f"[виды] готово: {out_path} — заполните пустой «Вид деятельности» "
          f"и передайте файл в «пересчитать --справочник».")
    if unknown:
        print("[виды] вид указан в регистрационном номере на карточке СРО: "
              "СРО-П-… проектирование, СРО-И-… изыскания")
    return 0


def cmd_reclassify(args) -> int:
    """Переписывает «Вид деятельности» в готовой выгрузке, не качая её заново."""
    path = Path(args.nopriz)
    if not path.exists():
        print(f"Файл {path} не найден.", file=sys.stderr)
        return 1
    guide = {}
    if args.справочник:
        guide_path = Path(args.справочник)
        if not guide_path.exists():
            print(f"Файл {guide_path} не найден.", file=sys.stderr)
            return 1
        guide = load_guide(guide_path)
        print(f"[пересчёт] в справочнике записей: {len(guide)}")

    header, rows = read_rows(path)
    changed = unknown = 0
    for row in rows:
        was = (row.get("Вид деятельности") or "").strip()
        now = kind_of_row(row, guide)
        if now != was:
            row["Вид деятельности"] = now
            changed += 1
        if not now:
            unknown += 1

    out_path = Path(args.out)
    write_rows(out_path, header, rows, "НОПРИЗ")
    print(f"[пересчёт] записей: {len(rows)}, изменено: {changed}, без вида: {unknown}")
    print(f"[пересчёт] готово: {out_path}")
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Реестр НОПРИЗ: проверка членства в проектных и изыскательских СРО")
    sub = parser.add_subparsers(dest="command")

    p = sub.add_parser("выгрузка", help="скачать реестр НОПРИЗ целиком (один раз)")
    p.add_argument("--out", default="нопориз_реестр.xlsx", help="куда сохранить")
    p.add_argument("--page-size", type=int, default=100, help="записей на страницу")
    p.add_argument("--delay", type=float, default=0.3, help="пауза между страницами, сек")
    p.add_argument("--resume", action="store_true", help="продолжить прерванную выгрузку")
    p.set_defaults(func=cmd_dump)

    p = sub.add_parser("сверить", help="проставить членство в НОПРИЗ своему файлу")
    p.add_argument("--file", required=True, help="ваш файл с колонкой ИНН")
    p.add_argument("--nopriz", default="нопориз_реестр.xlsx",
                   help="выгрузка реестра, сделанная командой «выгрузка»")
    p.add_argument("--out", default="результат.xlsx", help="куда сохранить результат")
    p.add_argument("--inn-column", default="ИНН", help="имя колонки с ИНН")
    p.add_argument("--мин-дней", type=int, default=1,
                   help="минимальный разрыв в днях для листа «Вступили после "
                        "стройки»: 30 отсеет тех, кто оформлял всё одним заходом")
    p.add_argument("--дата-стройки", default="Дата вступления НП",
                   help="колонка с датой вступления в строительную СРО — по ней "
                        "строится лист «Вступили после стройки»")
    p.set_defaults(func=cmd_match)

    p = sub.add_parser("виды", help="список СРО из выгрузки и их вид деятельности")
    p.add_argument("--nopriz", default="нопориз_реестр.xlsx", help="выгрузка реестра")
    p.add_argument("--out", default="виды_сро.xlsx", help="куда сохранить список")
    p.set_defaults(func=cmd_kinds)

    p = sub.add_parser("пересчитать",
                       help="переписать вид деятельности в готовой выгрузке")
    p.add_argument("--nopriz", default="нопориз_реестр.xlsx", help="выгрузка реестра")
    p.add_argument("--справочник", default=None,
                   help="файл из команды «виды» с заполненным видом деятельности")
    p.add_argument("--out", default="нопориз_реестр_исправленный.xlsx",
                   help="куда сохранить")
    p.set_defaults(func=cmd_reclassify)

    p = sub.add_parser("проверить",
                       help="пробить свои ИНН прямо в реестре (без общей выгрузки)")
    p.add_argument("--file", required=True, help="ваш файл с колонкой ИНН")
    p.add_argument("--out", default="проверка_НОПРИЗ.xlsx", help="куда сохранить")
    p.add_argument("--inn-column", default="ИНН", help="имя колонки с ИНН")
    p.add_argument("--delay", type=float, default=0.4, help="пауза между ИНН, сек")
    p.add_argument("--resume", action="store_true", help="продолжить прерванную проверку")
    p.set_defaults(func=cmd_recheck)

    p = sub.add_parser("образец", help="диагностика: сырая запись реестра")
    p.add_argument("--count", type=int, default=3, help="сколько записей разобрать")
    p.set_defaults(func=cmd_sample)

    args = parser.parse_args(argv)
    if not getattr(args, "func", None):
        parser.print_help()
        return 1
    try:
        return args.func(args) or 0
    except KeyboardInterrupt:
        print("\nПрервано пользователем.")
        return 130


if __name__ == "__main__":
    sys.exit(main())
