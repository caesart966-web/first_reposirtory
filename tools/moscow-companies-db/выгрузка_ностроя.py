"""Выгрузка из реестра НОСТРОЙ: члены выбранных СРО с датой вступления и контактами.

Самостоятельный скрипт: не зависит от пакета mosstroybase, нужны только
requests и openpyxl (уже стоят в .venv).

ПОРЯДОК РАБОТЫ (два шага)

Номер СРО — это число в адресе её страницы: у https://reestr.nostroy.ru/sro/247
номер 247.

1) Выгрузить членов нужных СРО, вступивших с начала 2026 года:

       .venv\\Scripts\\activate
       python выгрузка_ностроя.py выгрузка --sro-number С-230 С-306 ^
              --since 01.01.2026 --out новые_члены.xlsx

   Регистрационный номер («С-230» или полностью «СРО-С-230-07092010»)
   скрипт сам переведёт в ID реестра. Можно задать и напрямую:
       --sro-id 247 75            ID из адреса reestr.nostroy.ru/sro/247
       --sro-name "часть названия"

   Полный список СРО с номерами и ID:
       python выгрузка_ностроя.py список-сро --out сро.xlsx

2) Дотянуть телефоны и почту из карточек компаний на сайте НОСТРОЙ:

       python выгрузка_ностроя.py контакты --file новые_члены.xlsx

   Команда дописывает колонки «Телефоны (НОСТРОЙ)» и «E-mail (НОСТРОЙ)»
   в тот же файл. У кого контактов в карточке нет — добираются как обычно,
   через сайты компаний и Checko.

ЕСЛИ ЧТО-ТО ПОШЛО НЕ ТАК

Структура ответов реестра может измениться. Диагностика:

       python выгрузка_ностроя.py образец            # одна сырая запись списка
       python выгрузка_ностроя.py образец --member-id 19414719   # сырая карточка

Пришлите вывод — поправлю разбор.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator

import requests

BASE = "https://reestr.nostroy.ru"
MEMBER_LIST_URL = f"{BASE}/api/sro/all/member/list"
# Члены одной СРО: на сайте это страница /sro/247/member/list, поэтому у API
# ожидаем такой же путь. Если не сработает — скрипт сам перейдёт на полный
# перебор реестра с фильтром по СРО
SRO_MEMBER_LIST_URL = f"{BASE}/api/sro/{{id}}/member/list"

# Кандидаты адресов для списка СРО и карточки компании. Точный адрес заранее
# неизвестен, поэтому перебираем — рабочий запоминается на время запуска.
_SRO_LIST_URLS = (
    (f"{BASE}/api/sro/list", "post"),
    (f"{BASE}/api/sro/all/list", "post"),
    (f"{BASE}/api/sro", "get"),
    (f"{BASE}/api/sro/list", "get"),
)
_MEMBER_CARD_URLS = (
    f"{BASE}/api/member/{{id}}",
    f"{BASE}/api/sro/member/{{id}}",
    f"{BASE}/api/member/view/{{id}}",
    f"{BASE}/api/sro/all/member/{{id}}",
)

# Дата вступления. Порядок важен: сначала точное имя поля, каким его отдаёт
# реестр (проверено на живых данных: "registry_registration_date":
# "2018-02-12T00:00:00+03:00"), потом запасные варианты на случай изменений.
# Осторожно с общим "registration_date": в записи может оказаться и дата
# госрегистрации самой компании (ОГРН, 1990-е) — она нам не нужна, поэтому
# такие ключи отсекаются отдельно (_STATE_REG_MARKERS).
_START_KEY_HINTS = (
    "registry_registration_date",
    "member_right_start",
    "right_start",
    "вступ",
    "начал",
    "start_date",
)
# Ключи, которые похожи на дату вступления, но ей не являются
_STATE_REG_MARKERS = ("state_registration", "ogrn", "егрюл", "егрип")
_STOP_KEY_HINTS = ("member_right_stop", "right_stop", "stop_date", "прекращ", "исключ")
_FORMER_MARKERS = ("исключ", "прекращ")

_PHONE_KEY_HINTS = ("phone", "тел", "телефон", "contact")
_EMAIL_KEY_HINTS = ("email", "e-mail", "почт", "мэйл", "мейл")

# Телефон РФ: код города/оператора начинается с 3/4/8/9 — как в основном
# проекте, чтобы в контакты не попадали куски длинных номеров документов
_PHONE_RE = re.compile(
    r"(?<!\d)(?:\+7|8|7)?[\s\-(]*([3489]\d{2})[\s\-)]*(\d{3})[\s\-]*(\d{2})[\s\-]*(\d{2})(?!\d)"
)
_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")

COLUMNS = ["ИНН", "Наименование", "ОГРН", "Дата вступления", "СРО", "ID СРО",
           "Статус", "Регион", "Номер в реестре СРО", "ID записи"]
CONTACT_COLUMNS = ["Телефоны (НОСТРОЙ)", "E-mail (НОСТРОЙ)"]


# -- разбор ответов ----------------------------------------------------


def _strings(obj: Any) -> Iterator[str]:
    if isinstance(obj, str):
        yield obj
    elif isinstance(obj, dict):
        for value in obj.values():
            yield from _strings(value)
    elif isinstance(obj, list):
        for item in obj:
            yield from _strings(item)


def _parse_date(text: str) -> date | None:
    text = (text or "").strip()
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", text)
    if m:
        day, month, year = map(int, m.groups())
    else:
        m = re.match(r"(\d{4})-(\d{2})-(\d{2})", text)
        if not m:
            return None
        year, month, day = map(int, m.groups())
    try:
        return date(year, month, day)
    except ValueError:
        return None


# Ветки записи, которые НЕ описывают членство компании: даты внутри них
# относятся к самой СРО (когда она внесена в госреестр и т.п.) и не должны
# приниматься за дату вступления компании
_SKIP_BRANCHES = ("sro", "sro_info", "organization")


def _dates_by_hints(record: dict, hints: tuple[str, ...]) -> list[tuple[date, str]]:
    """Все подходящие даты записи как (дата, имя_поля).

    Блок сведений о самой СРО пропускается — иначе её собственные даты
    («внесена в реестр», «начало деятельности») подменяли бы дату вступления.
    """
    for hint in hints:
        found: list[tuple[date, str]] = []

        def rec(obj: Any, path: str = "") -> None:
            if isinstance(obj, dict):
                for key, value in obj.items():
                    if key.lower() in _SKIP_BRANCHES:
                        continue
                    full = f"{path}.{key}" if path else key
                    low = key.lower()
                    if (isinstance(value, str) and hint in low
                            and not any(m in low for m in _STATE_REG_MARKERS)):
                        parsed = _parse_date(value)
                        if parsed:
                            found.append((parsed, full))
                    elif isinstance(value, (dict, list)):
                        rec(value, full)
            elif isinstance(obj, list):
                for item in obj:
                    rec(item, path)

        rec(record)
        if found:
            return found
    return []


def _date_by_hints(record: dict, hints: tuple[str, ...]) -> date | None:
    """Дата из записи по «смысловым» именам ключей (имена полей могут меняться).

    Если дат несколько (у компании бывает несколько прав/специализаций,
    добавленных в разное время), берём самую раннюю — это и есть момент
    вступления, а не последующее расширение допуска.
    """
    found = _dates_by_hints(record, hints)
    return min(d for d, _key in found) if found else None


def _values_by_hints(obj: Any, hints: tuple[str, ...]) -> Iterator[str]:
    """Строки из полей, чьи имена похожи на подсказки (телефон/почта)."""
    if isinstance(obj, dict):
        for key, value in obj.items():
            matched = any(h in key.lower() for h in hints)
            if matched:
                yield from _strings(value)
            if isinstance(value, (dict, list)):
                yield from _values_by_hints(value, hints)
    elif isinstance(obj, list):
        for item in obj:
            yield from _values_by_hints(item, hints)


def extract_phones(payload: Any) -> list[str]:
    phones: list[str] = []
    for text in _values_by_hints(payload, _PHONE_KEY_HINTS):
        for match in _PHONE_RE.finditer(text):
            code, a, b, c = match.groups()
            phone = f"+7{code}{a}{b}{c}"
            if phone not in phones:
                phones.append(phone)
    return phones


def extract_emails(payload: Any) -> list[str]:
    emails: list[str] = []
    for text in _values_by_hints(payload, _EMAIL_KEY_HINTS):
        for match in _EMAIL_RE.finditer(text):
            email = match.group(0).lower()
            if email not in emails:
                emails.append(email)
    return emails


def _first_str(record: dict, *keys: str) -> str:
    for key in keys:
        value = record.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def _sro_block(record: dict) -> dict:
    sro = record.get("sro")
    return sro if isinstance(sro, dict) else {}


def sro_name(record: dict) -> str:
    sro = _sro_block(record)
    if sro:
        return _first_str(sro, "title", "full_description", "short_description", "name")
    value = record.get("sro")
    return value.strip() if isinstance(value, str) else ""


def sro_id(record: dict) -> str:
    sro = _sro_block(record)
    for key in ("id", "sro_id", "registry_number", "number"):
        value = sro.get(key)
        if isinstance(value, (str, int)) and str(value).strip():
            return str(value).strip()
    for key in ("sro_id", "sroId"):
        value = record.get(key)
        if isinstance(value, (str, int)) and str(value).strip():
            return str(value).strip()
    return ""


def sro_registration_number(record: dict) -> str:
    """Регистрационный номер СРО, например «СРО-С-230-07092010»."""
    sro = _sro_block(record) or record
    return _first_str(sro, "registration_number", "registry_number",
                      "reg_number", "number")


def number_core(text: Any) -> str:
    """Ядро номера СРО: «СРО-С-230-07092010», «С-230» и «230» -> «С-230».

    Люди называют СРО коротким номером, реестр хранит полный — сравнивать
    их можно только приведя к общему виду. Латинскую C заодно приводим к
    кириллической: на клавиатуре они неразличимы, а строки — различаются.
    """
    text = str(text or "").strip().upper().replace("C", "С")
    match = re.search(r"С\s*-\s*(\d+)", text)
    if match:
        return f"С-{int(match.group(1))}"
    digits = re.sub(r"\D", "", text)
    return f"С-{int(digits)}" if digits else ""


def member_status(record: dict) -> str:
    status = record.get("member_status")
    if isinstance(status, dict):
        return _first_str(status, "title", "name")
    return status.strip() if isinstance(status, str) else ""


# Формулировки статуса, означающие действующее членство
_ACTIVE_MARKERS = ("является членом", "действ", "член ")


def is_former(record: dict) -> bool:
    """Вышел/исключён — в выгрузку «действующих новых членов» не идёт.

    Решает прежде всего ТЕКСТ статуса. Дата прекращения сама по себе не
    доказательство: у действующих членов в записи может стоять плановая
    дата окончания права (в будущем), и если считать её признаком выхода,
    отсеются вообще все компании.
    """
    status = member_status(record).lower()
    if any(m in status for m in _FORMER_MARKERS):
        return True
    if any(m in status for m in _ACTIVE_MARKERS):
        return False
    stopped = _date_by_hints(record, _STOP_KEY_HINTS)
    # Статус непонятен: считаем бывшим только если дата прекращения уже прошла
    return stopped is not None and stopped <= date.today()


def member_ogrn(record: dict) -> str:
    """ОГРН/ОГРНИП: у ИП поле называется ogrnip."""
    for key in ("ogrn", "ogrnip", "ogrn_ip"):
        value = record.get(key)
        if isinstance(value, (str, int)) and str(value).strip():
            return str(value).strip()
    return ""


def member_region(record: dict) -> str:
    """Регион из записи реестра (точнее, чем гадание по первым цифрам ИНН)."""
    block = record.get("region_number")
    if isinstance(block, dict):
        name = _first_str(block, "title", "name")
        if name:
            return name
    return ""


def record_id(record: dict) -> str:
    for key in ("id", "member_id", "memberId",
                # запасные: по ним тоже можно открыть карточку в реестре
                "registration_number", "inventory_number"):
        value = record.get(key)
        if isinstance(value, (str, int)) and str(value).strip():
            return str(value).strip()
    return ""


def region_of(inn: str) -> str:
    return inn[:2] if len(inn) >= 2 else ""


def records_of(payload: Any) -> list[dict]:
    if not isinstance(payload, dict):
        return []
    data = payload.get("data")
    if isinstance(data, dict):
        data = data.get("data")
    if isinstance(data, list):
        return [r for r in data if isinstance(r, dict)]
    if isinstance(payload.get("data"), list):
        return [r for r in payload["data"] if isinstance(r, dict)]
    return []


def total_of(payload: Any) -> int | None:
    if not isinstance(payload, dict):
        return None
    data = payload.get("data")
    containers = [data, payload] if isinstance(data, dict) else [payload]
    for container in containers:
        for key in ("count", "total", "totalCount"):
            value = container.get(key)
            if isinstance(value, int):
                return value
    return None


# -- сеть ---------------------------------------------------------------


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "User-Agent": "mosstroybase/0.1 (+open-data harvester)",
        "Content-Type": "application/json",
        "Accept": "application/json",
    })
    return session


def fetch_page(session: requests.Session, page: int, page_size: int,
               timeout: int = 60, attempts: int = 5) -> dict | None:
    body = {"page": page, "pageSize": page_size}
    for attempt in range(1, attempts + 1):
        try:
            resp = session.post(MEMBER_LIST_URL, json=body, timeout=timeout)
            if resp.status_code == 200:
                return resp.json()
            print(f"[нострой] страница {page}: HTTP {resp.status_code} "
                  f"(попытка {attempt}/{attempts})", flush=True)
        except (requests.RequestException, ValueError) as exc:
            print(f"[нострой] страница {page}: {type(exc).__name__} "
                  f"(попытка {attempt}/{attempts})", flush=True)
        time.sleep(min(2 ** attempt, 30))
    return None


def fetch_sro_page(session: requests.Session, sro: str, page: int, page_size: int,
                   timeout: int = 60, attempts: int = 6) -> dict | None:
    """Страница членов ОДНОЙ СРО. None — если страницу так и не отдали.

    Реестр периодически отвечает 500 на отдельные страницы — это его
    временный сбой, поэтому повторяем с нарастающей паузой.
    """
    body = {"page": page, "pageSize": page_size}
    url = SRO_MEMBER_LIST_URL.format(id=sro)
    for attempt in range(1, attempts + 1):
        try:
            resp = session.post(url, json=body, timeout=timeout)
            if resp.status_code == 200:
                return resp.json()
            if resp.status_code in (404, 405):
                return None  # такого пути нет — уходим на полный перебор
            print(f"[сро {sro}] страница {page}: HTTP {resp.status_code} "
                  f"(попытка {attempt}/{attempts})", flush=True)
        except (requests.RequestException, ValueError) as exc:
            print(f"[сро {sro}] страница {page}: {type(exc).__name__} "
                  f"(попытка {attempt}/{attempts})", flush=True)
        time.sleep(min(2 ** attempt, 30))
    return None


def iter_sro_members(session: requests.Session, sro: str, page_size: int,
                     delay: float, failed_pages: list[int] | None = None,
                     declared_total: list[int] | None = None,
                     ) -> Iterator[dict] | None:
    """Все члены одной СРО постранично; None — если путь API не работает.

    Страницу, которую реестр так и не отдал, НЕ считаем концом списка:
    запоминаем её номер и идём дальше, иначе один сбой сервера молча
    обрезал бы выгрузку. Номера сбойных страниц уходят в failed_pages —
    вызывающий код повторит их и честно сообщит, если данные неполные.
    """
    first = fetch_sro_page(session, sro, 1, page_size)
    if first is None or not records_of(first):
        return None

    def generate() -> Iterator[dict]:
        payload = first
        page = 1
        total = total_of(payload)
        # ВАЖНО: реестр может игнорировать запрошенный pageSize и отдавать
        # своё количество (например, по 20 вместо 100). Считать страницы по
        # запрошенному размеру нельзя — иначе остановимся, прочитав лишь
        # часть списка. Берём фактический размер первой страницы.
        actual_size = len(records_of(first)) or page_size
        if actual_size != page_size:
            print(f"[сро {sro}] реестр отдаёт по {actual_size} записей на "
                  f"страницу (запрашивали {page_size})")
        total_pages = ((total + actual_size - 1) // actual_size) if total else None
        if total:
            if declared_total is not None:
                declared_total.append(total)
            print(f"[сро {sro}] членов в реестре: {total} (~{total_pages} страниц)")
        seen = 0
        misses = 0
        while True:
            records = records_of(payload) if payload is not None else []
            if payload is not None and not records:
                return  # пустая страница — реестр закончился
            yield from records
            seen += len(records)
            hint = f"/{total_pages}" if total_pages else ""
            print(f"[сро {sro}] страница {page}{hint}: прочитано {seen}", flush=True)

            page += 1
            # Предохранитель от бесконечного цикла, но с запасом: счётчик
            # реестра бывает неточным, а настоящий конец списка — пустая
            # страница, её и ждём в первую очередь
            if total_pages and page > total_pages + 5:
                print(f"[сро {sro}] дошли до расчётного конца списка "
                      f"({total_pages} страниц), останавливаюсь", flush=True)
                return
            time.sleep(delay)
            payload = fetch_sro_page(session, sro, page, page_size)
            if payload is None:
                misses += 1
                if failed_pages is not None:
                    failed_pages.append(page)
                print(f"[сро {sro}] страница {page} не отдалась — пропускаю, "
                      "повторю в конце", file=sys.stderr)
                if not total_pages and misses >= 3:
                    # Сколько всего страниц неизвестно, а сбои идут подряд —
                    # похоже, список кончился или реестр лёг
                    print(f"[сро {sro}] три страницы подряд не отдались — "
                          "останавливаюсь", file=sys.stderr)
                    return

    return generate()


def retry_failed_pages(session: requests.Session, sro: str, pages: list[int],
                       page_size: int, delay: float) -> tuple[list[dict], list[int]]:
    """Повторный заход на страницы, сбойнувшие в основном проходе."""
    recovered: list[dict] = []
    still_failed: list[int] = []
    for page in pages:
        payload = fetch_sro_page(session, sro, page, page_size)
        if payload is None:
            still_failed.append(page)
            continue
        recovered.extend(records_of(payload))
        time.sleep(delay)
    return recovered, still_failed


_card_url_template: str | None = None


def fetch_member_card(session: requests.Session, member_id: str,
                      timeout: int = 30) -> Any | None:
    """Карточка компании; адрес API подбирается один раз и запоминается."""
    global _card_url_template
    templates = ([_card_url_template] if _card_url_template
                 else list(_MEMBER_CARD_URLS))
    for template in templates:
        url = template.format(id=member_id)
        try:
            resp = session.get(url, timeout=timeout)
        except requests.RequestException:
            continue
        if resp.status_code != 200:
            continue
        try:
            payload = resp.json()
        except ValueError:
            continue
        if payload:
            _card_url_template = template
            return payload
    return None


# -- вывод в файл -------------------------------------------------------


def write_rows(path: Path, columns: list[str], rows: list[dict], sheet: str) -> None:
    if path.suffix.lower() == ".csv":
        import csv
        with open(path, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=columns, delimiter=";",
                                    extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        return
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(columns)
    for row in rows:
        ws.append([row.get(c, "") for c in columns])
    ws.freeze_panes = "A2"
    wb.save(path)


def read_rows(path: Path) -> tuple[list[str], list[dict]]:
    if path.suffix.lower() == ".csv":
        import csv
        with open(path, encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh, delimiter=";")
            return list(reader.fieldnames or []), list(reader)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    rows = []
    for values in rows_iter:
        row = {header[i]: ("" if v is None else str(v))
               for i, v in enumerate(values) if i < len(header)}
        if any(row.values()):
            rows.append(row)
    return header, rows


# -- команды ------------------------------------------------------------


def fetch_sro_list(session: requests.Session) -> Any:
    """Ответ реестра со списком всех СРО или None, если ни один адрес не отдал."""
    for url, method in _SRO_LIST_URLS:
        try:
            if method == "post":
                resp = session.post(url, json={"page": 1, "pageSize": 1000}, timeout=60)
            else:
                resp = session.get(url, timeout=60)
        except requests.RequestException:
            continue
        if resp.status_code != 200:
            continue
        try:
            candidate = resp.json()
        except ValueError:
            continue
        if records_of(candidate):
            print(f"[нострой] список СРО получен: {url}")
            return candidate
    return None


def sro_ids_by_number(session: requests.Session,
                      numbers: list[str]) -> dict[str, str]:
    """Номер СРО («С-230») -> её ID в реестре.

    Быстрый путь выгрузки работает по ID, а называют СРО номером — без
    этого сопоставления пришлось бы перебирать весь реестр (полчаса против
    секунд).
    """
    wanted = {number_core(n) for n in numbers} - {""}
    payload = fetch_sro_list(session)
    found: dict[str, str] = {}
    for record in records_of(payload) if payload else []:
        core = number_core(sro_registration_number(record))
        if core in wanted:
            sid = str(record.get("id") or "").strip() or sro_id(record)
            if sid:
                found[core] = sid
    return found


def cmd_list_sro(args) -> int:
    """Список всех СРО реестра — чтобы выбрать нужные и узнать их ID."""
    session = make_session()
    payload = fetch_sro_list(session)

    if payload is None:
        # Запасной путь: собрать СРО из первых страниц списка членов
        print("[нострой] отдельный список СРО не отдался — собираю названия "
              "из записей реестра (это медленнее, но работает)")
        seen: dict[str, str] = {}
        for page in range(1, args.scan_pages + 1):
            page_payload = fetch_page(session, page, 100)
            if page_payload is None:
                break
            records = records_of(page_payload)
            if not records:
                break
            for record in records:
                sid, name = sro_id(record), sro_name(record)
                if name and sid not in seen:
                    seen[sid] = name
            print(f"[нострой] страница {page}: разных СРО найдено {len(seen)}",
                  flush=True)
            time.sleep(args.delay)
        rows = [{"ID": sid, "Название": name} for sid, name in
                sorted(seen.items(), key=lambda kv: kv[1])]
    else:
        rows = []
        for record in records_of(payload):
            rows.append({
                "ID": str(record.get("id") or record.get("registry_number") or ""),
                "Рег. номер": sro_registration_number(record),
                "Название": _first_str(record, "title", "full_description",
                                       "short_description", "name"),
                "ИНН": str(record.get("inn") or ""),
                "Регион": _first_str(record, "region", "region_name"),
            })

    if not rows:
        print("Не удалось получить список СРО.", file=sys.stderr)
        return 1

    out = Path(args.out)
    columns = (["ID"] + [c for c in ("Рег. номер",) if c in rows[0]] + ["Название"]
               + [c for c in ("ИНН", "Регион") if c in rows[0]])
    write_rows(out, columns, rows, "СРО")
    print(f"[нострой] готово: {out} ({len(rows)} СРО). "
          "Найдите нужные две и возьмите их ID для команды «выгрузка».")
    return 0


def cmd_export(args) -> int:
    """Члены выбранных СРО, вступившие начиная с указанной даты."""
    since = _parse_date(args.since)
    if since is None:
        print(f"Не понял дату «{args.since}». Формат: 01.01.2026", file=sys.stderr)
        return 1
    until = _parse_date(args.until) if args.until else None
    if args.until and until is None:
        print(f"Не понял дату «{args.until}». Формат: 31.12.2026", file=sys.stderr)
        return 1
    if until and until < since:
        print("Верхняя граница периода раньше нижней — проверьте --since/--until",
              file=sys.stderr)
        return 1

    period = f"с {since:%d.%m.%Y}" + (f" по {until:%d.%m.%Y}" if until else "")
    wanted_ids = {str(i) for i in (args.sro_id or [])}
    wanted_names = [n.lower() for n in (args.sro_name or [])]
    wanted_numbers = {number_core(n) for n in (args.sro_number or [])} - {""}
    if not wanted_ids and not wanted_names and not wanted_numbers:
        print("Укажите СРО: --sro-number С-230 С-306, --sro-id 123 456 "
              "или --sro-name «часть названия»", file=sys.stderr)
        return 1

    session = make_session()

    # Номера переводим в ID заранее: по ID работает быстрый путь (секунды),
    # без него остаётся перебор всего реестра (полчаса)
    if wanted_numbers:
        найдено = sro_ids_by_number(session, sorted(wanted_numbers))
        for номер, sid in sorted(найдено.items()):
            print(f"[нострой] {номер} -> ID {sid}")
            wanted_ids.add(sid)
        пропали = sorted(wanted_numbers - set(найдено))
        if пропали:
            print(f"[нострой] не нашёл в списке СРО: {', '.join(пропали)} — "
                  "они будут искаться перебором реестра по номеру",
                  file=sys.stderr)

    def sro_matches(record: dict) -> bool:
        if wanted_ids and sro_id(record) in wanted_ids:
            return True
        if wanted_numbers and number_core(sro_registration_number(record)) in wanted_numbers:
            return True
        if wanted_names:
            name = sro_name(record).lower()
            return any(part in name for part in wanted_names)
        return False

    out_path = Path(args.out)
    progress_path = out_path.with_suffix(out_path.suffix + ".progress.json")

    def row_of(record: dict, started: date) -> dict:
        inn = str(record.get("inn") or "").strip()
        return {
            "ИНН": inn,
            "Наименование": _first_str(record, "short_description",
                                       "full_description", "name", "title"),
            "ОГРН": member_ogrn(record),
            "Дата вступления": started.strftime("%d.%m.%Y"),
            "СРО": sro_name(record),
            "ID СРО": sro_id(record),
            "Статус": member_status(record),
            "Регион": member_region(record) or region_of(inn),
            "Номер в реестре СРО": _first_str(record, "registration_number"),
            "ID записи": record_id(record),
        }

    def finish(rows_found: list[dict]) -> int:
        if not rows_found:
            print("[нострой] подходящих компаний не найдено — файл не создан")
            return 1
        unique: dict[str, dict] = {}
        for row in rows_found:
            unique.setdefault(f"{row['ИНН']}|{row['ID СРО']}", row)
        rows = sorted(unique.values(),
                      key=lambda r: _parse_date(r["Дата вступления"]) or date.min,
                      reverse=True)
        write_rows(out_path, COLUMNS, rows, "Новые члены СРО")
        print(f"[нострой] готово: {out_path} ({len(rows)} компаний). "
              f"Дальше телефоны: python выгрузка_ностроя.py контакты --file {out_path}")
        return 0

    # Быстрый путь: если известны ID СРО, берём членов каждой СРО напрямую —
    # это секунды вместо получаса перебора всего реестра
    if wanted_ids and not args.full_scan:
        fast: list[dict] = []
        all_ok = True
        lost_pages: dict[str, list[int]] = {}

        reported_field = {"done": False}
        # Счётчик причин отсева: без него «найдено 0» невозможно объяснить
        rejected = {"бывший/исключён": 0, "нет даты вступления": 0,
                    "вступил раньше периода": 0, "вступил позже периода": 0,
                    "нет ИНН": 0}
        examples: dict[str, str] = {}

        def take(record: dict) -> bool:
            """Отбор одной записи; True — попала в выгрузку."""
            if is_former(record):
                rejected["бывший/исключён"] += 1
                examples.setdefault(
                    "бывший/исключён",
                    f"ИНН {record.get('inn')}, статус «{member_status(record)}»")
                return False
            dates = _dates_by_hints(record, _START_KEY_HINTS)
            if not dates:
                rejected["нет даты вступления"] += 1
                examples.setdefault("нет даты вступления", f"ИНН {record.get('inn')}")
                return False
            started, field = min(dates, key=lambda pair: pair[0])
            # Один раз показываем, ОТКУДА взята дата — чтобы можно было
            # глазами убедиться, что фильтр смотрит в нужное поле
            if not reported_field["done"]:
                reported_field["done"] = True
                print(f"[нострой] дата вступления берётся из поля «{field}» "
                      f"(пример значения: {started:%d.%m.%Y})")
                others = {k for _d, k in dates if k != field}
                if others:
                    print(f"[нострой]   в записи есть и другие даты начала: "
                          f"{', '.join(sorted(others))} — беру самую раннюю")
            if started < since:
                rejected["вступил раньше периода"] += 1
                examples.setdefault("вступил раньше периода",
                                    f"ИНН {record.get('inn')}, {started:%d.%m.%Y}")
                return False
            if until and started > until:
                rejected["вступил позже периода"] += 1
                examples.setdefault("вступил позже периода",
                                    f"ИНН {record.get('inn')}, {started:%d.%m.%Y}")
                return False
            if not str(record.get("inn") or "").strip():
                rejected["нет ИНН"] += 1
                return False
            fast.append(row_of(record, started))
            return True

        for sro in sorted(wanted_ids):
            failed_pages: list[int] = []
            declared: list[int] = []   # сколько членов реестр обещал отдать
            members = iter_sro_members(session, sro, args.page_size, args.delay,
                                       failed_pages, declared)
            if members is None:
                print(f"[сро {sro}] прямой путь API не сработал — "
                      "перехожу на полный перебор реестра")
                all_ok = False
                break
            seen = kept = 0
            for record in members:
                seen += 1
                kept += take(record)

            # Страницы, сбойнувшие из-за ошибок реестра (HTTP 500), повторяем
            if failed_pages:
                print(f"[сро {sro}] повторяю {len(failed_pages)} сбойных страниц")
                recovered, still_failed = retry_failed_pages(
                    session, sro, failed_pages, args.page_size, args.delay)
                for record in recovered:
                    seen += 1
                    kept += take(record)
                if still_failed:
                    lost_pages[sro] = still_failed

            print(f"[сро {sro}] просмотрено {seen}, вступивших {period} — {kept}")
            # Сверка: прочитали ли мы столько, сколько реестр обещал
            if declared and seen < declared[0]:
                print(f"[сро {sro}] ВНИМАНИЕ: реестр заявил {declared[0]} членов, "
                      f"а прочитано {seen} — список неполный, часть компаний "
                      "могла не попасть в выгрузку!", file=sys.stderr)
                lost_pages.setdefault(sro, [])

        if all_ok:
            # Почему записи не попали в выгрузку — иначе «найдено 0»
            # невозможно отличить от «фильтр сломан»
            dropped = {k: v for k, v in rejected.items() if v}
            if dropped:
                print("\n[нострой] отсеяно по причинам:")
                for reason, count in sorted(dropped.items(), key=lambda kv: -kv[1]):
                    sample = examples.get(reason)
                    tail = f"  (например: {sample})" if sample else ""
                    print(f"[нострой]   {reason}: {count}{tail}")
            if not fast and rejected["бывший/исключён"] > sum(
                    v for k, v in rejected.items() if k != "бывший/исключён"):
                print("[нострой] Похоже, признак «бывший член» срабатывает слишком "
                      "часто. Проверьте: python выгрузка_ностроя.py образец "
                      f"--sro-id {sorted(wanted_ids)[0]}", file=sys.stderr)

        if all_ok and lost_pages:
            print("\n[нострой] ВНИМАНИЕ: выгрузка НЕПОЛНАЯ — реестр отдал не всё.",
                  file=sys.stderr)
            for sro, pages in lost_pages.items():
                if pages:
                    nums = ", ".join(str(p) for p in pages[:20])
                    more = f" и ещё {len(pages) - 20}" if len(pages) > 20 else ""
                    print(f"[нострой]   СРО {sro}: не отдались страницы {nums}{more}",
                          file=sys.stderr)
                else:
                    print(f"[нострой]   СРО {sro}: прочитано меньше, чем заявлено "
                          "в реестре (см. предупреждение выше)", file=sys.stderr)
            print("[нострой] Запустите ту же команду позже — реестр обычно "
                  "чинится за несколько часов.", file=sys.stderr)
        if all_ok:
            return finish(fast)

    found: list[dict] = []
    start_page = 1
    if args.resume and progress_path.exists():
        saved = json.loads(progress_path.read_text(encoding="utf-8"))
        found = saved.get("found", [])
        start_page = saved.get("next_page", 1)
        print(f"[нострой] продолжаю со страницы {start_page}, уже найдено {len(found)}")

    print(f"[нострой] беру вступивших {period}")
    if wanted_ids:
        print(f"[нострой] СРО по ID: {', '.join(sorted(wanted_ids))}")
    if wanted_numbers:
        print(f"[нострой] СРО по номеру: {', '.join(sorted(wanted_numbers))}")
    if wanted_names:
        print(f"[нострой] СРО по названию: {', '.join(args.sro_name)}")

    scanned = matched_sro = no_start_date = 0
    page = start_page
    total_pages: int | None = None
    completed = False

    def save_progress(next_page: int) -> None:
        progress_path.write_text(
            json.dumps({"next_page": next_page, "found": found}, ensure_ascii=False),
            encoding="utf-8",
        )

    try:
        while True:
            if args.max_pages and (page - start_page) >= args.max_pages:
                print(f"[нострой] достигнут лимит {args.max_pages} страниц")
                break

            payload = fetch_page(session, page, args.page_size)
            if payload is None:
                print(f"[нострой] страница {page} не отдалась — останавливаюсь. "
                      "Запустите ту же команду с --resume, чтобы продолжить.",
                      file=sys.stderr)
                break

            if total_pages is None:
                total = total_of(payload)
                if total:
                    total_pages = (total + args.page_size - 1) // args.page_size
                    print(f"[нострой] всего записей в реестре: {total} "
                          f"(~{total_pages} страниц)")

            records = records_of(payload)
            if not records:
                completed = True
                print(f"[нострой] страница {page} пустая — конец реестра")
                break

            for record in records:
                scanned += 1
                if not sro_matches(record):
                    continue
                matched_sro += 1
                inn = str(record.get("inn") or "").strip()
                if not inn or is_former(record):
                    continue
                started = _date_by_hints(record, _START_KEY_HINTS)
                if started is None:
                    no_start_date += 1
                    continue
                if started < since or (until and started > until):
                    continue
                found.append(row_of(record, started))

            if page % 10 == 0 or page == start_page:
                hint = f"/{total_pages}" if total_pages else ""
                print(f"[нострой] страница {page}{hint}: просмотрено {scanned}, "
                      f"в нужных СРО {matched_sro}, подходящих {len(found)}", flush=True)
                save_progress(page + 1)

            page += 1
            time.sleep(args.delay)
    except KeyboardInterrupt:
        print("\n[нострой] прервано — сохраняю прогресс "
              "(продолжить: та же команда с --resume)")

    # Прогресс удаляется только после полного прохода реестра: иначе
    # возобновление стало бы невозможным
    if completed:
        if progress_path.exists():
            progress_path.unlink()
    else:
        save_progress(page)
        print(f"[нострой] прогресс сохранён: {progress_path}")

    if matched_sro == 0:
        print("[нострой] ВНИМАНИЕ: ни одна запись не совпала с указанными СРО. "
              "Проверьте ID/названия командой «список-сро».", file=sys.stderr)
    if no_start_date and not found:
        print(f"[нострой] ВНИМАНИЕ: у {no_start_date} записей не нашлась дата "
              "вступления. Запустите «python выгрузка_ностроя.py образец» и "
              "пришлите вывод — поправлю разбор.", file=sys.stderr)

    return finish(found)


def cmd_contacts(args) -> int:
    """Дотягивает телефоны/почту из карточек компаний на сайте НОСТРОЙ."""
    path = Path(args.file)
    if not path.exists():
        print(f"Файл {path} не найден.", file=sys.stderr)
        return 1

    header, rows = read_rows(path)
    if not rows:
        print("В файле нет строк.", file=sys.stderr)
        return 1
    if "ID записи" not in header:
        print("В файле нет колонки «ID записи» — карточку не найти. "
              "Пересоберите файл командой «выгрузка».", file=sys.stderr)
        return 1

    session = make_session()
    columns = header + [c for c in CONTACT_COLUMNS if c not in header]

    processed = with_phone = failed = 0
    try:
        for idx, row in enumerate(rows, 1):
            # Уже заполненные пропускаем — чтобы можно было продолжить после обрыва
            if row.get("Телефоны (НОСТРОЙ)") or row.get("E-mail (НОСТРОЙ)"):
                continue
            member_id = (row.get("ID записи") or "").strip()
            if not member_id:
                continue
            card = fetch_member_card(session, member_id)
            if card is None:
                failed += 1
                row.setdefault("Телефоны (НОСТРОЙ)", "")
                row.setdefault("E-mail (НОСТРОЙ)", "")
            else:
                phones = extract_phones(card)
                emails = extract_emails(card)
                row["Телефоны (НОСТРОЙ)"] = "; ".join(phones)
                row["E-mail (НОСТРОЙ)"] = "; ".join(emails)
                if phones:
                    with_phone += 1
            processed += 1
            if processed % 20 == 0:
                print(f"[карточки] обработано {processed}/{len(rows)}, "
                      f"с телефоном {with_phone}", flush=True)
                write_rows(path, columns, rows, "Новые члены СРО")
            time.sleep(args.delay)
    except KeyboardInterrupt:
        print("\n[карточки] прервано — сохраняю то, что успели")

    write_rows(path, columns, rows, "Новые члены СРО")
    print(f"[карточки] готово: обработано {processed}, с телефоном {with_phone}, "
          f"карточка не открылась у {failed}")
    if failed and failed == processed:
        print("[карточки] ни одна карточка не открылась — возможно, изменился адрес "
              "API. Запустите: python выгрузка_ностроя.py образец --member-id <ID> "
              "и пришлите вывод.", file=sys.stderr)
    if with_phone < processed:
        print("[карточки] у кого телефона в карточке нет — добираются как обычно: "
              "через сайты компаний и Checko")
    return 0


def cmd_sample(args) -> int:
    """Сырой ответ реестра — для сверки имён полей, если разбор не сработал."""
    session = make_session()
    if args.member_id:
        card = fetch_member_card(session, args.member_id)
        if card is None:
            print("Карточка не открылась ни по одному из известных адресов:")
            for template in _MEMBER_CARD_URLS:
                print("   ", template.format(id=args.member_id))
            return 1
        print(f"Карточка компании (адрес: {_card_url_template}):")
        print(json.dumps(card, ensure_ascii=False, indent=2)[:6000])
        print("\nИзвлечено скриптом:")
        print("  телефоны:", extract_phones(card))
        print("  e-mail:  ", extract_emails(card))
        return 0

    if args.sro_id:
        payload = fetch_sro_page(session, args.sro_id, 1, 20)
        source = f"СРО {args.sro_id}"
    else:
        payload = fetch_page(session, 1, 5)
        source = "общий список реестра"
    if payload is None:
        print("Реестр недоступен — проверьте интернет и повторите.", file=sys.stderr)
        return 1
    records = records_of(payload)
    if not records:
        print("Реестр ответил, но записей нет. Сырой ответ:")
        print(json.dumps(payload, ensure_ascii=False)[:4000])
        return 1

    print(f"=== Разбор первых записей ({source}) ===\n")
    for record in records[:args.count]:
        starts = _dates_by_hints(record, _START_KEY_HINTS)
        stops = _dates_by_hints(record, _STOP_KEY_HINTS)
        started = min((d for d, _k in starts), default=None)
        print(f"ИНН {record.get('inn')} — {_first_str(record, 'full_description', 'name')[:60]}")
        print(f"  статус:            «{member_status(record)}»")
        print(f"  даты вступления:   {[(f'{d:%d.%m.%Y}', k) for d, k in starts] or '— НЕ НАЙДЕНЫ —'}")
        print(f"  даты прекращения:  {[(f'{d:%d.%m.%Y}', k) for d, k in stops] or '— нет —'}")
        print(f"  вердикт «бывший»:  {is_former(record)}"
              f"{'   <-- из-за этого запись отсеивается' if is_former(record) else ''}")
        print(f"  итоговая дата:     {started:%d.%m.%Y}" if started else
              "  итоговая дата:     —")
        print()

    print("=== Первая запись целиком (сырой ответ реестра) ===")
    print(json.dumps(records[0], ensure_ascii=False, indent=2)[:6000])
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Реестр НОСТРОЙ: члены выбранных СРО с датой вступления и контактами",
    )
    sub = parser.add_subparsers(dest="command")

    p = sub.add_parser("список-сро", help="выгрузить список всех СРО (чтобы узнать ID)")
    p.add_argument("--out", default="сро.xlsx", help="куда сохранить")
    p.add_argument("--scan-pages", type=int, default=50,
                   help="сколько страниц реестра просмотреть, если отдельного "
                        "списка СРО нет (по умолчанию 50)")
    p.add_argument("--delay", type=float, default=0.3, help="пауза между запросами, сек")
    p.set_defaults(func=cmd_list_sro)

    p = sub.add_parser("выгрузка", help="члены выбранных СРО, вступившие с даты")
    p.add_argument("--sro-id", nargs="+", default=None, help="ID нужных СРО")
    p.add_argument("--sro-number", nargs="+", default=None,
                   help="регистрационные номера СРО, например С-230 С-306")
    p.add_argument("--sro-name", nargs="+", default=None,
                   help="часть названия нужных СРО (если не знаете ID)")
    p.add_argument("--until", default=None,
                   help="дата вступления не позже (например 31.12.2026); "
                        "без флага верхней границы нет")
    p.add_argument("--since", default="01.01.2026",
                   help="дата вступления не раньше (по умолчанию 01.01.2026)")
    p.add_argument("--out", default="новые_члены_СРО.xlsx", help="куда сохранить")
    p.add_argument("--page-size", type=int, default=100, help="записей на страницу")
    p.add_argument("--delay", type=float, default=0.3, help="пауза между страницами, сек")
    p.add_argument("--max-pages", type=int, default=0, help="лимит страниц (0 — без лимита)")
    p.add_argument("--resume", action="store_true", help="продолжить прерванную выгрузку")
    p.add_argument("--full-scan", action="store_true",
                   help="не пытаться брать членов СРО напрямую, а перебрать весь "
                        "реестр (медленно; на случай, если прямой путь врёт)")
    p.set_defaults(func=cmd_export)

    p = sub.add_parser("контакты", help="дотянуть телефоны/почту из карточек НОСТРОЯ")
    p.add_argument("--file", required=True, help="файл, собранный командой «выгрузка»")
    p.add_argument("--delay", type=float, default=0.3, help="пауза между карточками, сек")
    p.set_defaults(func=cmd_contacts)

    p = sub.add_parser("образец", help="сырой ответ реестра для диагностики")
    p.add_argument("--member-id", default=None,
                   help="показать карточку компании по её «ID записи»")
    p.add_argument("--sro-id", default=None,
                   help="взять записи из конкретной СРО (например 247)")
    p.add_argument("--count", type=int, default=3,
                   help="сколько записей разобрать (по умолчанию 3)")
    p.set_defaults(func=cmd_sample)

    args = parser.parse_args(argv)
    if not getattr(args, "func", None):
        parser.print_help()
        return 1
    try:
        return args.func(args) or 0
    except KeyboardInterrupt:
        print("\nПрервано пользователем.")
        return 130


if __name__ == "__main__":
    sys.exit(main())
