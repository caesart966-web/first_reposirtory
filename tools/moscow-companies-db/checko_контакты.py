"""Дозапрос контактов по ИНН через Checko API — прямо в готовый Excel-файл.

Самостоятельный скрипт: не зависит от пакета mosstroybase. Нужны requests и
openpyxl (уже стоят в .venv).

Зачем отдельно: команда `enrich-checko` из основного проекта намеренно
пропускает членов строительных СРО (для продаж СРО они не лиды) и требует
строительный ОКВЭД. Здесь задача обратная — добрать контакты как раз для
свежих членов СРО, поэтому никаких таких фильтров нет.

ЗАПУСК (из папки C:\\mosstroybase):

    .venv\\Scripts\\activate
    set CHECKO_API_KEY=ваш_ключ
    python checko_контакты.py --file новые_члены.xlsx

Дописывает в тот же файл колонки: Телефоны (Checko), E-mail (Checko), Сайт,
Руководитель, Должность, Сотрудников (СЧР), Налоги уплачено, Банкротство,
Адрес (Checko), Статус (Checko). Пометка «(Checko)» — чтобы не путались
с колонками «Статус»/«Адрес», которые уже есть в выгрузке из НОСТРОЯ.

Про лимит бесплатного тарифа (~100 запросов в сутки):
  * уже заполненные строки пропускаются, поэтому можно запускать хоть
    каждый день — продолжит с того места, где остановился;
  * --limit 100 ограничит число запросов за один прогон;
  * если ключей два, второй прогон делайте с --key <второй ключ>.

Файл сохраняется каждые 10 компаний, так что обрыв связи или Ctrl+C
ничего не теряют.

Каждый прогон дополнительно кладёт рядом отдельный файл только со своей
порцией — `спб_для_checko_прогон_2026-09-01_1522.xlsx`. Общий файл при этом
продолжает пополняться: именно по нему скрипт понимает, кого уже пробил, и
не тратит лимит на повторы. Отключается флагом --без-порции.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Iterator

import requests

CHECKO_URL = "https://api.checko.ru/v2/company"
API_KEY_ENV = "CHECKO_API_KEY"

# Колонки, которые дописывает скрипт. Все с пометкой «(Checko)» там, где
# имя может совпасть с колонкой исходного файла: в выгрузке из НОСТРОЯ уже
# есть свои «Статус» и «Адрес», и без пометки скрипт принимал бы их за свои
# и считал строку заполненной.
CONTACT_COLUMNS = [
    "Телефоны (Checko)", "E-mail (Checko)", "Сайт", "Руководитель", "Должность",
    "Сотрудников (СЧР)", "Налоги уплачено, ₽", "Банкротство",
    "Адрес (Checko)", "Статус (Checko)",
]
# По этим колонкам определяется, обработана ли строка — только свои!
_DONE_MARKERS = ("Телефоны (Checko)", "E-mail (Checko)", "Статус (Checko)")

# Осечки, после которых компанию надо пробовать заново. Без этого списка
# любой таймаут навсегда исключал бы компанию из работы: пометка о неудаче
# ложится в «Статус (Checko)», а по нему строка считается обработанной.
_ВРЕМЕННЫЕ_ОСЕЧКИ = ("сеть:", "ответ не разобрался", "лимит запросов исчерпан",
                     "HTTP 5", "HTTP 429")


def нужно_повторить(row: dict) -> bool:
    """Строка помечена неудачей, которая может пройти сама собой."""
    статус = (row.get("Статус (Checko)") or "").strip()
    if not статус:
        return False
    return статус.startswith(_ВРЕМЕННЫЕ_ОСЕЧКИ)

# Телефон РФ: код начинается с 3/4/8/9. Границы (?<!\d) и (?!\d) не дают
# выхватить куски длинных номеров документов — на этом уже обжигались
_PHONE_RE = re.compile(
    r"(?<!\d)(?:\+7|8|7)?[\s\-(]*([3489]\d{2})[\s\-)]*(\d{3})[\s\-]*(\d{2})[\s\-]*(\d{2})(?!\d)"
)
_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")


def walk(obj: Any, parent_key: str = "") -> Iterator[tuple[str, Any]]:
    """Обход JSON. Элементы списков отдаются с ключом самого списка —
    без этого терялись телефоны, приходящие как {"Тел": ["+7...", ...]}.
    """
    if isinstance(obj, dict):
        for key, value in obj.items():
            yield key, value
            yield from walk(value, key)
    elif isinstance(obj, list):
        for item in obj:
            yield parent_key, item
            yield from walk(item, parent_key)


def deep_find(data: Any, keys: tuple[str, ...]) -> Iterator[Any]:
    wanted = tuple(k.lower() for k in keys)
    for key, value in walk(data):
        if isinstance(key, str) and key.lower() in wanted:
            yield value


def _strings(value: Any) -> Iterator[str]:
    if isinstance(value, str):
        yield value
    elif isinstance(value, (dict, list)):
        for _key, leaf in walk(value):
            if isinstance(leaf, str):
                yield leaf


def extract(data: dict) -> dict:
    """Контакты и признаки живости из ответа Checko."""
    info: dict = {}

    phones: list[str] = []
    for value in deep_find(data, ("Тел", "Телефон")):
        for text in _strings(value):
            for m in _PHONE_RE.finditer(text):
                code, a, b, c = m.groups()
                phone = f"+7{code}{a}{b}{c}"
                if phone not in phones:
                    phones.append(phone)
    info["Телефоны (Checko)"] = "; ".join(phones)

    emails: list[str] = []
    for value in deep_find(data, ("Емэйл", "Емейл", "Email", "E-mail", "ЭлПочта", "Почта")):
        for text in _strings(value):
            for m in _EMAIL_RE.finditer(text):
                email = m.group(0).lower()
                if email not in emails:
                    emails.append(email)
    info["E-mail (Checko)"] = "; ".join(emails)

    site = ""
    for value in deep_find(data, ("Сайт", "ВебСайт")):
        for text in _strings(value):
            text = text.strip()
            if text and "." in text:
                site = text
                break
        if site:
            break
    info["Сайт"] = site

    # Руководитель: берём ключ верхнего уровня, чтобы не зацепить
    # "СвязРуковод"/"МассРуковод"
    heads = data.get("Руковод")
    if isinstance(heads, dict):
        heads = [heads]
    fio = post = ""
    if isinstance(heads, list):
        for item in heads:
            if isinstance(item, dict) and isinstance(item.get("ФИО"), str):
                fio = item["ФИО"].strip()
                if isinstance(item.get("НаимДолжн"), str):
                    post = item["НаимДолжн"].strip().capitalize()
                break
    info["Руководитель"] = fio
    info["Должность"] = post

    schr = data.get("СЧР")
    if isinstance(schr, (int, float)):
        info["Сотрудников (СЧР)"] = int(schr)
    elif isinstance(schr, str) and schr.strip().isdigit():
        info["Сотрудников (СЧР)"] = int(schr.strip())
    else:
        info["Сотрудников (СЧР)"] = ""

    taxes = data.get("Налоги")
    paid = ""
    if isinstance(taxes, dict):
        try:
            paid = float(str(taxes.get("СумУпл")).replace(",", "."))
        except (TypeError, ValueError):
            paid = ""
    info["Налоги уплачено, ₽"] = paid

    efrsb = data.get("ЕФРСБ")
    info["Банкротство"] = "банкротство" if efrsb else ""

    address = ""
    for value in deep_find(data, ("АдресРФ",)):
        for text in _strings(value):
            if text.strip():
                address = text.strip()
                break
        if address:
            break
    info["Адрес (Checko)"] = address

    status = ""
    for value in deep_find(data, ("Статус",)):
        for text in _strings(value):
            low = text.lower()
            if "действ" in low or "ликвид" in low or "прекра" in low or "исключ" in low:
                status = text.strip()
                break
        if status:
            break
    info["Статус (Checko)"] = status

    return info


def fetch(inn: str, api_key: str, session: requests.Session,
          timeout: int = 60) -> tuple[dict | None, str]:
    """(данные, пометка). Пометка непустая — если запрос не удался."""
    try:
        resp = session.get(CHECKO_URL, params={"key": api_key, "inn": inn},
                           timeout=timeout)
    except requests.RequestException as exc:
        return None, f"сеть: {type(exc).__name__}"
    if resp.status_code == 404:
        return None, "не найдено в Checko"
    if resp.status_code == 429:
        return None, "лимит запросов исчерпан"
    if resp.status_code != 200:
        return None, f"HTTP {resp.status_code}"
    try:
        payload = resp.json()
    except ValueError:
        return None, "ответ не разобрался"
    data = payload.get("data") if isinstance(payload, dict) else None
    if not isinstance(data, dict) or not data:
        # Checko кладёт причину отказа в meta/message
        note = ""
        if isinstance(payload, dict):
            for key in ("message", "meta", "error"):
                value = payload.get(key)
                if isinstance(value, str) and value.strip():
                    note = value.strip()
                    break
                if isinstance(value, dict):
                    inner = value.get("message")
                    if isinstance(inner, str) and inner.strip():
                        note = inner.strip()
                        break
        return None, note or "пустой ответ"
    return data, ""


def read_rows(path: Path) -> tuple[list[str], list[dict]]:
    if path.suffix.lower() == ".csv":
        import csv
        with open(path, encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh, delimiter=";")
            return list(reader.fieldnames or []), list(reader)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    rows = []
    for values in rows_iter:
        row = {header[i]: ("" if v is None else str(v))
               for i, v in enumerate(values) if i < len(header)}
        if any(row.values()):
            rows.append(row)
    return header, rows


def write_rows(path: Path, columns: list[str], rows: list[dict]) -> None:
    if path.suffix.lower() == ".csv":
        import csv
        with open(path, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=columns, delimiter=";",
                                    extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        return
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Компании"
    ws.append(columns)
    for row in rows:
        ws.append([row.get(c, "") for c in columns])
    ws.freeze_panes = "A2"
    wb.save(path)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Дозапрос контактов по ИНН через Checko прямо в Excel-файл",
    )
    parser.add_argument("--file", required=True,
                        help="файл с колонкой ИНН (.xlsx или .csv)")
    parser.add_argument("--key", default=None,
                        help=f"API-ключ Checko (или переменная {API_KEY_ENV})")
    parser.add_argument("--limit", type=int, default=0,
                        help="максимум запросов за прогон (0 — без ограничения; "
                             "бесплатный тариф ~100 в сутки)")
    parser.add_argument("--delay", type=float, default=1.0,
                        help="пауза между запросами, сек")
    parser.add_argument("--inn-column", default="ИНН",
                        help="как называется колонка с ИНН (по умолчанию «ИНН»)")
    parser.add_argument("--порция", default=None,
                        help="куда сложить результат именно этого прогона; "
                             "по умолчанию рядом с общим файлом, с датой в имени")
    parser.add_argument("--без-порции", action="store_true",
                        help="не создавать отдельный файл на прогон")
    parser.add_argument("--redo", action="store_true",
                        help="перезапросить и те строки, что уже заполнены")
    args = parser.parse_args(argv)

    api_key = args.key or os.environ.get(API_KEY_ENV)
    if not api_key:
        print(f"Нужен ключ Checko: --key <ключ> или set {API_KEY_ENV}=<ключ>",
              file=sys.stderr)
        return 1

    path = Path(args.file)
    if not path.exists():
        print(f"Файл {path} не найден.", file=sys.stderr)
        return 1

    header, rows = read_rows(path)
    if not rows:
        print("В файле нет строк.", file=sys.stderr)
        return 1
    if args.inn_column not in header:
        print(f"В файле нет колонки «{args.inn_column}». Есть: {', '.join(header)}",
              file=sys.stderr)
        return 1

    columns = header + [c for c in CONTACT_COLUMNS if c not in header]

    # Что осталось сделать: строки без контактов (или все, если --redo)
    todo = []
    for row in rows:
        done = any((row.get(c) or "").strip() for c in _DONE_MARKERS)
        if args.redo or not done or нужно_повторить(row):
            inn = (row.get(args.inn_column) or "").strip()
            if inn:
                todo.append(row)
    повторно = sum(1 for row in todo if нужно_повторить(row))
    already = len(rows) - len(todo)
    if already:
        print(f"[checko] уже заполнено ранее: {already} (пропускаю)")
    if повторно:
        print(f"[checko] повторю после прошлых осечек: {повторно}")
    if not todo:
        print("[checko] все строки уже заполнены — работа не нужна")
        return 0

    plan = min(args.limit, len(todo)) if args.limit else len(todo)
    print(f"[checko] к запросу: {plan} из {len(todo)} оставшихся "
          f"(бесплатный тариф ~100/сутки — запускайте ежедневно, "
          "обработанные повторно не запрашиваются)")

    session = requests.Session()
    session.headers.update({"User-Agent": "mosstroybase/0.1 (+open-data harvester)"})

    done = with_phone = failed = 0
    stop_reason = ""
    за_прогон: list[dict] = []
    try:
        for row in todo:
            if args.limit and done >= args.limit:
                stop_reason = f"достигнут лимит прогона ({args.limit})"
                break
            inn = (row.get(args.inn_column) or "").strip()
            name = (row.get("Наименование") or "")[:40]
            data, note = fetch(inn, api_key, session)
            if data is None:
                failed += 1
                row.setdefault("Телефоны (Checko)", "")
                row["Статус (Checko)"] = note
                print(f"[checko] {done + 1}/{plan} {inn} {name}: {note}", flush=True)
                # Исчерпанный лимит/битый ключ — дальше идти бессмысленно
                if note in ("лимит запросов исчерпан",) or note.startswith("HTTP 40"):
                    stop_reason = note
                    break
                if failed >= 10 and failed > done:
                    stop_reason = "слишком много ошибок подряд"
                    break
            else:
                info = extract(data)
                row.update(info)
                if not info.get("Статус (Checko)"):
                    row["Статус (Checko)"] = ""   # затираем прошлую осечку
                if info["Телефоны (Checko)"]:
                    with_phone += 1
                print(f"[checko] {done + 1}/{plan} {inn} {name}: "
                      f"телефонов {len(info['Телефоны (Checko)'].split('; ')) if info['Телефоны (Checko)'] else 0}, "
                      f"e-mail {len(info['E-mail (Checko)'].split('; ')) if info['E-mail (Checko)'] else 0}",
                      flush=True)
            done += 1
            за_прогон.append(row)
            if done % 10 == 0:
                write_rows(path, columns, rows)
            time.sleep(args.delay)
    except KeyboardInterrupt:
        stop_reason = "прервано пользователем"
        print("\n[checko] прерываю, сохраняю собранное")

    write_rows(path, columns, rows)
    left = len(todo) - done
    print(f"\n[checko] готово: обработано {done}, с телефоном {with_phone}, "
          f"неудач {failed}")
    if stop_reason:
        print(f"[checko] остановка: {stop_reason}")
    if left > 0:
        print(f"[checko] осталось на следующий раз: {left} — запустите ту же "
              "команду завтра (или сегодня со вторым ключом: --key <второй>)")
    print(f"[checko] общий файл сохранён: {path}")

    if за_прогон and not args.без_порции:
        порция = Path(args.порция) if args.порция else path.with_name(
            f"{path.stem}_прогон_{datetime.now():%Y-%m-%d_%H%M}{path.suffix}")
        write_rows(порция, columns, за_прогон)
        с_телефоном = sum(1 for r in за_прогон
                          if (r.get("Телефоны (Checko)") or "").strip())
        print(f"[checko] порция этого прогона: {порция} "
              f"({len(за_прогон)} компаний, с телефоном {с_телефоном})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
