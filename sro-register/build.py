# -*- coding: utf-8 -*-
import json, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rows = json.load(open('rows.json'))
rows.sort(key=lambda x: x['n'])

def d(s):
    return datetime.datetime.strptime(s, '%d.%m.%Y').date() if s else None

FONT = 'Arial'
HDR_FILL   = PatternFill('solid', fgColor='1F3864')
TITLE_FONT = Font(name=FONT, size=14, bold=True, color='1F3864')
SUB_FONT   = Font(name=FONT, size=9, color='595959')
HDR_FONT   = Font(name=FONT, size=10, bold=True, color='FFFFFF')
BASE_FONT  = Font(name=FONT, size=10)
ZEBRA      = PatternFill('solid', fgColor='F2F5FA')
THIN       = Side(style='thin', color='BFBFBF')
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GREEN      = Font(name=FONT, size=10, bold=True, color='1E7B34')
RED        = Font(name=FONT, size=10, bold=True, color='B00020')

COLS = [
    ('№',                      7,  'center'),
    ('Полное наименование',    58, 'left'),
    ('Сокращённое наименование',30, 'left'),
    ('ИНН',                    14, 'center'),
    ('Контактные телефоны',    38, 'left'),
    ('Дата вступления',        16, 'center'),
    ('Дата прекращения',       17, 'center'),
    ('Статус',                 14, 'center'),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Реестр'

TITLE = 'Ассоциация «Петровское объединение строителей» (Ассоциация «ПОС»), СРО-С-303-22122020 — реестр членов'
ws['A1'] = TITLE
ws['A1'].font = TITLE_FONT
ws['A1'].alignment = Alignment(vertical='center')
ws.merge_cells('A1:H1')
ws.row_dimensions[1].height = 26

ws['A2'] = ('Источник: выгрузка реестра с сайта СРО (register-1-500, 501-1000, 1001-1500, 1501-1590). '
            'Телефоны приведены к единому формату +7 (код) номер; номера, не разбирающиеся однозначно, '
            'оставлены как в источнике. Пустая «Дата прекращения» = действующий член.')
ws['A2'].font = SUB_FONT
ws['A2'].alignment = Alignment(vertical='center', wrap_text=True)
ws.merge_cells('A2:H2')
ws.row_dimensions[2].height = 28

HDR = 3
for i, (name, width, _) in enumerate(COLS, start=1):
    c = ws.cell(row=HDR, column=i, value=name)
    c.font, c.fill = HDR_FONT, HDR_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[HDR].height = 30

for r, x in enumerate(rows, start=HDR + 1):
    active = not x['date_out']
    vals = [x['n'], x['full'], x['short'], x['inn'], x['phone_n'] or None,
            d(x['date_in']), d(x['date_out']), 'Действующий' if active else 'Прекращено']
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = BASE_FONT
        c.border = BORDER
        c.alignment = Alignment(horizontal=COLS[i-1][2], vertical='top',
                                wrap_text=i in (2, 3, 5))
        if r % 2 == 0:
            c.fill = ZEBRA
    ws.cell(row=r, column=4).number_format = '@'          # ИНН — текст, ведущий ноль не теряется
    ws.cell(row=r, column=6).number_format = 'DD.MM.YYYY'
    ws.cell(row=r, column=7).number_format = 'DD.MM.YYYY'
    ws.cell(row=r, column=8).font = GREEN if active else RED

last = HDR + len(rows)
ws.auto_filter.ref = 'A%d:H%d' % (HDR, last)
ws.freeze_panes = 'A%d' % (HDR + 1)
ws.sheet_view.showGridLines = False
ws.print_title_rows = '%d:%d' % (HDR, HDR)

# --- Сводка -----------------------------------------------------------------
s = wb.create_sheet('Сводка')
s.column_dimensions['A'].width = 46
s.column_dimensions['B'].width = 18
s['A1'] = 'Сводка по реестру'
s['A1'].font = TITLE_FONT
EPOCH = datetime.date(1899, 12, 30)                        # серийная дата Excel
serial = lambda dt: (dt - EPOCH).days
dates_in  = [d(x['date_in'])  for x in rows]
dates_out = [d(x['date_out']) for x in rows if x['date_out']]
n_act = sum(1 for x in rows if not x['date_out'])

# Формулы в файле хранятся с запятой-разделителем аргументов (Excel локализует её при показе).
# Кэш-значения считаем здесь же и вписываем в XML: без этого лист выглядит пустым,
# пока файл не пересчитает Excel.
items = [
    ('Всего записей в реестре',          '=COUNTA(Реестр!A4:A%d)' % last,                     len(rows)),
    ('Действующих членов',               '=COUNTIF(Реестр!H4:H%d,"Действующий")' % last,      n_act),
    ('Прекративших членство',            '=COUNTIF(Реестр!H4:H%d,"Прекращено")' % last,       len(rows) - n_act),
    ('Записей без контактного телефона', '=COUNTBLANK(Реестр!E4:E%d)' % last,                 sum(1 for x in rows if not x['phone_n'])),
    ('Доля действующих',                 '=IF(B3=0,0,B4/B3)',                                 n_act / len(rows)),
    ('Самая ранняя дата вступления',     '=MIN(Реестр!F4:F%d)' % last,                        serial(min(dates_in))),
    ('Самая поздняя дата вступления',    '=MAX(Реестр!F4:F%d)' % last,                        serial(max(dates_in))),
    ('Первое прекращение членства',      '=MIN(Реестр!G4:G%d)' % last,                        serial(min(dates_out))),
    ('Последнее прекращение членства',   '=MAX(Реестр!G4:G%d)' % last,                        serial(max(dates_out))),
]
CACHE = {}
for i, (label, formula, value) in enumerate(items, start=3):
    s.cell(row=i, column=1, value=label).font = BASE_FONT
    c = s.cell(row=i, column=2, value=formula)
    c.font = Font(name=FONT, size=10, bold=True)
    c.alignment = Alignment(horizontal='center')
    if 'дат' in label.lower() or 'прекращение' in label.lower(): c.number_format = 'DD.MM.YYYY'
    if label == 'Доля действующих': c.number_format = '0.0%'
    CACHE['B%d' % i] = value
s['A13'] = 'Повторные записи по одному ИНН — это компании, вступавшие в СРО дважды: каждая запись со своим номером и датами.'
s['A13'].font = SUB_FONT
s.sheet_view.showGridLines = False

OUT = 'Реестр_членов_СРО_ПОС.xlsx'
wb.save(OUT)

# --- кэш-значения формул прямо в XML ----------------------------------------
# openpyxl пишет формулы без результата, а пересчитать файл здесь нечем
# (LibreOffice в этом окружении не запускается). Поэтому дописываем <v> сами.
import re, shutil, zipfile

def inject(path, sheet_xml, cache):
    zin = zipfile.ZipFile(path)
    names = zin.namelist()
    data = {n: zin.read(n) for n in names}
    zin.close()
    xml = data[sheet_xml].decode('utf-8')
    for ref, value in cache.items():
        v = repr(round(value, 10)) if isinstance(value, float) else str(value)
        pat = re.compile(r'(<c r="%s"[^>]*>)(<f>.*?</f>)(?:<v\s*/>|<v>.*?</v>)?(</c>)' % ref)
        xml, n = pat.subn(lambda m: m.group(1) + m.group(2) + '<v>' + v + '</v>' + m.group(3), xml)
        if n != 1:
            raise SystemExit('не найдена формульная ячейка %s (совпадений: %d)' % (ref, n))
    data[sheet_xml] = xml.encode('utf-8')
    tmp = path + '.tmp'
    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(n, data[n])
    shutil.move(tmp, path)

inject(OUT, 'xl/worksheets/sheet2.xml', CACHE)
print('строк данных:', len(rows), '| последняя строка:', last, '| кэш формул:', len(CACHE))
