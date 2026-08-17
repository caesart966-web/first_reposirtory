import openpyxl, glob, json, re

FILES = ['register-1-500.xlsx','register-501-1000.xlsx','register-1001-1500.xlsx','register-1501-1590.xlsx']
C = {'num':1,'full':2,'short':3,'reg':4,'date_in':5,'date_gr':6,'date_out':7,'reason':8,'ogrn':9,'inn':10,'phone':11}

def clean(v):
    if v is None: return ''
    s = str(v).replace('\xa0',' ').strip()
    s = re.sub(r'\s+',' ', s)
    return s

rows = []
for f in FILES:
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for r in ws.iter_rows(min_row=4, values_only=True):
        n = r[C['num']]
        if isinstance(n,(int,float)) and clean(r[C['full']]):
            rows.append({
                'n': int(n),
                'full': clean(r[C['full']]),
                'short': clean(r[C['short']]),
                'reg': clean(r[C['reg']]),
                'date_in': clean(r[C['date_in']]),
                'date_out': clean(r[C['date_out']]),
                'reason': clean(r[C['reason']]),
                'ogrn': clean(r[C['ogrn']]),
                'inn': clean(r[C['inn']]),
                'phone': clean(r[C['phone']]),
            })
    wb.close()

print('total members:', len(rows))
print('unique n:', len(set(x['n'] for x in rows)))
inns = [x['inn'] for x in rows]
print('empty inn:', sum(1 for x in inns if not x))
print('inn lengths:', sorted(set(len(x) for x in inns)))
from collections import Counter
dup = [k for k,v in Counter(i for i in inns if i).items() if v>1]
print('duplicate INN count:', len(dup), dup[:10])
print('empty phone:', sum(1 for x in rows if not x['phone']))
print('empty date_in:', sum(1 for x in rows if not x['date_in']))
print('has date_out:', sum(1 for x in rows if x['date_out']))
print('date_in formats:', Counter(re.sub(r'\d','9',x['date_in']) for x in rows).most_common(6))
print('date_out formats:', Counter(re.sub(r'\d','9',x['date_out']) for x in rows if x['date_out']).most_common(6))
print('phone samples:')
for x in rows[:5]+rows[700:703]: print('   ', x['n'], '|', x['phone'], '|', x['inn'], '|', x['date_in'], '|', x['date_out'])
json.dump(rows, open('rows.json','w'), ensure_ascii=False)
