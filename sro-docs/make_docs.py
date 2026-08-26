#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Генератор документов дисциплинарной комиссии СРО по ежедневному эксель-файлу.

Запуск:  python make_docs.py [путь_к_файлу.xlsx]
Без аргумента берётся самый свежий .xlsx рядом со скриптом.

Что делает по каждой строке таблицы:
  * Решение «Отказ» или «Приостановить»  -> предписание о приостановке
    (вариант по причине: акт / страховка / акт+страховка / у изыскателей
    также взнос+акт+страховка), датой решения, сроком на 90 дней.
  * Результат проверки «Положительный»   -> положительный акт плановой
    проверки датой из колонки «Акт», номер = рег.номер/год.
  * Всё остальное (отрицательный результат текущей проверки, «Предписание
    на 14 дней», восстановления) пропускается и попадает в сводку.

Готовые документы раскладываются в папку «Документы <имя файла>» по СРО.
"""
import re
import sys
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from xml.sax.saxutils import escape

try:
    import openpyxl
except ImportError:
    print("Не установлена библиотека openpyxl. Выполните:  pip install openpyxl")
    sys.exit(1)

BASE = Path(__file__).resolve().parent
TEMPLATES = BASE / "templates"

MONTHS_GEN = ["января", "февраля", "марта", "апреля", "мая", "июня",
              "июля", "августа", "сентября", "октября", "ноября", "декабря"]

SUSPENSION_DAYS = 90

# Колонки таблицы: точное имя (после strip/lower) -> внутренний ключ
COLUMNS = {
    "сро": "sro",
    "рег номер": "reg",
    "контрагент": "company",
    "инн": "inn",
    "акт": "act_date",
    "результат проверки": "result",
    "решение": "decision",
    "дата решения": "decision_date",
    "дата отказа в возобновлении": "until",
}
COLUMNS_PREFIX = {
    "причина приостановки": "cause",
}

SRO_NAMES = {
    "izyskateli": "СФЕРА изыскатели",
    "proekt": "СФЕРА проект",
    "stroiteli": "СФЕРА-А",
}

OPF_FULL = [
    ("ООО ", "Общество с ограниченной ответственностью "),
    ("ИП ", "Индивидуальный предприниматель "),
    ("ЗАО ", "Закрытое акционерное общество "),
    ("ПАО ", "Публичное акционерное общество "),
    ("ОАО ", "Открытое акционерное общество "),
    ("НАО ", "Непубличное акционерное общество "),
    ("АО ", "Акционерное общество "),
]


def sro_key(value: str):
    v = (value or "").strip().lower()
    if not v:
        return None
    if "изыск" in v:
        return "izyskateli"
    if "проект" in v:
        return "proekt"
    if "сфера-а" in v or "цос" in v or "строит" in v:
        return "stroiteli"
    return None


def parse_date(value):
    """Дата из ячейки: datetime или строка ДД.ММ.ГГГГ."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip().strip(".")
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", s)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    return None


def fmt_dmy(d: date) -> str:
    return f"{d.day:02d}.{d.month:02d}.{d.year}"


def fmt_full(d: date) -> str:
    return f"{d.day:02d} {MONTHS_GEN[d.month - 1]} {d.year} года"


def cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def expand_opf(short_name: str) -> str:
    for abbr, full in OPF_FULL:
        if short_name.startswith(abbr):
            return full + short_name[len(abbr):]
    return short_name


def parse_causes(value: str):
    v = (value or "").lower()
    causes = set()
    if "акт" in v:
        causes.add("akt")
    if "страхов" in v:
        causes.add("strahovka")
    if "взнос" in v:
        causes.add("vznos")
    return causes


def pick_predpisanie_template(sro: str, causes: set):
    """Возвращает (имя шаблона, описание варианта) или (None, причина пропуска)."""
    c = set(causes)
    if sro == "izyskateli" and c == {"vznos", "akt", "strahovka"}:
        return "predpisanie_vznos_akt_strahovka.docx", "взнос + акт + страховка"
    c.discard("vznos")  # у стройки и проекта взнос в тексте не упоминается
    if c == {"akt", "strahovka"}:
        return "predpisanie_akt_strahovka.docx", "акт + страховка"
    if c == {"akt"}:
        return "predpisanie_akt.docx", "акт"
    if c == {"strahovka"}:
        return "predpisanie_strahovka.docx", "страховка"
    return None, "в причине не указаны ни акт, ни страховка"


def render(template: Path, out_path: Path, mapping: dict) -> None:
    with zipfile.ZipFile(template) as zin:
        items = [(item, zin.read(item.filename)) for item in zin.infolist()]
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item, data in items:
            if item.filename == "word/document.xml":
                xml = data.decode("utf-8")
                for key, value in mapping.items():
                    xml = xml.replace("{{" + key + "}}", escape(str(value)))
                rest = re.findall(r"\{\{\w+\}\}", xml)
                if rest:
                    raise RuntimeError(f"В шаблоне {template.name} остались поля: {rest}")
                data = xml.encode("utf-8")
            zout.writestr(item, data)


def safe_name(s: str) -> str:
    s = re.sub(r'[«»"\'\\/:*?<>|]', "", s)
    return re.sub(r"\s+", " ", s).strip()


def find_latest_xlsx() -> Path | None:
    files = [p for p in BASE.glob("*.xlsx") if not p.name.startswith("~$")]
    return max(files, key=lambda p: p.stat().st_mtime) if files else None


def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    if len(sys.argv) > 1:
        xlsx = Path(sys.argv[1])
    else:
        xlsx = find_latest_xlsx()
        if xlsx is None:
            print("Не найден .xlsx рядом со скриптом. Положите ежедневный файл в эту папку")
            print("или укажите путь: python make_docs.py путь\\к\\файлу.xlsx")
            sys.exit(1)
    if not xlsx.exists():
        print(f"Файл не найден: {xlsx}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.worksheets[0]

    # находим колонки по заголовкам первой строки
    col_idx = {}
    for j, cell in enumerate(ws[1]):
        header = cell_str(cell.value).lower()
        if not header:
            continue
        if header in COLUMNS:
            col_idx[COLUMNS[header]] = j
        else:
            for prefix, key in COLUMNS_PREFIX.items():
                if header.startswith(prefix):
                    col_idx[key] = j
    missing = [k for k in ("sro", "reg", "company", "inn", "result", "decision",
                           "decision_date") if k not in col_idx]
    if missing:
        print(f"В файле {xlsx.name} не найдены нужные колонки: {missing}")
        print("Проверьте, что первая строка листа — заголовки как в обычном ежедневном файле.")
        sys.exit(1)

    out_dir = BASE / f"Документы {xlsx.stem}"
    out_dir.mkdir(exist_ok=True)

    made, skipped, errors = [], [], []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v is not None and str(v).strip() for v in row):
            continue
        get = lambda key: row[col_idx[key]] if key in col_idx and col_idx[key] < len(row) else None
        company = cell_str(get("company"))
        if not company:
            continue
        sro = sro_key(cell_str(get("sro")))
        reg = cell_str(get("reg"))
        inn = cell_str(get("inn"))
        result = cell_str(get("result")).lower()
        decision = cell_str(get("decision")).lower()
        label = f"строка {i}: [{cell_str(get('sro')) or '?'}] {company}"

        if sro is None:
            errors.append(f"{label} — не распознана СРО «{cell_str(get('sro'))}»")
            continue
        sro_dir = out_dir / SRO_NAMES[sro]
        row_made = False

        # --- предписание о приостановке ---
        if decision in ("отказ", "приостановить"):
            causes_raw = cell_str(get("cause"))
            causes = parse_causes(causes_raw)
            tpl_name, variant = pick_predpisanie_template(sro, causes)
            if tpl_name is None:
                errors.append(f"{label} — решение «{decision}», но {variant} "
                              f"(в файле: «{causes_raw}»); оформите вручную")
            else:
                d = parse_date(get("decision_date"))
                if d is None:
                    errors.append(f"{label} — не заполнена «Дата решения», предписание не сделано")
                else:
                    until = parse_date(get("until")) or (d + timedelta(days=SUSPENSION_DAYS))
                    out = sro_dir / f"Предписание {safe_name(reg)} {safe_name(company)}.docx"
                    render(TEMPLATES / sro / tpl_name, out, {
                        "company": company,
                        "inn": inn,
                        "reg": reg,
                        "date_full": fmt_full(d),
                        "until": fmt_dmy(until),
                    })
                    made.append(f"{label} — ПРЕДПИСАНИЕ № ____/{reg} от {fmt_dmy(d)} "
                                f"до {fmt_dmy(until)} ({variant})")
                    row_made = True

        # --- положительный акт ---
        if result == "положительный":
            d = parse_date(get("act_date"))
            if d is None:
                errors.append(f"{label} — результат «Положительный», но не заполнена дата "
                              f"в колонке «Акт»; акт не сделан")
            else:
                out = sro_dir / f"Акт {safe_name(reg)}-{d.year} {safe_name(company)}.docx"
                render(TEMPLATES / sro / "akt.docx", out, {
                    "company": company,
                    "company_full": expand_opf(company),
                    "inn": inn,
                    "reg": reg,
                    "day": f"{d.day:02d}",
                    "month_gen": MONTHS_GEN[d.month - 1],
                    "year": str(d.year),
                })
                made.append(f"{label} — АКТ № {reg}/{d.year} от {fmt_dmy(d)} (положительный)")
                row_made = True

        if row_made:
            continue
        if decision.startswith("предписание"):
            skipped.append(f"{label} — решение «{cell_str(get('decision'))}»: предписание "
                           f"об устранении оформляется вручную")
        elif result == "отрицательный":
            skipped.append(f"{label} — отрицательный результат текущей проверки, "
                           f"без решения о приостановке (вручную)")
        elif not decision and not result:
            skipped.append(f"{label} — нет ни результата проверки, ни решения")
        else:
            skipped.append(f"{label} — ситуация не распознана (результат: "
                           f"«{cell_str(get('result'))}», решение: «{cell_str(get('decision'))}»)")

    # --- сводка ---
    lines = [f"Файл: {xlsx.name}", f"Дата запуска: {fmt_dmy(date.today())}", ""]
    lines.append(f"СДЕЛАНО ДОКУМЕНТОВ: {len(made)}")
    lines += [f"  + {s}" for s in made]
    lines.append("")
    lines.append(f"ПРОПУЩЕНО (ручная работа или не требуется): {len(skipped)}")
    lines += [f"  - {s}" for s in skipped]
    if errors:
        lines.append("")
        lines.append(f"ВНИМАНИЕ, ТРЕБУЕТ ПРОВЕРКИ: {len(errors)}")
        lines += [f"  ! {s}" for s in errors]
    lines.append("")
    lines.append(f"Документы сложены в папку: {out_dir}")
    lines.append("Не забудьте вписать первую часть номера в предписаниях (№ ____/…).")
    report = "\n".join(lines)
    (out_dir / "Сводка.txt").write_text(report, encoding="utf-8-sig")
    print(report)


if __name__ == "__main__":
    main()
